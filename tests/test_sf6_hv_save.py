import DBrun
import reports
from database import init_db
from openpyxl import Workbook, load_workbook

from reports import (
    _export_sf6_excel,
    _get_sf6_report_data,
    normalize_decimal_numeric_text,
)


def test_hv_sf6_leakage_persistence(tmp_path):
    db_path = tmp_path / "test.db"
    conn = init_db(str(db_path))
    cur = conn.cursor()

    # create minimal substation and element (HV breaker)
    cur.execute("INSERT INTO substations (id, name) VALUES (?, ?)", (1, "Test Sub"))
    conn.commit()

    cur.execute(
        "INSERT INTO elements (id, substation_id, element_type, name, breaker_category) VALUES (?, ?, ?, ?, ?)",
        (766, 1, "Διακόπτης ΥΤ", "Ρ-200", "SF6"),
    )
    conn.commit()

    # Insert maintenance record
    cur.execute(
        "INSERT INTO maintenance (id, substation_id, name, date_time) VALUES (?, ?, ?, ?)",
        (2912, 1, "Test Maint", "2026-01-05"),
    )
    conn.commit()

    # Insert maintenance_elements with SF6 leakage and methodology
    cur.execute(
        (
            "INSERT INTO maintenance_elements (maintenance_id, element_id, element_comments, sf6_leakage_kg, sf6_leak_methodology)"
            " VALUES (?, ?, ?, ?, ?)"
        ),
        (2912, 766, "Imported SF6", 0.5, "Πλήρωση"),
    )
    conn.commit()

    # Read back and assert
    cur.execute(
        "SELECT sf6_leakage_kg, sf6_leak_methodology FROM maintenance_elements WHERE maintenance_id=? AND element_id=?",
        (2912, 766),
    )
    row = cur.fetchone()
    assert row is not None
    assert row[0] == 0.5
    assert row[1] == "Πλήρωση"

    conn.close()


def test_normalize_decimal_numeric_text_accepts_comma_and_dot():
    assert normalize_decimal_numeric_text("1,15") == "1.15"
    assert normalize_decimal_numeric_text("1.15") == "1.15"
    assert normalize_decimal_numeric_text("2,50", decimal_separator=",") == "2,50"


def test_normalize_decimal_numeric_text_handles_mixed_separators_and_spaces():
    assert normalize_decimal_numeric_text("1.234,56") == "1234.56"
    assert normalize_decimal_numeric_text("1,234.56") == "1234.56"
    assert normalize_decimal_numeric_text(" 1 234,56 ") == "1234.56"
    assert normalize_decimal_numeric_text("-0,75") == "-0.75"


def test_sf6_report_data_groups_by_substation_and_keeps_ids(tmp_path):
    db_path = tmp_path / "test_sf6_report.db"
    conn = init_db(str(db_path))
    cur = conn.cursor()

    cur.execute("INSERT INTO substations (id, name) VALUES (?, ?)", (1, "ΥΣ Α"))
    cur.execute("INSERT INTO substations (id, name) VALUES (?, ?)", (2, "ΥΣ Β"))
    cur.execute(
        "INSERT INTO elements (id, substation_id, element_type, name, breaker_category, operating_status) VALUES (?, ?, ?, ?, ?, ?)",
        (10, 1, "Διακόπτης ΥΤ", "Q1", "SF6", "Ενεργή"),
    )
    cur.execute(
        "INSERT INTO elements (id, substation_id, element_type, name, breaker_category, operating_status) VALUES (?, ?, ?, ?, ?, ?)",
        (20, 2, "Διακόπτης ΜΤ", "Q2", "SF6", "Ενεργή"),
    )
    cur.execute(
        "INSERT INTO maintenance (id, substation_id, name, date_time) VALUES (?, ?, ?, ?)",
        (100, 1, "Leak 1", "2026-01-05"),
    )
    cur.execute(
        "INSERT INTO maintenance (id, substation_id, name, date_time) VALUES (?, ?, ?, ?)",
        (200, 2, "Leak 2", "2026-03-10"),
    )
    cur.execute(
        "INSERT INTO maintenance_elements (maintenance_id, element_id, sf6_leakage_kg, sf6_leak_methodology) VALUES (?, ?, ?, ?)",
        (100, 10, 0.25, "Πλήρωση"),
    )
    cur.execute(
        "INSERT INTO maintenance_elements (maintenance_id, element_id, sf6_leakage_kg, sf6_leak_methodology) VALUES (?, ?, ?, ?)",
        (200, 20, 1.10, "Αντικατάσταση"),
    )
    conn.commit()

    class DummyApp:
        def __init__(self, connection):
            self.conn = connection

        def _format_maintenance_date(self, value):
            return value

    data = _get_sf6_report_data(DummyApp(conn), "2026")

    assert set(data["available_substations"]) == {"ΥΣ Α", "ΥΣ Β"}
    assert len(data["substation_rows"]["ΥΣ Α"]) == 1
    assert len(data["substation_rows"]["ΥΣ Β"]) == 1
    assert data["substation_rows"]["ΥΣ Α"][0]["maintenance_id"] == 100
    assert data["substation_rows"]["ΥΣ Β"][0]["element_id"] == 20
    assert data["leakage_bands"]["max"] == 1.10

    conn.close()


def test_sf6_report_data_accumulates_leakage_per_element(tmp_path):
    db_path = tmp_path / "test_sf6_report_running_total.db"
    conn = init_db(str(db_path))
    cur = conn.cursor()

    cur.execute("INSERT INTO substations (id, name) VALUES (?, ?)", (1, "ΥΣ Α"))
    cur.execute(
        "INSERT INTO elements (id, substation_id, element_type, name, breaker_category, operating_status) VALUES (?, ?, ?, ?, ?, ?)",
        (10, 1, "Διακόπτης ΥΤ", "Q1", "SF6", "Ενεργή"),
    )
    cur.execute(
        "INSERT INTO maintenance (id, substation_id, name, date_time) VALUES (?, ?, ?, ?)",
        (100, 1, "Leak 1", "2026-01-05"),
    )
    cur.execute(
        "INSERT INTO maintenance (id, substation_id, name, date_time) VALUES (?, ?, ?, ?)",
        (101, 1, "Leak 2", "2026-02-10"),
    )
    cur.execute(
        "INSERT INTO maintenance_elements (maintenance_id, element_id, sf6_leakage_kg, sf6_leak_methodology) VALUES (?, ?, ?, ?)",
        (100, 10, 1.0, "Πλήρωση"),
    )
    cur.execute(
        "INSERT INTO maintenance_elements (maintenance_id, element_id, sf6_leakage_kg, sf6_leak_methodology) VALUES (?, ?, ?, ?)",
        (101, 10, 1.0, "Αντικατάσταση"),
    )
    conn.commit()

    class DummyApp:
        def __init__(self, connection):
            self.conn = connection

        def _format_maintenance_date(self, value):
            return value

    data = _get_sf6_report_data(DummyApp(conn), "2026")

    assert [row["element_total_leakage"] for row in data["rows"]] == [1.0, 2.0]
    assert data["leakage_bands"]["max"] == 2.0
    assert reports._classify_sf6_leakage(1.0, data["leakage_bands"]) == "green"
    assert reports._classify_sf6_leakage(2.0, data["leakage_bands"]) == "red"

    conn.close()


def test_sf6_report_excludes_methodology_only_rows(tmp_path):
    db_path = tmp_path / "test_sf6_methodology_only.db"
    conn = init_db(str(db_path))
    cur = conn.cursor()

    cur.execute("INSERT INTO substations (id, name) VALUES (?, ?)", (1, "ΥΣ Α"))
    cur.execute(
        "INSERT INTO elements (id, substation_id, element_type, name, breaker_category, operating_status) VALUES (?, ?, ?, ?, ?, ?)",
        (10, 1, "Διακόπτης ΜΤ", "Q1", "SF6", "Ενεργή"),
    )
    cur.execute(
        "INSERT INTO maintenance (id, substation_id, name, date_time) VALUES (?, ?, ?, ?)",
        (100, 1, "SF6 Method", "2026-04-20"),
    )
    cur.execute(
        "INSERT INTO maintenance_elements (maintenance_id, element_id, sf6_leak_methodology) VALUES (?, ?, ?)",
        (100, 10, "Πλήρωση"),
    )
    conn.commit()

    class DummyApp:
        def __init__(self, connection):
            self.conn = connection

        def _format_maintenance_date(self, value):
            return value

    data = _get_sf6_report_data(DummyApp(conn), "2026")

    assert data["rows"] == []

    conn.close()


def test_sf6_excel_export_uses_template_and_keeps_blank_leakage_rows(
    tmp_path, monkeypatch
):
    db_path = tmp_path / "test_sf6_export.db"
    conn = init_db(str(db_path))
    cur = conn.cursor()

    template_path = tmp_path / "ΚΣΜΘ ΥΣ SF6 2026.xlsx"
    template_substation_sheet = "ΥΣ ΔΟΚΙΜΗ"
    wb_template = Workbook()
    ws_summary = wb_template.active
    ws_summary.title = "Σύνοψη"
    ws_substation = wb_template.create_sheet(template_substation_sheet)
    for ws in (ws_summary, ws_substation):
        ws.cell(
            row=1, column=1, value="ΠΙΝΑΚΑΣ 4: ΠΗΓΗ ΕΚΠΟΜΠΩΝ ΑΠΌ ΕΞΟΠΛΙΣΜΟ ΧΡΗΣΗΣ SF6"
        )
        ws.cell(row=2, column=1, value="Α/Α")
        ws.cell(row=3, column=1, value=1)
        ws.cell(row=8, column=1, value="Σχόλια")
    wb_template.save(template_path)
    template_sheetnames = load_workbook(template_path, read_only=True).sheetnames
    monkeypatch.setattr(
        reports,
        "_resolve_sf6_template_path",
        lambda _year: str(template_path),
    )

    cur.execute(
        "INSERT INTO substations (id, name) VALUES (?, ?)",
        (1, template_substation_sheet),
    )
    cur.execute(
        "INSERT INTO elements (id, substation_id, element_type, name, breaker_category, operating_status) VALUES (?, ?, ?, ?, ?, ?)",
        (10, 1, "Διακόπτης ΜΤ", "Q1", "SF6", "Ενεργή"),
    )
    cur.execute(
        "INSERT INTO maintenance (id, substation_id, name, date_time) VALUES (?, ?, ?, ?)",
        (100, 1, "SF6 Method", "2026-04-20"),
    )
    cur.execute(
        "INSERT INTO maintenance_elements (maintenance_id, element_id, sf6_leak_methodology) VALUES (?, ?, ?)",
        (100, 10, "Πλήρωση"),
    )
    conn.commit()

    class DummyApp:
        def __init__(self, connection):
            self.conn = connection

        def _format_maintenance_date(self, value):
            return value

    output_path = _export_sf6_excel(DummyApp(conn), "2026")
    wb = load_workbook(output_path)

    assert wb.sheetnames == template_sheetnames

    ws = wb[template_substation_sheet]
    data_row = None
    for row_idx in range(3, 10):
        if ws.cell(row=row_idx, column=1).value == 1:
            data_row = row_idx
            break

    assert data_row == 3
    assert ws.cell(row=data_row, column=5).value == "Πλήρωση"
    assert ws.cell(row=data_row, column=7).value in (None, "")

    conn.close()


def test_normalize_sf6_leakage_fields_clears_methodology_without_value():
    app = DBrun.SubstationApp()

    leakage, methodology = app._normalize_sf6_leakage_fields(None, "Πλήρωση")

    assert leakage is None
    assert methodology is None


def test_normalize_sf6_leakage_fields_keeps_methodology_with_value():
    app = DBrun.SubstationApp()

    leakage, methodology = app._normalize_sf6_leakage_fields(0.5, " Πλήρωση ")

    assert leakage == 0.5
    assert methodology == "Πλήρωση"


def test_preserve_existing_measurement_payload_keeps_sf6_values_on_blank_resave():
    app = DBrun.SubstationApp()

    existing_data = {
        "sf6_leakage_kg": 0.9,
        "sf6_leak_methodology": "Πλήρωση",
        "sf6": {
            "sf6_n2_fa": None,
            "h2o_fa": None,
            "so2_fa": None,
            "sf6_n2_fb": None,
            "h2o_fb": None,
            "so2_fb": None,
            "sf6_n2_fc": None,
            "h2o_fc": None,
            "so2_fc": None,
        },
        "vidar": {},
        "extra_measurements": {},
    }
    current_payload = {
        "ins_closed_fa": None,
        "ins_closed_fa_unit": "GΩ",
        "ins_closed_fb": None,
        "ins_closed_fb_unit": "GΩ",
        "ins_closed_fc": None,
        "ins_closed_fc_unit": "GΩ",
        "ins_open_fa": None,
        "ins_open_fa_unit": "GΩ",
        "ins_open_fb": None,
        "ins_open_fb_unit": "GΩ",
        "ins_open_fc": None,
        "ins_open_fc_unit": "GΩ",
        "cont_fa": None,
        "cont_fb": None,
        "cont_fc": None,
        "ops_count": None,
        "sf6_leakage_kg": None,
        "sf6_leak_methodology": None,
        "sf6": {},
        "vidar": {},
        "extra_measurements": {},
    }

    merged = app._preserve_existing_measurement_payload(
        existing_data,
        current_payload,
        measurements_enabled=True,
    )

    assert merged["sf6_leakage_kg"] == 0.9
    assert merged["sf6_leak_methodology"] == "Πλήρωση"


def test_preserve_existing_measurement_payload_keeps_new_values_when_present():
    app = DBrun.SubstationApp()

    existing_data = {
        "sf6_leakage_kg": 0.9,
        "sf6_leak_methodology": "Πλήρωση",
    }
    current_payload = {
        "ins_closed_fa": None,
        "ins_closed_fa_unit": "GΩ",
        "ins_closed_fb": None,
        "ins_closed_fb_unit": "GΩ",
        "ins_closed_fc": None,
        "ins_closed_fc_unit": "GΩ",
        "ins_open_fa": None,
        "ins_open_fa_unit": "GΩ",
        "ins_open_fb": None,
        "ins_open_fb_unit": "GΩ",
        "ins_open_fc": None,
        "ins_open_fc_unit": "GΩ",
        "cont_fa": None,
        "cont_fb": None,
        "cont_fc": None,
        "ops_count": None,
        "sf6_leakage_kg": 0.4,
        "sf6_leak_methodology": "Αντικατάσταση",
        "sf6": {},
        "vidar": {},
        "extra_measurements": {},
    }

    merged = app._preserve_existing_measurement_payload(
        existing_data,
        current_payload,
        measurements_enabled=True,
    )

    assert merged["sf6_leakage_kg"] == 0.4
    assert merged["sf6_leak_methodology"] == "Αντικατάσταση"
