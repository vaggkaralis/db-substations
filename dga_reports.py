import os
import re
import shutil
import tempfile
import math
from functools import lru_cache
from datetime import datetime

from openpyxl import load_workbook
from strings_proxy import STRINGS as S


_DGA_FIELD_ROWS = [
    ("h2", 19, "gases", "H2"),
    ("c2h2", 20, "gases", "C2H2"),
    ("c2h4", 21, "gases", "C2H4"),
    ("c2h6", 22, "gases", "C2H6"),
    ("co", 23, "gases", "CO"),
    ("co2", 24, "gases", "CO2"),
    ("ch4", 25, "gases", "CH4"),
    ("o2", 26, "gases", "O2"),
    ("c3h8", 27, "gases", "C3H8"),
    ("n2", 28, "gases", "N2"),
    ("h2o", 29, "gases", "H2O"),
    ("density", 45, "physchem", "ΠΥΚΝΟΤΗΤΑ"),
    ("humidity", 46, "physchem", "ΥΓΡΑΣΙΑ"),
    ("dielectric_strength", 47, "physchem", "ΔΙΗΛΕΚΤΡΙΚΗ ΑΝΤΟΧΗ"),
    ("loss_factor", 48, "physchem", "ΣΥΝΤΕΛΕΣΤΗΣ ΑΠΩΛΕΙΩΝ"),
    ("surface_tension", 49, "physchem", "ΔΙΕΠΙΦΑΝΕΙΑΚΗ ΤΑΣΗ"),
]

_GAS_LABEL_OVERRIDES = {
    "h2": "ΥΔΡΟΓΟΝΟ (H2)",
    "c2h2": "ΑΚΕΤΥΛΕΝΙΟ (C2H2)",
    "c2h4": "ΑΙΘΥΛΕΝΙΟ (C2H4)",
    "c2h6": "ΑΙΘΑΝΙΟ (C2H6)",
    "co": "ΜΟΝ.ΑΝΘΡΑΚΑ (CO)",
    "co2": "ΔΙΟΞ. ΑΝΘΡΑΚΑ (CO2)",
    "ch4": "ΜΕΘΑΝΙΟ (CH4)",
    "o2": "ΟΞΥΓΟΝΟ (O2)",
    "c3h8": "ΠΡΟΠΑΝΙΟ (C3H8)",
    "n2": "ΑΖΩΤΟ (N2)",
    "h2o": "ΥΓΡΑΣΙΑ (H2O)",
}

_GAS_LIMIT_OVERRIDES = {
    "h2": "50-150",
    "c2h2": "2-20",
    "c2h4": "60-280",
    "c2h6": "20-90",
    "co": "400-600",
    "co2": "3800-14000",
    "ch4": "30-130",
    "o2": "",
    "c3h8": "",
    "n2": "",
    "h2o": "",
}

# All gas quantities are expressed in ppm.
_GAS_UNIT = "ppm"

_PHYSCHEM_LABEL_OVERRIDES = {
    "density": "ΠΥΚΝΟΤΗΤΑ (DIN 51517)",
    "humidity": "ΥΓΡΑΣΙΑ (IEC 60814)",
    "dielectric_strength": "ΔΙΗΛΕΚΤΡΙΚΗ ΑΝΤΟΧΗ (IEC 156/95)",
    "loss_factor": "ΣΥΝΤΕΛΕΣΤΗΣ ΑΠΩΛΕΙΩΝ (IEC 60247)",
    "surface_tension": "ΔΙΕΠΙΦΑΝΕΙΑΚΗ ΤΑΣΗ (ASTM 971)",
}

_PHYSCHEM_UNIT_OVERRIDES = {
    "density": "g/cm\u00b3",
    "humidity": "ppm",
    "dielectric_strength": "KV",
    "loss_factor": "",
    "surface_tension": "mN/m",
}

_DGA_SEVERITY_ORDER = {"ok": 0, "warn": 1, "bad": 2}
_RATIO_METHOD_MIN_TOTAL_PPM = 100.0
_DUVAL_TRIANGLE_MIN_TOTAL_PPM = 100.0

_RATIO_RULES = [
    {
        "code": "PD",
        "label": "Partial discharge",
        "severity": "warn",
        "summary": "Pattern matches partial discharge / corona.",
        "root_cause": "Electrical partial discharges inside gas voids or weak insulation clearances.",
        "predicate": lambda ratios: (
            ratios.get("ch4_h2") is not None
            and ratios.get("c2h2_c2h4") is not None
            and ratios.get("c2h4_c2h6") is not None
            and ratios["ch4_h2"] < 0.1
            and ratios["c2h2_c2h4"] < 0.1
            and ratios["c2h4_c2h6"] < 1.0
        ),
    },
    {
        "code": "T1",
        "label": "Thermal fault T1",
        "severity": "warn",
        "summary": "Pattern matches low-temperature thermal fault.",
        "root_cause": "Localized overheating typically below 300 C, often linked to hot spots, poor cooling, or mild oil overheating.",
        "predicate": lambda ratios: (
            ratios.get("ch4_h2") is not None
            and ratios.get("c2h2_c2h4") is not None
            and ratios.get("c2h4_c2h6") is not None
            and ratios["ch4_h2"] > 1.0
            and ratios["c2h2_c2h4"] < 0.1
            and ratios["c2h4_c2h6"] < 1.0
        ),
    },
    {
        "code": "T2",
        "label": "Thermal fault T2",
        "severity": "bad",
        "summary": "Pattern matches medium-temperature thermal fault.",
        "root_cause": "Sustained overheating roughly in the 300-700 C range, often involving oil decomposition and accelerated insulation aging.",
        "predicate": lambda ratios: (
            ratios.get("ch4_h2") is not None
            and ratios.get("c2h2_c2h4") is not None
            and ratios.get("c2h4_c2h6") is not None
            and ratios["ch4_h2"] > 1.0
            and ratios["c2h2_c2h4"] < 0.1
            and 1.0 <= ratios["c2h4_c2h6"] <= 3.0
        ),
    },
    {
        "code": "T3",
        "label": "Thermal fault T3",
        "severity": "bad",
        "summary": "Pattern matches high-temperature thermal fault.",
        "root_cause": "Severe overheating above about 700 C, commonly linked to intense oil cracking, metal hot spots, or major cooling failure.",
        "predicate": lambda ratios: (
            ratios.get("ch4_h2") is not None
            and ratios.get("c2h2_c2h4") is not None
            and ratios.get("c2h4_c2h6") is not None
            and ratios["ch4_h2"] > 1.0
            and ratios["c2h2_c2h4"] < 0.1
            and ratios["c2h4_c2h6"] > 3.0
        ),
    },
    {
        "code": "D1",
        "label": "Discharge fault D1",
        "severity": "bad",
        "summary": "Pattern matches low-energy discharge.",
        "root_cause": "Sparking or intermittent electrical discharge, often at bad contacts, floating potentials, or deteriorated connections.",
        "predicate": lambda ratios: (
            ratios.get("ch4_h2") is not None
            and ratios.get("c2h2_c2h4") is not None
            and ratios.get("c2h4_c2h6") is not None
            and 0.1 <= ratios["ch4_h2"] <= 1.0
            and 0.1 <= ratios["c2h2_c2h4"] < 3.0
            and ratios["c2h4_c2h6"] > 1.0
        ),
    },
    {
        "code": "D2",
        "label": "Discharge fault D2",
        "severity": "bad",
        "summary": "Pattern matches high-energy discharge / arcing.",
        "root_cause": "Severe arcing or high-energy electrical breakdown with significant acetylene production and acute failure risk.",
        "predicate": lambda ratios: (
            ratios.get("ch4_h2") is not None
            and ratios.get("c2h2_c2h4") is not None
            and ratios.get("c2h4_c2h6") is not None
            and 0.1 <= ratios["ch4_h2"] <= 1.0
            and ratios["c2h2_c2h4"] >= 3.0
            and ratios["c2h4_c2h6"] > 1.0
        ),
    },
]

_DUVAL_TRIANGLE_1_ZONES = {
    "PD": [(0.98, 0.02, 0.00), (0.98, 0.00, 0.02), (1.00, 0.00, 0.00)],
    "T1": [
        (0.80, 0.00, 0.20),
        (0.87, 0.00, 0.13),
        (0.98, 0.00, 0.02),
        (0.98, 0.02, 0.00),
        (0.80, 0.02, 0.18),
    ],
    "T2": [
        (0.50, 0.00, 0.50),
        (0.80, 0.00, 0.20),
        (0.80, 0.02, 0.18),
        (0.50, 0.10, 0.40),
    ],
    "T3": [
        (0.00, 0.00, 1.00),
        (0.50, 0.00, 0.50),
        (0.50, 0.10, 0.40),
        (0.00, 0.15, 0.85),
    ],
    "D1": [
        (0.00, 0.15, 0.85),
        (0.50, 0.10, 0.40),
        (0.80, 0.02, 0.18),
        (0.98, 0.02, 0.00),
        (0.35, 0.65, 0.00),
        (0.00, 0.65, 0.35),
    ],
    "D2": [(0.00, 0.65, 0.35), (0.35, 0.65, 0.00), (0.00, 1.00, 0.00)],
    "DT": [(0.00, 0.15, 0.85), (0.00, 0.65, 0.35), (0.50, 0.10, 0.40)],
}

_DIAGNOSTIC_DETAILS = {
    "PD": {
        "label": "Partial discharge",
        "severity": "warn",
        "summary": "The gas pattern points to partial discharge activity.",
        "root_cause": "Likely corona, void discharge, or weak local insulation clearances.",
    },
    "D1": {
        "label": "Low-energy discharge",
        "severity": "bad",
        "summary": "The gas pattern points to low-energy electrical discharge.",
        "root_cause": "Likely sparking, bad contacts, or floating potentials.",
    },
    "D2": {
        "label": "High-energy discharge",
        "severity": "bad",
        "summary": "The gas pattern points to high-energy discharge / arcing.",
        "root_cause": "Likely arcing or major electrical breakdown with immediate reliability risk.",
    },
    "T1": {
        "label": "Thermal fault T1",
        "severity": "warn",
        "summary": "The gas pattern points to low-temperature thermal stress.",
        "root_cause": "Likely mild overheating, hot spots, or reduced cooling effectiveness.",
    },
    "T2": {
        "label": "Thermal fault T2",
        "severity": "bad",
        "summary": "The gas pattern points to medium-temperature thermal stress.",
        "root_cause": "Likely sustained overheating and accelerated oil/paper degradation.",
    },
    "T3": {
        "label": "Thermal fault T3",
        "severity": "bad",
        "summary": "The gas pattern points to severe high-temperature overheating.",
        "root_cause": "Likely intense oil cracking, metal hot spots, or severe cooling failure.",
    },
    "DT": {
        "label": "Mixed thermal/electrical fault",
        "severity": "bad",
        "summary": "The gas pattern points to combined thermal and electrical fault activity.",
        "root_cause": "Likely simultaneous overheating and discharge phenomena requiring detailed inspection.",
    },
}


def _safe_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _parse_limit_rule(expr):
    text = _safe_text(expr).replace(" ", "")
    if not text:
        return None
    text = text.replace("−", "-")
    text = text.replace("≤", "<=").replace("≥", ">=")

    if "-" in text and not text.startswith("-"):
        parts = text.split("-", 1)
        left = _safe_float(parts[0])
        right = _safe_float(parts[1])
        if left is not None and right is not None:
            # Accept compact upper bounds like 3800-14 => 3800-14000.
            if left >= 1000 and 0 < right < 100:
                while right < left:
                    right *= 10
            lo, hi = (left, right) if left <= right else (right, left)
            return {
                "raw": _safe_text(expr),
                "min": lo,
                "max": hi,
                "min_inc": True,
                "max_inc": True,
            }

    m = re.match(r"^(<=|>=|<|>)(.+)$", text)
    if m:
        op = m.group(1)
        val = _safe_float(m.group(2))
        if val is None:
            return None
        rule = {
            "raw": _safe_text(expr),
            "min": None,
            "max": None,
            "min_inc": True,
            "max_inc": True,
        }
        if op == "<":
            rule["max"] = val
            rule["max_inc"] = False
        elif op == "<=":
            rule["max"] = val
            rule["max_inc"] = True
        elif op == ">":
            rule["min"] = val
            rule["min_inc"] = False
        else:
            rule["min"] = val
            rule["min_inc"] = True
        return rule

    value = _safe_float(text)
    if value is not None:
        return {
            "raw": _safe_text(expr),
            "min": value,
            "max": value,
            "min_inc": True,
            "max_inc": True,
        }
    return None


def _matches_rule(value, rule):
    if value is None or rule is None:
        return False
    lo = rule.get("min")
    hi = rule.get("max")
    if lo is not None:
        if rule.get("min_inc", True):
            if value < lo:
                return False
        elif value <= lo:
            return False
    if hi is not None:
        if rule.get("max_inc", True):
            if value > hi:
                return False
        elif value >= hi:
            return False
    return True


def _severity_max(left: str, right: str) -> str:
    return (
        left
        if _DGA_SEVERITY_ORDER.get(left, 0) >= _DGA_SEVERITY_ORDER.get(right, 0)
        else right
    )


def _ratio_value(num, den):
    if num is None or den is None or den <= 0:
        return None
    return num / den


def _format_ratio_number(value):
    if value is None:
        return "-"
    return f"{float(value):.4g}"


def _ternary_to_cartesian(ch4_frac, c2h2_frac, c2h4_frac):
    return (
        c2h4_frac + (0.5 * c2h2_frac),
        (math.sqrt(3.0) / 2.0) * c2h2_frac,
    )


def _point_on_segment(point, start, end, eps=1e-9):
    px, py = point
    x1, y1 = start
    x2, y2 = end
    cross = abs((py - y1) * (x2 - x1) - (px - x1) * (y2 - y1))
    if cross > eps:
        return False
    dot = (px - x1) * (x2 - x1) + (py - y1) * (y2 - y1)
    if dot < -eps:
        return False
    sq_len = (x2 - x1) ** 2 + (y2 - y1) ** 2
    if dot - sq_len > eps:
        return False
    return True


def _point_in_polygon(point, polygon):
    inside = False
    for idx in range(len(polygon)):
        start = polygon[idx]
        end = polygon[(idx + 1) % len(polygon)]
        if _point_on_segment(point, start, end):
            return True
        x1, y1 = start
        x2, y2 = end
        intersects = ((y1 > point[1]) != (y2 > point[1])) and (
            point[0] < (x2 - x1) * (point[1] - y1) / ((y2 - y1) or 1e-12) + x1
        )
        if intersects:
            inside = not inside
    return inside


def _build_diagnostic_result(
    method_key, method_label, code, reasoning, *, confidence="medium"
):
    details = _DIAGNOSTIC_DETAILS.get(code, {})
    summary = details.get("summary") or "DGA diagnostic pattern detected."
    root_cause = (
        details.get("root_cause") or "Potential abnormal transformer fault pattern."
    )
    return {
        "method": method_key,
        "method_label": method_label,
        "status": details.get("severity", "warn"),
        "code": code,
        "label": details.get("label") or code,
        "summary": summary,
        "root_cause": root_cause,
        "reasoning": reasoning,
        "confidence": confidence,
        "display_summary": f"{method_label}: {details.get('label') or code} - {summary}",
    }


def analyze_dga_diagnostics(values):
    gases = {
        key: _safe_float(values.get(key))
        for key in ("h2", "c2h2", "c2h4", "c2h6", "co", "co2", "ch4", "o2", "n2")
    }

    ratios = {
        "ch4_h2": _ratio_value(gases["ch4"], gases["h2"]),
        "c2h2_c2h4": _ratio_value(gases["c2h2"], gases["c2h4"]),
        "c2h4_c2h6": _ratio_value(gases["c2h4"], gases["c2h6"]),
        "co2_co": _ratio_value(gases["co2"], gases["co"]),
        "o2_n2": _ratio_value(gases["o2"], gases["n2"]),
        "c2h2_h2": _ratio_value(gases["c2h2"], gases["h2"]),
    }

    findings = []
    overall_level = "ok"
    primary = None

    ratio_total = sum(
        gases[key] or 0.0 for key in ("h2", "ch4", "c2h2", "c2h4", "c2h6")
    )
    ratio_diag = {
        "method": "iec_60599_ratios",
        "method_label": "IEC 60599 / Rogers ratios",
        "status": "ok",
        "code": None,
        "label": "No abnormal ratio diagnosis",
        "summary": "No abnormal IEC 60599 / Rogers ratio diagnosis.",
        "root_cause": "No standards-based ratio fault pattern detected.",
        "reasoning": [],
        "confidence": "low",
        "display_summary": "IEC 60599 / Rogers ratios: no abnormal diagnosis.",
        "insufficient_data": False,
    }
    if ratio_total < _RATIO_METHOD_MIN_TOTAL_PPM:
        ratio_diag.update(
            {
                "status": "ok",
                "label": "Insufficient data",
                "summary": "Key combustible gas total is too low for reliable ratio diagnosis.",
                "root_cause": "Use trend data or repeat sampling before drawing conclusions from ratios.",
                "reasoning": [
                    f"H2+CH4+C2H2+C2H4+C2H6 = {_format_ratio_number(ratio_total)} ppm < {_format_ratio_number(_RATIO_METHOD_MIN_TOTAL_PPM)} ppm"
                ],
                "display_summary": "IEC 60599 / Rogers ratios: insufficient gas volume for reliable diagnosis.",
                "insufficient_data": True,
            }
        )
    else:
        for rule in _RATIO_RULES:
            if rule["predicate"](ratios):
                ratio_diag = _build_diagnostic_result(
                    "iec_60599_ratios",
                    "IEC 60599 / Rogers ratios",
                    rule["code"],
                    [
                        f"CH4/H2 = {_format_ratio_number(ratios['ch4_h2'])}",
                        f"C2H2/C2H4 = {_format_ratio_number(ratios['c2h2_c2h4'])}",
                        f"C2H4/C2H6 = {_format_ratio_number(ratios['c2h4_c2h6'])}",
                    ],
                    confidence="medium",
                )
                break
        else:
            ratio_diag.update(
                {
                    "label": "Inconclusive ratio pattern",
                    "summary": "The ratios do not fall cleanly inside one IEC 60599 / Rogers class.",
                    "root_cause": "Possible mixed fault, early-stage fault, or measurement set outside the standard decision table.",
                    "reasoning": [
                        f"CH4/H2 = {_format_ratio_number(ratios['ch4_h2'])}",
                        f"C2H2/C2H4 = {_format_ratio_number(ratios['c2h2_c2h4'])}",
                        f"C2H4/C2H6 = {_format_ratio_number(ratios['c2h4_c2h6'])}",
                    ],
                    "display_summary": "IEC 60599 / Rogers ratios: inconclusive / mixed pattern.",
                }
            )

    duval_diag = {
        "method": "duval_triangle_1",
        "method_label": "Duval Triangle 1",
        "status": "ok",
        "code": None,
        "label": "No abnormal Duval diagnosis",
        "summary": "No abnormal Duval Triangle 1 diagnosis.",
        "root_cause": "No ternary fault zone assigned.",
        "reasoning": [],
        "confidence": "low",
        "display_summary": "Duval Triangle 1: no abnormal diagnosis.",
        "insufficient_data": False,
        "coordinates": None,
    }
    ch4 = gases["ch4"] or 0.0
    c2h2 = gases["c2h2"] or 0.0
    c2h4 = gases["c2h4"] or 0.0
    duval_total = ch4 + c2h2 + c2h4
    if duval_total < _DUVAL_TRIANGLE_MIN_TOTAL_PPM:
        duval_diag.update(
            {
                "label": "Insufficient data",
                "summary": "CH4, C2H2, and C2H4 total is too low for stable Duval Triangle classification.",
                "root_cause": "Repeat sampling or trend analysis is recommended before trusting a ternary diagnosis.",
                "reasoning": [
                    f"CH4+C2H2+C2H4 = {_format_ratio_number(duval_total)} ppm < {_format_ratio_number(_DUVAL_TRIANGLE_MIN_TOTAL_PPM)} ppm"
                ],
                "display_summary": "Duval Triangle 1: insufficient gas volume for reliable diagnosis.",
                "insufficient_data": True,
            }
        )
    elif duval_total > 0:
        ch4_frac = ch4 / duval_total
        c2h2_frac = c2h2 / duval_total
        c2h4_frac = c2h4 / duval_total
        point = _ternary_to_cartesian(ch4_frac, c2h2_frac, c2h4_frac)
        zone_code = None
        for candidate in ("PD", "D2", "DT", "D1", "T3", "T2", "T1"):
            polygon = [
                _ternary_to_cartesian(*vertex)
                for vertex in _DUVAL_TRIANGLE_1_ZONES[candidate]
            ]
            if _point_in_polygon(point, polygon):
                zone_code = candidate
                break
        duval_diag["coordinates"] = {
            "ch4_pct": round(ch4_frac * 100.0, 3),
            "c2h2_pct": round(c2h2_frac * 100.0, 3),
            "c2h4_pct": round(c2h4_frac * 100.0, 3),
        }
        if zone_code:
            duval_diag = _build_diagnostic_result(
                "duval_triangle_1",
                "Duval Triangle 1",
                zone_code,
                [
                    f"CH4={round(ch4_frac * 100.0, 2)}%",
                    f"C2H2={round(c2h2_frac * 100.0, 2)}%",
                    f"C2H4={round(c2h4_frac * 100.0, 2)}%",
                ],
                confidence="medium",
            )
            duval_diag["coordinates"] = {
                "ch4_pct": round(ch4_frac * 100.0, 3),
                "c2h2_pct": round(c2h2_frac * 100.0, 3),
                "c2h4_pct": round(c2h4_frac * 100.0, 3),
            }
        else:
            duval_diag.update(
                {
                    "label": "Inconclusive Duval point",
                    "summary": "The normalized point does not sit clearly inside a Duval Triangle 1 zone.",
                    "root_cause": "Possible mixed fault or boundary-condition case requiring trend review.",
                    "reasoning": [
                        f"CH4={round(ch4_frac * 100.0, 2)}%",
                        f"C2H2={round(c2h2_frac * 100.0, 2)}%",
                        f"C2H4={round(c2h4_frac * 100.0, 2)}%",
                    ],
                    "display_summary": "Duval Triangle 1: inconclusive / boundary condition.",
                }
            )

    paper_diag = None
    co = gases["co"]
    co2 = gases["co2"]
    if co is not None and co > 0 and co2 is not None:
        paper_ratio = co2 / co
        if paper_ratio < 3.0:
            paper_diag = {
                "method": "co2_co_ratio",
                "method_label": "CO2/CO ratio",
                "status": "bad",
                "code": "CELLULOSE_SEVERE",
                "label": "Severe cellulose degradation indication",
                "summary": "CO2/CO ratio is below the accepted paper-aging threshold.",
                "root_cause": "Possible severe paper overheating, oxidation, or cellulose decomposition.",
                "reasoning": [f"CO2/CO = {_format_ratio_number(paper_ratio)} < 3"],
                "confidence": "medium",
                "display_summary": "CO2/CO ratio: severe cellulose degradation indication.",
            }
        elif paper_ratio < 10.0:
            paper_diag = {
                "method": "co2_co_ratio",
                "method_label": "CO2/CO ratio",
                "status": "warn",
                "code": "CELLULOSE_WATCH",
                "label": "Cellulose aging watch",
                "summary": "CO2/CO ratio suggests possible paper aging or early cellulose degradation.",
                "root_cause": "Paper insulation may be under thermal stress and should be trended.",
                "reasoning": [
                    f"CO2/CO = {_format_ratio_number(paper_ratio)} between 3 and 10"
                ],
                "confidence": "medium",
                "display_summary": "CO2/CO ratio: watch cellulose condition.",
            }

    for diag in (ratio_diag, duval_diag, paper_diag):
        if not diag:
            continue
        if diag.get("code"):
            findings.append(diag)
            overall_level = _severity_max(overall_level, diag.get("status", "ok"))
            if primary is None or _DGA_SEVERITY_ORDER.get(
                diag.get("status", "ok"), 0
            ) > _DGA_SEVERITY_ORDER.get(primary.get("status", "ok"), 0):
                primary = diag

    consensus = None
    ratio_code = ratio_diag.get("code")
    duval_code = duval_diag.get("code")
    if ratio_code and duval_code:
        if ratio_code == duval_code:
            consensus = {
                "status": _severity_max(
                    ratio_diag.get("status", "ok"), duval_diag.get("status", "ok")
                ),
                "summary": f"IEC 60599 / Rogers and Duval Triangle 1 both indicate {ratio_code}.",
                "reasoning": [
                    ratio_diag.get("display_summary"),
                    duval_diag.get("display_summary"),
                ],
            }
            overall_level = _severity_max(overall_level, consensus["status"])
        else:
            consensus = {
                "status": "warn",
                "summary": "IEC 60599 / Rogers and Duval Triangle 1 do not fully agree.",
                "reasoning": [
                    ratio_diag.get("display_summary"),
                    duval_diag.get("display_summary"),
                ],
            }
            overall_level = _severity_max(overall_level, "warn")

    return {
        "ratios": ratios,
        "ratio_method": ratio_diag,
        "duval_triangle_1": duval_diag,
        "paper_condition": paper_diag,
        "findings": findings,
        "consensus": consensus,
        "primary": primary,
        "overall_level": overall_level,
    }


def _row_limit_rules(ws, section, row):
    if section == "gases":
        good_expr = _safe_text(ws[f"F{row}"].value)
        tolerable_expr = ""
        poor_expr = ""
    else:
        # Columns: K=ΚΑΛΟ, L=ΑΝΕΚΤΟ (optional), N=ΠΤΩΧΟ (optional, skips M).
        good_expr = _safe_text(ws[f"K{row}"].value)
        tolerable_expr = _safe_text(ws[f"L{row}"].value)
        poor_expr = _safe_text(ws[f"N{row}"].value)

    good_rules = []
    tolerable_rules = []
    poor_rules = []

    if good_expr:
        rule = _parse_limit_rule(good_expr)
        if rule:
            good_rules.append(rule)

    if tolerable_expr:
        rule = _parse_limit_rule(tolerable_expr)
        if rule:
            tolerable_rules.append(rule)

    if poor_expr:
        rule = _parse_limit_rule(poor_expr)
        if rule:
            poor_rules.append(rule)

    shown_exprs = [expr for expr in [good_expr, tolerable_expr] if expr]
    combined = good_rules + [r for r in tolerable_rules if r not in good_rules]
    return {
        "good_rules": good_rules,
        "tolerable_rules": tolerable_rules,
        "poor_rules": poor_rules,
        "all_rules": combined,
        "limit_text": " / ".join(shown_exprs),
        "good_text": good_expr,
        "tolerable_text": tolerable_expr,
        "poor_text": poor_expr,
    }


@lru_cache(maxsize=8)
def _load_dga_template_metadata_cached(template_path, mtime):
    try:
        wb = load_workbook(template_path, data_only=True)
    except PermissionError:
        temp_copy = os.path.join(tempfile.gettempdir(), "dga_template_meta_copy.xlsx")
        shutil.copy2(template_path, temp_copy)
        wb = load_workbook(temp_copy, data_only=True)
    ws = wb["Αναφορά"] if "Αναφορά" in wb.sheetnames else wb[wb.sheetnames[0]]

    fields = []
    by_key = {}

    for key, row, section, fallback_label in _DGA_FIELD_ROWS:
        label = _safe_text(ws[f"B{row}"].value) or fallback_label
        if section == "gases":
            label = _GAS_LABEL_OVERRIDES.get(key, label)
        elif section == "physchem":
            label = _PHYSCHEM_LABEL_OVERRIDES.get(key, label)
        unit = _safe_text(ws[f"F{row}"].value)
        if section == "gases":
            unit = _GAS_UNIT
        elif key in _PHYSCHEM_UNIT_OVERRIDES:
            unit = _PHYSCHEM_UNIT_OVERRIDES[key]
        if section == "gases":
            gas_expr = _GAS_LIMIT_OVERRIDES.get(key, "")
            gas_rule = _parse_limit_rule(gas_expr) if gas_expr else None
            rule_info = {
                "good_rules": [gas_rule] if gas_rule else [],
                "tolerable_rules": [],
                "all_rules": [gas_rule] if gas_rule else [],
                "limit_text": gas_expr,
                "good_text": gas_expr,
                "tolerable_text": "",
            }
        else:
            rule_info = _row_limit_rules(ws, section, row)
        field = {
            "key": key,
            "row": row,
            "section": section,
            "label": label,
            "unit": unit,
            "rules": rule_info.get("all_rules") or [],
            "good_rules": rule_info.get("good_rules") or [],
            "tolerable_rules": rule_info.get("tolerable_rules") or [],
            "poor_rules": rule_info.get("poor_rules") or [],
            "limit_text": rule_info.get("limit_text") or "",
            "good_text": rule_info.get("good_text") or "",
            "tolerable_text": rule_info.get("tolerable_text") or "",
            "poor_text": rule_info.get("poor_text") or "",
        }
        fields.append(field)
        by_key[key] = field

    return {
        "sections": {
            "meta": "Στοιχεία Δειγματοληψίας / Μέτρησης",
            "gases": _safe_text(ws["B18"].value) or "Αέρια",
            "physchem": _safe_text(ws["F38"].value) or "Φυσικοχημικές Μετρήσεις",
        },
        "fields": fields,
        "by_key": by_key,
    }


def load_dga_template_metadata(template_path):
    """Return field labels, categories and limit rules from the DGA template."""
    try:
        path = os.path.abspath(template_path)
        mtime = os.path.getmtime(path)
        return _load_dga_template_metadata_cached(path, mtime)
    except Exception:
        fallback = {
            "sections": {
                "meta": "Στοιχεία Δειγματοληψίας / Μέτρησης",
                "gases": "Αέρια",
                "physchem": "Φυσικοχημικές Μετρήσεις",
            },
            "fields": [],
            "by_key": {},
        }
        for key, row, section, label in _DGA_FIELD_ROWS:
            field = {
                "key": key,
                "row": row,
                "section": section,
                "label": label,
                "unit": "",
                "rules": [],
                "good_rules": [],
                "tolerable_rules": [],
                "poor_rules": [],
                "limit_text": "",
                "good_text": "",
                "tolerable_text": "",
                "poor_text": "",
            }
            fallback["fields"].append(field)
            fallback["by_key"][key] = field
        return fallback


def evaluate_dga_limits(values, template_path):
    """Evaluate one DGA measurement dict and return ok/warn/bad details."""
    metadata = load_dga_template_metadata(template_path)
    problems = []
    warnings = []
    checks = []

    for field in metadata.get("fields", []):
        rules = field.get("rules") or []
        if not rules:
            continue
        raw = values.get(field["key"])
        value = _safe_float(raw)
        if value is None:
            continue

        good_rules = field.get("good_rules") or []
        tolerable_rules = field.get("tolerable_rules") or []

        is_good = bool(good_rules) and any(
            _matches_rule(value, rule) for rule in good_rules
        )
        is_tolerable = bool(tolerable_rules) and any(
            _matches_rule(value, rule) for rule in tolerable_rules
        )
        is_allowed = any(_matches_rule(value, rule) for rule in rules)

        level = "bad"
        if is_good:
            level = "ok"
        elif is_tolerable:
            level = "warn"
        elif not good_rules and is_allowed:
            # Fields with only one limit column (e.g. gases) are either ok or bad.
            level = "ok"

        item = {
            "key": field["key"],
            "label": field.get("label") or field["key"],
            "value": value,
            "limit_text": field.get("limit_text") or "",
            "good_text": field.get("good_text") or "",
            "tolerable_text": field.get("tolerable_text") or "",
            "poor_text": field.get("poor_text") or "",
            "section": field.get("section") or "",
            "level": level,
        }
        checks.append(item)

        if level == "warn":
            warnings.append(item)
        elif level == "bad":
            problems.append(item)

    overall = "ok"
    if problems:
        overall = "bad"
    elif warnings:
        overall = "warn"

    diagnostics = analyze_dga_diagnostics(values)
    overall = _severity_max(overall, diagnostics.get("overall_level", "ok"))

    return {
        "is_problematic": bool(problems) or diagnostics.get("overall_level") == "bad",
        "has_warnings": bool(warnings)
        or diagnostics.get("overall_level") in {"warn", "bad"},
        "overall_level": overall,
        "checks": checks,
        "warnings": warnings,
        "problems": problems,
        "diagnostics": diagnostics,
        "metadata": metadata,
    }


def _safe_float(value):
    if value is None:
        return None
    txt = str(value).strip().replace(",", ".")
    if not txt:
        return None
    try:
        return float(txt)
    except Exception:
        return None


def _safe_date_text(value):
    txt = (value or "").strip()
    if not txt:
        return ""
    try:
        dt = datetime.fromisoformat(txt.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return txt


def _surname_upper(value):
    txt = (value or "").strip()
    if not txt:
        return ""
    parts = txt.split()
    return (parts[-1] if parts else txt).upper()


def generate_dga_excel_report(
    template_path: str, output_path: str, payload: dict
) -> str:
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            S["MESSAGES"]
            .get("DGA_TEMPLATE_NOT_FOUND_FMT", "DGA template not found: {path}")
            .format(path=template_path)
        )

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    try:
        wb = load_workbook(template_path)
    except PermissionError:
        # OneDrive placeholders or locked files may fail to open directly.
        temp_copy = os.path.join(tempfile.gettempdir(), "dga_template_copy.xlsx")
        shutil.copy2(template_path, temp_copy)
        wb = load_workbook(temp_copy)

    ws = wb["Αναφορά"] if "Αναφορά" in wb.sheetnames else wb[wb.sheetnames[0]]

    # Header section
    ws["B8"] = f"ΥΠΟΣΤΑΘΜΟΣ : {payload.get('substation_name', '')}"
    ws["B9"] = f"ΣΤΟΙΧΕΙΟ ΔΙΚΤΥΟΥ : {payload.get('element_name', '')}"
    ws["B10"] = f"ΚΑΤΑΣΚΕΥΑΣΤΗΣ : {payload.get('manufacturer', '')}"

    ws["I8"] = f"ΣΗΜΕΙΟ ΔΕΙΓΜΑΤΟΛΗΨΙΑΣ : {payload.get('sample_point', '')}"
    ws["I9"] = f"ΑΡΙΘ. ΚΑΤΑΣΚΕΥΑΣΤΗ : {payload.get('serial_number', '')}"
    ws["I10"] = f"ΜΕΘΟΔΟΣ : {payload.get('sampling_method', '')}"

    ws["H13"] = _surname_upper(payload.get("sampling_responsible", ""))
    ws["H14"] = _safe_date_text(payload.get("sampling_date", ""))
    ws["H15"] = _surname_upper(payload.get("measurement_responsible", ""))
    ws["H16"] = _safe_date_text(payload.get("measurement_date", ""))
    ws["H17"] = _safe_float(payload.get("sample_temperature"))

    # Gas analysis values
    ws["H19"] = _safe_float(payload.get("h2"))
    ws["H20"] = _safe_float(payload.get("c2h2"))
    ws["H21"] = _safe_float(payload.get("c2h4"))
    ws["H22"] = _safe_float(payload.get("c2h6"))
    ws["H23"] = _safe_float(payload.get("co"))
    ws["H24"] = _safe_float(payload.get("co2"))
    ws["H25"] = _safe_float(payload.get("ch4"))
    ws["H26"] = _safe_float(payload.get("o2"))
    ws["H27"] = _safe_float(payload.get("c3h8"))
    ws["H28"] = _safe_float(payload.get("n2"))
    ws["H29"] = _safe_float(payload.get("h2o"))

    # Physicochemical section
    ws["H39"] = _surname_upper(payload.get("sampling_responsible", ""))
    ws["H40"] = _safe_date_text(payload.get("sampling_date", ""))
    ws["H41"] = _surname_upper(payload.get("measurement_responsible", ""))
    ws["H42"] = _safe_date_text(payload.get("measurement_date", ""))
    ws["H43"] = _safe_float(payload.get("sample_temperature"))

    ws["H45"] = _safe_float(payload.get("density"))
    ws["H46"] = _safe_float(payload.get("humidity"))
    ws["H47"] = _safe_float(payload.get("dielectric_strength"))
    ws["H48"] = _safe_float(payload.get("loss_factor"))
    ws["H49"] = _safe_float(payload.get("surface_tension"))

    notes = (payload.get("notes") or "").strip()
    if notes:
        ws["C52"] = notes

    # Append diagnostics sheet with standards-based interpretation
    try:
        # Compute diagnostics from payload values (best-effort)
        diag = None
        try:
            diag = analyze_dga_diagnostics(payload)
        except Exception:
            diag = None

        # Create a Diagnostics sheet summarizing findings for human readers
        diag_sheet = wb.create_sheet("Diagnostics")
        diag_sheet["A1"] = "DGA Diagnostics Summary"
        row = 2
        if diag is None:
            diag_sheet[f"A{row}"] = "Diagnostics: unavailable"
        else:
            # Primary
            primary = diag.get("primary") or {}
            consensus = diag.get("consensus") or {}
            findings = diag.get("findings") or []

            diag_sheet[f"A{row}"] = "Overall level"
            diag_sheet[f"B{row}"] = diag.get("overall_level") or "ok"
            row += 1

            diag_sheet[f"A{row}"] = "Primary diagnosis"
            diag_sheet[f"B{row}"] = primary.get("display_summary") or "-"
            row += 1

            if consensus and consensus.get("summary"):
                diag_sheet[f"A{row}"] = "Consensus"
                diag_sheet[f"B{row}"] = consensus.get("summary")
                row += 1

            if findings:
                diag_sheet[f"A{row}"] = "Findings"
                row += 1
                diag_sheet[f"A{row}"] = "Code"
                diag_sheet[f"B{row}"] = "Label"
                diag_sheet[f"C{row}"] = "Status"
                diag_sheet[f"D{row}"] = "Summary"
                diag_sheet[f"E{row}"] = "Root cause / Reasoning"
                row += 1
                for f in findings:
                    diag_sheet[f"A{row}"] = f.get("code")
                    diag_sheet[f"B{row}"] = f.get("label")
                    diag_sheet[f"C{row}"] = f.get("status")
                    diag_sheet[f"D{row}"] = f.get("summary")
                    reasoning = "; ".join(f.get("reasoning") or [])
                    root = f.get("root_cause") or ""
                    diag_sheet[f"E{row}"] = root + (
                        " - " + reasoning if reasoning else ""
                    )
                    row += 1

    except Exception:
        # Non-fatal: if diagnostics writing fails, continue saving the Excel
        pass

    wb.save(output_path)
    return output_path
