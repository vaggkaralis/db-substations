import os

try:
    import openpyxl
except ImportError:
    openpyxl = None

from strings import STRINGS as S

TEMPLATE_SUBSTATIONS = "substations_import_template.xlsx"
TEMPLATE_ELEMENTS = "elements_import_template.xlsx"


def create_substations_template(base_dir: str) -> tuple[bool, str]:
    """Create substations import template. Returns (success, message/path)."""
    if openpyxl is None:
        return False, S["MESSAGES"].get("OPENPYXL_MISSING", "openpyxl δεν είναι εγκατεστημένο!")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Substations"

        headers = S["MESSAGES"].get("TEMPLATE_SUBSTATIONS_HEADERS", ["Name", "Location", "Adoption Date"])
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 15

        examples = [
            ("Υποσταθμός Α", "https://maps.google.com/?q=example1", "2025-01-15"),
            ("Υποσταθμός Β", "https://maps.google.com/?q=example2", "2025-01-20"),
        ]
        for idx, (name, location, date) in enumerate(examples, 2):
            ws.cell(row=idx, column=1, value=name)
            ws.cell(row=idx, column=2, value=location)
            ws.cell(row=idx, column=3, value=date)

        template_path = os.path.join(base_dir, TEMPLATE_SUBSTATIONS)
        wb.save(template_path)
        return True, template_path
    except Exception as exc:  # pragma: no cover - UI surface
        return False, f"Σφάλμα: {exc}"


def create_elements_template(base_dir: str) -> tuple[bool, str]:
    """Create elements import template. Returns (success, message/path)."""
    if openpyxl is None:
        return False, S["MESSAGES"].get("OPENPYXL_MISSING", "openpyxl δεν είναι εγκατεστημένο!")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Elements"

        # Add version info in first row
        ws.cell(row=1, column=1, value="TEMPLATE_VERSION: v2.0")
        ws.cell(row=1, column=1).font = Font(italic=True, color="999999")
        ws.row_dimensions[1].height = 15

        # The template requires model information; the element `manufacturer` is derived from the model.
        headers = S["MESSAGES"].get("TEMPLATE_ELEMENTS_HEADERS", [
            "Substation Name",
            "Element Type",
            "Name",
            "Serial Number",
            "Maintenance Date",
            "Τύπος Διακόπτη",
            "Breaker Role",
            "Operating Status",
            "Gate",
            "Model Name",
            "Model Manufacturer",
            "Model Installation Space",
        ])
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="70AD47", end_color="70AD47", fill_type="solid"
            )

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 20  # Breaker Role
        ws.column_dimensions["I"].width = 15  # Operating Status
        ws.column_dimensions["J"].width = 15  # Gate
        ws.column_dimensions["K"].width = 20
        ws.column_dimensions["L"].width = 20
        ws.column_dimensions["M"].width = 20
        ws.column_dimensions["N"].width = 25

        examples = [
            (
                "Υποσταθμός Α",
                "Διακόπτης ΜΤ",
                "Main Breaker",
                "SN-001",
                "2025-01-20",
                "SF6",
                "Κεντρικός",
                "Ενεργή",
                "ΠΥΛΗ 1",
                "SF6-400",
                "ABB",
                "Εσωτερικού",
            ),
            (
                "Υποσταθμός Α",
                "Μετασχηματιστής 150/20KV",
                "Transformer 1",
                "SN-002",
                "2025-01-18",
                "",
                "",
                "Ενεργή",
                "ΠΥΛΗ 1",
                "GEAFOL",
                "Siemens",
                "Εξωτερικού",
            ),
        ]
        for idx, row_data in enumerate(
            examples, 3
        ):  # Start at row 3 (after version and headers)
            for col, value in enumerate(row_data, 1):
                ws.cell(row=idx, column=col, value=value)

        template_path = os.path.join(base_dir, TEMPLATE_ELEMENTS)
        wb.save(template_path)
        return True, template_path
    except Exception as exc:  # pragma: no cover - UI surface
        return False, f"Σφάλμα: {exc}"
