import importlib
import hashlib
import os
import re
import subprocess
import sys
import tempfile
import shutil
import webbrowser

from pdf_reports import generate_sf6_leak_report
from popups import show_message_popup
from report_sync import safe_generate_and_store_report
from strings_proxy import STRINGS as S


def normalize_decimal_numeric_text(value, decimal_separator="."):
    text = "" if value is None else str(value).strip()
    if not text:
        return ""

    text = text.replace("\u00a0", " ")
    text = re.sub(r"\s+", "", text)
    if not text:
        return ""

    last_comma = text.rfind(",")
    last_dot = text.rfind(".")

    if last_comma != -1 and last_dot != -1:
        if last_comma > last_dot:
            canonical = text.replace(".", "").replace(",", ".")
        else:
            canonical = text.replace(",", "")
    else:
        canonical = text.replace(",", ".")

    if decimal_separator == ",":
        return canonical.replace(".", ",")
    return canonical


def _compute_sf6_leakage_bands(rows):
    values = [
        row.get("leakage") for row in (rows or []) if row.get("leakage") is not None
    ]
    if not values:
        return {"min": None, "low_max": None, "mid_max": None, "max": None}

    min_value = min(values)
    max_value = max(values)
    if max_value <= min_value:
        return {
            "min": min_value,
            "low_max": min_value,
            "mid_max": min_value,
            "max": max_value,
        }

    step = (max_value - min_value) / 3.0
    return {
        "min": min_value,
        "low_max": min_value + step,
        "mid_max": min_value + (2 * step),
        "max": max_value,
    }


def _classify_sf6_leakage(leakage, bands):
    if leakage is None or not bands or bands.get("min") is None:
        return "none"
    if bands.get("max") == bands.get("min"):
        return "green"
    # Inverted mapping: low -> green, mid -> yellow, high -> red
    if leakage <= bands.get("low_max"):
        return "green"
    if leakage <= bands.get("mid_max"):
        return "yellow"
    return "red"


def _sf6_row_background_rgba(leakage, bands):
    return {
        "red": (1.0, 0.90, 0.90, 1.0),
        "yellow": (1.0, 0.98, 0.86, 1.0),
        "green": (0.90, 1.0, 0.90, 1.0),
        "none": (1.0, 1.0, 1.0, 1.0),
    }[_classify_sf6_leakage(leakage, bands)]


def _format_display_date(app, date_time_value):
    try:
        return app._format_maintenance_date(date_time_value)
    except Exception:
        pass
    return str(date_time_value or "")


def _win_existing_path(path: str | None) -> str | None:
    if not path:
        return None
    try:
        abs_path = os.path.abspath(path)
    except Exception:
        abs_path = str(path)

    if os.path.exists(abs_path):
        return abs_path

    if os.name != "nt":
        return None

    if abs_path.startswith("\\\\?\\"):
        long_path = abs_path
    elif abs_path.startswith("\\\\"):
        long_path = "\\\\?\\UNC\\" + abs_path[2:]
    else:
        long_path = "\\\\?\\" + abs_path

    return long_path if os.path.exists(long_path) else None


def _normalize_open_path(path: str | None) -> str | None:
    if path is None:
        return None
    text = str(path).strip()
    if not text:
        return None
    try:
        text = text.replace("\r", "").replace("\n", os.sep)
    except Exception:
        pass
    try:
        return os.path.normpath(text)
    except Exception:
        return text


def _nearest_existing_folder(path: str | None) -> str | None:
    target = _normalize_open_path(path)
    if not target:
        return None

    existing = _win_existing_path(target)
    if existing:
        return existing if os.path.isdir(existing) else os.path.dirname(existing)

    current = target
    while current:
        parent = os.path.dirname(current.rstrip("\\/"))
        if not parent or parent == current:
            break
        existing_parent = _win_existing_path(parent)
        if existing_parent and os.path.isdir(existing_parent):
            return existing_parent
        current = parent
    return None


def open_folder_or_url(
    path,
    *,
    not_found_message="Ο φάκελος δεν βρέθηκε!",
    error_title="Σφάλμα",
    error_prefix="Αποτυχία ανοίγματος φακέλου ή συνδέσμου:\n",
):
    from popups import show_message_popup

    target = str(path or "").strip()
    if not target:
        show_message_popup(error_title, not_found_message)
        return False

    if target.startswith(("http://", "https://")):
        try:
            webbrowser.open(target)
            return True
        except Exception as exc:
            show_message_popup(error_title, f"{error_prefix}{str(exc)}")
            return False

    existing_folder = _nearest_existing_folder(target)
    if not existing_folder:
        show_message_popup(
            error_title,
            f"{not_found_message}\n{_normalize_open_path(target) or target}",
        )
        return False

    try:
        if sys.platform == "win32":
            os.startfile(existing_folder)
        elif sys.platform == "darwin":
            subprocess.call(["open", existing_folder])
        else:
            subprocess.call(["xdg-open", existing_folder])
        return True
    except Exception as exc:
        show_message_popup(error_title, f"{error_prefix}{str(exc)}")
        return False


def _short_temp_open_copy(path: str, *, existing_path: str | None = None) -> str | None:
    source_path = existing_path or _win_existing_path(path)
    if not source_path:
        return None

    suffix = os.path.splitext(path or source_path)[1]
    digest = hashlib.sha1(
        os.path.abspath(path or source_path).encode("utf-8", errors="ignore")
    ).hexdigest()[:12]
    temp_dir = os.path.join(tempfile.gettempdir(), "dbsub_open")
    os.makedirs(temp_dir, exist_ok=True)
    temp_path = os.path.join(temp_dir, f"open_{digest}{suffix}")
    shutil.copy2(source_path, temp_path)
    return temp_path


def show_sf6_management_popup(
    app,
    instance=None,
    preselected_year=None,
    preselected_substation=None,
):
    """Show SF6 leakage management report popup (delegated from DBrun)."""
    # Import Kivy widgets lazily to avoid top-level Kivy dependency in tests
    Popup = importlib.import_module("kivy.uix.popup").Popup
    BoxLayout = importlib.import_module("kivy.uix.boxlayout").BoxLayout
    Button = importlib.import_module("kivy.uix.button").Button
    Label = importlib.import_module("kivy.uix.label").Label
    Spinner = importlib.import_module("kivy.uix.spinner").Spinner
    GridLayout = importlib.import_module("kivy.uix.gridlayout").GridLayout
    ScrollView = importlib.import_module("kivy.uix.scrollview").ScrollView
    TextInput = importlib.import_module("kivy.uix.textinput").TextInput
    Widget = importlib.import_module("kivy.uix.widget").Widget
    Color = importlib.import_module("kivy.graphics").Color
    Rectangle = importlib.import_module("kivy.graphics").Rectangle
    IconOnlyButton = importlib.import_module("ui.shared").IconOnlyButton

    c = app.conn.cursor()
    c.execute(
        "SELECT DISTINCT substr(date_time, 1, 4) "
        "FROM maintenance "
        "WHERE date_time IS NOT NULL AND date_time != '' "
        "ORDER BY 1 DESC"
    )
    years = [row[0] for row in c.fetchall() if row[0] and row[0].isdigit()]
    if not years:
        years = [str(__import__("datetime").datetime.now().year)]

    popup = Popup(
        title=S["MESSAGES"].get("SF6_MANAGEMENT_TITLE", "Διαχείριση SF6"),
        size_hint=(0.95, 0.9),
    )
    main_layout = BoxLayout(orientation="vertical", padding=10, spacing=10)

    filter_row = BoxLayout(size_hint_y=None, height=40, spacing=10)
    filter_row.add_widget(
        Label(text=S["MESSAGES"].get("YEAR_LABEL", "Έτος:"), size_hint_x=0.15)
    )
    year_spinner = Spinner(
        text=years[0], values=years, size_hint_x=0.25, size_hint_y=None, height=35
    )
    if preselected_year in years:
        year_spinner.text = preselected_year
    filter_row.add_widget(year_spinner)

    show_all_label = S["MESSAGES"].get(
        "SHOW_ALL_SUBSTATIONS", "Προβολή Όλων των Υποσταθμών"
    )
    selected_substation = {
        "name": preselected_substation or show_all_label,
    }

    filter_row.add_widget(
        Label(
            text=S["MESSAGES"].get("SUBSTATION_LABEL", "Υποσταθμός:"), size_hint_x=0.18
        )
    )
    substation_filter_input = TextInput(
        text=selected_substation["name"],
        readonly=True,
        multiline=False,
        size_hint_x=0.25,
    )
    filter_row.add_widget(substation_filter_input)

    substation_picker_btn = Button(
        text=S["MESSAGES"].get("SELECT_PROMPT", "Επιλογή"),
        size_hint_x=0.11,
    )
    filter_row.add_widget(substation_picker_btn)

    reset_filter_btn = Button(
        text=S["BUTTONS"].get("CLEAR", "Καθαρισμός"),
        size_hint_x=0.16,
    )
    filter_row.add_widget(reset_filter_btn)

    main_layout.add_widget(filter_row)

    control_row = BoxLayout(size_hint_y=None, height=40, spacing=10)
    refresh_btn = Button(text=S["MESSAGES"].get("REFRESH", "Ανανέωση"), size_hint_x=0.2)
    control_row.add_widget(refresh_btn)
    print_btn = Button(text=S["MESSAGES"].get("PRINT", "Εκτύπωση"), size_hint_x=0.2)
    control_row.add_widget(print_btn)
    excel_btn = Button(text=S["MESSAGES"].get("EXCEL", "Excel"), size_hint_x=0.2)
    control_row.add_widget(excel_btn)
    main_layout.add_widget(control_row)

    summary_label = Label(text="", size_hint_y=None, height=60)
    summary_label.bind(
        width=lambda inst, val: setattr(inst, "text_size", (val, None)),
        texture_size=lambda inst, val: setattr(inst, "height", val[1] + 10),
    )
    main_layout.add_widget(summary_label)

    scroll = ScrollView(bar_width=10, scroll_type=["bars", "content"])
    table_layout = GridLayout(cols=1, spacing=5, size_hint_y=None, padding=5)
    table_layout.bind(minimum_height=table_layout.setter("height"))
    scroll.add_widget(table_layout)
    main_layout.add_widget(scroll)

    def _apply_background(widget, rgba):
        try:
            with widget.canvas.before:
                widget._sf6_bg_color = Color(*rgba)
                widget._sf6_bg_rect = Rectangle(pos=widget.pos, size=widget.size)

            def _update_bg(_inst, _value):
                if hasattr(widget, "_sf6_bg_rect"):
                    widget._sf6_bg_rect.pos = widget.pos
                    widget._sf6_bg_rect.size = widget.size

            widget.bind(pos=_update_bg, size=_update_bg)
        except Exception:
            pass

    def _make_section_header(title_text, subtitle_text=""):
        header_box = BoxLayout(size_hint_y=None, height=34, padding=(8, 0), spacing=8)
        _apply_background(header_box, (0.89, 0.92, 0.96, 1.0))
        header_box.add_widget(
            Label(
                text=f"[b]{title_text}[/b]",
                markup=True,
                halign="left",
                valign="middle",
                size_hint_x=0.6,
            )
        )
        header_box.add_widget(
            Label(
                text=subtitle_text,
                halign="right",
                valign="middle",
                size_hint_x=0.4,
            )
        )
        return header_box

    def _make_table_header():
        header = BoxLayout(size_hint_y=None, height=34, spacing=6, padding=(6, 0))
        _apply_background(header, (0.84, 0.88, 0.93, 1.0))
        columns = [
            (S["MESSAGES"].get("DATE_LABEL", "Ημερομηνία"), 0.17),
            (S["MESSAGES"].get("ELEMENT_LABEL", "Στοιχείο"), 0.31),
            (S["MESSAGES"].get("LEAKAGE_LABEL", "Διαρροή (kg)"), 0.13),
            ("Μεθοδολογία", 0.17),
            (S["MESSAGES"].get("RESPONSIBLE_LABEL", "Υπεύθυνος"), 0.16),
            ("", 0.06),
        ]
        for text, width in columns:
            header.add_widget(Label(text=text, bold=True, size_hint_x=width))
        return header

    def _open_maintenance_editor(maintenance_id):
        popup.dismiss()
        app.show_maintenance_menu(
            maintenance_id=maintenance_id,
            after_save_callback=lambda: show_sf6_management_popup(
                app,
                preselected_year=year_spinner.text,
                preselected_substation=(
                    None
                    if selected_substation["name"] == show_all_label
                    else selected_substation["name"]
                ),
            ),
        )

    def render_report(year_value: str):
        table_layout.clear_widgets()
        data = app._get_sf6_report_data(year_value)
        available_substations = data.get("available_substations") or []
        selected_name = selected_substation["name"]
        if (
            selected_name != show_all_label
            and selected_name not in available_substations
        ):
            selected_substation["name"] = show_all_label
            selected_name = show_all_label
        substation_filter_input.text = selected_name

        substation_rows = data.get("substation_rows") or {}
        substation_stats = data.get("substation_stats") or {}
        leakage_bands = data.get("leakage_bands") or {}

        if selected_name == show_all_label:
            displayed_groups = [
                (substation, substation_rows.get(substation, []))
                for substation in sorted(substation_rows)
            ]
            total_leakage = data["total_leakage"]
            installed_sf6 = data["installed_sf6"]
            percentage = data["percentage"]
            active_elements = data["active_elements"]
            active_substations = data["active_substations"]
            displayed_rows = [
                row for _substation, rows in displayed_groups for row in rows
            ]
        else:
            displayed_rows = substation_rows.get(selected_name, [])
            displayed_groups = [(selected_name, displayed_rows)]
            selected_stats = substation_stats.get(selected_name, {})
            total_leakage = sum((row.get("leakage") or 0.0) for row in displayed_rows)
            installed_sf6 = selected_stats.get("installed_sf6", 0.0)
            percentage = (total_leakage / installed_sf6 * 100) if installed_sf6 else 0.0
            active_elements = selected_stats.get("active_elements", 0)
            active_substations = 1 if selected_stats else 0

        summary_text = "\n".join(
            [
                (
                    f"Εγκατεστημένο SF6 (ενεργά): {installed_sf6:.2f} kg | "
                    f"Ενεργά στοιχεία SF6: {active_elements} | "
                    f"Υποσταθμοί με SF6: {active_substations}"
                ),
                (
                    f"Έτος: {year_value} | Διαρροές: {total_leakage:.2f} kg | "
                    f"Ποσοστό: {percentage:.2f}%"
                ),
                (f"Φίλτρο Υποσταθμού: {selected_name}"),
            ]
        )
        summary_label.text = summary_text

        if not displayed_rows:
            table_layout.add_widget(
                Label(
                    text=S["MESSAGES"].get(
                        "NO_LEAK_ENTRIES",
                        "Δεν υπάρχουν καταχωρήσεις διαρροών για το έτος.",
                    ),
                    size_hint_y=None,
                    height=30,
                )
            )
            return

        table_layout.add_widget(_make_table_header())

        for substation_name, rows in displayed_groups:
            installed_substation = substation_stats.get(substation_name, {}).get(
                "installed_sf6", 0.0
            )
            subtotal = sum((row.get("leakage") or 0.0) for row in rows)
            table_layout.add_widget(
                _make_section_header(
                    substation_name or "-",
                    f"Διαρροές: {subtotal:.2f} kg | Εγκατεστημένο SF6: {installed_substation:.2f} kg",
                )
            )
            for row in rows:
                leakage = row.get("leakage")
                leakage_text = "-" if leakage is None else f"{leakage:.2f}"
                row_layout = BoxLayout(
                    size_hint_y=None,
                    height=36,
                    spacing=6,
                    padding=(6, 0),
                )
                _apply_background(
                    row_layout,
                    _sf6_row_background_rgba(leakage, leakage_bands),
                )
                row_layout.add_widget(
                    Label(
                        text=_format_display_date(app, row.get("date_time")) or "-",
                        size_hint_x=0.17,
                        color=(0, 0, 0, 1),
                    )
                )
                row_layout.add_widget(
                    Label(
                        text=row.get("element") or "-",
                        size_hint_x=0.31,
                        color=(0, 0, 0, 1),
                    )
                )
                row_layout.add_widget(
                    Label(text=leakage_text, size_hint_x=0.13, color=(0, 0, 0, 1))
                )
                row_layout.add_widget(
                    Label(
                        text=row.get("methodology") or "-",
                        size_hint_x=0.17,
                        color=(0, 0, 0, 1),
                    )
                )
                row_layout.add_widget(
                    Label(
                        text=row.get("responsible") or "-",
                        size_hint_x=0.16,
                        color=(0, 0, 0, 1),
                    )
                )
                edit_btn = IconOnlyButton(
                    icon_type="edit",
                    icon_color=getattr(app, "theme", {}).get(
                        "primary", (0.2, 0.6, 1, 1)
                    ),
                    size=(30, 30),
                    tooltip=S["MESSAGES"].get("TOOLTIP_EDIT", "Επεξεργασία"),
                    size_hint_x=0.06,
                )
                edit_btn.bind(
                    on_press=lambda _instance, maintenance_id=row.get("maintenance_id"): (
                        _open_maintenance_editor(maintenance_id)
                    )
                )
                row_layout.add_widget(edit_btn)
                table_layout.add_widget(row_layout)

            table_layout.add_widget(Widget(size_hint_y=None, height=6))

    def _open_substation_picker(*_args):
        chooser_rows = [(-1, show_all_label)] + [
            (index, substation_name)
            for index, substation_name in enumerate(
                data_cache.get("available_substations") or [], start=1
            )
        ]

        def _on_select(substation_name):
            selected_substation["name"] = substation_name or show_all_label
            substation_filter_input.text = selected_substation["name"]
            render_report(year_spinner.text)

        app._show_substation_selection_window_with_callback(
            popup,
            chooser_rows,
            on_select=_on_select,
            title=S["MESSAGES"].get("FILTER_SUBSTATION", "Φίλτρο Υποσταθμού"),
        )

    def _reset_substation_filter(*_args):
        selected_substation["name"] = show_all_label
        substation_filter_input.text = show_all_label
        render_report(year_spinner.text)

    data_cache = {}

    def _refresh_and_cache(*_args):
        data_cache.clear()
        data_cache.update(app._get_sf6_report_data(year_spinner.text))
        render_report(year_spinner.text)

    def handle_print(*_args):
        try:
            pdf_path = generate_sf6_leak_report(app.conn, year_spinner.text)

            def _open_pdf():
                try:
                    if sys.platform == "win32":
                        os.startfile(pdf_path)
                    elif sys.platform == "darwin":
                        subprocess.call(["open", pdf_path])
                    else:
                        subprocess.call(["xdg-open", pdf_path])
                except Exception:
                    # ignore open-errors; show_message_popup already reports success
                    pass

            show_message_popup(
                S["TITLES"]["SUCCESS"],
                S["MESSAGES"]
                .get("PDF_CREATED", "Το PDF δημιουργήθηκε:\n{path}")
                .format(path=pdf_path),
                callback=_open_pdf,
            )
        except Exception as exc:
            show_message_popup(
                S["TITLES"]["ERROR"],
                S["MESSAGES"]
                .get("PDF_CREATE_FAILED", "Αποτυχία δημιουργίας PDF:\n{err}")
                .format(err=str(exc)),
            )

    def handle_excel(*_args):
        try:
            excel_path = app._export_sf6_excel(
                year_spinner.text,
                None
                if selected_substation["name"] == show_all_label
                else selected_substation["name"],
            )

            def _open_excel():
                try:
                    if sys.platform == "win32":
                        os.startfile(excel_path)
                    elif sys.platform == "darwin":
                        subprocess.call(["open", excel_path])
                    else:
                        subprocess.call(["xdg-open", excel_path])
                except Exception:
                    pass

            show_message_popup(
                S["TITLES"]["SUCCESS"],
                S["MESSAGES"]
                .get("EXCEL_CREATED", "Το Excel δημιουργήθηκε:\n{path}")
                .format(path=excel_path),
                callback=_open_excel,
            )
        except Exception as exc:
            show_message_popup(
                S["TITLES"]["ERROR"],
                S["MESSAGES"]
                .get("EXCEL_CREATE_FAILED", "Αποτυχία δημιουργίας Excel:\n{err}")
                .format(err=str(exc)),
            )

    refresh_btn.bind(on_press=_refresh_and_cache)
    year_spinner.bind(text=lambda _s, _t: _refresh_and_cache())
    print_btn.bind(on_press=handle_print)
    excel_btn.bind(on_press=handle_excel)
    substation_picker_btn.bind(on_press=_open_substation_picker)
    reset_filter_btn.bind(on_press=_reset_substation_filter)

    _refresh_and_cache()

    close_btn = Button(text=S["BUTTONS"]["CLOSE"], size_hint_y=None, height=40)
    close_btn.bind(on_press=popup.dismiss)
    main_layout.add_widget(close_btn)

    popup.content = main_layout
    popup.open()


def generate_pdf_report(app, maintenance_id, element_id, element_name):
    """Generate PDF maintenance report (UI wrapper)."""
    try:
        c = app.conn.cursor()
        c.execute(
            "SELECT gate FROM elements WHERE id=?",
            (element_id,),
        )
        elem_row = c.fetchone()
        gate_value = (
            elem_row[0]
            if elem_row and isinstance(elem_row, (tuple, list))
            else (elem_row["gate"] if elem_row else None)
        )

        def _open_pdf(path: str):
            try:
                if sys.platform == "win32":
                    os.startfile(path)
                elif sys.platform == "darwin":
                    subprocess.call(["open", path])
                else:
                    subprocess.call(["xdg-open", path])
            except Exception:
                pass

        def _show_success(path: str, action_taken: str):
            if action_taken == "replaced":
                msg = (
                    f'Το αρχείο PDF για το στοιχείο "{element_name}"\n'
                    "αντικαταστάθηκε επιτυχώς!\n\n"
                    f"Αποθηκεύτηκε στο:\n{path}"
                )
            elif action_taken == "opened":
                msg = (
                    f'Το αρχείο PDF για το στοιχείο "{element_name}"\n'
                    "υπάρχει ήδη.\n\n"
                    f"Το αρχείο είναι:\n{path}"
                )
            else:
                msg = (
                    f'Το αρχείο PDF για το στοιχείο "{element_name}"\n'
                    "δημιουργήθηκε επιτυχώς!\n\n"
                    f"Αποθηκεύτηκε στο:\n{path}"
                )
            show_message_popup(
                "PDF Έτοιμο",
                msg,
                callback=lambda: _open_pdf(path),
            )

        result = safe_generate_and_store_report(
            app.conn,
            maintenance_id=maintenance_id,
            element_id=element_id,
            gate_value=gate_value,
            db_path=getattr(app, "db_path", None),
        )

        if result.get("action_taken") == "prompt_user" and result.get("path"):
            existing_path = result["path"]

            def _replace_existing():
                replace_result = safe_generate_and_store_report(
                    app.conn,
                    maintenance_id=maintenance_id,
                    element_id=element_id,
                    gate_value=gate_value,
                    db_path=getattr(app, "db_path", None),
                    user_prompted_action="replace",
                )
                if not replace_result.get("success"):
                    show_message_popup(
                        S["TITLES"].get("ERROR", "Σφάλμα"),
                        replace_result.get("message") or "Αποτυχία αντικατάστασης PDF.",
                    )
                    return
                _show_success(
                    replace_result["path"],
                    replace_result.get("action_taken") or "replaced",
                )

            show_confirm(
                S["TITLES"].get("CONFIRM", "Επιβεβαίωση"),
                (
                    f"Το PDF υπάρχει ήδη:\n{existing_path}\n\n"
                    "Θέλετε αντικατάσταση ή άνοιγμα του υπάρχοντος;"
                ),
                yes_callback=_replace_existing,
                yes_text="ΑΝΤΙΚΑΤΑΣΤΑΣΗ",
                no_text="ΑΝΟΙΓΜΑ",
                no_callback=lambda: open_file(existing_path),
                cancel_text=S["BUTTONS"].get("CANCEL", "Ακύρωση"),
            )
            return

        if not result.get("success"):
            show_message_popup(
                S["TITLES"].get("ERROR", "Σφάλμα"),
                result.get("message") or "Αποτυχία δημιουργίας PDF.",
            )
            return

        _show_success(result["path"], result.get("action_taken") or "created")

    except Exception as e:
        show_message_popup(S["TITLES"]["ERROR"], f"Αποτυχία δημιουργίας PDF:\n{str(e)}")


def export_full_db_ui(app, parent_popup=None):
    """Invoke `excel_io.export_full_db` using `app.conn`.

    This keeps the UI wiring for the "Export DB" button in `reports.py`.
    """
    try:
        from excel_io import export_full_db
    except Exception:
        return False

    try:
        if parent_popup:
            parent_popup.dismiss()
    except Exception:
        pass

    return export_full_db(app.conn)


def open_file(
    path,
    *,
    not_found_message="Το αρχείο δεν βρέθηκε!",
    error_title="Σφάλμα",
    error_prefix="Αποτυχία ανοίγματος αρχείου:\n",
):
    """Open a file with the platform default application and handle errors with popups.

    Returns True on success, False on failure.
    """
    import subprocess
    import sys

    from popups import show_message_popup

    existing_path = _win_existing_path(path)
    if not existing_path:
        show_message_popup(error_title, not_found_message)
        return False
    try:
        if sys.platform == "win32":
            open_target = existing_path
            suffix = os.path.splitext(existing_path)[1].lower()
            if len(os.path.abspath(existing_path)) >= 220 and suffix in {
                ".xls",
                ".xlsx",
                ".xlsm",
                ".doc",
                ".docx",
                ".ppt",
                ".pptx",
            }:
                temp_copy = _short_temp_open_copy(path, existing_path=existing_path)
                if temp_copy:
                    open_target = temp_copy
            try:
                os.startfile(open_target)
            except Exception:
                fallback_target = (
                    open_target if open_target != existing_path else existing_path
                )
                os.startfile(fallback_target)
        elif sys.platform == "darwin":
            subprocess.call(["open", existing_path])
        else:
            subprocess.call(["xdg-open", existing_path])
        return True
    except Exception as exc:
        show_message_popup(error_title, f"{error_prefix}{str(exc)}")
        return False


def show_confirm(
    title: str,
    message: str,
    yes_callback=None,
    yes_text="ΝΑΙ",
    no_text="ΟΧΙ",
    yes_color=None,
    size_hint=(0.6, 0.3),
    no_callback=None,
    cancel_text=None,
):
    """Show a standardized confirmation popup and call `yes_callback` when confirmed.

    The callback is called after the popup is dismissed.
    """
    Popup = __import__("kivy.uix.popup", fromlist=["Popup"]).Popup
    BoxLayout = __import__("kivy.uix.boxlayout", fromlist=["BoxLayout"]).BoxLayout
    Label = __import__("kivy.uix.label", fromlist=["Label"]).Label
    Button = __import__("kivy.uix.button", fromlist=["Button"]).Button
    ScrollView = __import__("kivy.uix.scrollview", fromlist=["ScrollView"]).ScrollView

    popup = Popup(title=title, size_hint=size_hint)
    layout = BoxLayout(orientation="vertical", padding=10, spacing=10)

    scroll = ScrollView(size_hint_y=0.65, bar_width=10, scroll_type=["bars", "content"])
    warning_label = Label(
        text=message,
        size_hint_y=None,
        halign="left",
        valign="top",
    )

    def _sync_message_size(_instance=None, _value=None):
        try:
            label_width = max(10, scroll.width - 12)
            warning_label.text_size = (label_width, None)
            warning_label.texture_update()
            warning_label.height = max(60, warning_label.texture_size[1] + 8)
        except Exception:
            pass

    scroll.bind(size=_sync_message_size)
    warning_label.bind(texture_size=_sync_message_size)
    _sync_message_size()

    scroll.add_widget(warning_label)
    layout.add_widget(scroll)

    buttons_layout = BoxLayout(size_hint_y=0.3, spacing=10)

    def _on_yes(_instance=None):
        try:
            popup.dismiss()
        except Exception:
            pass
        try:
            if yes_callback:
                yes_callback()
        except Exception:
            pass

    yes_btn = Button(text=yes_text)
    if yes_color:
        try:
            yes_btn.color = yes_color
        except Exception:
            pass
    yes_btn.bind(on_press=lambda x: _on_yes(x))
    buttons_layout.add_widget(yes_btn)

    no_btn = Button(text=no_text)

    def _on_no(_instance=None):
        try:
            popup.dismiss()
        except Exception:
            pass
        try:
            if no_callback:
                no_callback()
        except Exception:
            pass

    no_btn.bind(on_press=lambda x: _on_no(x))
    buttons_layout.add_widget(no_btn)

    if cancel_text:
        cancel_btn = Button(text=cancel_text)
        cancel_btn.bind(on_press=lambda _x: popup.dismiss())
        buttons_layout.add_widget(cancel_btn)

    layout.add_widget(buttons_layout)
    popup.content = layout
    popup.open()
    return popup


def _get_sf6_report_data(app, year: str):
    """Return SF6 report data dictionary for `year` using `app.conn`."""
    c = app.conn.cursor()
    year_prefix = f"{year}%"
    sf6_filter = "me.sf6_leakage_kg IS NOT NULL"

    c.execute(
        f"""
             SELECT m.id, m.date_time, s.id, s.name, e.id, e.name, e.element_type, me.sf6_leakage_kg,
                 me.sf6_leak_methodology, p.name
        FROM maintenance_elements me
        JOIN maintenance m ON me.maintenance_id = m.id
        JOIN elements e ON me.element_id = e.id
        JOIN substations s ON m.substation_id = s.id
        LEFT JOIN people p ON m.responsible_id = p.id
        WHERE e.breaker_category = 'SF6'
          AND m.date_time LIKE ?
                    AND {sf6_filter}
        ORDER BY m.date_time ASC
                """,
        (year_prefix,),
    )
    leak_rows = c.fetchall()

    total_leakage = 0.0
    rows = []
    for (
        maintenance_id,
        date_time,
        substation_id,
        sub_name,
        element_id,
        elem_name,
        elem_type,
        leakage,
        methodology,
        responsible_name,
    ) in leak_rows:
        total_leakage += leakage or 0.0
        rows.append(
            {
                "maintenance_id": maintenance_id,
                "date_time": date_time,
                "substation_id": substation_id,
                "substation": sub_name or "-",
                "element_id": element_id,
                "element": elem_name or "-",
                "element_type": elem_type or "-",
                "leakage": leakage,
                "methodology": methodology or "",
                "responsible": responsible_name or "-",
            }
        )

    substation_rows = {}
    for row in rows:
        substation_rows.setdefault(row["substation"], []).append(row)

    c.execute("""
        SELECT
            COUNT(*),
            SUM(COALESCE(em.sf6_capacity_kg, 0))
        FROM elements e
        LEFT JOIN element_models em ON e.element_model_id = em.id
        WHERE e.operating_status = 'Ενεργή'
          AND e.breaker_category = 'SF6'
          AND e.element_type IN ('Διακόπτης ΥΤ', 'Διακόπτης ΜΤ')
        """)
    counts = c.fetchone()
    total_elements = counts[0] or 0
    installed_sf6 = counts[1] or 0.0

    c.execute("""
        SELECT COUNT(*), COUNT(DISTINCT s.id)
        FROM elements e
        JOIN substations s ON e.substation_id = s.id
        WHERE e.operating_status = 'Ενεργή'
          AND e.breaker_category = 'SF6'
          AND e.element_type IN ('Διακόπτης ΥΤ', 'Διακόπτης ΜΤ')
        """)
    active_counts = c.fetchone()
    active_elements = active_counts[0] or 0
    active_substations = active_counts[1] or 0

    c.execute("""
        SELECT s.name, COUNT(*), SUM(COALESCE(em.sf6_capacity_kg, 0))
        FROM elements e
        JOIN substations s ON e.substation_id = s.id
        LEFT JOIN element_models em ON e.element_model_id = em.id
        WHERE e.operating_status = 'Ενεργή'
          AND e.breaker_category = 'SF6'
          AND e.element_type IN ('Διακόπτης ΥΤ', 'Διακόπτης ΜΤ')
        GROUP BY s.name
        """)
    substation_stats = {
        row[0]: {
            "active_elements": row[1] or 0,
            "installed_sf6": row[2] or 0.0,
        }
        for row in c.fetchall()
    }

    available_substations = sorted(
        {substation for substation in substation_rows}
        | {substation for substation in substation_stats}
    )

    percentage = (total_leakage / installed_sf6 * 100) if installed_sf6 else 0.0

    return {
        "total_leakage": total_leakage,
        "installed_sf6": installed_sf6,
        "percentage": percentage,
        "total_elements": total_elements,
        "active_elements": active_elements,
        "active_substations": active_substations,
        "rows": rows,
        "substation_rows": substation_rows,
        "substation_installed": {
            name: stats.get("installed_sf6", 0.0)
            for name, stats in substation_stats.items()
        },
        "substation_stats": substation_stats,
        "available_substations": available_substations,
        "leakage_bands": _compute_sf6_leakage_bands(rows),
    }


def _export_sf6_excel(app, year: str, substation_filter=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except Exception as exc:
        raise RuntimeError(
            "Δεν βρέθηκε το πακέτο openpyxl. Εγκαταστήστε το για εξαγωγή Excel."
        ) from exc

    data = _get_sf6_report_data(app, year)
    selected_substation_stats = None
    if substation_filter:
        selected_rows = data["substation_rows"].get(substation_filter, [])
        selected_substation_stats = data.get("substation_stats", {}).get(
            substation_filter, {}
        )
        data = {
            **data,
            "rows": selected_rows,
            "substation_rows": {substation_filter: selected_rows},
            "total_leakage": sum((row.get("leakage") or 0.0) for row in selected_rows),
            "installed_sf6": selected_substation_stats.get("installed_sf6", 0.0),
            "active_elements": selected_substation_stats.get("active_elements", 0),
            "active_substations": 1 if selected_substation_stats else 0,
        }
        data["percentage"] = (
            (data["total_leakage"] / data["installed_sf6"] * 100)
            if data["installed_sf6"]
            else 0.0
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Σύνοψη"

    grey_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    blue_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    summary_titles = [
        "ΣΥΝΟΛΙΚΗ ΕΓΚΑΤΕΣΤΗΜΕΝΗ ΠΟΣΟΤΗΤΑ (kg)",
        f"ΔΙΑΡΡΟΕΣ {year} (kg)",
        f"ΠΟΣΟΣΤΟ ΔΙΑΡΡΟΩΝ {year}",
    ]
    summary_values = [
        f"{data['installed_sf6']:.2f}",
        f"{data['total_leakage']:.2f}",
        f"{data['percentage']:.2f}%",
    ]

    for col_idx, title in enumerate(summary_titles, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = grey_fill
        cell.font = bold_font
        cell.alignment = center
        cell.border = border

        val_cell = ws.cell(row=2, column=col_idx, value=summary_values[col_idx - 1])
        val_cell.alignment = center
        val_cell.border = border

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24

    for col in range(1, 4):
        ws.column_dimensions[chr(64 + col)].width = 35

    substation_sums = []
    for substation, rows in data["substation_rows"].items():
        total = sum([r.get("leakage") or 0 for r in rows])
        if total > 0:
            substation_sums.append((substation, total))

    start_row = 4
    if substation_sums:
        ws.cell(row=start_row, column=1, value="Υποσταθμός").font = bold_font
        ws.cell(row=start_row, column=2, value="Σύνολο Διαρροών (kg)").font = bold_font
        for col_idx in (1, 2):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.alignment = center
            cell.border = border

        row_ptr = start_row + 1
        for substation, total in sorted(substation_sums, key=lambda x: x[0] or ""):
            ws.cell(row=row_ptr, column=1, value=substation or "-").border = border
            ws.cell(row=row_ptr, column=2, value=f"{total:.2f}").border = border
            ws.cell(row=row_ptr, column=1).alignment = center
            ws.cell(row=row_ptr, column=2).alignment = center
            row_ptr += 1

    for substation, rows in data["substation_rows"].items():
        sheet_title = substation[:31] if substation else "Υποσταθμός"
        if sheet_title in wb.sheetnames:
            suffix = 1
            base = sheet_title[:28]
            while f"{base}_{suffix}" in wb.sheetnames:
                suffix += 1
            sheet_title = f"{base}_{suffix}"

        ws_sub = wb.create_sheet(title=sheet_title)
        ws_sub.merge_cells("A1:J1")
        title_cell = ws_sub["A1"]
        title_cell.value = "ΠΙΝΑΚΑΣ 4: ΠΗΓΗ ΕΚΠΟΜΠΩΝ ΑΠΌ ΕΞΟΠΛΙΣΜΟ ΧΡΗΣΗΣ SF6"
        title_cell.alignment = center

        for col_idx in range(1, 11):
            cell = ws_sub.cell(row=1, column=col_idx)
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = center
            cell.border = border

        headers = [
            "Α/Α",
            "ΒΟΚ ή ΠΕΡΙΟΧΗ",
            "ΕΓΚΑΤΑΣΤΑΣΗ (Πχ. Όνομα Υ/Σ)",
            "ΜΟΝΑΔΑ ΜΕΤΡΗΣΗΣ",
            "ΠΛΗΡΩΣΗ Ή ΑΝΤΙΚΑΤΑΣΤΑΣΗ (ΜΕΘΟΔΟΛΟΓΙΑ)",
            "ΣΥΝΟΛΙΚΗ ΕΓΚΑΤΕΣΤΗΜΕΝΗ ΠΟΣΟΤΗΤΑ (kg)",
            "ΠΟΣΟΤΗΤΑ ΔΙΑΡΡΟΩΝ (kg)",
            "ΗΜ/ΝΙΑ",
            "ΥΠΕΥΘΥΝΟΣ ΣΥΝΕΡΓΕΙΟΥ",
            "ΥΠΟΓΡΑΦΗ",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws_sub.cell(row=2, column=col_idx, value=header)
            cell.font = bold_font
            cell.alignment = center
            cell.border = border

        installed_sub = data["substation_installed"].get(substation, 0.0)
        start_row = 3
        for idx, row in enumerate(rows, start=1):
            values = [
                idx,
                "ΔΕΕΔ",
                substation,
                "kg",
                row.get("methodology", "") or "",
                f"{installed_sub:.2f}",
                f"{row['leakage']:.2f}",
                _format_display_date(app, row.get("date_time")) or "-",
                row.get("responsible", "-") or "-",
                "",
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws_sub.cell(row=start_row, column=col_idx, value=value)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                cell.border = border
            start_row += 1

        ws_sub.row_dimensions[1].height = 30
        ws_sub.row_dimensions[2].height = 28
        for col_idx in range(1, 11):
            ws_sub.column_dimensions[chr(64 + col_idx)].width = 22

    reports_dir = os.path.join(os.path.dirname(__file__), "reports")
    os.makedirs(reports_dir, exist_ok=True)
    timestamp = __import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(reports_dir, f"SF6_Leakages_{year}_{timestamp}.xlsx")
    wb.save(output_path)
    return output_path


def create_substations_template(base_dir: str) -> tuple[bool, str]:
    """Create substations import template. Returns (success, message/path)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception:
        return False, "openpyxl δεν είναι εγκατεστημένο!"

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Substations"

        headers = ["Name", "Location", "Adoption Date"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 15

        examples = [
            ("Υποσταθμός Α", "https://maps.google.com/?q=example1", "2025-01-15"),
            ("Υποσταθμός Β", "https://maps.google.com/?q=example2", "2025-01-20"),
        ]
        for idx, (name, location, date) in enumerate(examples, 2):
            ws.cell(row=idx, column=1, value=name)
            ws.cell(row=idx, column=2, value=location)
            ws.cell(row=idx, column=3, value=date)

        template_path = os.path.join(base_dir, "substations_import_template.xlsx")
        wb.save(template_path)
        return True, template_path
    except Exception as exc:  # pragma: no cover - UI surface
        return False, f"Σφάλμα: {exc}"


def create_elements_template(base_dir: str) -> tuple[bool, str]:
    """Create elements import template. Returns (success, message/path)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception:
        return False, "openpyxl δεν είναι εγκατεστημένο!"

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Elements"

        # Add version info in first row
        ws.cell(row=1, column=1, value="TEMPLATE_VERSION: v2.0")
        ws.cell(row=1, column=1).font = Font(italic=True, color="999999")
        ws.row_dimensions[1].height = 15

        headers = [
            "Substation Name",
            "Element Type",
            "Name",
            "Serial Number",
            "Maintenance Date",
            "Τύπος Διακόπτη",
            "Breaker Role",
            "Operating Status",
            "Gate",
            "Model Name",
            "Model Manufacturer",
            "Model Installation Space",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="70AD47", end_color="70AD47", fill_type="solid"
            )

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 20
        ws.column_dimensions["I"].width = 15
        ws.column_dimensions["J"].width = 15
        ws.column_dimensions["K"].width = 20
        ws.column_dimensions["L"].width = 20

        examples = [
            (
                "Υποσταθμός Α",
                "Διακόπτης ΜΤ",
                "Main Breaker",
                "SN-001",
                "2025-01-20",
                "SF6",
                "Κεντρικός",
                "Ενεργή",
                "ΠΥΛΗ 1",
                "SF6-400",
                "ABB",
                "Εσωτερικού",
            ),
            (
                "Υποσταθμός Α",
                "Μετασχηματιστής 150/20KV",
                "Transformer 1",
                "SN-002",
                "2025-01-18",
                "",
                "",
                "Ενεργή",
                "ΠΥΛΗ 1",
                "GEAFOL",
                "Siemens",
                "Εξωτερικού",
            ),
        ]
        for idx, row_data in enumerate(examples, 3):
            for col, value in enumerate(row_data, 1):
                ws.cell(row=idx, column=col, value=value)

        template_path = os.path.join(base_dir, "elements_import_template.xlsx")
        wb.save(template_path)
        return True, template_path
    except Exception as exc:  # pragma: no cover - UI surface
        return False, f"Σφάλμα: {exc}"
