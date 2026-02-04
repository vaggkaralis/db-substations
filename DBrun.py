import kivy
kivy.require('2.0.0')
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.spinner import Spinner
from kivy.uix.checkbox import CheckBox
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.image import Image
from kivy.uix.behaviors import ButtonBehavior
from kivy.uix.widget import Widget
from kivy.properties import StringProperty, ListProperty
from kivy.graphics import Color, Rectangle, Ellipse, Line
from kivy.core.window import Window
from kivy.core.image import Image as CoreImage
from kivy.core.clipboard import Clipboard
from kivy.clock import Clock
import webbrowser
import os
import re
import subprocess
import sys
import unicodedata
from datetime import datetime
import json

# Maximize window on startup
Window.maximize()

from database import init_db
from importers import (
    import_elements_from_csv,
    import_elements_from_excel,
    import_substations_from_csv,
    import_substations_from_excel,
)
from popups import show_message_popup
from templates import create_elements_template, create_substations_template
from model_management import show_models_management
from pdf_reports import generate_maintenance_report, generate_inspection_report, generate_sf6_leak_report
from import_wizard import ColumnMappingPopup, DataValidationPopup
from email_eml_parser import parse_eml_file


class IconWidget(Widget):
    """Simple vector pictogram drawn on canvas."""
    icon_type = StringProperty('database')
    icon_color = ListProperty([1, 1, 1, 1])

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.bind(pos=self._redraw, size=self._redraw, icon_type=self._redraw, icon_color=self._redraw)

    def _redraw(self, *_args):
        self.canvas.clear()
        with self.canvas:
            Color(*self.icon_color)
            x, y = self.x, self.y
            w, h = self.width, self.height
            if w <= 0 or h <= 0:
                return
            pad = min(w, h) * 0.12

            line_w = max(1.05, min(w, h) * 0.042)

            if self.icon_type == 'database':
                Line(ellipse=(x + pad, y + h * 0.58, w - 2 * pad, h * 0.35), width=line_w)
                Line(rectangle=(x + pad, y + pad, w - 2 * pad, h * 0.58), width=line_w)
                Line(ellipse=(x + pad, y + pad - h * 0.12, w - 2 * pad, h * 0.24), width=line_w)
                Line(points=[x + pad, y + h * 0.58, x + w - pad, y + h * 0.58], width=line_w)
            elif self.icon_type == 'import':
                Line(rectangle=(x + pad, y + pad, w - 2 * pad, h - 2 * pad), width=line_w)
                Line(points=[x + w * 0.5, y + h * 0.75, x + w * 0.5, y + h * 0.35], width=line_w)
                Line(points=[x + w * 0.38, y + h * 0.48, x + w * 0.5, y + h * 0.35, x + w * 0.62, y + h * 0.48], width=line_w)
            elif self.icon_type == 'models':
                Line(rectangle=(x + pad * 1.2, y + pad * 1.2, w - 2.4 * pad, h * 0.28), width=line_w)
                Line(rectangle=(x + pad, y + h * 0.38, w - 2 * pad, h * 0.28), width=line_w)
                Line(rectangle=(x + pad * 1.2, y + h * 0.6, w - 2.4 * pad, h * 0.28), width=line_w)
            elif self.icon_type == 'people':
                Line(circle=(x + w * 0.5, y + h * 0.7, w * 0.18), width=line_w)
                Line(rectangle=(x + w * 0.26, y + pad, w * 0.48, h * 0.35), width=line_w)
            elif self.icon_type == 'maintenance':
                Line(circle=(x + w * 0.35, y + h * 0.6, w * 0.15), width=line_w)
                Line(points=[x + w * 0.5, y + h * 0.3, x + w * 0.82, y + h * 0.62], width=line_w)
                Line(points=[x + w * 0.68, y + h * 0.5, x + w * 0.82, y + h * 0.62, x + w * 0.66, y + h * 0.66], width=line_w)
            elif self.icon_type == 'inspection':
                Line(circle=(x + w * 0.4, y + h * 0.55, w * 0.2), width=line_w)
                Line(points=[x + w * 0.56, y + h * 0.38, x + w * 0.82, y + h * 0.12], width=line_w)
            elif self.icon_type == 'sf6':
                # Gas cylinder pictogram
                body_x = x + w * 0.28
                body_w = w * 0.44
                body_y = y + h * 0.2
                body_h = h * 0.6
                Line(rectangle=(body_x, body_y, body_w, body_h), width=line_w)
                Line(ellipse=(body_x, body_y + body_h - h * 0.12, body_w, h * 0.2), width=line_w)
                Line(ellipse=(body_x, body_y - h * 0.08, body_w, h * 0.16), width=line_w)
                # Valve
                Line(circle=(x + w * 0.5, y + h * 0.86, w * 0.06), width=line_w)
                Line(points=[x + w * 0.5, y + h * 0.8, x + w * 0.5, y + h * 0.74], width=line_w)
            elif self.icon_type == 'isolation':
                Line(rectangle=(x + w * 0.26, y + pad, w * 0.48, h * 0.45), width=line_w)
                Line(points=[x + w * 0.32, y + h * 0.48, x + w * 0.32, y + h * 0.7, x + w * 0.68, y + h * 0.7, x + w * 0.68, y + h * 0.48], width=line_w)
            elif self.icon_type == 'info':
                Line(circle=(x + w * 0.5, y + h * 0.5, w * 0.3), width=line_w)
                Line(points=[x + w * 0.5, y + h * 0.38, x + w * 0.5, y + h * 0.62], width=line_w)
                Ellipse(pos=(x + w * 0.46, y + h * 0.68), size=(w * 0.08, h * 0.08))


class ShiftSelectableTextInput(TextInput):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._shift_select_anchor = None

    def on_touch_down(self, touch):
        if self.collide_point(*touch.pos):
            if not self.focus:
                self.focus = True
            if 'shift' in Window.modifiers:
                if self._shift_select_anchor is None:
                    self._shift_select_anchor = self.cursor_index()
                self.cursor = self.get_cursor_from_xy(*touch.pos)
                self.select_text(self._shift_select_anchor, self.cursor_index())
                return True

        result = super().on_touch_down(touch)
        if self.collide_point(*touch.pos) and 'shift' not in Window.modifiers:
            self._shift_select_anchor = self.cursor_index()
        return result


class IconButton(ButtonBehavior, BoxLayout):
    """Button with a simple pictogram and text."""
    text = StringProperty('')
    icon_type = StringProperty('database')
    bg_color = ListProperty([0.05, 0.18, 0.36, 1])
    bg_color_down = ListProperty([0.03, 0.12, 0.25, 1])
    text_color = ListProperty([1, 1, 1, 1])

    def __init__(self, **kwargs):
        theme = kwargs.pop('theme', None)
        super().__init__(**kwargs)
        if theme:
            self.bg_color = list(theme.get('primary', self.bg_color))
            self.bg_color_down = list(theme.get('primary_dark', self.bg_color_down))
            self.text_color = list(theme.get('text_on_primary', self.text_color))

        self.orientation = 'horizontal'
        self.spacing = 10
        self.padding = (12, 8)

        self.icon = IconWidget(icon_type=self.icon_type, icon_color=self.text_color, size_hint=(None, None))
        self.icon.size = (23, 23)
        self.icon.pos_hint = {'center_y': 0.5}
        self.label = Label(text=self.text, color=self.text_color, halign='left', valign='middle')
        self.label.font_size = '26sp'
        self.label.bind(size=self._sync_text_size)

        self.add_widget(self.icon)
        self.add_widget(self.label)

        with self.canvas.before:
            self._bg_color_inst = Color(*self.bg_color)
            self._bg_rect = Rectangle(pos=self.pos, size=self.size)

        self.bind(pos=self._update_bg, size=self._update_bg)
        self.bind(pos=self._update_icon_pos, size=self._update_icon_pos)
        self.bind(size=self._update_icon_size)
        self.bind(text=self._update_text)
        self.bind(icon_type=self._update_icon)
        self.bind(text_color=self._update_colors)

    def _sync_text_size(self, _instance, _value):
        self.label.text_size = (self.label.width, self.label.height)

    def _update_icon_size(self, *_args):
        icon_dim = max(22, int(self.height * 0.6))
        icon_dim = int(icon_dim * 0.64)
        self.icon.size = (icon_dim, icon_dim)

    def _update_icon_pos(self, *_args):
        self.icon.center_y = self.center_y

    def _update_bg(self, *_args):
        self._bg_rect.pos = self.pos
        self._bg_rect.size = self.size

    def _update_text(self, *_args):
        self.label.text = self.text

    def _update_icon(self, *_args):
        self.icon.icon_type = self.icon_type

    def _update_colors(self, *_args):
        self.label.color = self.text_color
        self.icon.icon_color = self.text_color

    def on_press(self):
        self._bg_color_inst.rgba = self.bg_color_down

    def on_release(self):
        self._bg_color_inst.rgba = self.bg_color

class SubstationApp(App):
    # Define element types as a class variable
    ELEMENT_TYPES = [
        'Διακόπτης ΥΤ',
        'Διακόπτης ΜΤ',
        'Μετασχηματιστής 150/20KV', 
        'Motor Drive',
        'Μ/Σ Εγχύσεως',
        'Μ/Σ Έντασης',
        'Μ/Σ Τάσης',
        'Μ/Σ ΧΤ/ΜΤ (ΒΜΣ)',
        'Αποζεύκτης',
        'Ασφαλειοαποζεύκτης',
        'Γειωτής',
        'Συστοιχία Πυκνωτών',
        'Αντίσταση Κόμβου',
        'Αλεξικέραυνο',
        'Συστοιχία Συσσωρευτών'
    ]
    BREAKER_CATEGORIES_ALL = ['SF6', 'Κενού', 'Πτωχού Ελαίου', 'Ελαίου']  # All breaker categories
    BREAKER_CATEGORIES_HV = ['SF6', 'Κενού', 'Ελαίου']  # HV breaker categories
    BREAKER_CATEGORIES_MV = ['SF6', 'Κενού', 'Πτωχού Ελαίου', 'Ελαίου']  # MV breaker categories
    BREAKER_TYPES = ['Κεντρικός', 'Γραμμής', 'Διασυνδετικός', 'Διακόπτης Πυκνωτών']  # Main, Line, Interconnection, or Capacitor breaker
    OPERATING_STATUS = ['Ενεργή', 'Ανενεργή']
    INSTALLATION_SPACE = ['Εσωτερικός', 'Εξωτερικός']
    VOLTAGE_LEVELS = ['(Κενό)', '150/20KV', '20KV', '150KV', '20KV/400V']
    THEME_FALLBACK = {
        'primary': (0.05, 0.36, 0.64, 1),
        'primary_dark': (0.03, 0.28, 0.5, 1),
        'accent': (0.12, 0.52, 0.86, 1),
        'background': (0.97, 0.98, 0.99, 1),
        'popup_bg': (1, 1, 1, 1),
        'input_bg': (1, 1, 1, 1),
        'text': (0.12, 0.12, 0.12, 1),
        'text_on_primary': (1, 1, 1, 1),
    }
    # Central definition of element fields for easy future extension
    ELEMENT_FIELD_DEFS = [
        {'key': 'name', 'label': 'Όνομα Στοιχείου', 'type': 'text', 'hint': 'Όνομα Στοιχείου'},
        {'key': 'serial_number', 'label': 'Σειριακός Αριθμός', 'type': 'text', 'hint': 'Σειριακός Αριθμός'},
        {'key': 'manufacture_year', 'label': 'Έτος κατασκευής', 'type': 'text', 'hint': 'YYYY'},
        {'key': 'maintenance_date', 'label': 'Τελευταία Συντ.', 'type': 'text', 'hint': 'YYYY-MM-DD'},
        {'key': 'manufacturer', 'label': 'Κατασκευαστής', 'type': 'text', 'hint': 'Κατασκευαστής'},
        {'key': 'model', 'label': 'Μοντέλο', 'type': 'text', 'hint': 'Μοντέλο'},
        {'key': 'model_version', 'label': 'Έκδοση Μοντέλου', 'type': 'text', 'hint': 'Έκδοση'},
        {'key': 'installation_space', 'label': 'Χώρος Εγκατ.', 'type': 'spinner', 'values': INSTALLATION_SPACE},
        {'key': 'operating_status', 'label': 'Λειτ. Κατάσταση', 'type': 'spinner', 'values': OPERATING_STATUS},
        {'key': 'maintenance_cycle', 'label': 'Κύκλος Συντ.', 'type': 'text', 'hint': 'Αριθμός'},
    ]

    def _get_breaker_categories_for_element_type(self, element_type: str):
        if element_type == 'Διακόπτης ΜΤ':
            return list(self.BREAKER_CATEGORIES_MV)
        if element_type == 'Διακόπτης ΥΤ':
            return list(self.BREAKER_CATEGORIES_HV)
        return list(self.BREAKER_CATEGORIES_ALL)
    
    def build(self):
        self.title = 'Υποσταθμοί ΔΕΔΔΗΕ ΔΕΕΔ/ΚΣΜΘ/ΤΕΙ'
        self._apply_theme()
        self.root_layout = BoxLayout(orientation='vertical')
        Window.bind(on_key_down=self._handle_tab_navigation)
        Window.bind(on_request_close=self._handle_request_close)

        loading_label = Label(text='Φόρτωση...', font_size='22sp')
        self.root_layout.add_widget(loading_label)

        from kivy.clock import Clock
        Clock.schedule_once(self._finish_build, 0)
        return self.root_layout

    def _finish_build(self, *_args):
        layout = self.root_layout
        layout.clear_widgets()

        self._add_logo_to_layout(layout, height=120, reserve=True)

        # Top-right small app info button
        top_bar = BoxLayout(orientation='horizontal', size_hint_y=None, height=40, padding=10)
        top_bar.add_widget(Widget())
        self.app_info_btn = Button(text='Πληρ. Εφαρμ.', size_hint=(None, None), height=30, width=130, font_size='12sp')
        self.app_info_btn.bind(on_press=self.show_app_info_popup)
        top_bar.add_widget(self.app_info_btn)
        layout.add_widget(top_bar)

        self.show_btn = IconButton(text='Εμφάνιση βάσης υποσταθμών', icon_type='database', theme=self.theme)
        self.show_btn.bind(on_press=self.show_records)
        self.import_btn = IconButton(text='Εισαγωγή από αρχείο', icon_type='import', theme=self.theme)
        self.import_btn.bind(on_press=self.show_import_menu)
        self.maintenance_btn = IconButton(text='Συντηρήσεις', icon_type='maintenance', theme=self.theme)
        self.maintenance_btn.bind(on_press=self.show_maintenance_menu_popup)
        self.inspection_btn = IconButton(text='Επιθεωρήσεις', icon_type='inspection', theme=self.theme)
        self.inspection_btn.bind(on_press=self.show_inspection_menu_popup)
        self.isolation_btn = IconButton(text='Αιτήσεις Απομόνωσης', icon_type='isolation', theme=self.theme)
        self.isolation_btn.bind(on_press=self.show_isolation_requests)
        self.models_btn = IconButton(text='Διαχείριση Τύπων Στοιχείων', icon_type='models', theme=self.theme)
        self.models_btn.bind(on_press=self.show_models_management)
        self.people_btn = IconButton(text='Διαχείριση Προσωπικού', icon_type='people', theme=self.theme)
        self.people_btn.bind(on_press=self.show_people_management)
        self.sf6_btn = IconButton(text='Διαχείριση SF6', icon_type='sf6', theme=self.theme)
        self.sf6_btn.bind(on_press=self.show_sf6_management_popup)

        buttons_layout = BoxLayout(orientation='horizontal', spacing=10, padding=10)
        left_col = BoxLayout(orientation='vertical', spacing=10)
        right_col = BoxLayout(orientation='vertical', spacing=10)

        left_col.add_widget(self.show_btn)
        left_col.add_widget(self.import_btn)
        left_col.add_widget(self.models_btn)
        left_col.add_widget(self.people_btn)

        right_col.add_widget(self.maintenance_btn)
        right_col.add_widget(self.inspection_btn)
        right_col.add_widget(self.isolation_btn)
        right_col.add_widget(self.sf6_btn)

        buttons_layout.add_widget(left_col)
        buttons_layout.add_widget(right_col)
        layout.add_widget(buttons_layout)

        self.conn = init_db()

    def _handle_request_close(self, *args):
        self._cleanup_before_exit()
        return False

    def on_stop(self):
        self._cleanup_before_exit()

    def _cleanup_before_exit(self):
        try:
            if getattr(self, 'conn', None):
                self.conn.close()
                self.conn = None
        except Exception:
            pass
        try:
            if self.root:
                self.root.clear_widgets()
        except Exception:
            pass

    def _apply_theme(self):
        """Apply a logo-based theme to common UI widgets."""
        theme = self._get_modern_theme()
        self.theme = theme

        Window.clearcolor = theme['background']

        Button.background_normal = 'atlas://data/images/defaulttheme/button'
        Button.background_down = 'atlas://data/images/defaulttheme/button_pressed'
        Button.background_color = theme['primary']
        Button.color = theme['text_on_primary']

        Spinner.background_normal = ''
        Spinner.background_down = ''
        Spinner.background_color = theme['primary']
        Spinner.color = theme['text_on_primary']

        # Spinner dropdown options (opaque background)
        from kivy.uix.spinner import SpinnerOption
        SpinnerOption.background_normal = ''
        SpinnerOption.background_down = ''
        SpinnerOption.background_color = theme['primary']
        SpinnerOption.color = theme['text_on_primary']

        Label.color = theme['text']

        TextInput.background_color = theme['input_bg']
        TextInput.foreground_color = theme['text']
        TextInput.cursor_color = theme['primary_dark']
        TextInput.selection_color = theme['accent']

        Popup.background = ''
        Popup.background_color = theme['popup_bg']

    def _add_logo_to_layout(self, layout, height=80, reserve=False):
        """Add logo to the top of a layout if available."""
        logo_path = os.path.join(os.path.dirname(__file__), 'logo_deddie.png')
        fallback_path = os.path.join(os.path.dirname(__file__), 'deddie_logo.png')
        if os.path.exists(logo_path) or os.path.exists(fallback_path):
            logo = Image(
                source=logo_path if os.path.exists(logo_path) else fallback_path,
                size_hint_y=None,
                height=height,
                allow_stretch=True,
                keep_ratio=True
            )
            layout.add_widget(logo)
            return
        if reserve:
            layout.add_widget(Label(text='', size_hint_y=None, height=height))

    def _get_modern_theme(self):
        """Return a modern dark-blue palette."""
        primary = (0.05, 0.18, 0.36, 1)
        primary_dark = (0.03, 0.12, 0.25, 1)
        accent = (0.12, 0.42, 0.85, 1)
        background = (0.94, 0.96, 0.99, 1)
        popup_bg = (0.98, 0.99, 1, 1)

        return {
            'primary': primary,
            'primary_dark': primary_dark,
            'accent': accent,
            'background': background,
            'popup_bg': popup_bg,
            'input_bg': (1, 1, 1, 1),
            'text': (0.12, 0.12, 0.12, 1),
            'text_on_primary': (1, 1, 1, 1),
        }

    def _load_logo_theme(self):
        """Extract a color theme from deddie_logo.png if available."""
        logo_path = os.path.join(os.path.dirname(__file__), 'logo_deddie.png')
        fallback_path = os.path.join(os.path.dirname(__file__), 'deddie_logo.png')
        if not os.path.exists(logo_path) and not os.path.exists(fallback_path):
            return dict(self.THEME_FALLBACK)

        if not os.path.exists(logo_path):
            logo_path = fallback_path

        try:
            image = CoreImage(logo_path)
            texture = image.texture
            if not texture or not texture.pixels:
                return dict(self.THEME_FALLBACK)

            pixels = texture.pixels
            total_pixels = max(1, texture.size[0] * texture.size[1])
            step = max(1, total_pixels // 5000)

            r_sum = g_sum = b_sum = 0
            count = 0
            for i in range(0, len(pixels), 4 * step):
                r, g, b, a = pixels[i], pixels[i + 1], pixels[i + 2], pixels[i + 3]
                if a < 10:
                    continue
                if r > 245 and g > 245 and b > 245:
                    continue
                r_sum += r
                g_sum += g
                b_sum += b
                count += 1

            if count == 0:
                return dict(self.THEME_FALLBACK)

            primary = (r_sum / count / 255, g_sum / count / 255, b_sum / count / 255, 1)
            primary_dark = self._adjust_color(primary, 0.8)
            accent = self._adjust_color(primary, 1.15)
            background = self._blend_color(primary, (1, 1, 1, 1), 0.92)
            popup_bg = self._blend_color(primary, (1, 1, 1, 1), 0.96)

            brightness = (primary[0] * 0.299) + (primary[1] * 0.587) + (primary[2] * 0.114)
            text_on_primary = (0, 0, 0, 1) if brightness > 0.6 else (1, 1, 1, 1)

            return {
                'primary': primary,
                'primary_dark': primary_dark,
                'accent': accent,
                'background': background,
                'popup_bg': popup_bg,
                'input_bg': (1, 1, 1, 1),
                'text': (0.12, 0.12, 0.12, 1),
                'text_on_primary': text_on_primary,
            }
        except Exception:
            return dict(self.THEME_FALLBACK)

    def _adjust_color(self, color, factor):
        return (
            min(max(color[0] * factor, 0), 1),
            min(max(color[1] * factor, 0), 1),
            min(max(color[2] * factor, 0), 1),
            color[3] if len(color) > 3 else 1
        )

    def _blend_color(self, color, target, ratio):
        return (
            color[0] * (1 - ratio) + target[0] * ratio,
            color[1] * (1 - ratio) + target[1] * ratio,
            color[2] * (1 - ratio) + target[2] * ratio,
            1
        )

    def _handle_tab_navigation(self, window, key, scancode, codepoint, modifiers):
        """Use Tab/Shift+Tab to move focus between text inputs."""
        if key != 9:  # Tab
            return False

        text_inputs = []
        for widget in Window.children:
            for child in widget.walk():
                if isinstance(child, TextInput):
                    child.write_tab = False
                    text_inputs.append(child)

        focused = next((ti for ti in text_inputs if ti.focus), None)
        if focused:
            next_widget = focused.get_focus_previous() if 'shift' in modifiers else focused.get_focus_next()
            if next_widget:
                focused.focus = False
                next_widget.focus = True
        return True

    def show_maintenance_menu_popup(self, instance=None):
        """Show maintenance main menu with entry + global history."""
        menu_popup = Popup(title='Συντηρήσεις', size_hint=(0.6, 0.4))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        self._add_logo_to_layout(layout, height=70)

        layout.add_widget(Label(text='Επιλέξτε ενέργεια:', size_hint_y=0.2))

        add_btn = Button(text='Καταχώρηση Συντήρησης', size_hint_y=0.3)
        add_btn.bind(on_press=lambda x: self.show_maintenance_menu(parent_popup=menu_popup))
        layout.add_widget(add_btn)

        import_email_btn = Button(text='Εισαγωγή συντήρησης από e-mail', size_hint_y=0.3)
        import_email_btn.bind(on_press=lambda x: self._show_import_maintenance_email_dialog(menu_popup))
        layout.add_widget(import_email_btn)

        history_btn = Button(text='Ιστορικό Συντηρήσεων', size_hint_y=0.3)
        history_btn.bind(on_press=lambda x: (menu_popup.dismiss(), self.show_maintenance_history(None)))
        layout.add_widget(history_btn)

        cancel_btn = Button(text='Ακύρωση', size_hint_y=0.2)
        cancel_btn.bind(on_press=menu_popup.dismiss)
        layout.add_widget(cancel_btn)

        menu_popup.content = layout
        menu_popup.open()

    def _show_import_maintenance_email_dialog(self, parent_popup=None):
        popup = Popup(title='Εισαγωγή Συντήρησης από E-mail', size_hint=(0.9, 0.9))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        path_label = Label(text='Διαδρομή αρχείου (.eml):', size_hint_y=0.1)
        layout.add_widget(path_label)

        path_input = TextInput(
            hint_text='Διαδρομή αρχείου .eml',
            size_hint_y=0.12,
            multiline=False
        )
        layout.add_widget(path_input)

        layout.add_widget(Label(text='Ή επιλέξτε από τη λίστα:', size_hint_y=0.1))
        chooser = FileChooserListView(filters=['*.eml'], path=os.path.dirname(__file__))
        layout.add_widget(chooser)

        buttons_layout = BoxLayout(size_hint_y=0.12, spacing=10)

        def import_email_file():
            file_path = path_input.text.strip() if path_input.text.strip() else (chooser.selection[0] if chooser.selection else None)

            if not file_path:
                show_message_popup('Σφάλμα', 'Παρακαλώ εισάγετε διαδρομή ή επιλέξτε αρχείο!')
                return

            if not os.path.exists(file_path):
                show_message_popup('Σφάλμα', 'Το αρχείο δεν βρέθηκε!')
                return

            if not file_path.lower().endswith('.eml'):
                show_message_popup('Σφάλμα', 'Παρακαλώ επιλέξτε αρχείο .eml!')
                return

            popup.dismiss()
            if parent_popup:
                parent_popup.dismiss()
            self._import_maintenance_from_email_file(file_path)

        import_btn = Button(text='Εισαγωγή')
        import_btn.bind(on_press=lambda x: import_email_file())
        buttons_layout.add_widget(import_btn)

        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        buttons_layout.add_widget(cancel_btn)

        layout.add_widget(buttons_layout)
        popup.content = layout
        popup.open()

    def _import_maintenance_from_email_file(self, file_path):
        try:
            payload = parse_eml_file(file_path)
        except Exception as exc:
            show_message_popup('Σφάλμα', f'Αποτυχία ανάγνωσης .eml:\n{str(exc)}')
            return

        self._open_maintenance_from_email_payload(payload)

    def _normalize_text(self, value: str) -> str:
        if not value:
            return ''
        value = unicodedata.normalize('NFKD', value)
        value = ''.join(ch for ch in value if not unicodedata.combining(ch))
        value = value.replace('ς', 'σ').lower()
        return value

    def _tokenize_text(self, value: str):
        normalized = self._normalize_text(value)
        normalized = re.sub(r'[^0-9a-zα-ω]+', ' ', normalized)
        return [tok for tok in normalized.split() if tok]

    def _tokens_match(self, left_tokens, right_tokens):
        if len(left_tokens) != len(right_tokens):
            return False
        for left, right in zip(left_tokens, right_tokens):
            if left == right:
                continue
            common_len = min(len(left), len(right))
            if common_len >= 4 and (left.startswith(right) or right.startswith(left)):
                continue
            if common_len >= 4 and left[:-1] == right[:-1]:
                continue
            return False
        return True

    def _find_substation_in_text(self, text: str, substations):
        tokens = self._tokenize_text(text)
        if not tokens:
            return None

        for sub_id, sub_name in substations:
            name_tokens = self._tokenize_text(sub_name)
            if not name_tokens:
                continue
            for i in range(len(tokens) - len(name_tokens) + 1):
                candidate = tokens[i:i + len(name_tokens)]
                if self._tokens_match(candidate, name_tokens):
                    return sub_id, sub_name

        return None

    def _match_person_by_sender(self, sender_name: str, people):
        sender_tokens = self._tokenize_text(sender_name)
        if not sender_tokens:
            return None
        sender_full = ' '.join(sender_tokens)
        for pid, name, _role in people:
            person_tokens = self._tokenize_text(name)
            if not person_tokens:
                continue
            if sender_full == ' '.join(person_tokens):
                return pid
        for pid, name, _role in people:
            person_tokens = self._tokenize_text(name)
            if not person_tokens:
                continue
            surname = person_tokens[-1]
            if surname and surname in sender_tokens:
                return pid
        return None

    def _find_people_in_body(self, body_text: str, people, exclude_ids=None):
        exclude_ids = exclude_ids or set()
        tokens = self._tokenize_text(body_text)
        token_set = set(tokens)
        compact = re.sub(r'[^0-9a-zα-ω]+', '', self._normalize_text(body_text))

        found = set()
        for pid, name, _role in people:
            if pid in exclude_ids:
                continue
            person_tokens = self._tokenize_text(name)
            if not person_tokens:
                continue
            surname = person_tokens[-1]
            first = person_tokens[0]
            initial = first[0] if first else ''

            if surname and surname in token_set:
                found.add(pid)
                continue

            if initial and surname:
                if f'{initial}{surname}' in compact:
                    found.add(pid)

        return found

    def _find_elements_in_body(self, body_text: str, substation_id: int):
        c = self.conn.cursor()
        c.execute("SELECT id, name, element_type FROM elements WHERE substation_id=?", (substation_id,))
        elements = c.fetchall()

        normalized_body = self._normalize_text(body_text)
        normalized_body = re.sub(r'μ\s*[\./-]\s*σ', 'μσ', normalized_body)
        normalized_body = re.sub(r'm\s*[\./-]\s*s', 'ms', normalized_body)
        normalized_body = re.sub(r'(μσ|ms)\s*[ilι]\b', r'\g<1>1', normalized_body)
        compact_body = re.sub(r'[^0-9a-zα-ω]+', '', normalized_body)
        compact_body = re.sub(r'(μσ|ms)[ilι](?![0-9])', r'\g<1>1', compact_body)
        has_satyf = 'σατυφ' in compact_body
        transformer_numbers = set(re.findall(r'(?:μσ|μετασχηματιστησ)[^0-9]{0,3}([0-9]+)', normalized_body))
        transformer_numbers.update(re.findall(r'(?:ms|transformer)[^0-9]{0,3}([0-9]+)', normalized_body))
        ms_numbers = set(re.findall(r'μσ([0-9]+)', compact_body))
        ms_numbers.update(re.findall(r'ms([0-9]+)', compact_body))

        matched = set()
        motor_drive_candidates = []
        for elem_id, elem_name, elem_type in elements:
            base = self._normalize_text(elem_name)
            compact = re.sub(r'[^0-9a-zα-ω]+', '', base)
            variants = {compact}

            digits = ''.join(ch for ch in compact if ch.isdigit())
            if digits:
                if 'μετασχηματιστης' in self._normalize_text(elem_type) or compact.startswith('μσ'):
                    variants.add(f'μσ{digits}')
                    variants.add(f'μετασχηματιστης{digits}')
                if compact.startswith('ρ'):
                    variants.add(f'ρ{digits}')

            element_ms_numbers = set(re.findall(r'(?:μσ|μετασχηματιστησ)[^0-9]{0,3}([0-9]+)', base))
            element_ms_numbers.update(re.findall(r'(?:ms|transformer)[^0-9]{0,3}([0-9]+)', base))
            element_ms_numbers.update(re.findall(r'μσ([0-9]+)', compact))
            element_ms_numbers.update(re.findall(r'ms([0-9]+)', compact))

            elem_type_norm = self._normalize_text(elem_type)
            is_motor_drive = (
                'motor drive' in elem_type_norm or
                elem_type_norm == 'motordrive' or
                'md' in compact or
                'σατυφ' in base
            )

            if is_motor_drive:
                motor_drive_candidates.append((elem_id, element_ms_numbers, compact, elem_type_norm))

            if is_motor_drive and has_satyf:
                if element_ms_numbers and (element_ms_numbers & (transformer_numbers | ms_numbers)):
                    matched.add(elem_id)
                    continue
                if (transformer_numbers or ms_numbers) and (not element_ms_numbers or element_ms_numbers & (transformer_numbers | ms_numbers)):
                    matched.add(elem_id)
                    continue
                if digits and (digits in transformer_numbers or digits in ms_numbers):
                    matched.add(elem_id)
                    continue
                if digits and (f'μσ{digits}' in compact_body or f'μετασχηματιστης{digits}' in compact_body):
                    matched.add(elem_id)
                    continue
                if not digits and len(transformer_numbers | ms_numbers) == 1:
                    matched.add(elem_id)
                    continue

            if any(var and var in compact_body for var in variants):
                matched.add(elem_id)

        if has_satyf and (transformer_numbers or ms_numbers):
            md_matched = any(eid in matched for eid, _, _, _ in motor_drive_candidates)
            if not md_matched:
                for elem_id, element_ms_numbers, compact, elem_type_norm in motor_drive_candidates:
                    if element_ms_numbers & (transformer_numbers | ms_numbers):
                        matched.add(elem_id)
                if not any(eid in matched for eid, _, _, _ in motor_drive_candidates) and len(motor_drive_candidates) == 1:
                    matched.add(motor_drive_candidates[0][0])

        if has_satyf and not (transformer_numbers or ms_numbers):
            md_matched = any(eid in matched for eid, _, _, _ in motor_drive_candidates)
            if not md_matched and len(motor_drive_candidates) == 1:
                matched.add(motor_drive_candidates[0][0])

        return matched

    def _get_previous_maintenance_defaults(self, substation_id: int, date_time_value: str):
        c = self.conn.cursor()
        c.execute(
            """
            SELECT id, maintenance_type, overall_comments, responsible_id
            FROM maintenance
            WHERE substation_id = ? AND date_time < ?
            ORDER BY date_time DESC
            LIMIT 1
            """,
            (substation_id, date_time_value)
        )
        row = c.fetchone()
        if not row:
            return {}

        maintenance_id, maint_type, comments, responsible_id = row

        c.execute("SELECT person_id, role FROM maintenance_people WHERE maintenance_id=?", (maintenance_id,))
        people_rows = c.fetchall()
        crew_ids = {pid for pid, role in people_rows if role == 'crew'}
        if not responsible_id:
            for pid, role in people_rows:
                if role == 'responsible':
                    responsible_id = pid
                    break

        c.execute("SELECT element_id FROM maintenance_elements WHERE maintenance_id=?", (maintenance_id,))
        element_ids = {row[0] for row in c.fetchall()}

        return {
            'maintenance_type': maint_type,
            'overall_comments': comments,
            'responsible_id': responsible_id,
            'crew_ids': crew_ids,
            'element_ids': element_ids
        }

    def _open_maintenance_from_email_payload(self, payload, forced_substation=None):
        subject = payload.get('subject', '')
        body = payload.get('body', '')
        sender_name = payload.get('sender_name', '')
        received_at = payload.get('received_at', '')

        c = self.conn.cursor()
        c.execute("SELECT id, name FROM substations ORDER BY name")
        substations = c.fetchall()
        if not substations:
            show_message_popup('Σφάλμα', 'Δεν υπάρχουν υποσταθμοί!')
            return

        substation = None
        if forced_substation:
            for sub_id, sub_name in substations:
                if sub_name == forced_substation:
                    substation = (sub_id, sub_name)
                    break
        if not substation:
            substation = self._find_substation_in_text(subject, substations)
        if not substation:
            substation = self._find_substation_in_text(body, substations)
        if not substation:
            self._prompt_substation_selection(substations, payload)
            return

        substation_id, substation_name = substation

        c.execute("SELECT COUNT(*) FROM elements WHERE substation_id=?", (substation_id,))
        if c.fetchone()[0] == 0:
            self._prompt_add_elements_then_continue(substation_id, substation_name, payload)
            return

        c.execute("SELECT id, name, role FROM people WHERE active=1 ORDER BY name")
        people = c.fetchall()
        if not people:
            show_message_popup('Σφάλμα', 'Δεν υπάρχουν καταχωρημένα άτομα. Παρακαλώ προσθέστε προσωπικό.')
            return

        responsible_id = self._match_person_by_sender(sender_name, people)
        crew_ids = self._find_people_in_body(body, people, exclude_ids={responsible_id} if responsible_id else set())

        element_ids = self._find_elements_in_body(body, substation_id)
        incomplete_elements = set(element_ids)

        date_time_value = ''
        if received_at:
            try:
                dt = datetime.fromisoformat(received_at.replace('Z', '+00:00'))
                date_time_value = dt.strftime('%Y-%m-%d %H:%M')
            except Exception:
                date_time_value = ''
        if not date_time_value:
            date_time_value = datetime.now().strftime('%Y-%m-%d %H:%M')

        prefill = {
            'substation_id': substation_id,
            'substation_name': substation_name,
            'maintenance_type': 'Επαναληπτική συντήρηση',
            'date_time': date_time_value,
            'overall_comments': body,
            'responsible_id': responsible_id,
            'crew_ids': crew_ids,
            'element_ids': element_ids,
            'incomplete_elements': incomplete_elements
        }

        prev = self._get_previous_maintenance_defaults(substation_id, date_time_value)
        if prev:
            if not prefill['responsible_id'] and prev.get('responsible_id'):
                prefill['responsible_id'] = prev.get('responsible_id')
            if not prefill['crew_ids'] and prev.get('crew_ids'):
                prefill['crew_ids'] = prev.get('crew_ids')
            if not prefill['element_ids'] and prev.get('element_ids'):
                prefill['element_ids'] = prev.get('element_ids')
                prefill['incomplete_elements'] = set(prefill['element_ids'])
            if not prefill['maintenance_type'] and prev.get('maintenance_type'):
                prefill['maintenance_type'] = prev.get('maintenance_type')
            if not prefill['overall_comments'] and prev.get('overall_comments'):
                prefill['overall_comments'] = prev.get('overall_comments')

        if not prefill['responsible_id']:
            self._prompt_responsible_selection(people, prefill)
            return

        self.show_maintenance_menu(
            preselected_substation_name=substation_name,
            parent_popup=None,
            maintenance_id=None,
            after_save_callback=None,
            prefill_data=prefill
        )

    def _prompt_substation_selection(self, substations, payload):
        popup = Popup(title='Ο υποσταθμός δε βρέθηκε', size_hint=(0.7, 0.5))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        layout.add_widget(Label(text='Επιλέξτε υποσταθμό για την εισαγωγή:', size_hint_y=None, height=40))
        substation_names = [s[1] for s in substations]
        spinner = Spinner(text=substation_names[0] if substation_names else '', values=substation_names, size_hint_y=None, height=40)
        layout.add_widget(spinner)

        layout.add_widget(Label(text='Ή προσθέστε νέο υποσταθμό:', size_hint_y=None, height=30))
        new_name_input = TextInput(hint_text='Όνομα νέου υποσταθμού', size_hint_y=None, height=40, multiline=False)
        layout.add_widget(new_name_input)

        buttons = BoxLayout(size_hint_y=None, height=40, spacing=10)

        def confirm():
            if not spinner.text:
                show_message_popup('Σφάλμα', 'Παρακαλώ επιλέξτε ή προσθέστε υποσταθμό.')
                return
            popup.dismiss()
            self._open_maintenance_from_email_payload(payload, forced_substation=spinner.text)

        def add_new_substation():
            new_name = new_name_input.text.strip()
            if not new_name:
                show_message_popup('Σφάλμα', 'Παρακαλώ εισάγετε όνομα υποσταθμού.')
                return
            c = self.conn.cursor()
            c.execute("SELECT id FROM substations WHERE name=?", (new_name,))
            if c.fetchone():
                show_message_popup('Πληροφορία', 'Ο υποσταθμός υπάρχει ήδη.')
                spinner.text = new_name
                return
            c.execute("INSERT INTO substations (name, location, adoption_date, division) VALUES (?, ?, ?, ?)",
                      (new_name, '', '', 'ΤΜΘ'))
            self.conn.commit()
            new_substation_id = c.lastrowid
            substation_names.append(new_name)
            spinner.values = substation_names
            spinner.text = new_name
            new_name_input.text = ''
            popup.dismiss()
            self._prompt_add_elements_then_continue(new_substation_id, new_name, payload)

        ok_btn = Button(text='Επιβεβαίωση')
        ok_btn.bind(on_press=lambda x: confirm())
        buttons.add_widget(ok_btn)

        add_btn = Button(text='Προσθήκη')
        add_btn.bind(on_press=lambda x: add_new_substation())
        buttons.add_widget(add_btn)

        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        buttons.add_widget(cancel_btn)

        layout.add_widget(buttons)
        popup.content = layout
        popup.open()

    def _prompt_add_elements_then_continue(self, substation_id, substation_name, payload):
        popup = Popup(title='Προσθήκη στοιχείων', size_hint=(0.7, 0.5))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        layout.add_widget(Label(
            text='Προσθέστε στοιχεία για τον νέο υποσταθμό πριν τη συνέχεια:',
            size_hint_y=None,
            height=50
        ))

        def add_element():
            self.show_add_element_popup_for_substation(substation_id, substation_name, popup)

        def continue_import():
            c = self.conn.cursor()
            c.execute("SELECT COUNT(*) FROM elements WHERE substation_id=?", (substation_id,))
            count = c.fetchone()[0]
            if count == 0:
                show_message_popup('Σφάλμα', 'Προσθέστε τουλάχιστον ένα στοιχείο πριν τη συνέχεια.')
                return
            popup.dismiss()
            self._open_maintenance_from_email_payload(payload, forced_substation=substation_name)

        buttons = BoxLayout(size_hint_y=None, height=40, spacing=10)
        add_btn = Button(text='Προσθήκη Στοιχείου')
        add_btn.bind(on_press=lambda x: add_element())
        buttons.add_widget(add_btn)

        continue_btn = Button(text='Συνέχεια')
        continue_btn.bind(on_press=lambda x: continue_import())
        buttons.add_widget(continue_btn)

        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        buttons.add_widget(cancel_btn)

        layout.add_widget(buttons)
        popup.content = layout
        popup.open()

    def _prompt_responsible_selection(self, people, prefill):
        popup = Popup(title='Ο υπεύθυνος δε βρέθηκε', size_hint=(0.7, 0.5))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        layout.add_widget(Label(text='Επιλέξτε υπεύθυνο συντήρησης:', size_hint_y=None, height=40))
        labels = [f"{p[1]} ({p[2]})" for p in people]
        spinner = Spinner(text=labels[0], values=labels, size_hint_y=None, height=40)
        layout.add_widget(spinner)

        buttons = BoxLayout(size_hint_y=None, height=40, spacing=10)

        def confirm():
            selected_index = labels.index(spinner.text)
            prefill['responsible_id'] = people[selected_index][0]
            popup.dismiss()
            self.show_maintenance_menu(
                preselected_substation_name=prefill['substation_name'],
                parent_popup=None,
                maintenance_id=None,
                after_save_callback=None,
                prefill_data=prefill
            )

        ok_btn = Button(text='Επιβεβαίωση')
        ok_btn.bind(on_press=lambda x: confirm())
        buttons.add_widget(ok_btn)

        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        buttons.add_widget(cancel_btn)

        layout.add_widget(buttons)
        popup.content = layout
        popup.open()

    def _get_ui_font_kwargs(self):
        """Return font kwargs for UI symbols if bundled font exists."""
        font_path = os.path.join(os.path.dirname(__file__), 'DejaVuSans.ttf')
        if os.path.exists(font_path):
            return {'font_name': font_path}
        return {}

    def show_app_info_popup(self, instance=None):
        """Show application information."""
        version = self._get_app_version()

        app_dir = os.path.dirname(__file__)
        info_text_plain = (
            'Υποσταθμοί ΔΕΔΔΗΕ ΔΕΕΔ/ΚΣΜΘ/ΤΕΙ\n'
            f'Έκδοση: {version}\n\n'
            'Λειτουργίες εφαρμογής:\n'
            '• Προβολή και διαχείριση βάσης υποσταθμών\n'
            '• Προσθήκη/επεξεργασία/διαγραφή υποσταθμών και στοιχείων\n'
            '• Κατηγορίες διακοπτών (SF6/Κενού/Ελαίου/Πτωχού Ελαίου)\n'
            '• Διαχείριση τύπων στοιχείων (μοντέλα/κατασκευαστές/κύκλοι)\n'
            '• Καταχώρηση συντηρήσεων\n'
            '• Εισαγωγή συντήρησης από e-mail (.eml)\n'
            '• Ιστορικό συντηρήσεων (όλων/ανά υποσταθμό)\n'
            '• Μετρήσεις διακοπτών (μόνωση/διέλευση/χειρισμοί)\n'
            '• Ποιότητα αερίου SF6 & διαρροές (kg)\n'
            '• Διαχείριση SF6 (αναφορά διαρροών ανά έτος)\n'
            '• Εξαγωγή Excel αναφορών SF6 (σύνοψη & ανά υποσταθμό)\n'
            '• Εκτύπωση PDF αναφορών συντήρησης\n'
            '• Επιθεωρήσεις (καταχώρηση/προβολή/ιστορικό)\n'
            '• Αιτήσεις απομόνωσης\n'
            '• Εισαγωγή υποσταθμών/στοιχείων από CSV/Excel\n'
            '• Αναφορές PDF & Excel\n\n'
            f'Φάκελος εφαρμογής: {app_dir}'
        )

        popup = Popup(title='Πληροφορίες Εφαρμογής', size_hint=(0.7, 0.6))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        info_field = ShiftSelectableTextInput(
            text=info_text_plain,
            readonly=True,
            multiline=True,
            size_hint_y=None,
            background_normal='',
            background_active='',
            background_color=(0, 0, 0, 0),
            foreground_color=self.theme.get('text', (0.12, 0.12, 0.12, 1)),
            selection_color=(0.3, 0.5, 1, 0.3),
            cursor_blink=False,
            cursor_width=1,
            write_tab=False,
            is_focusable=True,
            allow_copy=True,
            keyboard_mode='managed',
            padding=(5, 5)
        )
        info_field.bind(minimum_height=info_field.setter('height'))
        scroll = ScrollView(bar_width=10, scroll_type=['bars'])
        scroll.add_widget(info_field)
        layout.add_widget(scroll)
        def _sync_field_width(*_args):
            info_field.width = max(10, scroll.width - 10)

        scroll.bind(size=_sync_field_width)
        _sync_field_width()

        def _handle_info_keys(window, key, scancode, codepoint, modifiers):
            key_action_map = {
                276: 'cursor_left',
                275: 'cursor_right',
                273: 'cursor_up',
                274: 'cursor_down',
                278: 'cursor_home',
                279: 'cursor_end',
                280: 'cursor_pgup',
                281: 'cursor_pgdown',
            }
            if (('ctrl' in modifiers) or ('meta' in modifiers)) and (key == ord('c')):
                if info_field.selection_text:
                    Clipboard.copy(info_field.selection_text)
                    return True
            if key not in key_action_map:
                return False

            ctrl = 'ctrl' in modifiers or 'meta' in modifiers
            alt = 'alt' in modifiers

            if 'shift' in modifiers and info_field._shift_select_anchor is None:
                info_field._shift_select_anchor = info_field.cursor_index()

            info_field.do_cursor_movement(key_action_map[key], control=ctrl, alt=alt)

            if 'shift' in modifiers:
                info_field.select_text(info_field._shift_select_anchor, info_field.cursor_index())
            else:
                info_field.cancel_selection()
                info_field._shift_select_anchor = info_field.cursor_index()
            return True

        buttons_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=40, spacing=10)
        copy_btn = Button(text='Αντιγραφή')
        copy_btn.bind(on_press=lambda *_: Clipboard.copy(info_field.selection_text or info_text_plain))
        close_btn = Button(text='Κλείσιμο')
        close_btn.bind(on_press=popup.dismiss)
        buttons_layout.add_widget(copy_btn)
        buttons_layout.add_widget(close_btn)
        layout.add_widget(buttons_layout)

        popup.content = layout

        def _on_open(*_args):
            Window.bind(on_key_down=_handle_info_keys)
            info_field.focus = True
            info_field.cursor = (0, 0)

        def _on_dismiss(*_args):
            Window.unbind(on_key_down=_handle_info_keys)

        popup.bind(on_open=_on_open, on_dismiss=_on_dismiss)
        popup.open()

    def _get_app_version(self):
        version = os.environ.get('APP_VERSION')
        if version:
            return version.strip()

        version_path = os.path.join(os.path.dirname(__file__), 'VERSION')
        try:
            if os.path.exists(version_path):
                with open(version_path, 'r', encoding='utf-8') as vf:
                    return vf.read().strip() or '-'
        except Exception:
            return '-'

        return '-'

    def _get_sf6_report_data(self, year: str):
        c = self.conn.cursor()
        year_prefix = f"{year}%"

        c.execute(
            """
                 SELECT m.date_time, s.name, e.name, e.element_type, me.sf6_leakage_kg,
                     me.sf6_leak_methodology, p.name
            FROM maintenance_elements me
            JOIN maintenance m ON me.maintenance_id = m.id
            JOIN elements e ON me.element_id = e.id
            JOIN substations s ON m.substation_id = s.id
            LEFT JOIN people p ON m.responsible_id = p.id
            WHERE e.breaker_category = 'SF6'
              AND m.date_time LIKE ?
              AND me.sf6_leakage_kg IS NOT NULL
              AND me.sf6_leakage_kg > 0
            ORDER BY m.date_time ASC
            """,
            (year_prefix,)
        )
        leak_rows = c.fetchall()

        total_leakage = 0.0
        rows = []
        for date_time, sub_name, elem_name, elem_type, leakage, methodology, responsible_name in leak_rows:
            total_leakage += leakage
            rows.append({
                'date_time': date_time or '-',
                'substation': sub_name or '-',
                'element': elem_name or '-',
                'leakage': leakage,
                'methodology': methodology or '',
                'responsible': responsible_name or '-'
            })

        substation_rows = {}
        for row in rows:
            substation_rows.setdefault(row['substation'], []).append(row)

        c.execute(
            """
            SELECT
                COUNT(*),
                SUM(COALESCE(em.sf6_capacity_kg, 0))
            FROM elements e
            LEFT JOIN element_models em ON e.element_model_id = em.id
            WHERE e.operating_status = 'Ενεργή'
              AND e.breaker_category = 'SF6'
              AND e.element_type IN ('Διακόπτης ΥΤ', 'Διακόπτης ΜΤ')
            """
        )
        counts = c.fetchone()
        total_elements = counts[0] or 0
        installed_sf6 = counts[1] or 0.0

        c.execute(
            """
            SELECT COUNT(*), COUNT(DISTINCT s.id)
            FROM elements e
            JOIN substations s ON e.substation_id = s.id
            WHERE e.operating_status = 'Ενεργή'
              AND e.breaker_category = 'SF6'
              AND e.element_type IN ('Διακόπτης ΥΤ', 'Διακόπτης ΜΤ')
            """
        )
        active_counts = c.fetchone()
        active_elements = active_counts[0] or 0
        active_substations = active_counts[1] or 0

        c.execute(
            """
            SELECT s.name, SUM(COALESCE(em.sf6_capacity_kg, 0))
            FROM elements e
            JOIN substations s ON e.substation_id = s.id
            LEFT JOIN element_models em ON e.element_model_id = em.id
            WHERE e.operating_status = 'Ενεργή'
              AND e.breaker_category = 'SF6'
              AND e.element_type IN ('Διακόπτης ΥΤ', 'Διακόπτης ΜΤ')
            GROUP BY s.name
            """
        )
        substation_installed = {row[0]: (row[1] or 0.0) for row in c.fetchall()}

        percentage = (total_leakage / installed_sf6 * 100) if installed_sf6 else 0.0

        return {
            'total_leakage': total_leakage,
            'installed_sf6': installed_sf6,
            'percentage': percentage,
            'total_elements': total_elements,
            'active_elements': active_elements,
            'active_substations': active_substations,
            'rows': rows,
            'substation_rows': substation_rows,
            'substation_installed': substation_installed,
        }

    def _export_sf6_excel(self, year: str):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
        except Exception as exc:
            raise RuntimeError('Δεν βρέθηκε το πακέτο openpyxl. Εγκαταστήστε το για εξαγωγή Excel.') from exc

        data = self._get_sf6_report_data(year)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Σύνοψη'

        grey_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
        blue_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
        white_font = Font(color='FFFFFF', bold=True)
        bold_font = Font(bold=True)
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        summary_titles = [
            'ΣΥΝΟΛΙΚΗ ΕΓΚΑΤΕΣΤΗΜΕΝΗ ΠΟΣΟΤΗΤΑ (kg)',
            f'ΔΙΑΡΡΟΕΣ {year} (kg)',
            f'ΠΟΣΟΣΤΟ ΔΙΑΡΡΟΩΝ {year}'
        ]
        summary_values = [
            f"{data['installed_sf6']:.2f}",
            f"{data['total_leakage']:.2f}",
            f"{data['percentage']:.2f}%"
        ]

        for col_idx, title in enumerate(summary_titles, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.fill = grey_fill
            cell.font = bold_font
            cell.alignment = center
            cell.border = border

            val_cell = ws.cell(row=2, column=col_idx, value=summary_values[col_idx - 1])
            val_cell.alignment = center
            val_cell.border = border

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 24

        for col in range(1, 4):
            ws.column_dimensions[chr(64 + col)].width = 35

        substation_sums = []
        for substation, rows in data['substation_rows'].items():
            total = sum([r.get('leakage') or 0 for r in rows])
            if total > 0:
                substation_sums.append((substation, total))

        start_row = 4
        if substation_sums:
            ws.cell(row=start_row, column=1, value='Υποσταθμός').font = bold_font
            ws.cell(row=start_row, column=2, value='Σύνολο Διαρροών (kg)').font = bold_font
            for col_idx in (1, 2):
                cell = ws.cell(row=start_row, column=col_idx)
                cell.alignment = center
                cell.border = border

            row_ptr = start_row + 1
            for substation, total in sorted(substation_sums, key=lambda x: x[0] or ''):
                ws.cell(row=row_ptr, column=1, value=substation or '-').border = border
                ws.cell(row=row_ptr, column=2, value=f"{total:.2f}").border = border
                ws.cell(row=row_ptr, column=1).alignment = center
                ws.cell(row=row_ptr, column=2).alignment = center
                row_ptr += 1

        for substation, rows in data['substation_rows'].items():
            sheet_title = substation[:31] if substation else 'Υποσταθμός'
            if sheet_title in wb.sheetnames:
                suffix = 1
                base = sheet_title[:28]
                while f"{base}_{suffix}" in wb.sheetnames:
                    suffix += 1
                sheet_title = f"{base}_{suffix}"

            ws_sub = wb.create_sheet(title=sheet_title)
            ws_sub.merge_cells('A1:J1')
            title_cell = ws_sub['A1']
            title_cell.value = 'ΠΙΝΑΚΑΣ 4: ΠΗΓΗ ΕΚΠΟΜΠΩΝ ΑΠΌ ΕΞΟΠΛΙΣΜΟ ΧΡΗΣΗΣ SF6'
            title_cell.alignment = center

            for col_idx in range(1, 11):
                cell = ws_sub.cell(row=1, column=col_idx)
                cell.fill = blue_fill
                cell.font = white_font
                cell.alignment = center
                cell.border = border

            headers = [
                'Α/Α',
                'ΒΟΚ ή ΠΕΡΙΟΧΗ',
                'ΕΓΚΑΤΑΣΤΑΣΗ (Πχ. Όνομα Υ/Σ)',
                'ΜΟΝΑΔΑ ΜΕΤΡΗΣΗΣ',
                'ΠΛΗΡΩΣΗ Ή ΑΝΤΙΚΑΤΑΣΤΑΣΗ (ΜΕΘΟΔΟΛΟΓΙΑ)',
                'ΣΥΝΟΛΙΚΗ ΕΓΚΑΤΕΣΤΗΜΕΝΗ ΠΟΣΟΤΗΤΑ (kg)',
                'ΠΟΣΟΤΗΤΑ ΔΙΑΡΡΟΩΝ (kg)',
                'ΗΜ/ΝΙΑ',
                'ΥΠΕΥΘΥΝΟΣ ΣΥΝΕΡΓΕΙΟΥ',
                'ΥΠΟΓΡΑΦΗ'
            ]

            for col_idx, header in enumerate(headers, start=1):
                cell = ws_sub.cell(row=2, column=col_idx, value=header)
                cell.font = bold_font
                cell.alignment = center
                cell.border = border

            installed_sub = data['substation_installed'].get(substation, 0.0)
            start_row = 3
            for idx, row in enumerate(rows, start=1):
                values = [
                    idx,
                    'ΔΕΕΔ',
                    substation,
                    'kg',
                    row.get('methodology', '') or '',
                    f"{installed_sub:.2f}",
                    f"{row['leakage']:.2f}",
                    row['date_time'],
                    row.get('responsible', '-') or '-',
                    ''
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws_sub.cell(row=start_row, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border
                start_row += 1

            ws_sub.row_dimensions[1].height = 30
            ws_sub.row_dimensions[2].height = 28
            for col_idx in range(1, 11):
                ws_sub.column_dimensions[chr(64 + col_idx)].width = 22

        reports_dir = os.path.join(os.path.dirname(__file__), 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = os.path.join(reports_dir, f'SF6_Leakages_{year}_{timestamp}.xlsx')
        wb.save(output_path)
        return output_path

    def show_sf6_management_popup(self, instance=None):
        """Show SF6 leakage management report popup."""
        c = self.conn.cursor()
        c.execute("SELECT DISTINCT substr(date_time, 1, 4) FROM maintenance WHERE date_time IS NOT NULL AND date_time != '' ORDER BY 1 DESC")
        years = [row[0] for row in c.fetchall() if row[0] and row[0].isdigit()]
        if not years:
            years = [str(datetime.now().year)]

        popup = Popup(title='Διαχείριση SF6', size_hint=(0.95, 0.9))
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        control_row = BoxLayout(size_hint_y=None, height=40, spacing=10)
        control_row.add_widget(Label(text='Έτος:', size_hint_x=0.15))
        year_spinner = Spinner(text=years[0], values=years, size_hint_x=0.25, size_hint_y=None, height=35)
        control_row.add_widget(year_spinner)

        refresh_btn = Button(text='Ανανέωση', size_hint_x=0.2)
        control_row.add_widget(refresh_btn)
        print_btn = Button(text='Εκτύπωση', size_hint_x=0.2)
        control_row.add_widget(print_btn)
        excel_btn = Button(text='Excel', size_hint_x=0.2)
        control_row.add_widget(excel_btn)
        main_layout.add_widget(control_row)

        summary_label = Label(text='', size_hint_y=None, height=60)
        summary_label.bind(
            width=lambda inst, val: setattr(inst, 'text_size', (val, None)),
            texture_size=lambda inst, val: setattr(inst, 'height', val[1] + 10)
        )
        main_layout.add_widget(summary_label)

        scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        table_layout = GridLayout(cols=1, spacing=5, size_hint_y=None, padding=5)
        table_layout.bind(minimum_height=table_layout.setter('height'))
        scroll.add_widget(table_layout)
        main_layout.add_widget(scroll)

        def render_report(year_value: str):
            table_layout.clear_widgets()
            data = self._get_sf6_report_data(year_value)
            total_leakage = data['total_leakage']
            installed_sf6 = data['installed_sf6']
            percentage = data['percentage']
            total_elements = data['total_elements']
            active_elements = data['active_elements']
            active_substations = data['active_substations']

            summary_text = (
                f"Εγκατεστημένο SF6 (ενεργά): {installed_sf6:.2f} kg | "
                f"Ενεργά στοιχεία SF6: {active_elements} | Υποσταθμοί με SF6: {active_substations}\n"
                f"Έτος: {year_value} | Διαρροές: {total_leakage:.2f} kg | Ποσοστό: {percentage:.2f}%"
            )
            summary_label.text = summary_text

            header = GridLayout(cols=4, size_hint_y=None, height=30)
            header.add_widget(Label(text='Ημερομηνία', bold=True))
            header.add_widget(Label(text='Υποσταθμός', bold=True))
            header.add_widget(Label(text='Στοιχείο', bold=True))
            header.add_widget(Label(text='Διαρροή (kg)', bold=True))
            table_layout.add_widget(header)

            if not data['rows']:
                table_layout.add_widget(Label(text='Δεν υπάρχουν καταχωρήσεις διαρροών για το έτος.', size_hint_y=None, height=30))
                return

            for row in data['rows']:
                rlayout = GridLayout(cols=4, size_hint_y=None, height=30)
                rlayout.add_widget(Label(text=row['date_time'] or '-'))
                rlayout.add_widget(Label(text=row['substation'] or '-'))
                rlayout.add_widget(Label(text=row['element'] or '-'))
                leakage_text = '-' if row['leakage'] is None else f"{row['leakage']:.2f}"
                rlayout.add_widget(Label(text=leakage_text))
                table_layout.add_widget(rlayout)

        def handle_print(*_args):
            try:
                pdf_path = generate_sf6_leak_report(self.conn, year_spinner.text)
                confirm_popup = Popup(title='PDF Δημιουργήθηκε', size_hint=(0.6, 0.4))
                layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
                layout.add_widget(Label(text=f'Το PDF δημιουργήθηκε:\n{pdf_path}', size_hint_y=0.6))
                buttons = BoxLayout(size_hint_y=0.4, spacing=10)

                def open_pdf():
                    if sys.platform == 'win32':
                        os.startfile(pdf_path)
                    elif sys.platform == 'darwin':
                        subprocess.call(['open', pdf_path])
                    else:
                        subprocess.call(['xdg-open', pdf_path])
                    confirm_popup.dismiss()

                open_btn = Button(text='Άνοιγμα PDF')
                open_btn.bind(on_press=lambda _x: open_pdf())
                close_btn = Button(text='Κλείσιμο')
                close_btn.bind(on_press=confirm_popup.dismiss)
                buttons.add_widget(open_btn)
                buttons.add_widget(close_btn)
                layout.add_widget(buttons)
                confirm_popup.content = layout
                confirm_popup.open()
            except Exception as exc:
                show_message_popup('Σφάλμα', f'Αποτυχία δημιουργίας PDF:\n{str(exc)}')

        def handle_excel(*_args):
            try:
                excel_path = self._export_sf6_excel(year_spinner.text)
                confirm_popup = Popup(title='Excel Δημιουργήθηκε', size_hint=(0.6, 0.4))
                layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
                layout.add_widget(Label(text=f'Το Excel δημιουργήθηκε:\n{excel_path}', size_hint_y=0.6))
                buttons = BoxLayout(size_hint_y=0.4, spacing=10)

                def open_excel():
                    if sys.platform == 'win32':
                        os.startfile(excel_path)
                    elif sys.platform == 'darwin':
                        subprocess.call(['open', excel_path])
                    else:
                        subprocess.call(['xdg-open', excel_path])
                    confirm_popup.dismiss()

                open_btn = Button(text='Άνοιγμα Excel')
                open_btn.bind(on_press=lambda _x: open_excel())
                close_btn = Button(text='Κλείσιμο')
                close_btn.bind(on_press=confirm_popup.dismiss)
                buttons.add_widget(open_btn)
                buttons.add_widget(close_btn)
                layout.add_widget(buttons)
                confirm_popup.content = layout
                confirm_popup.open()
            except Exception as exc:
                show_message_popup('Σφάλμα', f'Αποτυχία δημιουργίας Excel:\n{str(exc)}')

        refresh_btn.bind(on_press=lambda _x: render_report(year_spinner.text))
        year_spinner.bind(text=lambda _s, _t: render_report(year_spinner.text))
        print_btn.bind(on_press=handle_print)
        excel_btn.bind(on_press=handle_excel)

        render_report(year_spinner.text)

        close_btn = Button(text='Κλείσιμο', size_hint_y=None, height=40)
        close_btn.bind(on_press=popup.dismiss)
        main_layout.add_widget(close_btn)

        popup.content = main_layout
        popup.open()

    def _format_maintenance_date(self, date_time_str):
        """Format maintenance date to DD/MM/YYYY for naming."""
        try:
            dt = datetime.strptime(date_time_str, '%Y-%m-%d %H:%M')
            return dt.strftime('%d/%m/%Y')
        except Exception:
            try:
                dt = datetime.strptime(date_time_str, '%Y-%m-%d')
                return dt.strftime('%d/%m/%Y')
            except Exception:
                return date_time_str

    def _build_maintenance_name(self, substation_name, date_time_str):
        formatted_date = self._format_maintenance_date(date_time_str)
        return f'Υ/Σ {substation_name} - DATE {formatted_date}'

    def _derive_voltage_level(self, element_type: str) -> str:
        if element_type == 'Μετασχηματιστής 150/20KV':
            return '150/20KV'
        if element_type == 'Διακόπτης ΜΤ':
            return '20KV'
        if element_type == 'Διακόπτης ΥΤ':
            return '150KV'
        if element_type == 'Μ/Σ ΧΤ/ΜΤ (ΒΜΣ)':
            return '20KV/400V'
        if element_type in ['Συστοιχία Συσσωρευτών', 'Μ/Σ Εγχύσεως']:
            return '20KV'
        return ''

    def _get_maintenance_people(self, maintenance_id):
        c = self.conn.cursor()
        c.execute("""
            SELECT p.name, mp.role
            FROM maintenance_people mp
            JOIN people p ON mp.person_id = p.id
            WHERE mp.maintenance_id = ?
            ORDER BY mp.role, p.name
        """, (maintenance_id,))
        rows = c.fetchall()
        responsible = None
        crew = []
        for name, role in rows:
            if role == 'responsible' and responsible is None:
                responsible = name
            elif role == 'crew':
                crew.append(name)
        return responsible, crew

    def show_people_management(self, instance=None):
        """Manage people (roles) used in maintenance records."""
        popup = Popup(title='Διαχείριση Προσωπικού', size_hint=(0.7, 0.8))
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        form_layout = GridLayout(cols=2, size_hint_y=None, height=120, spacing=5)
        form_layout.add_widget(Label(text='Όνομα:', size_hint_x=0.3))
        name_input = TextInput(multiline=False, size_hint_x=0.7)
        form_layout.add_widget(name_input)

        form_layout.add_widget(Label(text='Ρόλος:', size_hint_x=0.3))
        role_input = TextInput(multiline=False, size_hint_x=0.7)
        form_layout.add_widget(role_input)

        form_layout.add_widget(Label(text='Email:', size_hint_x=0.3))
        email_input = TextInput(multiline=False, size_hint_x=0.7)
        form_layout.add_widget(email_input)

        main_layout.add_widget(form_layout)

        receiver_layout = BoxLayout(size_hint_y=None, height=30, spacing=5)
        receiver_checkbox = CheckBox(size_hint_x=0.1, color=self.theme.get('primary', (0.05, 0.18, 0.36, 1)))
        receiver_layout.add_widget(receiver_checkbox)
        receiver_layout.add_widget(Label(text='Παραλήπτης email αναφοράς', size_hint_x=0.9))
        main_layout.add_widget(receiver_layout)

        add_btn = Button(text='Προσθήκη', size_hint_y=None, height=40)

        list_scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        list_layout = GridLayout(cols=1, spacing=5, size_hint_y=None, padding=5)
        list_layout.bind(minimum_height=list_layout.setter('height'))

        def refresh_list():
            list_layout.clear_widgets()
            c = self.conn.cursor()
            c.execute("SELECT id, name, role, email, report_receiver, active FROM people ORDER BY active DESC, name")
            for person_id, name, role, email, report_receiver, active in c.fetchall():
                row = BoxLayout(size_hint_y=None, height=35, spacing=5)
                status = 'Ενεργός' if active else 'Ανενεργός'
                email_text = email if email else '-'
                receiver_text = 'Ναι' if report_receiver else 'Όχι'
                row.add_widget(Label(text=f'{name} ({role}) | {email_text} | Παραλήπτης: {receiver_text} | {status}', size_hint_x=0.8))

                edit_btn = Button(text='Επεξ.', size_hint_x=0.1)
                delete_btn = Button(text='Διαγραφή', size_hint_x=0.1)

                def make_delete(pid, pname):
                    return lambda x: self._confirm_delete_person(pid, pname, refresh_list)

                def make_edit(pid):
                    return lambda x: self._show_edit_person_popup(pid, refresh_list)

                row.add_widget(edit_btn)
                row.add_widget(delete_btn)
                edit_btn.bind(on_press=make_edit(person_id))
                delete_btn.bind(on_press=make_delete(person_id, name))
                list_layout.add_widget(row)

        def add_person(instance):
            name = name_input.text.strip()
            role = role_input.text.strip()
            email = email_input.text.strip()
            if not name or not role:
                show_message_popup('Σφάλμα', 'Το όνομα και ο ρόλος είναι υποχρεωτικά!')
                return
            c = self.conn.cursor()
            c.execute("INSERT INTO people (name, role, email, report_receiver, active) VALUES (?, ?, ?, ?, 1)", (name, role, email, 1 if receiver_checkbox.active else 0))
            self.conn.commit()
            name_input.text = ''
            role_input.text = ''
            email_input.text = ''
            receiver_checkbox.active = False
            refresh_list()

        add_btn.bind(on_press=add_person)
        main_layout.add_widget(add_btn)

        refresh_list()
        list_scroll.add_widget(list_layout)
        main_layout.add_widget(list_scroll)

        close_btn = Button(text='Κλείσιμο', size_hint_y=None, height=40)
        close_btn.bind(on_press=popup.dismiss)
        main_layout.add_widget(close_btn)

        popup.content = main_layout
        popup.open()

    def _toggle_person_active(self, person_id, active, refresh_cb):
        c = self.conn.cursor()
        c.execute("UPDATE people SET active=? WHERE id=?", (active, person_id))
        self.conn.commit()
        if refresh_cb:
            refresh_cb()

    def _toggle_person_receiver(self, person_id, report_receiver, refresh_cb):
        c = self.conn.cursor()
        c.execute("UPDATE people SET report_receiver=? WHERE id=?", (report_receiver, person_id))
        self.conn.commit()
        if refresh_cb:
            refresh_cb()

    def _confirm_delete_person(self, person_id, person_name, refresh_cb):
        """Confirm and delete person if not referenced in maintenance records."""
        c = self.conn.cursor()
        c.execute("SELECT COUNT(*) FROM maintenance_people WHERE person_id=?", (person_id,))
        usage_count = c.fetchone()[0]
        if usage_count > 0:
            show_message_popup('Πληροφορία', 'Το άτομο έχει χρησιμοποιηθεί σε συντηρήσεις. Διαγράψτε το μόνο αφού αφαιρεθεί από το ιστορικό ή απενεργοποιήστε το.')
            return

        confirm_popup = Popup(title='Επιβεβαίωση Διαγραφής', size_hint=(0.6, 0.3))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        warning_label = Label(
            text=f'Είστε σίγουροι ότι θέλετε να διαγράψετε\nτο άτομο "{person_name}";',
            size_hint_y=0.5
        )
        layout.add_widget(warning_label)

        buttons_layout = BoxLayout(size_hint_y=0.3, spacing=10)

        def confirm_delete():
            confirm_popup.dismiss()
            c.execute("DELETE FROM people WHERE id=?", (person_id,))
            self.conn.commit()
            if refresh_cb:
                refresh_cb()

        yes_btn = Button(text='ΝΑΙ', color=(1, 0, 0, 1))
        yes_btn.bind(on_press=lambda x: confirm_delete())
        buttons_layout.add_widget(yes_btn)

        no_btn = Button(text='ΟΧΙ')
        no_btn.bind(on_press=confirm_popup.dismiss)
        buttons_layout.add_widget(no_btn)

        layout.add_widget(buttons_layout)
        confirm_popup.content = layout
        confirm_popup.open()

    def _show_edit_person_popup(self, person_id, refresh_cb):
        """Edit person details."""
        c = self.conn.cursor()
        c.execute("SELECT name, role, email, report_receiver, active FROM people WHERE id=?", (person_id,))
        row = c.fetchone()
        if not row:
            show_message_popup('Σφάλμα', 'Το άτομο δεν βρέθηκε!')
            return

        name, role, email, report_receiver, active = row

        popup = Popup(title='Επεξεργασία Προσώπου', size_hint=(0.6, 0.5))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        form = GridLayout(cols=2, size_hint_y=None, height=140, spacing=5)
        form.add_widget(Label(text='Όνομα:', size_hint_x=0.3))
        name_input = TextInput(text=name or '', multiline=False, size_hint_x=0.7)
        form.add_widget(name_input)

        form.add_widget(Label(text='Ρόλος:', size_hint_x=0.3))
        role_input = TextInput(text=role or '', multiline=False, size_hint_x=0.7)
        form.add_widget(role_input)

        form.add_widget(Label(text='Email:', size_hint_x=0.3))
        email_input = TextInput(text=email or '', multiline=False, size_hint_x=0.7)
        form.add_widget(email_input)

        layout.add_widget(form)

        receiver_layout = BoxLayout(size_hint_y=None, height=30, spacing=5)
        receiver_checkbox = CheckBox(size_hint_x=0.1, active=bool(report_receiver), color=self.theme.get('primary', (0.05, 0.18, 0.36, 1)))
        receiver_layout.add_widget(receiver_checkbox)
        receiver_layout.add_widget(Label(text='Παραλήπτης email αναφοράς', size_hint_x=0.9))
        layout.add_widget(receiver_layout)

        active_layout = BoxLayout(size_hint_y=None, height=30, spacing=5)
        active_checkbox = CheckBox(size_hint_x=0.1, active=bool(active), color=self.theme.get('primary', (0.05, 0.18, 0.36, 1)))
        active_layout.add_widget(active_checkbox)
        active_layout.add_widget(Label(text='Ενεργός', size_hint_x=0.9))
        layout.add_widget(active_layout)

        buttons_layout = BoxLayout(size_hint_y=None, height=40, spacing=10)

        def save_changes():
            new_name = name_input.text.strip()
            new_role = role_input.text.strip()
            new_email = email_input.text.strip()
            if not new_name or not new_role:
                show_message_popup('Σφάλμα', 'Το όνομα και ο ρόλος είναι υποχρεωτικά!')
                return
            c.execute(
                "UPDATE people SET name=?, role=?, email=?, report_receiver=?, active=? WHERE id=?",
                (new_name, new_role, new_email, 1 if receiver_checkbox.active else 0, 1 if active_checkbox.active else 0, person_id)
            )
            self.conn.commit()
            popup.dismiss()
            if refresh_cb:
                refresh_cb()

        save_btn = Button(text='Αποθήκευση')
        save_btn.bind(on_press=lambda x: save_changes())
        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)

        buttons_layout.add_widget(save_btn)
        buttons_layout.add_widget(cancel_btn)
        layout.add_widget(buttons_layout)

        popup.content = layout
        popup.open()

    def _element_type_report_label(self, element_type):
        mapping = {
            'Διακόπτης ΜΤ': 'διακόπτες Μέσης Τάσης',
            'Διακόπτης ΥΤ': 'διακόπτες Υψηλής Τάσης',
            'Μετασχηματιστής 150/20KV': 'μετασχηματιστές 150/20KV',
            'Motor Drive': 'motor drives',
            'Μ/Σ Εγχύσεως': 'Μ/Σ Εγχύσεως',
            'Μ/Σ Έντασης': 'Μ/Σ Έντασης',
            'Μ/Σ Τάσης': 'Μ/Σ Τάσης',
            'Μ/Σ ΧΤ/ΜΤ (ΒΜΣ)': 'Μ/Σ ΧΤ/ΜΤ (ΒΜΣ)',
            'Αποζεύκτης': 'αποζεύκτες',
            'Ασφαλειοαποζεύκτης': 'ασφαλειοαποζεύκτες',
            'Γειωτής': 'γειωτές',
            'Συστοιχία Πυκνωτών': 'συστοιχίες πυκνωτών',
            'Αντίσταση Κόμβου': 'αντιστάσεις κόμβου',
            'Αλεξικέραυνο': 'αλεξικέραυνα',
            'Συστοιχία Συσσωρευτών': 'συστοιχίες συσσωρευτών'
        }
        return mapping.get(element_type, element_type)

    def send_maintenance_email_report(self, maintenance_id):
        """Compose and open an email report for a maintenance instance."""
        c = self.conn.cursor()
        c.execute("""
            SELECT m.name, m.date_time, s.name
            FROM maintenance m
            JOIN substations s ON m.substation_id = s.id
            WHERE m.id = ?
        """, (maintenance_id,))
        maint_row = c.fetchone()
        if not maint_row:
            show_message_popup('Σφάλμα', 'Δεν βρέθηκε η συντήρηση.')
            return

        maint_name, date_time, substation_name = maint_row
        display_name = maint_name or self._build_maintenance_name(substation_name, date_time)

        c.execute("""
            SELECT e.element_type, e.name
            FROM maintenance_elements me
            JOIN elements e ON me.element_id = e.id
            WHERE me.maintenance_id = ?
            ORDER BY e.element_type, e.name
        """, (maintenance_id,))
        elements = c.fetchall()

        if not elements:
            show_message_popup('Πληροφορία', 'Δεν υπάρχουν στοιχεία για αυτή τη συντήρηση.')
            return

        # Group by element type
        grouped = {}
        for elem_type, elem_name in elements:
            grouped.setdefault(elem_type, []).append(elem_name)

        responsible, crew = self._get_maintenance_people(maintenance_id)
        crew_text = ', '.join(crew) if crew else '-'
        resp_text = responsible if responsible else '-'

        lines = []
        lines.append(f'Αναφορά Συντήρησης: {display_name}')
        lines.append(f'Υποσταθμός: {substation_name}')
        lines.append(f'Ημερομηνία: {date_time}')
        lines.append(f'Υπεύθυνος: {resp_text}')
        lines.append(f'Ομάδα: {crew_text}')
        lines.append('')

        for elem_type, names in grouped.items():
            label = self._element_type_report_label(elem_type)
            lines.append(f'Συντηρήθηκαν οι παρακάτω {label}:')
            for name in names:
                lines.append(f' - {name}')
            lines.append('')

        body = '\n'.join(lines).strip()
        subject = f'Αναφορά Συντήρησης - {display_name}'

        c.execute("SELECT email FROM people WHERE active=1 AND report_receiver=1 AND email IS NOT NULL AND email != ''")
        recipients = [row[0] for row in c.fetchall()]

        if not recipients:
            show_message_popup('Σφάλμα', 'Δεν υπάρχουν παραλήπτες email. Προσθέστε παραλήπτες από τη Διαχείριση Προσωπικού.')
            return

        import urllib.parse
        mailto = 'mailto:' + ','.join(recipients)
        subject_encoded = urllib.parse.quote(subject or '', safe='')
        body_encoded = urllib.parse.quote(body or '', safe='')
        webbrowser.open(f'{mailto}?subject={subject_encoded}&body={body_encoded}')

    def show_inspection_menu_popup(self, instance=None):
        """Show inspections main menu with import + history."""
        menu_popup = Popup(title='Επιθεωρήσεις', size_hint=(0.6, 0.4))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        self._add_logo_to_layout(layout, height=70)

        layout.add_widget(Label(text='Επιλέξτε ενέργεια:', size_hint_y=0.2))

        add_btn = Button(text='Καταχώρηση Επιθεώρησης', size_hint_y=0.3)
        add_btn.bind(on_press=lambda x: (menu_popup.dismiss(), self.show_inspection_entry_popup(None)))
        layout.add_widget(add_btn)

        import_btn = Button(text='Εισαγωγή Επιθεώρησης από Αρχείο', size_hint_y=0.3)
        import_btn.bind(on_press=lambda x: (menu_popup.dismiss(), self.show_import_inspections_dialog(None)))
        layout.add_widget(import_btn)

        history_btn = Button(text='Ιστορικό Επιθεωρήσεων', size_hint_y=0.3)
        history_btn.bind(on_press=lambda x: (menu_popup.dismiss(), self.show_inspection_history(None)))
        layout.add_widget(history_btn)

        cancel_btn = Button(text='Ακύρωση', size_hint_y=0.2)
        cancel_btn.bind(on_press=menu_popup.dismiss)
        layout.add_widget(cancel_btn)

        menu_popup.content = layout
        menu_popup.open()

    def show_import_inspections_dialog(self, instance):
        self._create_file_import_dialog(
            'Εισαγωγή επιθεωρήσεων από αρχείο',
            self.import_inspections_from_file
        )

    def _read_inspection_template_columns(self):
        """Read inspection template columns if available."""
        template_path = os.path.join(os.path.dirname(__file__), 'επιθεωρήσεις_template.xlsx')
        if not os.path.exists(template_path):
            return self._get_inspection_fallback_fields()

        def is_valid_label(value):
            text = str(value).strip()
            if not text or text.lower() in {'nan', 'none'}:
                return False
            lower = text.lower()
            if lower.startswith('unnamed'):
                return False
            if 'σελ' in lower or 'page' in lower:
                return False
            if 'version:' in lower or 'template_version' in lower:
                return False
            if not re.search(r'[A-Za-zΑ-Ωα-ω]', text):
                return False
            return True

        try:
            import pandas as pd
            df_peek = pd.read_excel(template_path, nrows=1, header=None)
            first_cell = str(df_peek.iloc[0, 0]) if len(df_peek) > 0 else ''
            if 'Version:' in first_cell or 'TEMPLATE_VERSION:' in first_cell:
                df = pd.read_excel(template_path, skiprows=1, nrows=0)
            else:
                df = pd.read_excel(template_path, nrows=0)
            cols_raw = [c for c in list(df.columns) if str(c).strip()]
            cols = [str(c).strip() for c in cols_raw if is_valid_label(c)]
            if cols and (len(cols) >= 5 or len(cols) >= max(1, len(cols_raw) // 2)):
                return cols
        except Exception:
            pass

        # Fallback: read header row with openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(template_path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True, min_row=1, max_row=2))
            if not rows:
                return []
            header_row = rows[0]
            if header_row and isinstance(header_row[0], str) and ('Version:' in header_row[0] or 'TEMPLATE_VERSION:' in header_row[0]):
                header_row = rows[1] if len(rows) > 1 else None
            if not header_row:
                return []
            header_cols_raw = [c for c in header_row if c is not None and str(c).strip()]
            header_cols = [str(c).strip() for c in header_cols_raw if is_valid_label(c)]
            if header_cols and (len(header_cols) >= 5 or len(header_cols) >= max(1, len(header_cols_raw) // 2)):
                return header_cols
        except Exception:
            pass

        # Fallback: read form labels from layout (first non-empty in A-D)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(template_path, read_only=True, data_only=True)
            ws = wb.active
            fields = []
            for row in ws.iter_rows(values_only=True, min_row=1, max_row=200):
                for cell in row:
                    if not isinstance(cell, str):
                        continue
                    label = cell.strip()
                    if not label:
                        continue
                    label_lower = label.lower()
                    if 'πιν' in label_lower or 'παρατηρ' in label_lower:
                        continue
                    if not is_valid_label(label):
                        continue
                    if label not in fields:
                        fields.append(label)
            if fields:
                return fields
        except Exception:
            pass

        return self._get_inspection_fallback_fields()

    def _get_inspection_fallback_fields(self):
        """Fallback inspection fields based on the standard report template."""
        return [
            'Υποσταθμός',
            'Αρ. Δελτίου',
            'Μήνας',
            'Ονομ. Επιθεωρητή',
            'Περιοχή',
            'Ημέρα',
            'Έτος',
            'Ημερομηνία',
            {'type': 'section', 'title': '1. Έλεγχος Χώρων ΥΣ'},
            'Παρατηρήσεις (1. Έλεγχος Χώρων ΥΣ)',
            'Έλεγχος εξωτερικών & εσωτερικών Θυρών ΥΣ',
            'Έλεγχος εσωτερικού Χώρου κτηρίου (Φωτισμός, κλιματισμός κλπ)',
            'Έλεγχος περιβάλλοντος χώρου (βλάστηση, δένδρα, φωτισμός κλπ)',
            'Έλεγχος μέσων πυρόσβεσης γενικά',
            {'type': 'section', 'title': '2. Μ/Σ 150/20kV & Διακόπτες 150kV & 20kV'},
            'Παρατηρήσεις (2. Μ/Σ 150/20kV & Διακόπτες 150kV & 20kV)',
            'Οπτικός έλεγχος, διαρροής/στάθμης/θερμοκρασίας λαδιού, silica gel στον Μ/Σ',
            'Οπτικός έλεγχος διαρροής λαδιού ή πίεσης SF6 ή πίεσης αέρα στους Διακόπτες Ισχύος 150kV & 20kV',
            'Έλεγχος λειτουργίας ανεμιστήρων Μ/Σ',
            'Οπτικός έλεγχος Μ/Σ εγχύσεως, ΜΣΕ, ΜΣΤ, Μ/Σ εσωτ. Υπηρ., αντίστασης κόμβου (θερμοκρασία)',
            'Οπτικός έλεγχος Μονωτήρων (ρύπανση, εκδορές κ.α.)',
            'Οπτικός έλεγχος τηκτών πυκνωτών',
            'Έλεγχος σημάνσεων στους Πίνακες Μ/Σ , Α/Δ 150kV & 20kV',
            'Λήψη φωτογραφίας όταν απαιτείται',
            {'type': 'section', 'title': '3α. Υπαίθριες πύλες 20 kV'},
            'Παρατηρήσεις (3α. Υπαίθριες πύλες 20 kV)',
            'Οπτικός έλεγχος των πυλών, A/Z και γενικά του ικριώματος για τυχόν φωλιές από πτηνά, σπασίματα, μονωτήρες, κλαδιά, σύρματα κλπ',
            {'type': 'section', 'title': '3β. Πίνακες 20 kV'},
            'Παρατηρήσεις (3β. Πίνακες 20 kV)',
            'Οπτικός έλεγχος στους πίνακες Διακοπτών 20kV (αναγγελίες, ενδείξεις οργάνων, πόρτες) και έλεγχος θορύβων, ιονισμών',
            'Έλεγχοι υγρασίας (υπόγειο, κανάλια καλωδίων), αφυγραντήρων, θερμαντικών, φορητών πυροσβεστήρων',
            {'type': 'section', 'title': '4. Κτίριο χειρισμών & Τ.Α.Σ.'},
            'Παρατηρήσεις (4. Κτίριο χειρισμών & Τ.Α.Σ.)',
            'Έλεγχος φορτιστή 110 V οπτικά με έλεγχο της τάσης, έντασης και καταγραφή',
            'Έλεγχος για alarm έλλειψης DC στον γενικό πίνακα DC',
            'Οπτικός έλεγχος διαρροών στοιχείων συσσωρευτών',
            {'type': 'section', 'title': '5. Αποζεύκτες Γραμμών'},
            'Παρατηρήσεις (5. Αποζεύκτες Γραμμών)',
            'Οπτικός έλεγχος των ΑΠ/Ζ και των "γεφυρών" αυτών στον 1ο Στύλο κάθε Γραμμής (σπασμένοι ΑΠ/Ζ, μονωτήρες, εκτονωμένα Α/Ξ κλπ)',
            {'type': 'section', 'title': '6. PC ΧΕΙΡΙΣΜΩΝ'},
            'Παρατηρήσεις (6. PC ΧΕΙΡΙΣΜΩΝ)',
            'Έλεγχος λειτουργίας ψηφιακού συστήματος (χειρισμοί, ενδείξεις, σημάνσεις)',
            'Τροφοδοσία υπολογιστή',
            {'type': 'section', 'title': '7. Απόψεις'},
            'Απόψεις - Προτάσεις'
        ]

    def _is_inspection_meta_column(self, col_name, keywords):
        col_text = str(col_name).strip().lower()
        return any(key in col_text for key in keywords)

    def show_inspection_entry_popup(self, instance=None, preselected_substation_name=None, parent_popup=None):
        """Manual inspection entry using template fields."""
        if parent_popup:
            parent_popup.dismiss()
        c = self.conn.cursor()
        c.execute("SELECT id, name FROM substations ORDER BY name")
        substations = c.fetchall()

        if not substations:
            show_message_popup('Σφάλμα', 'Δεν υπάρχουν υποσταθμοί!')
            return

        popup = Popup(title='Καταχώρηση Επιθεώρησης', size_hint=(0.9, 0.95))
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        content_layout = GridLayout(cols=1, spacing=10, size_hint_y=None, padding=10)
        content_layout.bind(minimum_height=content_layout.setter('height'))

        content_layout.add_widget(Label(text='Επιλογή Υποσταθμού:', size_hint_y=None, height=35))
        substation_map = {s[1]: s[0] for s in substations}
        initial_substation = preselected_substation_name if preselected_substation_name in substation_map else substations[0][1]
        substation_row = BoxLayout(size_hint_y=None, height=40, spacing=5)
        substation_label = Label(text='Υποσταθμός:', size_hint_x=0.18)
        substation_picker = BoxLayout(size_hint_x=0.42, spacing=5)
        substation_input = TextInput(text=initial_substation, readonly=True, size_hint_x=0.7, multiline=False)
        select_sub_btn = Button(text='Επιλογή', size_hint_x=0.3)
        substation_picker.add_widget(substation_input)
        substation_picker.add_widget(select_sub_btn)
        form_number_label = Label(text='Αρ. Δελτίου:', size_hint_x=0.18)
        form_number_input = TextInput(hint_text='Αρ. Δελτίου', size_hint_x=0.22, multiline=False)
        substation_row.add_widget(substation_label)
        substation_row.add_widget(substation_picker)
        substation_row.add_widget(form_number_label)
        substation_row.add_widget(form_number_input)
        content_layout.add_widget(substation_row)

        # People list for inspector
        c.execute("SELECT name FROM people WHERE active=1 ORDER BY name")
        people = [row[0] for row in c.fetchall()]
        if not people:
            show_message_popup('Σφάλμα', 'Δεν υπάρχουν καταχωρημένα άτομα. Παρακαλώ προσθέστε προσωπικό.', callback=lambda: self.show_people_management(None))
            return

        row_two = BoxLayout(size_hint_y=None, height=40, spacing=5)
        date_label = Label(text='Ημερομηνία:', size_hint_x=0.18)
        date_input = TextInput(
            text=datetime.now().strftime('%Y-%m-%d'),
            hint_text='YYYY-MM-DD',
            size_hint_x=0.32,
            height=40,
            multiline=False
        )
        region_label = Label(text='Περιοχή:', size_hint_x=0.14)
        region_input = TextInput(hint_text='Περιοχή', size_hint_x=0.16, multiline=False)
        inspector_label = Label(text='Ονομ. Επιθεωρητή:', size_hint_x=0.12)
        inspector_spinner = Spinner(text=people[0], values=people, size_hint_x=0.18, height=40)
        row_two.add_widget(date_label)
        row_two.add_widget(date_input)
        row_two.add_widget(region_label)
        row_two.add_widget(region_input)
        row_two.add_widget(inspector_label)
        row_two.add_widget(inspector_spinner)
        content_layout.add_widget(row_two)

        row_three = BoxLayout(size_hint_y=None, height=40, spacing=5)
        month_label = Label(text='Μήνας:', size_hint_x=0.18)
        month_input = TextInput(readonly=True, size_hint_x=0.32, multiline=False)
        day_label = Label(text='Ημέρα:', size_hint_x=0.18)
        day_input = TextInput(readonly=True, size_hint_x=0.32, multiline=False)
        year_label = Label(text='Έτος:', size_hint_x=0.18)
        year_input = TextInput(readonly=True, size_hint_x=0.18, multiline=False)
        row_three.add_widget(month_label)
        row_three.add_widget(month_input)
        row_three.add_widget(day_label)
        row_three.add_widget(day_input)
        row_three.add_widget(year_label)
        row_three.add_widget(year_input)
        content_layout.add_widget(row_three)

        fields_inputs = []

        greek_months = [
            'Ιανουάριος', 'Φεβρουάριος', 'Μάρτιος', 'Απρίλιος', 'Μάιος', 'Ιούνιος',
            'Ιούλιος', 'Αύγουστος', 'Σεπτέμβριος', 'Οκτώβριος', 'Νοέμβριος', 'Δεκέμβριος'
        ]
        greek_days = ['Δευτέρα', 'Τρίτη', 'Τετάρτη', 'Πέμπτη', 'Παρασκευή', 'Σάββατο', 'Κυριακή']

        def update_date_meta(_instance=None, text=None):
            parsed = self._parse_inspection_date(date_input.text.strip())
            try:
                dt = datetime.strptime(parsed, '%Y-%m-%d')
                month_input.text = greek_months[dt.month - 1]
                day_input.text = greek_days[dt.weekday()]
                year_input.text = f'{dt.year}'
            except Exception:
                month_input.text = ''
                day_input.text = ''
                year_input.text = ''

        def open_substation_selection(_instance=None):
            self._show_substation_selection_window_with_callback(
                popup,
                substations,
                lambda selected_name: setattr(substation_input, 'text', selected_name)
            )

        select_sub_btn.bind(on_press=open_substation_selection)
        date_input.bind(text=update_date_meta)
        update_date_meta()

        # Chapter 2: Έλεγχος Χώρων ΥΣ
        content_layout.add_widget(Label(text='[b]Έλεγχος Χώρων ΥΣ[/b]', markup=True, size_hint_y=None, height=35))

        def add_inspection_row(label_text):
            row = BoxLayout(size_hint_y=None, height=60, spacing=5)
            label = Label(text=label_text, size_hint_x=0.7, size_hint_y=None)
            label.bind(width=lambda inst, val: setattr(inst, 'text_size', (val, None)))

            ti = TextInput(hint_text='Παρατηρήσεις', size_hint_x=0.3, size_hint_y=None, height=60, multiline=True)

            def sync_row_height(_instance=None, _value=None):
                row.height = max(label.texture_size[1] if label.texture_size else 0, ti.height, 60)
                label.height = row.height

            label.bind(texture_size=sync_row_height)
            ti.bind(height=sync_row_height)
            sync_row_height()

            row.add_widget(label)
            row.add_widget(ti)
            content_layout.add_widget(row)
            fields_inputs.append((label_text, ti))

        add_inspection_row('Έλεγχος εξωτερικών & εσωτερικών Θυρών ΥΣ')
        add_inspection_row('Έλεγχος εσωτερικού Χώρου κτηρίου (Φωτισμός, κλιματισμός κλπ)')
        add_inspection_row('Έλεγχος περιβάλλοντος χώρου (βλάστηση, δένδρα, φωτισμός κλπ)')
        add_inspection_row('Έλεγχος μέσων πυρόσβεσης γενικά.')

        # Chapter 3: Μ/Σ 150/20kV & Διακόπτες 150kV & 20kV
        content_layout.add_widget(Label(text='[b]Μ/Σ 150/20kV & Διακόπτες 150kV & 20kV[/b]', markup=True, size_hint_y=None, height=35))

        add_inspection_row('Οπτικός έλεγχος, διαρροής/στάθμης/θερμοκρασίας λαδιού, silica gel στον Μ/Σ')
        add_inspection_row('Οπτικός έλεγχος διαρροής λαδιού ή πίεσης SF6 ή πίεσης αέρα στους Διακόπτες Ισχύος 150kV & 20kV')
        add_inspection_row('Έλεγχος λειτουργίας ανεμιστήρων  Μ/Σ')
        add_inspection_row('Οπτικός έλεγχος Μ/Σ εγχύσεως, ΜΣΕ, ΜΣΤ, Μ/Σ εσωτ. Υπηρ., αντίστασης κόμβου (θερμοκρασία)')
        add_inspection_row('Οπτικός έλεγχος Μονωτήρων (ρύπανση, εκδορές κ.α.)')
        add_inspection_row('Οπτικός έλεγχος τηκτών πυκνωτών')
        add_inspection_row('Έλεγχος σημάνσεων στους Πίνακες Μ/Σ , Α/Δ 150kV & 20kV')
        add_inspection_row('Λήψη φωτογραφίας όταν απαιτείται')

        # Chapter 3: Υπαίθριες πύλες 20 kV
        content_layout.add_widget(Label(text='[b]Υπαίθριες πύλες 20 kV[/b]', markup=True, size_hint_y=None, height=35))
        add_inspection_row('Οπτικός έλεγχος των πυλών, A/Z  και γενικά του ικριώματος για τυχόν φωλιές από πτηνα, σπασιματά, μονωτήρων, κλαδιά, σύρματα κλπ')

        # Chapter 4: Υπαίθριες πύλες 20 kV
        content_layout.add_widget(Label(text='[b]Υπαίθριες πύλες 20 kV[/b]', markup=True, size_hint_y=None, height=35))
        add_inspection_row('Οπτικός έλεγχος στους πίνακες Διακοπτών 20kV (αναγγελίες, ενδείξεις οργάνων, πόρτες) και έλεγχος θορύβων, ιονισμών.')
        add_inspection_row('Έλεγχοι υγρασίας (υπόγειο, κανάλια καλωδίων), αφυγραντήρων, θερμαντικών, φορητών πυροσβεστήρων.')

        # Chapter 5: Κτίριο χειρισμών & Τ.Α.Σ.
        content_layout.add_widget(Label(text='[b]Κτίριο χειρισμών & Τ.Α.Σ.[/b]', markup=True, size_hint_y=None, height=35))
        add_inspection_row('Έλεγχος φορτιστή 110 V οπτικά με έλεγχο της τάσης, έντασης και καταγραφή')
        add_inspection_row('Έλεγχος για alarm έλλειψης DC στον γενικό πίνακα DC.')
        add_inspection_row('Οπτικός έλεγχος διαρροών στοιχείων συσσωρευτών.')

        # Chapter 6: Αποζευκτες Γραμμών
        content_layout.add_widget(Label(text='[b]Αποζευκτες Γραμμών[/b]', markup=True, size_hint_y=None, height=35))
        add_inspection_row('Οπτικός έλεγχος των ΑΠ/Ζ και των "γεφυρών" αυτών στον 1ο Στύλο κάθε Γραμμής (σπασμένοι ΑΠ/Ζ, μονωτήρες, εκτονωμένα Α/Ξ κλπ)')

        # Chapter 7: PC ΧΕΙΡΙΣΜΩΝ
        content_layout.add_widget(Label(text='[b]PC ΧΕΙΡΙΣΜΩΝ[/b]', markup=True, size_hint_y=None, height=35))
        add_inspection_row('Έλεγχος λειτουργίας ψηφιακού συστήματος (χειρισμοί, ενδείξεις, σημάνσεις')
        add_inspection_row('Τροφοδοσία υπολογιστή.')

        # Chapter 8: Απόψεις
        content_layout.add_widget(Label(text='[b]Απόψεις[/b]', markup=True, size_hint_y=None, height=35))
        add_inspection_row('Απόψεις και τυχόν προτάσεις  για την καλύτερη λειτουργία τόσο του εξοπλισμού, όσο και του κτηρίου γενικά του Υ/Σ.')

        scroll.add_widget(content_layout)
        main_layout.add_widget(scroll)

        buttons_layout = BoxLayout(size_hint_y=0.1, spacing=10)

        def save_inspection():
            substation_name = substation_input.text
            substation_id = substation_map.get(substation_name)
            inspection_date = self._parse_inspection_date(date_input.text.strip()) or datetime.now().strftime('%Y-%m-%d')
            month_key = self._derive_month_key(inspection_date)

            fields = []
            fields.append({'label': 'Υποσταθμός', 'value': substation_name})
            fields.append({'label': 'Αρ. Δελτίου', 'value': self._format_inspection_value(form_number_input.text)})
            fields.append({'label': 'Περιοχή', 'value': self._format_inspection_value(region_input.text)})
            fields.append({'label': 'Ονομ. Επιθεωρητή', 'value': self._format_inspection_value(inspector_spinner.text)})
            fields.append({'label': 'Μήνας', 'value': self._format_inspection_value(month_input.text)})
            fields.append({'label': 'Ημέρα', 'value': self._format_inspection_value(day_input.text)})
            fields.append({'label': 'Έτος', 'value': self._format_inspection_value(year_input.text)})
            fields.append({'label': 'Ημερομηνία', 'value': self._format_inspection_value(inspection_date)})
            for label, input_widget in fields_inputs:
                fields.append({
                    'label': label,
                    'value': self._format_inspection_value(input_widget.text)
                })

            data_json = json.dumps({'fields': fields}, ensure_ascii=False)
            created_at = datetime.now().strftime('%Y-%m-%d %H:%M')

            c.execute("""
                INSERT INTO inspections (
                    substation_id, substation_name, inspection_date,
                    month_key, data_json, source_file, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                substation_id,
                substation_name,
                inspection_date,
                month_key,
                data_json,
                'manual-entry',
                created_at
            ))
            self.conn.commit()
            popup.dismiss()
            if parent_popup:
                show_message_popup('Επιτυχία', 'Η επιθεώρηση καταχωρήθηκε!', callback=lambda: self.show_substation_inspection_history(substation_id, substation_name))
            else:
                show_message_popup('Επιτυχία', 'Η επιθεώρηση καταχωρήθηκε!', callback=lambda: self.show_inspection_history(None))

        save_btn = Button(text='Αποθήκευση')
        save_btn.bind(on_press=lambda x: save_inspection())
        buttons_layout.add_widget(save_btn)

        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        buttons_layout.add_widget(cancel_btn)

        main_layout.add_widget(buttons_layout)
        popup.content = main_layout
        popup.open()

    def _detect_inspection_column(self, columns, keywords):
        for col in columns:
            col_text = str(col).strip().lower()
            for key in keywords:
                if key in col_text:
                    return col
        return None

    def _format_inspection_value(self, value):
        if value is None:
            return ''
        try:
            import math
            if isinstance(value, float) and math.isnan(value):
                return ''
        except Exception:
            pass

        if hasattr(value, 'to_pydatetime'):
            try:
                value = value.to_pydatetime()
            except Exception:
                pass

        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d')

        if isinstance(value, float) and value.is_integer():
            value = int(value)

        return str(value).strip()

    def _parse_inspection_date(self, value):
        if value is None:
            return ''
        if hasattr(value, 'to_pydatetime'):
            try:
                value = value.to_pydatetime()
            except Exception:
                pass
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d')

        text = str(value).strip()
        if not text:
            return ''

        for fmt in (
            '%Y-%m-%d', '%Y-%m-%d %H:%M',
            '%d/%m/%Y', '%d/%m/%Y %H:%M',
            '%d-%m-%Y', '%Y/%m/%d'
        ):
            try:
                return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
            except Exception:
                pass

        return text

    def _derive_month_key(self, date_str):
        if not date_str:
            return datetime.now().strftime('%Y-%m')

        for fmt in (
            '%Y-%m-%d', '%Y-%m-%d %H:%M',
            '%d/%m/%Y', '%d/%m/%Y %H:%M',
            '%d-%m-%Y', '%Y/%m/%d'
        ):
            try:
                return datetime.strptime(date_str, fmt).strftime('%Y-%m')
            except Exception:
                pass

        if len(date_str) >= 7 and date_str[4] == '-':
            return date_str[:7]

        return datetime.now().strftime('%Y-%m')

    def import_inspections_from_file(self, file_path):
        import pandas as pd

        try:
            if file_path.endswith('.xlsx'):
                df = pd.read_excel(file_path)
            elif file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                show_message_popup('Σφάλμα', 'Μη υποστηριζόμενη μορφή αρχείου')
                return
        except Exception as e:
            show_message_popup('Σφάλμα', f'Σφάλμα κατά την ανάγνωση αρχείου: {e}')
            return

        if df.empty:
            show_message_popup('Σφάλμα', 'Το αρχείο δεν περιέχει δεδομένα.')
            return

        columns = list(df.columns)
        date_col = self._detect_inspection_column(columns, ['ημερομην', 'ημ/ν', 'date'])
        substation_col = self._detect_inspection_column(columns, ['υποσταθ', 'substation'])

        inserted = 0
        created_at = datetime.now().strftime('%Y-%m-%d %H:%M')
        c = self.conn.cursor()

        for _, row in df.iterrows():
            if row.isna().all():
                continue

            date_value = row.get(date_col) if date_col else None
            inspection_date = self._parse_inspection_date(date_value) or datetime.now().strftime('%Y-%m-%d')
            month_key = self._derive_month_key(inspection_date)

            substation_name = ''
            if substation_col:
                substation_name = self._format_inspection_value(row.get(substation_col))

            substation_id = None
            if substation_name:
                c.execute("SELECT id FROM substations WHERE name=?", (substation_name,))
                sub_row = c.fetchone()
                substation_id = sub_row[0] if sub_row else None

            fields = []
            for col in columns:
                value = row.get(col)
                if pd.isna(value):
                    value = ''
                fields.append({
                    'label': str(col),
                    'value': self._format_inspection_value(value)
                })

            data_json = json.dumps({'fields': fields}, ensure_ascii=False)

            c.execute("""
                INSERT INTO inspections (
                    substation_id, substation_name, inspection_date,
                    month_key, data_json, source_file, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                substation_id,
                substation_name,
                inspection_date,
                month_key,
                data_json,
                os.path.basename(file_path),
                created_at
            ))
            inserted += 1

        self.conn.commit()
        show_message_popup('Εισαγωγή Επιθεωρήσεων', f'Ολοκληρώθηκε η εισαγωγή ({inserted} εγγραφές).', callback=lambda: self.show_inspection_history(None))

    def show_inspection_history(self, instance=None):
        """Show inspections list grouped by month."""
        font_kwargs = self._get_ui_font_kwargs()
        c = self.conn.cursor()
        c.execute("SELECT DISTINCT month_key FROM inspections ORDER BY month_key DESC")
        months = [row[0] for row in c.fetchall() if row[0]]

        if not months:
            show_message_popup('Πληροφορία', 'Δεν υπάρχουν καταχωρημένες επιθεωρήσεις')
            return

        popup = Popup(title='Ιστορικό Επιθεωρήσεων', size_hint=(0.95, 0.9))
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        top_bar = BoxLayout(size_hint_y=None, height=40, spacing=10)
        top_bar.add_widget(Label(text='Μήνας:', size_hint_x=0.2))
        month_spinner = Spinner(text=months[0], values=months, size_hint_x=0.4)
        top_bar.add_widget(month_spinner)

        add_btn = Button(text='Καταχώρηση', size_hint_x=0.2)
        add_btn.bind(on_press=lambda x: self.show_inspection_entry_popup(None))
        top_bar.add_widget(add_btn)

        import_btn = Button(text='Εισαγωγή από Αρχείο', size_hint_x=0.2)
        import_btn.bind(on_press=lambda x: self.show_import_inspections_dialog(None))
        top_bar.add_widget(import_btn)
        main_layout.add_widget(top_bar)

        scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        list_layout = GridLayout(cols=1, spacing=10, size_hint_y=None, padding=10)
        list_layout.bind(minimum_height=list_layout.setter('height'))

        def load_month(month_key):
            list_layout.clear_widgets()
            c.execute("""
                SELECT id, substation_name, inspection_date, data_json
                FROM inspections
                WHERE month_key = ?
                ORDER BY inspection_date DESC
            """, (month_key,))
            rows = c.fetchall()

            if not rows:
                list_layout.add_widget(Label(text='Δεν υπάρχουν επιθεωρήσεις για τον μήνα αυτό.', size_hint_y=None, height=30))
                return

            for insp_id, substation_name, inspection_date, _data_json in rows:
                card = BoxLayout(orientation='horizontal', size_hint_y=None, height=40, spacing=5)
                card.add_widget(Label(
                    text=f'{substation_name} | Ημερομηνία: {inspection_date}',
                    size_hint_x=0.6
                ))

                buttons_box = BoxLayout(size_hint_x=0.4, spacing=5)
                view_btn = Button(text='Εμφ.', size_hint_x=0.34, **font_kwargs)
                pdf_btn = Button(text='PDF', size_hint_x=0.33, **font_kwargs)
                email_btn = Button(text='Email', size_hint_x=0.33)
                view_btn.bind(on_press=lambda x, iid=insp_id: self.show_inspection_details(iid))
                pdf_btn.bind(on_press=lambda x, iid=insp_id, sname=substation_name: self.generate_inspection_pdf(iid, sname))
                email_btn.bind(on_press=lambda x, iid=insp_id: self.send_inspection_email_report(iid))
                buttons_box.add_widget(view_btn)
                buttons_box.add_widget(pdf_btn)
                buttons_box.add_widget(email_btn)
                card.add_widget(buttons_box)

                list_layout.add_widget(card)

        load_month(month_spinner.text)
        month_spinner.bind(text=lambda _spinner, text: load_month(text))

        scroll.add_widget(list_layout)
        main_layout.add_widget(scroll)

        close_btn = Button(text='Κλείσιμο', size_hint_y=0.1)
        close_btn.bind(on_press=popup.dismiss)
        main_layout.add_widget(close_btn)

        popup.content = main_layout
        popup.open()

    def get_available_gates(self, substation_id, is_interconnection=False):
        """Get available gates (ΠΥΛΗ) based on existing transformers in the substation
        
        Args:
            substation_id: The ID of the substation
            is_interconnection: If True, returns interconnection gates (1-2, 2-3, etc.)
                               If False, returns regular gates (1, 2, 3, etc.)
        """
        c = self.conn.cursor()
        # Get all transformers for this substation, ordered by name
        c.execute("""SELECT name FROM elements 
                    WHERE substation_id=? AND element_type='Μετασχηματιστής 150/20KV' 
                    ORDER BY name""", (substation_id,))
        transformers = c.fetchall()
        
        num_gates = len(transformers)
        
        if is_interconnection:
            # Generate interconnection gates: ΠΥΛΗ 1-2, ΠΥΛΗ 2-3, etc.
            gates = [f"ΠΥΛΗ {i}-{i+1}" for i in range(1, num_gates)]
        else:
            # Generate regular gates: ΠΥΛΗ 1, ΠΥΛΗ 2, etc.
            gates = [f"ΠΥΛΗ {i+1}" for i in range(num_gates)]
        
        # Always include option for unassigned
        return ['(Μη καταχωρημένο)'] + gates

    def show_import_menu(self, instance):
        # Show menu for import options
        menu_popup = Popup(title='Εισαγωγή από αρχείο', size_hint=(0.6, 0.55))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        self._add_logo_to_layout(layout, height=70)

        layout.add_widget(Label(text='Επιλέξτε τι θέλετε να εισάγετε:', size_hint_y=0.2))

        # Import elements button
        import_elements_btn = Button(text='Εισαγωγή Στοιχείων από Αρχείο', size_hint_y=0.2)
        import_elements_btn.bind(on_press=lambda x: self._show_import_elements_from_menu(menu_popup))
        layout.add_widget(import_elements_btn)

        # Import Android changes button
        import_android_btn = Button(text='Εισαγωγή αλλαγών από Android', size_hint_y=0.2)
        import_android_btn.bind(on_press=lambda x: self._show_import_android_changes_from_menu(menu_popup))
        layout.add_widget(import_android_btn)

        # Separator
        layout.add_widget(Label(text='Ή δημιουργήστε πρότυπο εισαγωγής:', size_hint_y=0.15))

        # Template button
        template_elements_btn = Button(text='Δημιουργία Template Εισαγωγής', size_hint_y=0.2)
        template_elements_btn.bind(on_press=self.create_elements_template)
        layout.add_widget(template_elements_btn)

        # Cancel button
        cancel_btn = Button(text='Ακύρωση', size_hint_y=0.15)
        cancel_btn.bind(on_press=menu_popup.dismiss)
        layout.add_widget(cancel_btn)

        menu_popup.content = layout
        menu_popup.open()
    
    def _show_import_substations_from_menu(self, menu_popup):
        menu_popup.dismiss()
        self.show_import_substations_dialog(None)
    
    def _show_import_elements_from_menu(self, menu_popup):
        menu_popup.dismiss()
        self.show_import_elements_dialog(None)

    def _show_import_android_changes_from_menu(self, menu_popup):
        menu_popup.dismiss()
        self.show_import_android_changes_dialog(None)

    def show_add_menu(self, instance):
        # Show intermediate menu for adding substation or element
        menu_popup = Popup(title='Προσθήκη υποσταθμών και στοιχείων', size_hint=(0.6, 0.4))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        self._add_logo_to_layout(layout, height=70)
        
        layout.add_widget(Label(text='Επιλέξτε τι θέλετε να προσθέσετε:', size_hint_y=0.3))
        
        # Add substation button
        add_substation_btn = Button(text='Προσθήκη Νέου Υποσταθμού', size_hint_y=0.3)
        add_substation_btn.bind(on_press=lambda x: self._show_add_substation_from_menu(menu_popup))
        layout.add_widget(add_substation_btn)
        
        # Add element button
        add_element_btn = Button(text='Προσθήκη Νέου Στοιχείου', size_hint_y=0.3)
        add_element_btn.bind(on_press=lambda x: self._show_add_element_from_menu(menu_popup))
        layout.add_widget(add_element_btn)
        
        # Cancel button
        cancel_btn = Button(text='Ακύρωση', size_hint_y=0.2)
        cancel_btn.bind(on_press=menu_popup.dismiss)
        layout.add_widget(cancel_btn)
        
        menu_popup.content = layout
        menu_popup.open()
    
    def _show_add_substation_from_menu(self, menu_popup):
        menu_popup.dismiss()
        self.show_add_substation_popup(None)
    
    def _show_add_element_from_menu(self, menu_popup):
        menu_popup.dismiss()
        self.show_add_element_popup(None)

    def show_add_substation_popup(self, instance):
        # Create popup
        popup = Popup(title='Προσθήκη Νέου Υποσταθμού', size_hint=(0.8, 0.5))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # Name input
        name_input = TextInput(
            hint_text='Όνομα Υποσταθμού',
            size_hint_y=0.25,
            multiline=False
        )
        layout.add_widget(Label(text='Όνομα Υποσταθμού:', size_hint_y=0.15))
        layout.add_widget(name_input)
        
        # Division spinner
        division_spinner = Spinner(
            text='ΤΜΘ',
            values=['ΤΜΘ'],
            size_hint_y=0.25
        )
        layout.add_widget(Label(text='Τομέας:', size_hint_y=0.15))
        layout.add_widget(division_spinner)
        
        # Buttons layout
        buttons_layout = BoxLayout(size_hint_y=0.3, spacing=10)
        
        def add_substation():
            if not name_input.text:
                show_message_popup('Σφάλμα', 'Παρακαλώ εισάγετε όνομα υποσταθμού!')
                return
            
            c = self.conn.cursor()
            c.execute("INSERT INTO substations (name, location, adoption_date, division) VALUES (?, ?, ?, ?)", 
                     (name_input.text, '', '', division_spinner.text))
            self.conn.commit()
            popup.dismiss()
            show_message_popup('Επιτυχία', 'Υποσταθμός προστέθηκε!', callback=lambda: self.show_records(None))
        
        add_btn = Button(text='Προσθήκη')
        add_btn.bind(on_press=lambda x: add_substation())
        buttons_layout.add_widget(add_btn)
        
        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        buttons_layout.add_widget(cancel_btn)
        
        layout.add_widget(buttons_layout)
        popup.content = layout
        popup.open()
    
    def show_add_substation_popup_from_db_view(self, parent_popup):
        """Add substation from within the database view, and refresh the view after"""
        # Create popup
        popup = Popup(title='Προσθήκη Νέου Υποσταθμού', size_hint=(0.8, 0.5))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # Name input
        name_input = TextInput(
            hint_text='Όνομα Υποσταθμού',
            size_hint_y=0.25,
            multiline=False
        )
        layout.add_widget(Label(text='Όνομα Υποσταθμού:', size_hint_y=0.15))
        layout.add_widget(name_input)
        
        # Division spinner
        division_spinner = Spinner(
            text='ΤΜΘ',
            values=['ΤΜΘ'],
            size_hint_y=0.25
        )
        layout.add_widget(Label(text='Τομέας:', size_hint_y=0.15))
        layout.add_widget(division_spinner)
        
        # Buttons layout
        buttons_layout = BoxLayout(size_hint_y=0.3, spacing=10)
        
        def add_substation():
            if not name_input.text:
                show_message_popup('Σφάλμα', 'Παρακαλώ εισάγετε όνομα υποσταθμού!')
                return
            
            c = self.conn.cursor()
            c.execute("INSERT INTO substations (name, location, adoption_date, division) VALUES (?, ?, ?, ?)", 
                     (name_input.text, '', '', division_spinner.text))
            self.conn.commit()
            popup.dismiss()
            parent_popup.dismiss()
            show_message_popup('Επιτυχία', 'Υποσταθμός προστέθηκε!', callback=lambda: self.show_records(None))
        
        add_btn = Button(text='Προσθήκη')
        add_btn.bind(on_press=lambda x: add_substation())
        buttons_layout.add_widget(add_btn)
        
        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        buttons_layout.add_widget(cancel_btn)
        
        layout.add_widget(buttons_layout)
        popup.content = layout
        popup.open()

    def show_records(self, instance):
        # Show intermediate selection dialog
        c = self.conn.cursor()
        c.execute("SELECT id, name FROM substations ORDER BY name")
        all_substations = c.fetchall()
        
        if not all_substations:
            show_message_popup('Σφάλμα', 'Δεν υπάρχουν υποσταθμοί στη βάση!')
            return
        
        # Create selection popup
        selection_popup = Popup(title='Επιλογή Προβολής', size_hint=(0.6, 0.4))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        self._add_logo_to_layout(layout, height=70)
        
        prompt_field = TextInput(
            text='Επιλέξτε τι θέλετε να δείτε:',
            readonly=True,
            multiline=False,
            size_hint_y=None,
            height=35,
            background_normal='',
            background_active='',
            background_color=(0, 0, 0, 0),
            foreground_color=self.theme.get('text', (0.12, 0.12, 0.12, 1)),
            selection_color=(0.3, 0.5, 1, 0.3),
            cursor_blink=False,
            cursor_width=0,
            write_tab=False,
            halign='center',
            is_focusable=True,
            allow_copy=True,
            padding=(5, 5)
        )
        layout.add_widget(prompt_field)
        
        # "Show All" button
        show_all_btn = Button(text='Εμφάνιση Όλων των Υποσταθμών', size_hint_y=0.35)
        show_all_btn.bind(on_press=lambda x: self._show_all_substations(selection_popup))
        layout.add_widget(show_all_btn)
        
        # "Select Specific Substation" button
        select_specific_btn = Button(text='Επιλογή Υποσταθμού', size_hint_y=0.35)
        select_specific_btn.bind(on_press=lambda x: self._show_substation_selection_window(selection_popup, all_substations))
        layout.add_widget(select_specific_btn)
        
        # Cancel button
        cancel_btn = Button(text='Ακύρωση', size_hint_y=0.2)
        cancel_btn.bind(on_press=selection_popup.dismiss)
        layout.add_widget(cancel_btn)
        
        selection_popup.content = layout
        selection_popup.open()
    
    def _show_substation_selection_window(self, parent_popup, all_substations):
        """Show a scrollable window with a 5x14 matrix of substation buttons"""
        parent_popup.dismiss()
        
        # Create selection popup
        selection_popup = Popup(title='Επιλογή Υποσταθμού', size_hint=(0.9, 0.85))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # Create scrollable area
        scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        grid = GridLayout(cols=5, spacing=5, size_hint_y=None, padding=10)
        grid.bind(minimum_height=grid.setter('height'))
        
        # Create 5x14 matrix (70 positions total)
        total_positions = 70
        
        # Add buttons for registered substations and empty boxes for remaining positions
        for i in range(total_positions):
            if i < len(all_substations):
                sub_id, sub_name = all_substations[i]
                sub_btn = Button(
                    text=sub_name,
                    size_hint_y=None,
                    height=50
                )
                sub_btn.bind(on_press=lambda x, name=sub_name, popup=selection_popup: self._show_specific_substation_from_window(name, popup))
                grid.add_widget(sub_btn)
            else:
                # Empty box for unregistered positions
                empty_btn = Button(
                    text='',
                    size_hint_y=None,
                    height=50,
                    disabled=True,
                    background_color=(0.3, 0.3, 0.3, 0.5)
                )
                grid.add_widget(empty_btn)
        
        scroll.add_widget(grid)
        layout.add_widget(scroll)
        
        # Cancel button
        cancel_btn = Button(text='Ακύρωση', size_hint_y=0.08)
        cancel_btn.bind(on_press=selection_popup.dismiss)
        layout.add_widget(cancel_btn)
        
        selection_popup.content = layout
        selection_popup.open()

    def _show_substation_selection_window_with_callback(self, parent_popup, all_substations, on_select, title='Επιλογή Υποσταθμού'):
        """Show a selection window and call on_select with the chosen substation name."""
        selection_popup = Popup(title=title, size_hint=(0.9, 0.85))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        grid = GridLayout(cols=5, spacing=5, size_hint_y=None, padding=10)
        grid.bind(minimum_height=grid.setter('height'))

        total_positions = 70

        def handle_select(sub_name):
            selection_popup.dismiss()
            if parent_popup:
                parent_popup.open()
            if on_select:
                on_select(sub_name)

        for i in range(total_positions):
            if i < len(all_substations):
                _sub_id, sub_name = all_substations[i]
                sub_btn = Button(text=sub_name, size_hint_y=None, height=50)
                sub_btn.bind(on_press=lambda x, name=sub_name: handle_select(name))
                grid.add_widget(sub_btn)
            else:
                empty_btn = Button(text='', size_hint_y=None, height=50, disabled=True, background_color=(0.3, 0.3, 0.3, 0.5))
                grid.add_widget(empty_btn)

        scroll.add_widget(grid)
        layout.add_widget(scroll)

        cancel_btn = Button(text='Ακύρωση', size_hint_y=0.08)
        cancel_btn.bind(on_press=selection_popup.dismiss)
        layout.add_widget(cancel_btn)

        if parent_popup:
            parent_popup.dismiss()

        selection_popup.content = layout
        selection_popup.open()
    
    def _show_all_substations(self, selection_popup):
        selection_popup.dismiss()
        self._display_substations(None)
    
    def _show_specific_substation_from_window(self, substation_name, selection_popup):
        selection_popup.dismiss()
        self._display_substations(substation_name)
    
    def _display_substations(self, filter_name=None, reuse_popup=None, element_type_filter=None, gate_filter=None):
        c = self.conn.cursor()
        if filter_name:
            c.execute("SELECT id, name, location, adoption_date, division, monogram_pdf FROM substations WHERE name=?", (filter_name,))
            title = f'Υποσταθμός: {filter_name}'
        else:
            c.execute("SELECT id, name, location, adoption_date, division, monogram_pdf FROM substations ORDER BY name")
            title = 'Εγγραφές Υποσταθμών'
        
        substations = c.fetchall()
        show_elements = filter_name is not None

        sub_ids = [row[0] for row in substations]
        elem_count_map = {}
        gate_count_map = {}
        capacitor_count_map = {}
        maint_count_map = {}
        last_maint_map = {}
        inactive_count_map = {}

        if sub_ids:
            placeholders = ','.join(['?'] * len(sub_ids))

            c.execute(f"SELECT substation_id, COUNT(*) FROM elements WHERE substation_id IN ({placeholders}) GROUP BY substation_id", sub_ids)
            elem_count_map = {sid: cnt for sid, cnt in c.fetchall()}

            c.execute(
                f"SELECT substation_id, COUNT(DISTINCT gate) FROM elements WHERE substation_id IN ({placeholders}) "
                "AND gate IS NOT NULL AND gate != '' AND gate NOT LIKE '%-%' GROUP BY substation_id",
                sub_ids
            )
            gate_count_map = {sid: cnt for sid, cnt in c.fetchall()}

            c.execute(
                f"SELECT substation_id, COUNT(*) FROM elements WHERE substation_id IN ({placeholders}) "
                "AND element_type='Διακόπτης ΜΤ' AND is_main_switch=3 GROUP BY substation_id",
                sub_ids
            )
            capacitor_count_map = {sid: cnt for sid, cnt in c.fetchall()}

            c.execute(f"SELECT substation_id, COUNT(*) FROM maintenance WHERE substation_id IN ({placeholders}) GROUP BY substation_id", sub_ids)
            maint_count_map = {sid: cnt for sid, cnt in c.fetchall()}

            c.execute(f"SELECT substation_id, MAX(date_time) FROM maintenance WHERE substation_id IN ({placeholders}) GROUP BY substation_id", sub_ids)
            last_maint_map = {sid: dt for sid, dt in c.fetchall()}

            c.execute(
                f"SELECT substation_id, COUNT(*) FROM elements WHERE substation_id IN ({placeholders}) "
                "AND operating_status='Ανενεργή' GROUP BY substation_id",
                sub_ids
            )
            inactive_count_map = {sid: cnt for sid, cnt in c.fetchall()}
        
        # Create popup window
        popup = reuse_popup if reuse_popup else Popup(title=title, size_hint=(0.95, 0.9))
        popup.title = title
        
        # Create main layout
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # Create scrollable grid for records
        scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        grid = GridLayout(cols=1, spacing=5, size_hint_y=None)
        grid.bind(minimum_height=grid.setter('height'))
        
        if substations:
            for sub_id, sub_name, location, adoption_date, division, monogram_pdf in substations:
                # Substation title in bigger letters
                substation_title = Label(
                    text=f'[b][size=22]{sub_name}[/size][/b]',
                    size_hint_y=None,
                    height=45,
                    markup=True
                )
                grid.add_widget(substation_title)
                
                # Add header for each substation
                header_layout = BoxLayout(size_hint_y=None, height=35, spacing=5)
                header_layout.add_widget(Label(text='Τοποθεσία', bold=True, size_hint_x=0.17))
                header_layout.add_widget(Label(text='Ανάληψη', bold=True, size_hint_x=0.1))
                header_layout.add_widget(Label(text='Στοιχεία', bold=True, size_hint_x=0.07))
                header_layout.add_widget(Label(text='Πύλες', bold=True, size_hint_x=0.07))
                header_layout.add_widget(Label(text='Πυκνωτές', bold=True, size_hint_x=0.07))
                header_layout.add_widget(Label(text='Συντηρήσεις', bold=True, size_hint_x=0.1))
                header_layout.add_widget(Label(text='Τελευταία', bold=True, size_hint_x=0.1))
                header_layout.add_widget(Label(text='Μονογραμμικό', bold=True, size_hint_x=0.12))
                header_layout.add_widget(Label(text='', size_hint_x=0.2))  # Space for buttons
                grid.add_widget(header_layout)
                
                elem_count = elem_count_map.get(sub_id, 0)
                gate_count = gate_count_map.get(sub_id, 0)
                capacitor_count = capacitor_count_map.get(sub_id, 0)
                maint_count = maint_count_map.get(sub_id, 0)
                last_maint = last_maint_map.get(sub_id)
                last_maint_display = last_maint if last_maint else '-'
                
                # Substation row (removed name since it's now a title)
                sub_row_layout = BoxLayout(size_hint_y=None, height=40, spacing=5)
                
                # Location button (clickable)
                if location:
                    location_btn = Button(
                        text='Google Maps Link', 
                        size_hint_x=0.17,
                        font_size='11sp',
                        padding=(5, 5)
                    )
                    # Bind text_size to button size for proper text wrapping
                    location_btn.bind(size=lambda btn, size: setattr(btn, 'text_size', size))
                    location_btn.bind(on_press=lambda x, url=location: webbrowser.open(url))
                    sub_row_layout.add_widget(location_btn)
                else:
                    sub_row_layout.add_widget(Label(text='-', size_hint_x=0.17))
                
                sub_row_layout.add_widget(Label(text=adoption_date or '-', size_hint_x=0.1))
                sub_row_layout.add_widget(Label(text=str(elem_count), size_hint_x=0.07))
                sub_row_layout.add_widget(Label(text=str(gate_count), size_hint_x=0.07))
                sub_row_layout.add_widget(Label(text=str(capacitor_count), size_hint_x=0.07))
                sub_row_layout.add_widget(Label(text=str(maint_count), size_hint_x=0.1))
                sub_row_layout.add_widget(Label(text=last_maint_display, size_hint_x=0.1))

                monogram_btn = Button(
                    text='Άνοιγμα' if monogram_pdf else 'Προσθήκη',
                    size_hint_x=0.12
                )
                if monogram_pdf and os.path.exists(monogram_pdf):
                    monogram_btn.bind(on_press=lambda x, path=monogram_pdf: self._open_monogram_pdf(path))
                else:
                    monogram_btn.bind(on_press=lambda x, sid=sub_id, p=popup, f=filter_name: self._select_monogram_pdf(sid, p, f))
                sub_row_layout.add_widget(monogram_btn)

                sub_row_layout.add_widget(Label(text='', size_hint_x=0.2))  # Empty space to match header
                grid.add_widget(sub_row_layout)
                
                # Edit, Delete, and Maintenance History buttons
                buttons_layout = BoxLayout(size_hint_y=None, height=35, spacing=5)
                
                edit_btn = Button(text='Επεξεργασία', size_hint_x=0.25)
                edit_btn.bind(on_press=lambda x, sid=sub_id, sname=sub_name, loc=location, adate=adoption_date, div=division, p=popup: self.show_edit_substation_popup(sid, sname, loc, adate, div, p))
                buttons_layout.add_widget(edit_btn)
                
                maint_hist_btn = Button(text='Ιστορικό Συντ.', size_hint_x=0.25)
                maint_hist_btn.bind(on_press=lambda x, sid=sub_id, sname=sub_name, p=popup: self.show_substation_maintenance_history(sid, sname, p))
                buttons_layout.add_widget(maint_hist_btn)

                insp_hist_btn = Button(text='Ιστορικό Επιθ.', size_hint_x=0.25)
                insp_hist_btn.bind(on_press=lambda x, sid=sub_id, sname=sub_name, p=popup: self.show_substation_inspection_history(sid, sname, p))
                buttons_layout.add_widget(insp_hist_btn)
                
                delete_sub_btn = Button(text='Διαγραφή', size_hint_x=0.25)
                delete_sub_btn.bind(on_press=lambda x, sid=sub_id, sname=sub_name, p=popup: self.confirm_delete_substation(sid, sname, p))
                buttons_layout.add_widget(delete_sub_btn)
                
                grid.add_widget(buttons_layout)
                
                # Add element button for this substation
                add_elem_btn = Button(
                    text=f"   + Προσθήκη Στοιχείου",
                    size_hint_y=None,
                    height=35
                )
                add_elem_btn.bind(on_press=lambda x, sid=sub_id, sname=sub_name, p=popup: self.show_add_element_popup_for_substation(sid, sname, p))
                grid.add_widget(add_elem_btn)
                
                # Inactive elements button with count
                inactive_count = inactive_count_map.get(sub_id, 0)
                inactive_elem_btn = Button(
                    text=f"   Ανενεργά Στοιχεία ({inactive_count})",
                    size_hint_y=None,
                    height=35
                )
                inactive_elem_btn.bind(on_press=lambda x, sid=sub_id, sname=sub_name, p=popup: self.show_inactive_elements(sid, sname, p))
                grid.add_widget(inactive_elem_btn)

                if not show_elements:
                    view_elements_btn = Button(
                        text=f"   Εμφάνιση ενεργών στοιχείων ({elem_count})",
                        size_hint_y=None,
                        height=35
                    )
                    view_elements_btn.bind(on_press=lambda x, sname=sub_name, p=popup: self._display_substations(sname, p))
                    grid.add_widget(view_elements_btn)
                    continue
                
                # Elements section (only active elements)
                c.execute(
                    "SELECT DISTINCT element_type FROM elements WHERE substation_id=? AND (operating_status IS NULL OR operating_status='Ενεργή') ORDER BY element_type",
                    (sub_id,)
                )
                type_values = ['(Όλα)'] + [row[0] for row in c.fetchall() if row[0]]
                current_type_filter = element_type_filter or '(Όλα)'
                if current_type_filter not in type_values:
                    current_type_filter = '(Όλα)'

                gate_query = "SELECT DISTINCT gate FROM elements WHERE substation_id=? AND (operating_status IS NULL OR operating_status='Ενεργή')"
                gate_params = [sub_id]
                if current_type_filter != '(Όλα)':
                    gate_query += " AND element_type=?"
                    gate_params.append(current_type_filter)
                c.execute(gate_query, gate_params)
                raw_gates = [row[0] for row in c.fetchall()]
                has_unassigned = any(gate is None or str(gate).strip() == '' for gate in raw_gates)
                gate_set = {str(gate).strip() for gate in raw_gates if gate is not None and str(gate).strip() != ''}
                gate_values = ['(Όλα)']
                gate_values.extend(sorted([g for g in gate_set if g.startswith('ΠΥΛΗ')]))
                gate_values.extend(sorted([g for g in gate_set if not g.startswith('ΠΥΛΗ')]))
                if has_unassigned:
                    gate_values.append('(Μη καταχωρημένο)')
                current_gate_filter = gate_filter or '(Όλα)'
                if current_gate_filter not in gate_values:
                    current_gate_filter = '(Όλα)'

                filter_layout = BoxLayout(size_hint_y=None, height=40, spacing=10)
                filter_layout.add_widget(Label(text='Φίλτρο Τύπου:', size_hint_x=0.2))
                type_spinner = Spinner(text=current_type_filter, values=type_values, size_hint_x=0.35)
                filter_layout.add_widget(type_spinner)
                filter_layout.add_widget(Label(text='Φίλτρο Πύλης:', size_hint_x=0.2))
                gate_spinner = Spinner(text=current_gate_filter, values=gate_values, size_hint_x=0.35)
                filter_layout.add_widget(gate_spinner)

                def refresh_filters(_spinner, _text):
                    self._display_substations(sub_name, popup, type_spinner.text, gate_spinner.text)

                type_spinner.bind(text=refresh_filters)
                gate_spinner.bind(text=refresh_filters)
                grid.add_widget(filter_layout)

                # Fetch model data from element_models table
                query = """
                          SELECT e.id, e.element_type, e.name, e.serial_number, e.maintenance_date, 
                              e.voltage_level, e.manufacturer, e.manufacture_year, e.gate, e.is_main_switch,
                           em.breaker_category, em.model_name, em.manufacturer as model_manufacturer, 
                           em.maintenance_cycle, em.installation_space
                    FROM elements e 
                    LEFT JOIN element_models em ON e.element_model_id = em.id 
                    WHERE e.substation_id=? AND (e.operating_status IS NULL OR e.operating_status='Ενεργή') 
                """
                params = [sub_id]
                if current_type_filter != '(Όλα)':
                    query += " AND e.element_type=?"
                    params.append(current_type_filter)
                if current_gate_filter != '(Όλα)':
                    if current_gate_filter == '(Μη καταχωρημένο)':
                        query += " AND (e.gate IS NULL OR e.gate='')"
                    else:
                        query += " AND e.gate=?"
                        params.append(current_gate_filter)
                query += " ORDER BY e.gate"
                c.execute(query, params)
                elements = c.fetchall()
                
                if elements:
                    # Define sort priority for element types
                    def get_element_priority(elem):
                        elem_id, elem_type, elem_name, serial_number, maintenance_date, voltage_level, manufacturer, manufacture_year, gate, is_main_switch, breaker_category, model_name, model_manufacturer, maintenance_cycle, installation_space = elem
                        
                        # Priority order: HV breaker, Transformer, Motor Drive, MV main breaker, MV line breakers, MV capacitor breakers, rest
                        if elem_type == 'Διακόπτης ΥΤ':
                            return (1, elem_name)
                        elif elem_type == 'Μετασχηματιστής 150/20KV':
                            return (2, elem_name)
                        elif elem_type == 'Motor Drive':
                            return (3, elem_name)
                        elif elem_type == 'Διακόπτης ΜΤ' and is_main_switch == 1:  # Main breaker
                            return (4, elem_name)
                        elif elem_type == 'Διακόπτης ΜΤ' and is_main_switch == 2:  # Interconnection breaker
                            return (5, elem_name)
                        elif elem_type == 'Διακόπτης ΜΤ' and is_main_switch == 0:  # Line breaker
                            return (6, elem_name)
                        elif elem_type == 'Διακόπτης ΜΤ' and is_main_switch == 3:  # Capacitor breaker
                            return (7, elem_name)
                        else:
                            return (8, elem_name)
                    
                    # Group elements by gate
                    gates_dict = {}
                    for elem in elements:
                        elem_id, elem_type, elem_name, serial_number, maintenance_date, voltage_level, manufacturer, manufacture_year, gate, is_main_switch, breaker_category, model_name, model_manufacturer, maintenance_cycle, installation_space = elem
                        
                        gate_key = gate if gate else '(Μη καταχωρημένο)'
                        if gate_key not in gates_dict:
                            gates_dict[gate_key] = []
                        gates_dict[gate_key].append(elem)
                    
                    # Sort elements within each gate according to priority
                    for gate_key in gates_dict:
                        gates_dict[gate_key].sort(key=get_element_priority)
                    
                    # Display elements grouped by gate
                    # Show gates in order: ΠΥΛΗ 1, ΠΥΛΗ 2, etc., then unassigned
                    sorted_gates = sorted([g for g in gates_dict.keys() if g.startswith('ΠΥΛΗ')])
                    if '(Μη καταχωρημένο)' in gates_dict:
                        sorted_gates.append('(Μη καταχωρημένο)')
                    
                    for gate_name in sorted_gates:
                        gate_elements = gates_dict[gate_name]
                        
                        # Gate header with count
                        element_count = len(gate_elements)
                        gate_label = Label(
                            text=f"   {gate_name} ({element_count} στοιχεία)",
                            size_hint_y=None,
                            height=35,
                            bold=True,
                            color=(0.2, 0.6, 1, 1)  # Blue color for gate headers
                        )
                        grid.add_widget(gate_label)
                        
                        # Display elements in this gate
                        for j, elem in enumerate(gate_elements, 1):
                            elem_id, elem_type, elem_name, serial_number, maintenance_date, voltage_level, manufacturer, manufacture_year, gate, is_main_switch, breaker_category, model_name, model_manufacturer, maintenance_cycle, installation_space = elem
                            
                            # Check if maintenance is overdue or missing
                            from datetime import datetime, timedelta
                            is_overdue = False
                            if maintenance_cycle and maintenance_cycle > 0:
                                if not maintenance_date or maintenance_date.strip() == '':
                                    # Missing maintenance date when cycle is defined
                                    is_overdue = True
                                else:
                                    try:
                                        last_maint = datetime.strptime(maintenance_date.split()[0], '%Y-%m-%d')
                                        years_ago = datetime.now() - timedelta(days=maintenance_cycle * 365)
                                        if last_maint < years_ago:
                                            is_overdue = True
                                    except:
                                        pass
                            
                            # Format maintenance date with color if overdue or missing
                            if is_overdue:
                                maint_display = f"[color=ff0000][b]Τελ. Συντ.: {maintenance_date or '-'}[/b][/color]"
                            else:
                                maint_display = f"Τελ. Συντ.: {maintenance_date or '-'}"
                            
                            # Create element text with multiple lines for better readability
                            # Add breaker type label for circuit breakers
                            if elem_type in ['Διακόπτης ΥΤ', 'Διακόπτης ΜΤ']:
                                if elem_type == 'Διακόπτης ΥΤ':
                                    breaker_type_label = 'Κεντρικός'
                                elif is_main_switch == 1:
                                    breaker_type_label = 'Κεντρικός'
                                elif is_main_switch == 2:
                                    breaker_type_label = 'Διασυνδετικός'
                                elif is_main_switch == 3:
                                    breaker_type_label = 'Διακόπτης Πυκνωτών'
                                else:
                                    breaker_type_label = 'Γραμμής'
                                elem_type = f"{elem_type} ({breaker_type_label})"
                            
                            breaker_info = f" | {breaker_category}" if breaker_category else ""
                            manufacture_info = f" | Έτος: {manufacture_year}" if manufacture_year else ""
                            elem_text = f"   {j}. [b][size=18]{elem_name}[/size][/b] - {elem_type}{breaker_info}\n      S/N: {serial_number or '-'}{manufacture_info}\n      Κατ.: {model_manufacturer or manufacturer or '-'} | Μοντ.: {model_name or '-'} | Χώρος: {installation_space or '-'} | Τάση: {voltage_level or '-'}\n      Κύκλος: {maintenance_cycle or '-'} έτη | {maint_display}"
                            
                            # Create a horizontal layout for element and buttons
                            elem_layout = BoxLayout(size_hint_y=None, spacing=5)
                            elem_layout.bind(minimum_height=elem_layout.setter('height'))
                            
                            elem_label = Label(
                                text=elem_text,
                                size_hint=(0.75, None),
                                markup=True
                            )
                            # Enable text wrapping and automatic height calculation
                            elem_label.bind(
                                width=lambda instance, value: setattr(instance, 'text_size', (value, None)),
                                texture_size=lambda instance, value: (
                                    setattr(instance, 'height', max(70, value[1] + 10)),
                                    setattr(elem_layout, 'height', max(70, value[1] + 10))
                                )
                            )
                            elem_layout.add_widget(elem_label)
                            
                            # Button container
                            btn_box = BoxLayout(size_hint_x=0.25, spacing=3)
                            
                            edit_elem_btn = Button(
                                text="Επεξ.",
                                size_hint_x=0.5
                            )
                            edit_elem_btn.bind(on_press=lambda x, eid=elem_id, sid=sub_id, sname=sub_name, p=popup: self.show_edit_element_popup(eid, sid, p, sname))
                            btn_box.add_widget(edit_elem_btn)
                            
                            delete_elem_btn = Button(
                                text="Διαγρ.",
                                size_hint_x=0.5
                            )
                            delete_elem_btn.bind(on_press=lambda x, eid=elem_id, ename=elem_name, sid=sub_id, sname=sub_name, p=popup: self.confirm_delete_element(eid, ename, sid, p, sname))
                            btn_box.add_widget(delete_elem_btn)
                            
                            elem_layout.add_widget(btn_box)
                            
                            grid.add_widget(elem_layout)
                else:
                    no_elem_label = Label(
                        text="   (Χωρίς στοιχεία)",
                        size_hint_y=None,
                        height=30
                    )
                    grid.add_widget(no_elem_label)
                
                # Add spacing between substations
                spacing_widget = Label(text='', size_hint_y=None, height=30)
                grid.add_widget(spacing_widget)
        else:
            empty_label = Label(
                text='Κενή βάση',
                size_hint_y=None,
                height=40
            )
            grid.add_widget(empty_label)
        
        scroll.add_widget(grid)
        main_layout.add_widget(scroll)
        
        # Add buttons layout
        buttons_bottom_layout = BoxLayout(size_hint_y=0.1, spacing=10)
        
        add_substation_btn = Button(text='+ Προσθήκη Υποσταθμού')
        add_substation_btn.bind(on_press=lambda x: self.show_add_substation_popup_from_db_view(popup))
        buttons_bottom_layout.add_widget(add_substation_btn)
        
        close_btn = Button(text='Κλείσιμο')
        close_btn.bind(on_press=popup.dismiss)
        buttons_bottom_layout.add_widget(close_btn)
        
        main_layout.add_widget(buttons_bottom_layout)
        
        popup.content = main_layout
        if not reuse_popup:
            popup.open()

    def create_substations_template(self, instance):
        success, message = create_substations_template(os.path.dirname(__file__))
        title = 'Template Υποσταθμών' if success else 'Σφάλμα'
        show_message_popup(title, message)
    
    def _create_file_import_dialog(self, title, import_callback):
        """Generic file import dialog for substations and elements
        
        Args:
            title: Popup title
            import_callback: Function to call with file_path when import is confirmed
        """
        popup = Popup(title=title, size_hint=(0.9, 0.9))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # Path input
        path_label = Label(text='Διαδρομή αρχείου:', size_hint_y=0.1)
        layout.add_widget(path_label)
        
        path_input = TextInput(
            hint_text='Διαδρομή αρχείου',
            size_hint_y=0.15,
            multiline=False
        )
        layout.add_widget(path_input)
        
        # File chooser with default path
        layout.add_widget(Label(text='Ή επιλέξτε από τη λίστα:', size_hint_y=0.1))
        chooser = FileChooserListView(filters=['*.xlsx', '*.csv'], path=os.path.dirname(__file__))
        layout.add_widget(chooser)
        
        # Buttons
        buttons_layout = BoxLayout(size_hint_y=0.1, spacing=10)
        
        def import_file():
            file_path = path_input.text.strip() if path_input.text.strip() else (chooser.selection[0] if chooser.selection else None)
            
            if not file_path:
                show_message_popup('Σφάλμα', 'Παρακαλώ εισάγετε διαδρομή ή επιλέξτε αρχείο!')
                return
            
            if not os.path.exists(file_path):
                show_message_popup('Σφάλμα', 'Το αρχείο δεν βρέθηκε!')
                return
            
            import_callback(file_path)
            popup.dismiss()
        
        import_btn = Button(text='Εισαγωγή')
        import_btn.bind(on_press=lambda x: import_file())
        buttons_layout.add_widget(import_btn)
        
        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        buttons_layout.add_widget(cancel_btn)
        
        layout.add_widget(buttons_layout)
        popup.content = layout
        popup.open()

    def _open_monogram_pdf(self, pdf_path):
        if not pdf_path or not os.path.exists(pdf_path):
            show_message_popup('Σφάλμα', 'Το αρχείο δεν βρέθηκε!')
            return
        try:
            import subprocess
            import sys
            if sys.platform == 'win32':
                os.startfile(pdf_path)
            elif sys.platform == 'darwin':
                subprocess.call(['open', pdf_path])
            else:
                subprocess.call(['xdg-open', pdf_path])
        except Exception as e:
            show_message_popup('Σφάλμα', f'Αποτυχία ανοίγματος PDF:\n{str(e)}')

    def _select_monogram_pdf(self, substation_id, parent_popup=None, filter_name=None):
        popup = Popup(title='Επιλογή Μονογραμμικού PDF', size_hint=(0.9, 0.9))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        path_label = Label(text='Διαδρομή αρχείου:', size_hint_y=0.1)
        layout.add_widget(path_label)

        path_input = TextInput(
            hint_text='Διαδρομή αρχείου',
            size_hint_y=0.12,
            multiline=False
        )
        layout.add_widget(path_input)

        layout.add_widget(Label(text='Ή επιλέξτε από τη λίστα:', size_hint_y=0.1))
        chooser = FileChooserListView(filters=['*.pdf'], path=os.path.dirname(__file__))
        layout.add_widget(chooser)

        buttons_layout = BoxLayout(size_hint_y=0.12, spacing=10)

        def save_file():
            file_path = path_input.text.strip() if path_input.text.strip() else (chooser.selection[0] if chooser.selection else None)

            if not file_path:
                show_message_popup('Σφάλμα', 'Παρακαλώ εισάγετε διαδρομή ή επιλέξτε αρχείο!')
                return

            if not os.path.exists(file_path):
                show_message_popup('Σφάλμα', 'Το αρχείο δεν βρέθηκε!')
                return

            if not file_path.lower().endswith('.pdf'):
                show_message_popup('Σφάλμα', 'Παρακαλώ επιλέξτε αρχείο PDF!')
                return

            c = self.conn.cursor()
            c.execute("UPDATE substations SET monogram_pdf=? WHERE id=?", (file_path, substation_id))
            self.conn.commit()
            popup.dismiss()

            if parent_popup:
                parent_popup.dismiss()
            self._display_substations(filter_name)

        save_btn = Button(text='Αποθήκευση')
        save_btn.bind(on_press=lambda x: save_file())
        buttons_layout.add_widget(save_btn)

        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        buttons_layout.add_widget(cancel_btn)

        layout.add_widget(buttons_layout)
        popup.content = layout
        popup.open()
    
    def create_elements_template(self, instance):
        success, message = create_elements_template(os.path.dirname(__file__))
        title = 'Template Στοιχείων' if success else 'Σφάλμα'
        show_message_popup(title, message)
    
    def show_import_substations_dialog(self, instance):
        self._create_file_import_dialog(
            'Εισαγωγή υποσταθμών από αρχείο',
            self.import_substations_from_file
        )
    
    def show_import_elements_dialog(self, instance):
        self._create_file_import_dialog(
            'Εισαγωγή στοιχείων από αρχείο',
            self.import_elements_from_file
        )

    def show_import_android_changes_dialog(self, instance):
        self._create_android_changes_import_dialog(
            'Εισαγωγή αλλαγών από Android',
            self.import_android_changes_from_file
        )

    def _create_android_changes_import_dialog(self, title, import_callback):
        popup = Popup(title=title, size_hint=(0.9, 0.9))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        path_label = Label(text='Διαδρομή αρχείου change log (.jsonl):', size_hint_y=0.1)
        layout.add_widget(path_label)

        path_input = TextInput(
            hint_text='Διαδρομή αρχείου',
            size_hint_y=0.15,
            multiline=False
        )
        layout.add_widget(path_input)

        layout.add_widget(Label(text='Ή επιλέξτε από τη λίστα:', size_hint_y=0.1))
        chooser = FileChooserListView(filters=['*.jsonl', '*.json'], path=os.path.dirname(__file__))
        layout.add_widget(chooser)

        buttons_layout = BoxLayout(size_hint_y=0.1, spacing=10)

        def import_file():
            file_path = path_input.text.strip() if path_input.text.strip() else (chooser.selection[0] if chooser.selection else None)

            if not file_path:
                show_message_popup('Σφάλμα', 'Παρακαλώ εισάγετε διαδρομή ή επιλέξτε αρχείο!')
                return

            if not os.path.exists(file_path):
                show_message_popup('Σφάλμα', 'Το αρχείο δεν βρέθηκε!')
                return

            import_callback(file_path)
            popup.dismiss()

        import_btn = Button(text='Εισαγωγή')
        import_btn.bind(on_press=lambda x: import_file())
        buttons_layout.add_widget(import_btn)

        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        buttons_layout.add_widget(cancel_btn)

        layout.add_widget(buttons_layout)
        popup.content = layout
        popup.open()

    def import_android_changes_from_file(self, file_path):
        show_message_popup('Εισαγωγή αλλαγών από Android', f'Λήφθηκε αρχείο:\n{file_path}\n\n(Η εφαρμογή αλλαγών θα υλοποιηθεί στο επόμενο βήμα)')
    
    def import_substations_from_file(self, file_path):
        def on_success(message):
            show_message_popup('Εισαγωγή Υποσταθμών', message, callback=lambda: self.show_records(None))

        def on_error(message):
            show_message_popup('Σφάλμα', message)

        if file_path.endswith('.xlsx'):
            import_substations_from_excel(self.conn, file_path, on_success, on_error)
        elif file_path.endswith('.csv'):
            import_substations_from_csv(self.conn, file_path, on_success, on_error)
        else:
            on_error('Μη υποστηριζόμενη μορφή αρχείου')
    
    def _read_elements_file(self, file_path):
        """Read elements file, handling TEMPLATE_VERSION row if present
        
        Returns:
            pd.DataFrame: The elements dataframe with proper headers
            
        Raises:
            ValueError: If file format is not supported
        """
        import pandas as pd
        
        if file_path.endswith('.xlsx'):
            # Peek at first row to check for version
            df_peek = pd.read_excel(file_path, sheet_name='Elements', nrows=1, header=None)
            first_cell = str(df_peek.iloc[0, 0]) if len(df_peek) > 0 else ''
            
            if 'Version:' in first_cell or 'TEMPLATE_VERSION:' in first_cell:
                return pd.read_excel(file_path, sheet_name='Elements', skiprows=1)
            else:
                return pd.read_excel(file_path, sheet_name='Elements')
                
        elif file_path.endswith('.csv'):
            # Peek at first row to check for version
            df_peek = pd.read_csv(file_path, nrows=1, header=None)
            first_cell = str(df_peek.iloc[0, 0]) if len(df_peek) > 0 else ''
            
            if 'Version:' in first_cell or 'TEMPLATE_VERSION:' in first_cell:
                return pd.read_csv(file_path, skiprows=1)
            else:
                return pd.read_csv(file_path)
        else:
            raise ValueError('Μη υποστηριζόμενη μορφή αρχείου')
    
    def _load_models_for_element_type(self, element_category, breaker_category=None, selected_model_id=None):
        """Load and filter models for a specific element category
        
        Args:
            element_category: The element type (e.g., 'Διακόπτης ΜΤ')
            breaker_category: Optional breaker category filter for circuit breakers
            selected_model_id: Optional model ID to pre-select
            
        Returns:
            tuple: (models_data dict, display_names list, selected_display_name)
        """
        c = self.conn.cursor()
        c.execute("SELECT id, model_name, manufacturer, maintenance_cycle, installation_space, breaker_category FROM element_models WHERE element_category=? ORDER BY model_name", (element_category,))
        models = c.fetchall()
        
        models_data = {}
        display_names = []
        selected_display_name = None
        
        # Filter models for circuit breakers by breaker category
        if element_category in ['Διακόπτης ΜΤ', 'Διακόπτης ΥΤ'] and breaker_category:
            filtered_models = [
                m for m in models 
                if (m[5] or 'Other').strip().lower() == breaker_category.lower()
            ]
        else:
            filtered_models = models
        
        # Build display names and data dictionary
        for m in filtered_models:
            display_name = f"{m[1]} - {m[2] or 'N/A'}"
            display_names.append(display_name)
            models_data[display_name] = {
                'id': m[0],
                'model_name': m[1],
                'manufacturer': m[2] or '',
                'maintenance_cycle': m[3] or 0,
                'installation_space': m[4] or '',
                'breaker_category': m[5] or ''
            }
            if m[0] == selected_model_id:
                selected_display_name = display_name
        
        return models_data, display_names, selected_display_name
    
    def import_elements_from_file(self, file_path):
        """Import elements with validation wizard: Step 1 - Column Mapping, Step 2 - Data Validation"""
        try:
            df_elem = self._read_elements_file(file_path)
            
            # Show Step 1: Column Mapping Wizard
            column_wizard = ColumnMappingPopup(
                df_columns=list(df_elem.columns),
                df=df_elem,
                conn=self.conn,
                on_continue=lambda mapping: self._on_column_mapping_complete(file_path, df_elem, mapping),
                on_cancel=lambda: None  # Just close
            )
            column_wizard.show()

        except Exception as e:
            show_message_popup('Σφάλμα', f'Σφάλμα κατά τον έλεγχο: {e}')
    
    def _on_column_mapping_complete(self, file_path, df, column_mapping):
        """Callback after column mapping is complete - show validation wizard"""
        # Show Step 2: Data Validation Wizard
        validation_wizard = DataValidationPopup(
            df=df,
            column_mapping=column_mapping,
            conn=self.conn,
            on_continue=lambda corrected_df, mapping: self._on_validation_complete(file_path, corrected_df, mapping),
            on_cancel=lambda: None,  # Just close
            on_back=lambda: self.import_elements_from_file(file_path)  # Go back to step 1
        )
        validation_wizard.show()
    
    def _on_validation_complete(self, file_path, df, column_mapping):
        """Callback after validation is complete - proceed with traditional flow"""
        # Rename columns to match expected names
        reverse_mapping = {v: k for k, v in column_mapping.items()}
        df_renamed = df.rename(columns=reverse_mapping)
        
        # Save the corrected dataframe back to temporary file
        import tempfile
        import pandas as pd
        
        # Create a temporary file with corrected data
        with tempfile.NamedTemporaryFile(mode='w', suffix='.xlsx', delete=False, encoding='utf-8') as tmp_file:
            temp_path = tmp_file.name
        
        # Write corrected data
        with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
            df_renamed.to_excel(writer, sheet_name='Elements', index=False)
        
        # Now proceed with traditional flow (check substations, models, duplicates)
        self._check_substations_and_proceed(temp_path, original_file=file_path)

    def _check_substations_and_proceed(self, file_path, original_file=None):
        """Check for new substations after validation"""
        try:
            import pandas as pd
            cursor = self.conn.cursor()
            df_elem = self._read_elements_file(file_path)

            # Check for new substations
            new_substations = set()
            for _, row in df_elem.iterrows():
                sub_name = str(row.get('Substation Name', '')).strip() if pd.notna(row.get('Substation Name', '')) else ''
                if sub_name:
                    cursor.execute('SELECT id FROM substations WHERE name=?', (sub_name,))
                    if not cursor.fetchone():
                        new_substations.add(sub_name)
            
            # If new substations found, prompt user
            if new_substations:
                self._show_new_substations_prompt(file_path, new_substations)
            else:
                # No new substations, proceed to check duplicates
                self._check_duplicates_and_import(file_path)

        except Exception as e:
            show_message_popup('Σφάλμα', f'Σφάλμα κατά τον έλεγχο: {e}')

    def _show_new_substations_prompt(self, file_path, new_substations):
        """Prompt user to confirm creation of new substations"""
        from kivy.uix.popup import Popup
        from kivy.uix.boxlayout import BoxLayout
        from kivy.uix.button import Button
        from kivy.uix.label import Label
        from kivy.uix.scrollview import ScrollView
        
        popup = Popup(title='Νέοι Υποσταθμοί Εντοπίστηκαν', size_hint=(0.6, 0.5))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        sub_list = '\\n'.join(f'• {sub}' for sub in sorted(new_substations))
        message = f'Οι παρακάτω υποσταθμοί δεν υπάρχουν και θα δημιουργηθούν:\\n\\n{sub_list}\\n\\nΣυνέχεια;'
        
        scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        label = Label(text=message, size_hint_y=None)
        label.bind(texture_size=label.setter('size'))
        scroll.add_widget(label)
        layout.add_widget(scroll)
        
        # Buttons
        btn_layout = BoxLayout(size_hint_y=0.2, spacing=10)
        
        yes_btn = Button(text='Ναι, Δημιουργία')
        yes_btn.bind(on_press=lambda x: self._create_substations_and_continue(file_path, new_substations, popup))
        btn_layout.add_widget(yes_btn)
        
        no_btn = Button(text='Ακύρωση')
        no_btn.bind(on_press=popup.dismiss)
        btn_layout.add_widget(no_btn)
        
        layout.add_widget(btn_layout)
        popup.content = layout
        popup.open()
    
    def _create_substations_and_continue(self, file_path, new_substations, prompt_popup):
        """Create new substations and continue with import"""
        cursor = self.conn.cursor()
        for sub_name in new_substations:
            cursor.execute('INSERT INTO substations (name) VALUES (?)', (sub_name,))
        self.conn.commit()
        prompt_popup.dismiss()
        
        # Now proceed to check duplicates
        self._check_duplicates_and_import(file_path)
    
    def _check_duplicates_and_import(self, file_path):
        """Check for models first, then duplicate elements, and proceed with import"""
        try:
            import pandas as pd
            cursor = self.conn.cursor()
            df_elem = self._read_elements_file(file_path)

            # First check models
            models_to_check = {}  # Key: (element_type, model_name, manufacturer), Value: {cycle, space}
            for _, row in df_elem.iterrows():
                element_type = str(row.get('Element Type', '')).strip() if pd.notna(row.get('Element Type', '')) else ''
                model_name = str(row.get('Model Name', '')).strip() if pd.notna(row.get('Model Name', '')) else ''
                model_manufacturer = str(row.get('Model Manufacturer', '')).strip() if pd.notna(row.get('Model Manufacturer', '')) else ''
                model_cycle = int(row.get('Model Maintenance Cycle', 0)) if pd.notna(row.get('Model Maintenance Cycle', '')) else 0
                model_space = str(row.get('Model Installation Space', '')).strip() if pd.notna(row.get('Model Installation Space', '')) else ''
                
                if model_name:  # Only check if model info provided
                    key = (element_type, model_name, model_manufacturer)
                    if key not in models_to_check:
                        models_to_check[key] = {
                            'cycle': model_cycle,
                            'space': model_space
                        }
            
            # Check which models exist and which need to be added/updated
            new_models = []
            conflicting_models = []
            
            for (elem_type, model_name, manufacturer), model_data in models_to_check.items():
                cursor.execute(
                    'SELECT id, maintenance_cycle, installation_space FROM element_models WHERE element_category=? AND model_name=? AND manufacturer=?',
                    (elem_type, model_name, manufacturer)
                )
                existing = cursor.fetchone()
                
                if existing:
                    existing_id, existing_cycle, existing_space = existing
                    # Check if data differs
                    if (existing_cycle != model_data['cycle'] or 
                        (existing_space or '') != (model_data['space'] or '')):
                        conflicting_models.append({
                            'category': elem_type,
                            'name': model_name,
                            'manufacturer': manufacturer,
                            'existing': {'cycle': existing_cycle, 'space': existing_space},
                            'new': model_data
                        })
                else:
                    new_models.append({
                        'category': elem_type,
                        'name': model_name,
                        'manufacturer': manufacturer,
                        'data': model_data
                    })
            
            # If there are model issues, prompt user
            if new_models or conflicting_models:
                self._show_model_check_popup(file_path, new_models, conflicting_models)
            else:
                # No model issues, proceed to check element duplicates
                self._check_element_duplicates(file_path)

        except Exception as e:
            show_message_popup('Σφάλμα', f'Σφάλμα κατά τον έλεγχο: {e}')
    
    def _show_model_check_popup(self, file_path, new_models, conflicting_models):
        """Show popup for user to review and approve model changes"""
        from kivy.uix.popup import Popup
        from kivy.uix.boxlayout import BoxLayout
        from kivy.uix.button import Button
        from kivy.uix.label import Label
        from kivy.uix.scrollview import ScrollView
        from kivy.uix.gridlayout import GridLayout
        
        popup = Popup(title='Έλεγχος Μοντέλων', size_hint=(0.85, 0.85))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        content = GridLayout(cols=1, spacing=10, size_hint_y=None, padding=10)
        content.bind(minimum_height=content.setter('height'))
        
        if new_models:
            content.add_widget(Label(
                text='[b]Νέα Μοντέλα (θα προστεθούν):[/b]',
                size_hint_y=None,
                height=30,
                markup=True
            ))
            for model in new_models:
                text = f"• {model['category']} - {model['name']} ({model['manufacturer']})\n  Κύκλος: {model['data']['cycle']} μήνες, Χώρος: {model['data']['space'] or 'N/A'}"
                content.add_widget(Label(
                    text=text,
                    size_hint_y=None,
                    height=50
                ))
        
        if conflicting_models:
            content.add_widget(Label(
                text='[b]Υπάρχοντα Μοντέλα με Διαφορετικά Δεδομένα:[/b]',
                size_hint_y=None,
                height=30,
                markup=True,
                color=(1, 0.5, 0, 1)
            ))
            for model in conflicting_models:
                text = f"• {model['category']} - {model['name']} ({model['manufacturer']})\n  Υπάρχον: Κύκλος {model['existing']['cycle']}, Χώρος {model['existing']['space'] or 'N/A'}\n  Νέο: Κύκλος {model['new']['cycle']}, Χώρος {model['new']['space'] or 'N/A'}"
                content.add_widget(Label(
                    text=text,
                    size_hint_y=None,
                    height=80,
                    color=(1, 0.7, 0, 1)
                ))
        
        scroll.add_widget(content)
        layout.add_widget(scroll)
        
        # Instructions
        if conflicting_models:
            layout.add_widget(Label(
                text='Επιλέξτε "Ενημέρωση" για να αντικαταστήσετε τα υπάρχοντα δεδομένα ή "Χρήση Υπαρχόντων" για να τα κρατήσετε.',
                size_hint_y=0.1
            ))
        
        # Buttons
        btn_layout = BoxLayout(size_hint_y=0.1, spacing=10)
        
        if conflicting_models:
            update_btn = Button(text='Ενημέρωση Μοντέλων')
            update_btn.bind(on_press=lambda x: self._apply_models_and_continue(file_path, new_models, conflicting_models, True, popup))
            btn_layout.add_widget(update_btn)
            
            keep_btn = Button(text='Χρήση Υπαρχόντων')
            keep_btn.bind(on_press=lambda x: self._apply_models_and_continue(file_path, new_models, conflicting_models, False, popup))
            btn_layout.add_widget(keep_btn)
        else:
            continue_btn = Button(text='Συνέχεια')
            continue_btn.bind(on_press=lambda x: self._apply_models_and_continue(file_path, new_models, [], False, popup))
            btn_layout.add_widget(continue_btn)
        
        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        btn_layout.add_widget(cancel_btn)
        
        layout.add_widget(btn_layout)
        popup.content = layout
        popup.open()
    
    def _apply_models_and_continue(self, file_path, new_models, conflicting_models, update_conflicts, prompt_popup):
        """Apply model changes and continue with element import"""
        cursor = self.conn.cursor()
        
        # Add new models
        for model in new_models:
            cursor.execute(
                'INSERT INTO element_models (element_category, model_name, manufacturer, maintenance_cycle, installation_space) VALUES (?, ?, ?, ?, ?)',
                (model['category'], model['name'], model['manufacturer'], model['data']['cycle'], model['data']['space'])
            )
        
        # Update conflicting models if user chose to
        if update_conflicts:
            for model in conflicting_models:
                cursor.execute(
                    'UPDATE element_models SET maintenance_cycle=?, installation_space=? WHERE element_category=? AND model_name=? AND manufacturer=?',
                    (model['new']['cycle'], model['new']['space'], model['category'], model['name'], model['manufacturer'])
                )
        
        self.conn.commit()
        prompt_popup.dismiss()
        
        # Now check element duplicates
        self._check_element_duplicates(file_path)
    
    def _check_element_duplicates(self, file_path):
        """Check for duplicate elements after models are handled"""
        try:
            import pandas as pd
            cursor = self.conn.cursor()
            df_elem = self._read_elements_file(file_path)

            duplicates = []  # list of tuples (sub_name, name, serial)
            for _, row in df_elem.iterrows():
                sub_name = row.get('Substation Name', '')
                name = str(row.get('Name', '')) if pd.notna(row.get('Name', '')) else ''
                serial_number = str(row.get('Serial Number', '')) if pd.notna(row.get('Serial Number', '')) else ''

                if sub_name and name:
                    cursor.execute('SELECT id FROM substations WHERE name=?', (str(sub_name),))
                    result = cursor.fetchone()
                    if result:
                        sub_id = result[0]
                        cursor.execute(
                            'SELECT id FROM elements WHERE substation_id=? AND name=? AND serial_number=?',
                            (sub_id, name, serial_number)
                        )
                        if cursor.fetchone():
                            duplicates.append((str(sub_name), name, serial_number))

            if duplicates:
                self._show_duplicate_choice_popup(file_path, duplicates)
            else:
                self._proceed_with_import(file_path, default_choice=None, decisions={})

        except Exception as e:
            show_message_popup('Σφάλμα', f'Σφάλμα κατά τον έλεγχο: {e}')

    def _show_duplicate_choice_popup(self, file_path, duplicates_list):
        # User chooses per-duplicate replace/skip, plus replace-all / skip-all shortcuts
        popup = Popup(title='Διπλότυπα Στοιχεία Εντοπίστηκαν', size_hint=(0.9, 0.85))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        instructions = Label(
            text='Επιλέξτε για κάθε διπλότυπο αν θα αντικατασταθεί ή θα παραλειφθεί.\nΜπορείτε να επιλέξετε "Αντικατάσταση Όλων" ή "Παράλειψη Όλων".',
            size_hint_y=None,
            height=60
        )
        layout.add_widget(instructions)

        # State
        decisions = {}  # key: (sub_name, name, serial) -> True/False
        default_choice = {'value': None}  # True replace all, False skip all
        manual_choice_made = {'value': False}  # track if any per-item choice occurred

        # Scrollable list
        scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        grid = GridLayout(cols=1, spacing=6, size_hint_y=None)
        grid.bind(minimum_height=grid.setter('height'))

        def update_continue_state():
            # Enable continue only if all decisions made or default chosen
            if default_choice['value'] is not None:
                continue_btn.disabled = False
                return
            continue_btn.disabled = len(decisions) < len(duplicates_list)

        def disable_global_buttons():
            btn_replace_all.disabled = True
            btn_skip_all.disabled = True

        def make_row(sub_name, name, serial):
            row = BoxLayout(size_hint_y=None, height=50, spacing=8)
            label_text = f"{name} (S/N: {serial or '-'}), Υποστ.: {sub_name}"
            row.add_widget(Label(text=label_text, size_hint_x=0.6))

            key = (sub_name, name, serial)

            def set_decision(val, btn_replace, btn_skip):
                decisions[key] = val
                manual_choice_made['value'] = True
                # Gray out both buttons after selection and color to show choice
                btn_replace.disabled = True
                btn_skip.disabled = True
                if val:
                    btn_replace.background_color = (0.6, 1, 0.6, 1)  # light green
                    btn_skip.background_color = (0.7, 0.7, 0.7, 1)
                else:
                    btn_skip.background_color = (1, 0.6, 0.6, 1)    # light red
                    btn_replace.background_color = (0.7, 0.7, 0.7, 1)
                disable_global_buttons()
                update_continue_state()

            replace_btn = Button(text='Αντικατάσταση', size_hint_x=0.2)
            skip_btn = Button(text='Παράλειψη', size_hint_x=0.2)
            replace_btn.bind(on_press=lambda _x, br=replace_btn, bs=skip_btn: set_decision(True, br, bs))
            skip_btn.bind(on_press=lambda _x, br=replace_btn, bs=skip_btn: set_decision(False, br, bs))

            row.add_widget(replace_btn)
            row.add_widget(skip_btn)
            return row

        for sub_name, name, serial in duplicates_list:
            grid.add_widget(make_row(sub_name, name, serial))

        scroll.add_widget(grid)
        layout.add_widget(scroll)

        # Global buttons
        buttons_all = BoxLayout(size_hint_y=None, height=50, spacing=10)

        def choose_all(val):
            default_choice['value'] = val
            # set all decisions too
            for tup in duplicates_list:
                decisions[tup] = val
            # Gray out all buttons visually by disabling continue gating
            update_continue_state()
            # disable global buttons once used
            btn_replace_all.disabled = True
            btn_skip_all.disabled = True

        btn_replace_all = Button(text='Αντικατάσταση Όλων')
        btn_replace_all.bind(on_press=lambda _x: choose_all(True))
        buttons_all.add_widget(btn_replace_all)

        btn_skip_all = Button(text='Παράλειψη Όλων')
        btn_skip_all.bind(on_press=lambda _x: choose_all(False))
        buttons_all.add_widget(btn_skip_all)

        layout.add_widget(buttons_all)

        # Action buttons
        actions = BoxLayout(size_hint_y=None, height=50, spacing=10)

        def on_continue(_x):
            if default_choice['value'] is None and len(decisions) < len(duplicates_list):
                show_message_popup('Σφάλμα', 'Ολοκληρώστε τις επιλογές για όλα τα διπλότυπα ή χρησιμοποιήστε "Αντικατάσταση Όλων" / "Παράλειψη Όλων".')
                return
            popup.dismiss()
            self._proceed_with_import(file_path, default_choice=default_choice['value'], decisions=decisions)

        def on_cancel(_x):
            popup.dismiss()

        continue_btn = Button(text='Συνέχεια', disabled=True)
        continue_btn.bind(on_press=on_continue)
        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=on_cancel)

        actions.add_widget(continue_btn)
        actions.add_widget(cancel_btn)
        layout.add_widget(actions)

        popup.content = layout
        popup.open()

        # Initial state
        update_continue_state()

    def _proceed_with_import(self, file_path, default_choice=None, decisions=None):
        decisions = decisions or {}

        def on_success(message):
            show_message_popup('Εισαγωγή Στοιχείων', message, callback=lambda: self.show_records(None))

        def on_error(message):
            show_message_popup('Σφάλμα', message)

        # Resolver passed to importer per duplicate
        def on_duplicate(sub_name, name, serial_number):
            key = (sub_name, name, serial_number)
            if key in decisions:
                return decisions[key]
            if default_choice is not None:
                return default_choice
            return False  # safe default

        if file_path.endswith('.xlsx'):
            import_elements_from_excel(self.conn, file_path, on_success, on_error, on_duplicate)
        elif file_path.endswith('.csv'):
            import_elements_from_csv(self.conn, file_path, on_success, on_error, on_duplicate)
        else:
            on_error('Μη υποστηριζόμενη μορφή αρχείου')

    def show_edit_element_popup(self, element_id, substation_id, parent_popup, substation_name=None, grandparent_popup=None):
        """Show popup to edit an existing element"""
        from kivy.uix.popup import Popup
        from kivy.uix.boxlayout import BoxLayout
        from kivy.uix.button import Button
        from kivy.uix.label import Label
        from kivy.uix.textinput import TextInput
        from kivy.uix.spinner import Spinner
        from kivy.uix.scrollview import ScrollView
        
        # Fetch element data
        c = self.conn.cursor()
        c.execute("SELECT element_type, name, serial_number, maintenance_date, voltage_level, manufacturer, model, model_version, installation_space, operating_status, maintenance_cycle, manufacture_year, element_model_id, gate, is_main_switch, breaker_category FROM elements WHERE id=?", (element_id,))
        element = c.fetchone()
        
        if not element:
            show_message_popup('Σφάλμα', 'Το στοιχείο δεν βρέθηκε!')
            return
        
        elem_type, name, serial_num, maint_date, voltage_level, manufacturer, model, model_version, install_space, op_status, maint_cycle, manuf_year, model_id, gate, is_main_switch, breaker_category = element
        
        popup = Popup(title=f'Επεξεργασία: {name}', size_hint=(0.9, 0.9))
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        layout = BoxLayout(orientation='vertical', size_hint_y=None, padding=5, spacing=8)
        layout.bind(minimum_height=layout.setter('height'))
        
        # Element type (read-only)
        layout.add_widget(Label(text=f'Τύπος: {elem_type}', size_hint_y=None, height=30, bold=True))

        # Voltage level (dropdown)
        layout.add_widget(Label(text='Επίπεδο Τάσης:', size_hint_y=None, height=30))
        current_voltage = voltage_level or self._derive_voltage_level(elem_type) or '(Κενό)'
        voltage_options = list(self.VOLTAGE_LEVELS)
        if current_voltage not in voltage_options:
            voltage_options.append(current_voltage)
        voltage_level_spinner = Spinner(
            text=current_voltage,
            values=voltage_options,
            size_hint_y=None,
            height=40
        )
        layout.add_widget(voltage_level_spinner)
        
        # Model selection
        # Breaker category filter (only for circuit breakers)
        breaker_category_label = Label(text='Κατηγορία Διακόπτη:', size_hint_y=None, height=30)
        breaker_category_options = self._get_breaker_categories_for_element_type(elem_type)
        breaker_category_text = breaker_category if breaker_category in breaker_category_options else (breaker_category_options[0] if breaker_category_options else 'SF6')
        breaker_category_spinner = Spinner(
            text=breaker_category_text,
            values=breaker_category_options,
            size_hint_y=None,
            height=40
        )
        
        if elem_type in ['Διακόπτης ΜΤ', 'Διακόπτης ΥΤ']:
            layout.add_widget(breaker_category_label)
            layout.add_widget(breaker_category_spinner)
        
        layout.add_widget(Label(text='Μοντέλο:', size_hint_y=None, height=30))
        
        # Load all models for this element type
        c.execute("SELECT id, model_name, manufacturer, breaker_category FROM element_models WHERE element_category=? ORDER BY model_name", (elem_type,))
        all_models = c.fetchall()
        
        models_data = {}
        model_spinner = Spinner(
            text='Επιλέξτε μοντέλο',
            values=['Επιλέξτε μοντέλο'],
            size_hint_y=None,
            height=40
        )
        
        def load_models_for_breaker_category(selected_category):
            """Filter and load models based on selected breaker category"""
            models_data_temp, display_names, selected_display_name = self._load_models_for_element_type(
                elem_type, 
                selected_category, 
                model_id
            )
            models_data.clear()
            models_data.update(models_data_temp)
            
            model_spinner.values = display_names if display_names else ['Δεν υπάρχουν μοντέλα']
            model_spinner.text = selected_display_name if selected_display_name and selected_display_name in model_spinner.values else model_spinner.values[0]
        
        # Bind breaker category change to reload models
        if elem_type in ['Διακόπτης ΜΤ', 'Διακόπτης ΥΤ']:
            breaker_category_spinner.bind(text=lambda spinner, text: load_models_for_breaker_category(text))
            load_models_for_breaker_category(breaker_category_spinner.text)
        else:
            load_models_for_breaker_category(None)
        
        layout.add_widget(model_spinner)
        
        # Gate selection
        layout.add_widget(Label(text='Πύλη (Gate):', size_hint_y=None, height=30))
        # Determine if current element is an interconnection breaker (only MV breakers can be interconnecting)
        is_interconnection = (elem_type == 'Διακόπτης ΜΤ' and is_main_switch == 2)
        available_gates = self.get_available_gates(substation_id, is_interconnection)
        current_gate_text = gate if gate else '(Μη καταχωρημένο)'
        # Ensure current gate is in the list
        if current_gate_text not in available_gates:
            available_gates.append(current_gate_text)
        gate_spinner = Spinner(
            text=current_gate_text,
            values=available_gates,
            size_hint_y=None,
            height=40
        )
        layout.add_widget(gate_spinner)
        
        # Breaker type selection (MV: selectable, HV: fixed to Κεντρικός)
        breaker_type_label = Label(text='Τύπος Διακόπτη:', size_hint_y=None, height=30)
        if is_main_switch == 1:
            current_breaker_type = 'Κεντρικός'
        elif is_main_switch == 2:
            current_breaker_type = 'Διασυνδετικός'
        elif is_main_switch == 3:
            current_breaker_type = 'Διακόπτης Πυκνωτών'
        else:
            current_breaker_type = 'Γραμμής'

        if elem_type == 'Διακόπτης ΥΤ':
            breaker_type_spinner = Spinner(
                text='Κεντρικός',
                values=['Κεντρικός'],
                size_hint_y=None,
                height=40,
                disabled=True
            )
        else:
            breaker_type_spinner = Spinner(
                text=current_breaker_type,
                values=self.BREAKER_TYPES,
                size_hint_y=None,
                height=40
            )
        
        # Handler to refresh gates when breaker type changes
        def on_breaker_type_change(spinner, text):
            is_interconnection = (text == 'Διασυνδετικός')
            available_gates = self.get_available_gates(substation_id, is_interconnection)
            gate_spinner.values = available_gates
            if gate_spinner.text not in available_gates:
                gate_spinner.text = available_gates[0] if available_gates else '(Μη καταχωρημένο)'
        
        breaker_type_spinner.bind(text=on_breaker_type_change)
        
        # Show breaker type selector for MV and HV (HV locked to Κεντρικός)
        if elem_type in ['Διακόπτης ΜΤ', 'Διακόπτης ΥΤ']:
            layout.add_widget(breaker_type_label)
            layout.add_widget(breaker_type_spinner)
        
        # Dynamic fields
        field_inputs = {}
        for field in self.ELEMENT_FIELD_DEFS:
            layout.add_widget(Label(text=f"{field['label']}:", size_hint_y=None, height=30))
            
            # Get current value
            current_value = ''
            if field['key'] == 'name':
                current_value = name or ''
            elif field['key'] == 'serial_number':
                current_value = serial_num or ''
            elif field['key'] == 'manufacture_year':
                current_value = manuf_year or ''
            elif field['key'] == 'maintenance_date':
                current_value = maint_date or ''
            elif field['key'] == 'manufacturer':
                current_value = manufacturer or ''
            elif field['key'] == 'model':
                current_value = model or ''
            elif field['key'] == 'model_version':
                current_value = model_version or ''
            elif field['key'] == 'installation_space':
                current_value = install_space or self.INSTALLATION_SPACE[0]
            elif field['key'] == 'operating_status':
                current_value = op_status or self.OPERATING_STATUS[0]
            elif field['key'] == 'maintenance_cycle':
                current_value = str(maint_cycle) if maint_cycle else '0'
            
            if field.get('type') == 'spinner':
                spinner = Spinner(text=current_value, values=field['values'], size_hint_y=None, height=40)
                field_inputs[field['key']] = spinner
                layout.add_widget(spinner)
            else:
                ti = TextInput(text=current_value, hint_text=field.get('hint', ''), size_hint_y=None, height=40, multiline=False)
                field_inputs[field['key']] = ti
                layout.add_widget(ti)
        
        scroll.add_widget(layout)
        main_layout.add_widget(scroll)
        
        # Buttons
        buttons_layout = BoxLayout(size_hint_y=0.1, spacing=10)
        
        def save_changes():
            name_val = field_inputs['name'].text.strip()
            if not name_val:
                show_message_popup('Σφάλμα', 'Το όνομα είναι υποχρεωτικό!')
                return
            
            # Validate maintenance_cycle
            try:
                cycle_val = int(field_inputs['maintenance_cycle'].text) if field_inputs['maintenance_cycle'].text else 0
            except ValueError:
                show_message_popup('Σφάλμα', 'Ο κύκλος συντήρησης πρέπει να είναι αριθμός!')
                return
            
            # Check for duplicate name (excluding current element)
            c = self.conn.cursor()
            c.execute("SELECT id FROM elements WHERE substation_id=? AND name=? AND id!=?", (substation_id, name_val, element_id))
            if c.fetchone():
                show_message_popup('Σφάλμα', f'Υπάρχει ήδη στοιχείο με όνομα "{name_val}" σε αυτόν τον υποσταθμό!')
                return
            
            # Get model_id if selected
            new_model_id = models_data[model_spinner.text]['id'] if model_spinner.text in models_data else None
            
            # Get gate value
            gate_value = gate_spinner.text if gate_spinner.text != '(Μη καταχωρημένο)' else ''
            
            # Get breaker category for circuit breakers
            breaker_category_value = None
            if elem_type in ['Διακόπτης ΥΤ', 'Διακόπτης ΜΤ']:
                breaker_category_value = breaker_category_spinner.text
            
            # Update is_main_switch based on element type and breaker type selection
            if elem_type == 'Διακόπτης ΥΤ':
                # HV breakers are always main breakers
                new_is_main_switch = 1
            elif elem_type == 'Διακόπτης ΜΤ':
                if breaker_type_spinner.text == 'Κεντρικός':
                    new_is_main_switch = 1
                elif breaker_type_spinner.text == 'Διασυνδετικός':
                    new_is_main_switch = 2
                elif breaker_type_spinner.text == 'Διακόπτης Πυκνωτών':
                    new_is_main_switch = 3
                else:
                    new_is_main_switch = 0
            else:
                new_is_main_switch = 0

            voltage_level_value = voltage_level_spinner.text if voltage_level_spinner.text != '(Κενό)' else ''

            c.execute("""UPDATE elements SET 
                            name=?, serial_number=?, maintenance_date=?, voltage_level=?, manufacturer=?, model=?, model_version=?,
                            installation_space=?, operating_status=?, 
                            maintenance_cycle=?, manufacture_year=?, element_model_id=?, gate=?, is_main_switch=?, breaker_category=?
                            WHERE id=?""",
                        (name_val,
                         field_inputs['serial_number'].text.strip(),
                         field_inputs['maintenance_date'].text.strip(),
                         voltage_level_value,
                         field_inputs['manufacturer'].text.strip(),
                         field_inputs['model'].text.strip(),
                         field_inputs['model_version'].text.strip(),
                         field_inputs['installation_space'].text,
                         field_inputs['operating_status'].text,
                         cycle_val,
                         field_inputs['manufacture_year'].text.strip(),
                         new_model_id,
                         gate_value,
                         new_is_main_switch,
                         breaker_category_value,
                         element_id))
            self.conn.commit()
            popup.dismiss()
            parent_popup.dismiss()
            if grandparent_popup:
                grandparent_popup.dismiss()
            if substation_name:
                show_message_popup('Επιτυχία', 'Οι αλλαγές αποθηκεύτηκαν!', callback=lambda: self._display_substations(substation_name))
            else:
                show_message_popup('Επιτυχία', 'Οι αλλαγές αποθηκεύτηκαν!', callback=lambda: self.show_records(None))
        
        save_btn = Button(text='Αποθήκευση')
        save_btn.bind(on_press=lambda x: save_changes())
        buttons_layout.add_widget(save_btn)
        
        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        buttons_layout.add_widget(cancel_btn)
        
        main_layout.add_widget(buttons_layout)
        popup.content = main_layout
        popup.open()
    
    def confirm_delete_element(self, element_id, element_name, substation_id, parent_popup, substation_name=None):
        """Show confirmation popup before deleting an element"""
        from kivy.uix.popup import Popup
        from kivy.uix.boxlayout import BoxLayout
        from kivy.uix.button import Button
        from kivy.uix.label import Label
        
        confirm_popup = Popup(title='Επιβεβαίωση Διαγραφής', size_hint=(0.6, 0.3))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        warning_label = Label(
            text=f'Είστε σίγουροι ότι θέλετε να διαγράψετε\nτο στοιχείο "{element_name}";',
            size_hint_y=0.5
        )
        layout.add_widget(warning_label)
        
        buttons_layout = BoxLayout(size_hint_y=0.3, spacing=10)
        
        def confirm():
            confirm_popup.dismiss()
            self.delete_element(element_id, substation_id, parent_popup, substation_name)
        
        yes_btn = Button(text='ΝΑΙ', color=(1, 0, 0, 1))
        yes_btn.bind(on_press=lambda x: confirm())
        buttons_layout.add_widget(yes_btn)
        
        no_btn = Button(text='ΟΧΙ')
        no_btn.bind(on_press=confirm_popup.dismiss)
        buttons_layout.add_widget(no_btn)
        
        layout.add_widget(buttons_layout)
        confirm_popup.content = layout
        confirm_popup.open()

    def delete_element(self, element_id, substation_id, parent_popup, substation_name=None):
        c = self.conn.cursor()
        c.execute("DELETE FROM elements WHERE id=?", (element_id,))
        self.conn.commit()
        parent_popup.dismiss()
        if substation_name:
            show_message_popup('Ολοκληρώθηκε', 'Το στοιχείο διαγράφηκε!', callback=lambda: self._display_substations(substation_name))
        else:
            show_message_popup('Ολοκληρώθηκε', 'Το στοιχείο διαγράφηκε!', callback=lambda: self.show_records(None))

    def show_inactive_elements(self, substation_id, substation_name, parent_popup):
        """Show list of inactive elements for a substation"""
        from kivy.uix.popup import Popup
        from kivy.uix.boxlayout import BoxLayout
        from kivy.uix.button import Button
        from kivy.uix.label import Label
        from kivy.uix.scrollview import ScrollView
        from kivy.uix.gridlayout import GridLayout
        
        c = self.conn.cursor()
        # Fetch model data from element_models table
        c.execute("""
            SELECT e.id, e.element_type, e.name, e.serial_number, 
                   em.manufacturer as model_manufacturer, em.model_name, e.is_main_switch
            FROM elements e 
            LEFT JOIN element_models em ON e.element_model_id = em.id
            WHERE e.substation_id=? AND e.operating_status='Ανενεργή' 
            ORDER BY e.name
        """, (substation_id,))
        inactive_elements = c.fetchall()
        
        popup = Popup(title=f'Ανενεργά Στοιχεία - {substation_name}', size_hint=(0.8, 0.8))
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        if not inactive_elements:
            main_layout.add_widget(Label(
                text='Δεν υπάρχουν ανενεργά στοιχεία σε αυτόν τον υποσταθμό',
                size_hint_y=0.8
            ))
        else:
            scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
            grid = GridLayout(cols=1, spacing=10, size_hint_y=None, padding=10)
            grid.bind(minimum_height=grid.setter('height'))
            
            for elem_id, elem_type, elem_name, serial_number, model_manufacturer, model_name, is_main_switch in inactive_elements:
                elem_layout = BoxLayout(size_hint_y=None, height=80, spacing=5, orientation='vertical')
                
                # Add breaker type label for circuit breakers
                display_elem_type = elem_type
                if elem_type in ['Διακόπτης ΥΤ', 'Διακόπτης ΜΤ']:
                    if elem_type == 'Διακόπτης ΥΤ':
                        breaker_type_label = 'Κεντρικός'
                    elif is_main_switch == 1:
                        breaker_type_label = 'Κεντρικός'
                    elif is_main_switch == 2:
                        breaker_type_label = 'Διασυνδετικός'
                    elif is_main_switch == 3:
                        breaker_type_label = 'Διακόπτης Πυκνωτών'
                    else:
                        breaker_type_label = 'Γραμμής'
                    display_elem_type = f"{elem_type} ({breaker_type_label})"
                
                # Element info
                info_text = f'[b]{elem_name}[/b] - {display_elem_type}\nS/N: {serial_number or "-"} | Κατ.: {model_manufacturer or "-"} | Μοντ.: {model_name or "-"}'
                elem_label = Label(
                    text=info_text,
                    size_hint_y=None,
                    height=50,
                    markup=True
                )
                elem_layout.add_widget(elem_label)
                
                # Buttons
                btn_layout = BoxLayout(size_hint_y=None, height=30, spacing=5)
                
                edit_btn = Button(text='Επεξεργασία')
                edit_btn.bind(on_press=lambda x, eid=elem_id, sid=substation_id, sname=substation_name, p=popup, gp=parent_popup: self.show_edit_element_popup(eid, sid, p, sname, gp))
                btn_layout.add_widget(edit_btn)
                
                elem_layout.add_widget(btn_layout)
                grid.add_widget(elem_layout)
            
            scroll.add_widget(grid)
            main_layout.add_widget(scroll)
        
        # Close button
        close_btn = Button(text='Κλείσιμο', size_hint_y=0.1)
        close_btn.bind(on_press=popup.dismiss)
        main_layout.add_widget(close_btn)
        
        popup.content = main_layout
        popup.open()

    def confirm_delete_substation(self, substation_id, substation_name, parent_popup):
        """Confirm before deleting a substation and its elements."""
        confirm_popup = Popup(title='Επιβεβαίωση Διαγραφής', size_hint=(0.65, 0.35))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        warning_label = Label(
            text=f'Είστε σίγουροι ότι θέλετε να διαγράψετε\nτον υποσταθμό "{substation_name}"\nκαι ΟΛΑ τα στοιχεία του;',
            size_hint_y=0.6
        )
        layout.add_widget(warning_label)

        buttons_layout = BoxLayout(size_hint_y=0.3, spacing=10)

        def confirm():
            confirm_popup.dismiss()
            self.delete_substation(substation_id, parent_popup)

        yes_btn = Button(text='ΝΑΙ', color=(1, 0, 0, 1))
        yes_btn.bind(on_press=lambda x: confirm())
        buttons_layout.add_widget(yes_btn)

        no_btn = Button(text='ΟΧΙ')
        no_btn.bind(on_press=confirm_popup.dismiss)
        buttons_layout.add_widget(no_btn)

        layout.add_widget(buttons_layout)
        confirm_popup.content = layout
        confirm_popup.open()

    def delete_substation(self, substation_id, parent_popup):
        c = self.conn.cursor()
        # Delete all elements for this substation first
        c.execute("DELETE FROM elements WHERE substation_id=?", (substation_id,))
        # Then delete the substation
        c.execute("DELETE FROM substations WHERE id=?", (substation_id,))
        self.conn.commit()
        parent_popup.dismiss()
        show_message_popup('Ολοκληρώθηκε', 'Ο υποσταθμός και όλα τα στοιχεία του διαγράφηκαν!', callback=lambda: self.show_records(None))
    
    def show_edit_substation_popup(self, substation_id, substation_name, location, adoption_date, division, parent_popup):
        # Create popup
        popup = Popup(title=f'Επεξεργασία Υποσταθμού: {substation_name}', size_hint=(0.8, 0.7))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # Division spinner
        division_spinner = Spinner(
            text=division or 'ΤΜΘ',
            values=['ΤΜΘ'],
            size_hint_y=0.15
        )
        layout.add_widget(Label(text='Τομέας:', size_hint_y=0.08))
        layout.add_widget(division_spinner)
        
        # Location input
        location_input = TextInput(
            text=location or '',
            hint_text='Τοποθεσία (Google Maps link)',
            size_hint_y=0.15,
            multiline=False
        )
        layout.add_widget(Label(text='Τοποθεσία:', size_hint_y=0.08))
        layout.add_widget(location_input)
        
        # Adoption date input
        date_input = TextInput(
            text=adoption_date or '',
            hint_text='Ημερομηνία Ανάληψης (YYYY-MM-DD)',
            size_hint_y=0.2,
            multiline=False
        )
        layout.add_widget(Label(text='Ανάληψη:', size_hint_y=0.1))
        layout.add_widget(date_input)
        
        # Buttons layout
        buttons_layout = BoxLayout(size_hint_y=0.2, spacing=10)
        
        def save_changes():
            c = self.conn.cursor()
            c.execute("UPDATE substations SET location=?, adoption_date=?, division=? WHERE id=?", 
                     (location_input.text, date_input.text, division_spinner.text, substation_id))
            self.conn.commit()
            popup.dismiss()
            parent_popup.dismiss()
            show_message_popup('Ολοκληρώθηκε', 'Υποσταθμός ενημερώθηκε!', callback=lambda: self.show_records(None))
        
        save_btn = Button(text='Αποθήκευση')
        save_btn.bind(on_press=lambda x: save_changes())
        buttons_layout.add_widget(save_btn)
        
        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        buttons_layout.add_widget(cancel_btn)
        
        layout.add_widget(buttons_layout)
        popup.content = layout
        popup.open()

    def show_add_element_popup(self, instance):
        # Get list of substations
        c = self.conn.cursor()
        c.execute("SELECT id, name FROM substations ORDER BY name")
        substations = c.fetchall()
        
        if not substations:
            show_message_popup('Σφάλμα', 'Δεν υπάρχουν υποσταθμοί!')
            return

        # Get active people for responsible/crew selection
        c.execute("SELECT id, name, role FROM people WHERE active=1 ORDER BY name")
        people = c.fetchall()
        if not people:
            show_message_popup('Σφάλμα', 'Δεν υπάρχουν καταχωρημένα άτομα. Παρακαλώ προσθέστε προσωπικό.', callback=lambda: self.show_people_management(None))
            return
        
        # Store substations mapping for later use
        self.substations_map = {s[1]: s[0] for s in substations}
        
        # Create popup
        popup = Popup(title='Προσθήκη Στοιχείου', size_hint=(0.8, 0.9))
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # Create scrollable area for inputs
        scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        layout = BoxLayout(orientation='vertical', size_hint_y=None, padding=5, spacing=8)
        layout.bind(minimum_height=layout.setter('height'))
        
        # Substation spinner
        substation_names = list(self.substations_map.keys())
        layout.add_widget(Label(text='Επιλέξτε Υποσταθμό:', size_hint_y=None, height=30))
        substation_spinner = Spinner(
            text=substation_names[0],
            values=substation_names,
            size_hint_y=None,
            height=40
        )
        layout.add_widget(substation_spinner)
        
        # Element type spinner
        layout.add_widget(Label(text='Επιλέξτε Στοιχείο:', size_hint_y=None, height=30))
        element_spinner = Spinner(
            text=self.ELEMENT_TYPES[0],
            values=self.ELEMENT_TYPES,
            size_hint_y=None,
            height=40
        )
        layout.add_widget(element_spinner)

        # Voltage level selection
        layout.add_widget(Label(text='Επίπεδο Τάσης:', size_hint_y=None, height=30))
        initial_voltage = self._derive_voltage_level(element_spinner.text) or '(Κενό)'
        voltage_level_spinner = Spinner(
            text=initial_voltage,
            values=self.VOLTAGE_LEVELS,
            size_hint_y=None,
            height=40
        )
        layout.add_widget(voltage_level_spinner)
        
        # Gate selection (auto-populated from transformers)
        gate_label = Label(text='Πύλη (Gate):', size_hint_y=None, height=30)
        layout.add_widget(gate_label)
        
        # Get initial gates for the first substation
        initial_gates = self.get_available_gates(self.substations_map[substation_names[0]])
        gate_spinner = Spinner(
            text=initial_gates[0] if initial_gates else '(Μη καταχωρημένο)',
            values=initial_gates if initial_gates else ['(Μη καταχωρημένο)'],
            size_hint_y=None,
            height=40
        )
        layout.add_widget(gate_spinner)
        
        # Update gates when substation changes
        def on_substation_change(spinner, text):
            substation_id = self.substations_map[text]
            # Check if current element type is a breaker and breaker type is Διασυνδετικός
            is_interconnection = (element_spinner.text in ['Διακόπτης ΥΤ', 'Διακόπτης ΜΤ'] and 
                                 breaker_type_spinner.text == 'Διασυνδετικός')
            available_gates = self.get_available_gates(substation_id, is_interconnection)
            gate_spinner.values = available_gates
            gate_spinner.text = available_gates[0] if available_gates else '(Μη καταχωρημένο)'
        
        substation_spinner.bind(text=on_substation_change)
        
        # Breaker type selection (Main or Line or Interconnection) - only for circuit breakers
        breaker_type_label = Label(text='Τύπος Διακόπτη:', size_hint_y=None, height=30)
        breaker_type_spinner = Spinner(
            text=self.BREAKER_TYPES[0],
            values=self.BREAKER_TYPES,
            size_hint_y=None,
            height=40
        )
        
        # Update gates when breaker type changes
        def on_breaker_type_change(spinner, text):
            substation_id = self.substations_map[substation_spinner.text]
            is_interconnection = (element_spinner.text in ['Διακόπτης ΥΤ', 'Διακόπτης ΜΤ'] and 
                                 text == 'Διασυνδετικός')
            available_gates = self.get_available_gates(substation_id, is_interconnection)
            gate_spinner.values = available_gates
            gate_spinner.text = available_gates[0] if available_gates else '(Μη καταχωρημένο)'
        
        breaker_type_spinner.bind(text=on_breaker_type_change)
        
        # Breaker category filter (only for circuit breakers)
        breaker_category_label = Label(text='Κατηγορία Διακόπτη:', size_hint_y=None, height=30)
        initial_breaker_categories = self._get_breaker_categories_for_element_type(element_spinner.text)
        breaker_category_spinner = Spinner(
            text=initial_breaker_categories[0] if initial_breaker_categories else 'SF6',
            values=initial_breaker_categories,
            size_hint_y=None,
            height=40
        )
        breaker_category_spinner.bind(text=on_breaker_category_change)
        
        # Model selection with "Add New" button
        model_header = BoxLayout(size_hint_y=None, height=30, spacing=5)
        model_header.add_widget(Label(text='Μοντέλο:', size_hint_x=0.7))
        add_model_btn = Button(text='+ Νέο Μοντέλο', size_hint_x=0.3, size_hint_y=None, height=30)
        model_header.add_widget(add_model_btn)
        layout.add_widget(model_header)
        
        # Model spinner (will be populated based on element type)
        model_spinner = Spinner(
            text='Επιλέξτε μοντέλο',
            values=['Επιλέξτε μοντέλο'],
            size_hint_y=None,
            height=40
        )
        layout.add_widget(model_spinner)
        
        # Store model data
        models_data = {}
        all_models_cache = {}
        
        def load_models_for_category(category, selected_breaker_category=None):
            """Load models for selected element category"""
            models_data_temp, display_names, _ = self._load_models_for_element_type(
                category, 
                selected_breaker_category
            )
            models_data.clear()
            models_data.update(models_data_temp)
            
            if display_names:
                model_spinner.values = display_names
                model_spinner.text = display_names[0]
            else:
                model_spinner.values = ['Επιλέξτε μοντέλο']
                model_spinner.text = 'Επιλέξτε μοντέλο'
        
        def on_breaker_category_change(spinner, text):
            """Reload models when breaker category changes"""
            current_element_type = element_spinner.text
            load_models_for_category(current_element_type, text)
        
        # Function to load models when element type changes
        def on_element_type_change(spinner, text):
            # Show/hide breaker category filter for circuit breakers
            if text in ['Διακόπτης ΥΤ', 'Διακόπτης ΜΤ']:
                breaker_category_options = self._get_breaker_categories_for_element_type(text)
                breaker_category_spinner.values = breaker_category_options
                if breaker_category_spinner.text not in breaker_category_options:
                    breaker_category_spinner.text = breaker_category_options[0] if breaker_category_options else 'SF6'
                if breaker_category_label not in layout.children:
                    idx = layout.children.index(model_header)
                    layout.add_widget(breaker_category_spinner, index=idx+1)
                    layout.add_widget(breaker_category_label, index=idx+2)
                    # Bind the breaker category change event
                    breaker_category_spinner.bind(text=on_breaker_category_change)
                load_models_for_category(text, breaker_category_spinner.text)
            else:
                if breaker_category_label in layout.children:
                    breaker_category_spinner.unbind(text=on_breaker_category_change)
                    layout.remove_widget(breaker_category_label)
                    layout.remove_widget(breaker_category_spinner)
                load_models_for_category(text, None)
            
            # Show breaker type selector for circuit breakers
            if text in ['Διακόπτης ΥΤ', 'Διακόπτης ΜΤ']:
                if breaker_type_label not in layout.children:
                    idx = layout.children.index(model_spinner)
                    layout.add_widget(breaker_type_spinner, index=idx)
                    layout.add_widget(breaker_type_label, index=idx+1)
                # Update gates based on breaker type
                substation_id = self.substations_map[substation_spinner.text]
                is_interconnection = (breaker_type_spinner.text == 'Διασυνδετικός')
                available_gates = self.get_available_gates(substation_id, is_interconnection)
                gate_spinner.values = available_gates
                gate_spinner.text = available_gates[0] if available_gates else '(Μη καταχωρημένο)'
            else:
                if breaker_type_label in layout.children:
                    layout.remove_widget(breaker_type_label)
                    layout.remove_widget(breaker_type_spinner)
                # Reset to regular gates for non-breaker elements
                substation_id = self.substations_map[substation_spinner.text]
                available_gates = self.get_available_gates(substation_id, False)
                gate_spinner.values = available_gates
                gate_spinner.text = available_gates[0] if available_gates else '(Μη καταχωρημένο)'

            # Auto-select voltage level based on element type
            derived_voltage = self._derive_voltage_level(text) or '(Κενό)'
            voltage_level_spinner.text = derived_voltage
        
        element_spinner.bind(text=on_element_type_change)
        on_element_type_change(element_spinner, element_spinner.text)
        
        # Dynamic element fields (auto-filled from model, can be overridden)
        field_inputs = {}
        for field in self.ELEMENT_FIELD_DEFS:
            layout.add_widget(Label(text=f"{field['label']}:", size_hint_y=None, height=30))
            if field.get('type') == 'spinner':
                spinner = Spinner(text=field['values'][0], values=field['values'], size_hint_y=None, height=40)
                field_inputs[field['key']] = spinner
                layout.add_widget(spinner)
            else:
                ti = TextInput(hint_text=field.get('hint', ''), size_hint_y=None, height=40, multiline=False)
                field_inputs[field['key']] = ti
                layout.add_widget(ti)
        
        # Auto-fill fields when model is selected
        def on_model_selected(spinner, text):
            if text in models_data:
                model = models_data[text]
                # Auto-fill fields from model
                field_inputs['manufacturer'].text = model['manufacturer']
                field_inputs['model'].text = model['model_name']
                field_inputs['maintenance_cycle'].text = str(model['maintenance_cycle'])
                field_inputs['installation_space'].text = model['installation_space']
        
        model_spinner.bind(text=on_model_selected)
        
        # Add model button action
        def open_add_model():
            from model_management import show_add_model_popup
            def reload_models():
                load_models_for_category(element_spinner.text)
            show_add_model_popup(self, callback=reload_models, category=element_spinner.text)
        
        add_model_btn.bind(on_press=lambda x: open_add_model())
        
        scroll.add_widget(layout)
        main_layout.add_widget(scroll)

        # Buttons layout
        buttons_layout = BoxLayout(size_hint_y=0.1, spacing=10)

        def add_element():
            substation_name = substation_spinner.text
            substation_id = self.substations_map[substation_name]
            element_type = element_spinner.text

            name_val = field_inputs['name'].text if hasattr(field_inputs['name'], 'text') else field_inputs['name'].text
            if not name_val:
                show_message_popup('Σφάλμα', 'Παρακαλώ εισάγετε όνομα στοιχείου!')
                return

            # Gather values
            values = {
                key: (field_inputs[key].text if hasattr(field_inputs[key], 'text') else field_inputs[key].text)
                for key in field_inputs
            }
            if 'operating_status' in values and hasattr(field_inputs['operating_status'], 'text'):
                values['operating_status'] = field_inputs['operating_status'].text

            # Determine is_main_switch based on element type and breaker type
            # HV breakers are always main breakers (is_main_switch=1)
            if element_type == 'Διακόπτης ΥΤ':
                is_main_switch = 1
            elif element_type == 'Διακόπτης ΜΤ':
                if breaker_type_spinner.text == 'Κεντρικός':
                    is_main_switch = 1
                elif breaker_type_spinner.text == 'Διασυνδετικός':
                    is_main_switch = 2
                elif breaker_type_spinner.text == 'Διακόπτης Πυκνωτών':
                    is_main_switch = 3
                else:
                    is_main_switch = 0
            else:
                is_main_switch = 0
            
            # Get gate assignment
            gate_value = gate_spinner.text if gate_spinner.text != '(Μη καταχωρημένο)' else ''
            
            # Get breaker category for circuit breakers
            breaker_category_value = None
            if element_type in ['Διακόπτης ΥΤ', 'Διακόπτης ΜΤ']:
                breaker_category_value = breaker_category_spinner.text
            
            # Get model_id if selected
            model_id = None
            if model_spinner.text in models_data:
                model_id = models_data[model_spinner.text]['id']
            
            # Validate maintenance_cycle is a number
            maintenance_cycle = values.get('maintenance_cycle', '0')
            try:
                maintenance_cycle_int = int(maintenance_cycle) if maintenance_cycle else 0
            except ValueError:
                show_message_popup('Σφάλμα', 'Ο κύκλος συντήρησης πρέπει να είναι αριθμός!')
                return

            # Check for unique name within substation
            c = self.conn.cursor()
            c.execute("SELECT id FROM elements WHERE substation_id=? AND name=?", (substation_id, name_val))
            if c.fetchone():
                show_message_popup('Σφάλμα', f'Υπάρχει ήδη στοιχείο με όνομα "{name_val}" σε αυτόν τον υποσταθμό!')
                return

            voltage_level_value = voltage_level_spinner.text if voltage_level_spinner.text != '(Κενό)' else ''

            c.execute(
                "INSERT INTO elements (substation_id, element_type, name, serial_number, maintenance_date, voltage_level, manufacturer, model, model_version, installation_space, operating_status, maintenance_cycle, element_model_id, manufacture_year, gate, is_main_switch, breaker_category) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    substation_id,
                    element_type,
                    values.get('name', ''),
                    values.get('serial_number', ''),
                    values.get('maintenance_date', ''),
                    voltage_level_value,
                    values.get('manufacturer', ''),
                    values.get('model', ''),
                    values.get('model_version', ''),
                    values.get('installation_space', 'Εσωτερικός'),
                    values.get('operating_status', 'Ενεργή'),
                    maintenance_cycle_int,
                    model_id,
                    values.get('manufacture_year', ''),
                    gate_value,
                    is_main_switch,
                    breaker_category_value,
                ),
            )
            new_element_id = c.lastrowid
            self.conn.commit()
            
            popup.dismiss()
            show_message_popup('Επιτυχία', f'Στοιχείο προστέθηκε στον {substation_name}!', callback=lambda: self._display_substations(substation_name))
        
        add_btn = Button(text='Προσθήκη')
        add_btn.bind(on_press=lambda x: add_element())
        buttons_layout.add_widget(add_btn)
        
        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        buttons_layout.add_widget(cancel_btn)
        
        main_layout.add_widget(buttons_layout)
        popup.content = main_layout
        popup.open()

    def show_add_element_popup_for_substation(self, substation_id, substation_name, parent_popup):
        # Create popup
        popup = Popup(title=f'Προσθήκη Στοιχείου', size_hint=(0.8, 0.9))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # Create scrollable area for inputs
        scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        input_layout = BoxLayout(orientation='vertical', size_hint_y=None, padding=10, spacing=10)
        input_layout.bind(minimum_height=input_layout.setter('height'))
        
        # Substation selection dropdown at the top
        input_layout.add_widget(Label(text='Υποσταθμός:', size_hint_y=None, height=30))
        c = self.conn.cursor()
        c.execute("SELECT id, name FROM substations ORDER BY name")
        all_substations = c.fetchall()
        
        substation_spinner = Spinner(
            text=substation_name,
            values=[sub[1] for sub in all_substations],
            size_hint_y=None,
            height=40
        )
        input_layout.add_widget(substation_spinner)
        
        # Store substation mapping (name -> id)
        substation_map = {sub[1]: sub[0] for sub in all_substations}
        
        # Element type spinner
        element_spinner = Spinner(
            text=self.ELEMENT_TYPES[0],
            values=self.ELEMENT_TYPES,
            size_hint_y=None,
            height=40
        )
        input_layout.add_widget(Label(text='Επιλέξτε Τύπο Στοιχείου:', size_hint_y=None, height=30))
        input_layout.add_widget(element_spinner)

        # Voltage level selection
        input_layout.add_widget(Label(text='Επίπεδο Τάσης:', size_hint_y=None, height=30))
        initial_voltage = self._derive_voltage_level(element_spinner.text) or '(Κενό)'
        voltage_level_spinner = Spinner(
            text=initial_voltage,
            values=self.VOLTAGE_LEVELS,
            size_hint_y=None,
            height=40
        )
        input_layout.add_widget(voltage_level_spinner)
        
        # Gate selection (auto-populated from transformers)
        gate_label = Label(text='Πύλη (Gate):', size_hint_y=None, height=30)
        input_layout.add_widget(gate_label)
        
        # Get initial gates for the selected substation
        initial_gates = self.get_available_gates(substation_id)
        gate_spinner = Spinner(
            text=initial_gates[0] if initial_gates else '(Μη καταχωρημένο)',
            values=initial_gates if initial_gates else ['(Μη καταχωρημένο)'],
            size_hint_y=None,
            height=40
        )
        input_layout.add_widget(gate_spinner)
        
        # Update gates when substation changes
        def on_substation_change(spinner, text):
            selected_substation_id = substation_map[text]
            # Check if current element type is a breaker and breaker type is Διασυνδετικός
            is_interconnection = (element_spinner.text in ['Διακόπτης ΥΤ', 'Διακόπτης ΜΤ'] and 
                                 breaker_type_spinner.text == 'Διασυνδετικός')
            available_gates = self.get_available_gates(selected_substation_id, is_interconnection)
            gate_spinner.values = available_gates
            gate_spinner.text = available_gates[0] if available_gates else '(Μη καταχωρημένο)'
        
        substation_spinner.bind(text=on_substation_change)
        
        # Breaker type selection (Main, Line, or Interconnection) - only for circuit breakers
        breaker_type_label = Label(text='Τύπος Διακόπτη:', size_hint_y=None, height=30)
        breaker_type_spinner = Spinner(
            text=self.BREAKER_TYPES[0],
            values=self.BREAKER_TYPES,
            size_hint_y=None,
            height=40
        )
        
        # Update gates when breaker type changes
        def on_breaker_type_change(spinner, text):
            selected_substation_id = substation_map[substation_spinner.text]
            is_interconnection = (element_spinner.text in ['Διακόπτης ΥΤ', 'Διακόπτης ΜΤ'] and 
                                 text == 'Διασυνδετικός')
            available_gates = self.get_available_gates(selected_substation_id, is_interconnection)
            gate_spinner.values = available_gates
            gate_spinner.text = available_gates[0] if available_gates else '(Μη καταχωρημένο)'
        
        breaker_type_spinner.bind(text=on_breaker_type_change)
        
        # Breaker category filter (only for circuit breakers)
        breaker_category_label = Label(text='Κατηγορία Διακόπτη:', size_hint_y=None, height=30)
        initial_breaker_categories = self._get_breaker_categories_for_element_type(element_spinner.text)
        breaker_category_spinner = Spinner(
            text=initial_breaker_categories[0] if initial_breaker_categories else 'SF6',
            values=initial_breaker_categories,
            size_hint_y=None,
            height=40
        )
        
        # Model selection with "Add New" button
        model_header = BoxLayout(size_hint_y=None, height=30, spacing=5)
        model_header.add_widget(Label(text='Μοντέλο:', size_hint_x=0.7))
        add_model_btn = Button(text='+ Νέο Μοντέλο', size_hint_x=0.3, size_hint_y=None, height=30)
        model_header.add_widget(add_model_btn)
        input_layout.add_widget(model_header)
        
        # Model spinner (will be populated based on element type)
        model_spinner = Spinner(
            text='Επιλέξτε μοντέλο',
            values=['Επιλέξτε μοντέλο'],
            size_hint_y=None,
            height=40
        )
        input_layout.add_widget(model_spinner)
        
        # Store model data
        models_data = {}
        all_models_cache = {}
        
        def on_breaker_category_change(spinner, text):
            """Reload models when breaker category changes"""
            current_element_type = element_spinner.text
            load_models_for_category(current_element_type, text)
        
        def load_models_for_category(category, selected_breaker_category=None):
            """Load models for selected element category"""
            c = self.conn.cursor()
            c.execute("SELECT id, model_name, manufacturer, maintenance_cycle, installation_space, breaker_category FROM element_models WHERE element_category=? ORDER BY model_name", (category,))
            models = c.fetchall()
            all_models_cache[category] = models
            
            models_data.clear()
            if models:
                # For circuit breakers, filter by selected breaker category
                if category in ['Διακόπτης ΜΤ', 'Διακόπτης ΥΤ']:
                    if selected_breaker_category:
                        # Filter models by breaker category (case-insensitive, trimmed)
                        filtered_models = []
                        for m in models:
                            model_category = (m[5] or 'Other').strip()
                            if model_category.lower() == selected_breaker_category.lower():
                                filtered_models.append(m)
                    else:
                        filtered_models = models
                    
                    display_names = []
                    for m in filtered_models:
                        display_name = f"{m[1]} - {m[2] or 'N/A'}"
                        display_names.append(display_name)
                        models_data[display_name] = {
                            'id': m[0],
                            'model_name': m[1],
                            'manufacturer': m[2] or '',
                            'maintenance_cycle': m[3] or 0,
                            'installation_space': m[4] or '',
                            'breaker_category': m[5] or ''
                        }
                    
                    model_spinner.values = display_names if display_names else ['Δεν υπάρχουν μοντέλα']
                    model_spinner.text = display_names[0] if display_names else 'Δεν υπάρχουν μοντέλα'
                else:
                    # For non-breaker elements, show all models
                    display_names = []
                    for m in models:
                        display_name = f"{m[1]} - {m[2] or 'N/A'}"
                        display_names.append(display_name)
                        models_data[display_name] = {
                            'id': m[0],
                            'model_name': m[1],
                            'manufacturer': m[2] or '',
                            'maintenance_cycle': m[3] or 0,
                            'installation_space': m[4] or '',
                            'breaker_category': m[5] or ''
                        }
                    model_spinner.values = display_names
                    model_spinner.text = display_names[0] if display_names else 'Επιλέξτε μοντέλο'
            else:
                model_spinner.values = ['Επιλέξτε μοντέλο']
                model_spinner.text = 'Επιλέξτε μοντέλο'
        
        # Function to load models when element type changes
        def on_element_type_change(spinner, text):
            # Show/hide breaker category filter for circuit breakers
            if text in ['Διακόπτης ΥΤ', 'Διακόπτης ΜΤ']:
                breaker_category_options = self._get_breaker_categories_for_element_type(text)
                breaker_category_spinner.values = breaker_category_options
                if breaker_category_spinner.text not in breaker_category_options:
                    breaker_category_spinner.text = breaker_category_options[0] if breaker_category_options else 'SF6'
                if breaker_category_label not in input_layout.children:
                    idx = input_layout.children.index(model_header)
                    input_layout.add_widget(breaker_category_spinner, index=idx+1)
                    input_layout.add_widget(breaker_category_label, index=idx+2)
                    # Bind the breaker category change event
                    breaker_category_spinner.bind(text=on_breaker_category_change)
                load_models_for_category(text, breaker_category_spinner.text)
            else:
                if breaker_category_label in input_layout.children:
                    breaker_category_spinner.unbind(text=on_breaker_category_change)
                    input_layout.remove_widget(breaker_category_label)
                    input_layout.remove_widget(breaker_category_spinner)
                load_models_for_category(text, None)
            
            # Show/hide breaker type spinner based on element type (only for MV breakers)
            if text == 'Διακόπτης ΜΤ':
                if breaker_type_label not in input_layout.children:
                    input_layout.add_widget(breaker_type_spinner, index=input_layout.children.index(gate_spinner) + 2)
                    input_layout.add_widget(breaker_type_label, index=input_layout.children.index(breaker_type_spinner) + 1)
                # Refresh gates based on current breaker type
                is_interconnection = (breaker_type_spinner.text == 'Διασυνδετικός')
                available_gates = self.get_available_gates(substation_id, is_interconnection)
                gate_spinner.values = available_gates
                if gate_spinner.text not in available_gates:
                    gate_spinner.text = available_gates[0] if available_gates else '(Μη καταχωρημένο)'
            else:
                if breaker_type_label in input_layout.children:
                    input_layout.remove_widget(breaker_type_label)
                    input_layout.remove_widget(breaker_type_spinner)
                # Reset to regular gates for non-MV breaker elements (HV breakers also use regular gates)
                available_gates = self.get_available_gates(substation_id, False)
                gate_spinner.values = available_gates
                if gate_spinner.text not in available_gates:
                    gate_spinner.text = available_gates[0] if available_gates else '(Μη καταχωρημένο)'

            # Auto-select voltage level based on element type
            derived_voltage = self._derive_voltage_level(text) or '(Κενό)'
            voltage_level_spinner.text = derived_voltage
        
        element_spinner.bind(text=on_element_type_change)
        on_element_type_change(element_spinner, element_spinner.text)
        
        # Auto-fill callback when model is selected
        def on_model_selected(spinner, text):
            if text in models_data:
                model = models_data[text]
                # Auto-fill fields from model
                if 'manufacturer' in field_inputs:
                    field_inputs['manufacturer'].text = model['manufacturer']
                if 'maintenance_cycle' in field_inputs:
                    field_inputs['maintenance_cycle'].text = str(model['maintenance_cycle'])
                if 'installation_space' in field_inputs:
                    field_inputs['installation_space'].text = model['installation_space']
                if 'model' in field_inputs:
                    field_inputs['model'].text = model['model_name']
        
        model_spinner.bind(text=on_model_selected)
        
        # Bind "Add New Model" button
        def open_add_model(instance):
            from model_management import show_add_model_popup
            def reload_models():
                load_models_for_category(element_spinner.text)
            show_add_model_popup(self, callback=reload_models, category=element_spinner.text)
        
        add_model_btn.bind(on_press=open_add_model)

        # Dynamic element fields
        field_inputs = {}
        for field in self.ELEMENT_FIELD_DEFS:
            input_layout.add_widget(Label(text=f"{field['label']}:", size_hint_y=None, height=30))
            if field.get('type') == 'spinner':
                spinner = Spinner(
                    text=field['values'][0],
                    values=field['values'],
                    size_hint_y=None,
                    height=40
                )
                field_inputs[field['key']] = spinner
                input_layout.add_widget(spinner)
            else:
                ti = TextInput(
                    hint_text=field.get('hint', ''),
                    size_hint_y=None,
                    height=40,
                    multiline=False
                )
                field_inputs[field['key']] = ti
                input_layout.add_widget(ti)
        
        # Trigger initial auto-fill for default selection
        if model_spinner.text in models_data:
            on_model_selected(model_spinner, model_spinner.text)
        
        scroll.add_widget(input_layout)
        layout.add_widget(scroll)
        
        # Buttons layout
        buttons_layout = BoxLayout(size_hint_y=0.1, spacing=10)
        
        def add_element():
            element_type = element_spinner.text

            # Gather values
            values = {
                key: (field_inputs[key].text if hasattr(field_inputs[key], 'text') else field_inputs[key].text)
                for key in field_inputs
            }
            if 'operating_status' in values and hasattr(field_inputs['operating_status'], 'text'):
                values['operating_status'] = field_inputs['operating_status'].text

            if not values.get('name'):
                show_message_popup('Σφάλμα', 'Παρακαλώ εισάγετε όνομα στοιχείου!')
                return
            
            # Determine breaker type value based on element type and selection
            # HV breakers are always main breakers (is_main_switch=1)
            if element_type == 'Διακόπτης ΥΤ':
                is_main_switch = 1
            elif element_type == 'Διακόπτης ΜΤ':
                if breaker_type_spinner.text == 'Κεντρικός':
                    is_main_switch = 1
                elif breaker_type_spinner.text == 'Διασυνδετικός':
                    is_main_switch = 2
                elif breaker_type_spinner.text == 'Διακόπτης Πυκνωτών':
                    is_main_switch = 3
                else:
                    is_main_switch = 0
            else:
                is_main_switch = 0
            
            # Get gate assignment
            gate_value = gate_spinner.text if gate_spinner.text != '(Μη καταχωρημένο)' else ''
            
            # Get breaker category for circuit breakers
            breaker_category_value = None
            if element_type in ['Διακόπτης ΥΤ', 'Διακόπτης ΜΤ']:
                breaker_category_value = breaker_category_spinner.text
            
            # Validate maintenance_cycle is a number
            maintenance_cycle = values.get('maintenance_cycle', '0')
            try:
                maintenance_cycle_int = int(maintenance_cycle) if maintenance_cycle else 0
            except ValueError:
                show_message_popup('Σφάλμα', 'Ο κύκλος συντήρησης πρέπει να είναι αριθμός!')
                return
            
            # Get selected substation from dropdown
            selected_substation_name = substation_spinner.text
            selected_substation_id = substation_map[selected_substation_name]
            
            # Check for unique name within substation
            c = self.conn.cursor()
            c.execute("SELECT id FROM elements WHERE substation_id=? AND name=?", (selected_substation_id, values.get('name')))
            if c.fetchone():
                show_message_popup('Σφάλμα', f'Υπάρχει ήδη στοιχείο με όνομα "{values.get("name")}" σε αυτόν τον υποσταθμό!')
                return
            
            # Get model_id if a model is selected
            model_id = models_data[model_spinner.text]['id'] if model_spinner.text in models_data else None

            voltage_level_value = voltage_level_spinner.text if voltage_level_spinner.text != '(Κενό)' else ''

            c.execute("INSERT INTO elements (substation_id, element_type, name, serial_number, maintenance_date, voltage_level, manufacturer, model, model_version, installation_space, operating_status, maintenance_cycle, element_model_id, manufacture_year, gate, is_main_switch, breaker_category) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", 
                     (
                        selected_substation_id,
                        element_type,
                        values.get('name', ''),
                        values.get('serial_number', ''),
                        values.get('maintenance_date', ''),
                        voltage_level_value,
                        values.get('manufacturer', ''),
                        values.get('model', ''),
                        values.get('model_version', ''),
                        values.get('installation_space', 'Εσωτερικός'),
                        values.get('operating_status', 'Ενεργή'),
                        maintenance_cycle_int,
                        model_id,
                        values.get('manufacture_year', ''),
                        gate_value,
                        is_main_switch,
                        breaker_category_value,
                     ))
            new_element_id = c.lastrowid
            self.conn.commit()
            
            popup.dismiss()
            parent_popup.dismiss()
            show_message_popup('Επιτυχία', 'Στοιχείο προστέθηκε!', callback=lambda: self._display_substations(selected_substation_name))
        
        add_btn = Button(text='Προσθήκη')
        add_btn.bind(on_press=lambda x: add_element())
        buttons_layout.add_widget(add_btn)
        
        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        buttons_layout.add_widget(cancel_btn)
        
        layout.add_widget(buttons_layout)
        popup.content = layout
        popup.open()

    def show_maintenance_menu(self, instance=None, preselected_substation_name=None, parent_popup=None, maintenance_id=None, after_save_callback=None, prefill_data=None):
        """Show maintenance recording dialog
        
        Args:
            instance: Button instance (optional, for compatibility)
            preselected_substation_name: Name of substation to preselect (optional)
            parent_popup: Parent popup to dismiss when opening this one (optional)
        """
        # Dismiss parent popup if provided
        if parent_popup:
            parent_popup.dismiss()
        
        # Get list of substations
        c = self.conn.cursor()
        c.execute("SELECT id, name FROM substations ORDER BY name")
        substations = c.fetchall()
        
        if not substations:
            show_message_popup('Σφάλμα', 'Δεν υπάρχουν υποσταθμοί!')
            return
        
        maintenance_record = None
        maintenance_people = []
        existing_elements_data = {}
        responsible_person_id = None
        prefill_data = prefill_data or {}

        if maintenance_id:
            c.execute("""
                SELECT substation_id, name, date_time, overall_comments, maintenance_type, user_name, responsible_id
                FROM maintenance
                WHERE id = ?
            """, (maintenance_id,))
            maintenance_record = c.fetchone()
            if not maintenance_record:
                show_message_popup('Σφάλμα', 'Η συντήρηση δεν βρέθηκε.')
                return

            maint_substation_id = maintenance_record[0]
            c.execute("SELECT name FROM substations WHERE id=?", (maint_substation_id,))
            sub_row = c.fetchone()
            if sub_row:
                preselected_substation_name = sub_row[0]

            c.execute("SELECT person_id, role FROM maintenance_people WHERE maintenance_id=?", (maintenance_id,))
            maintenance_people = c.fetchall()
            responsible_person_id = maintenance_record[6]
            if not responsible_person_id:
                for pid, role in maintenance_people:
                    if role == 'responsible':
                        responsible_person_id = pid
                        break

            c.execute("""
                SELECT element_id, element_comments,
                       insulation_closed_fa_ground, insulation_closed_fa_unit,
                       insulation_closed_fb_ground, insulation_closed_fb_unit,
                       insulation_closed_fc_ground, insulation_closed_fc_unit,
                       insulation_open_fa_fa, insulation_open_fa_unit,
                       insulation_open_fb_fb, insulation_open_fb_unit,
                       insulation_open_fc_fc, insulation_open_fc_unit,
                       contact_resistance_fa_fa, contact_resistance_fb_fb, contact_resistance_fc_fc,
                       operations_count,
                      sf6_leakage_kg, sf6_leak_methodology,
                       sf6_n2_fa, h2o_fa, so2_fa, sf6_n2_fb, h2o_fb, so2_fb, sf6_n2_fc, h2o_fc, so2_fc,
                       vidar_fa, vidar_fb, vidar_fc
                FROM maintenance_elements
                WHERE maintenance_id = ?
            """, (maintenance_id,))
            for row in c.fetchall():
                existing_elements_data[row[0]] = {
                    'element_comments': row[1] or '',
                    'ins_closed_fa': row[2], 'ins_closed_fa_unit': row[3] or 'GΩ',
                    'ins_closed_fb': row[4], 'ins_closed_fb_unit': row[5] or 'GΩ',
                    'ins_closed_fc': row[6], 'ins_closed_fc_unit': row[7] or 'GΩ',
                    'ins_open_fa': row[8], 'ins_open_fa_unit': row[9] or 'GΩ',
                    'ins_open_fb': row[10], 'ins_open_fb_unit': row[11] or 'GΩ',
                    'ins_open_fc': row[12], 'ins_open_fc_unit': row[13] or 'GΩ',
                    'cont_fa': row[14], 'cont_fb': row[15], 'cont_fc': row[16],
                    'ops_count': row[17],
                    'sf6_leakage_kg': row[18],
                    'sf6_leak_methodology': row[19] or '',
                    'sf6': {
                        'sf6_n2_fa': row[20], 'h2o_fa': row[21], 'so2_fa': row[22],
                        'sf6_n2_fb': row[23], 'h2o_fb': row[24], 'so2_fb': row[25],
                        'sf6_n2_fc': row[26], 'h2o_fc': row[27], 'so2_fc': row[28]
                    },
                    'vidar': {
                        'vidar_fa': row[29], 'vidar_fb': row[30], 'vidar_fc': row[31]
                    }
                }

        popup_title = 'Επεξεργασία Συντήρησης' if maintenance_id else 'Καταχώρηση Συντήρησης'
        popup = Popup(title=popup_title, size_hint=(0.9, 0.95))
        
        # Create a scrollable container for all content
        scroll_view = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        content_layout = GridLayout(cols=1, spacing=10, size_hint_y=None, padding=10)
        content_layout.bind(minimum_height=content_layout.setter('height'))
        
        # Substation selection
        content_layout.add_widget(Label(text='Επιλογή Υποσταθμού:', size_hint_y=None, height=40))
        substation_map = {s[1]: s[0] for s in substations}
        
        # Use preselected substation if provided, otherwise use first in list
        prefill_substation_name = preselected_substation_name
        if not prefill_substation_name and prefill_data.get('substation_id'):
            for sid, sname in substations:
                if sid == prefill_data['substation_id']:
                    prefill_substation_name = sname
                    break
        if not prefill_substation_name and prefill_data.get('substation_name'):
            prefill_substation_name = prefill_data.get('substation_name')
        initial_substation = prefill_substation_name if prefill_substation_name else substations[0][1]
        
        substation_spinner = Spinner(
            text=initial_substation,
            values=[s[1] for s in substations],
            size_hint_y=None,
            height=40
        )
        if maintenance_id:
            substation_spinner.disabled = True
        content_layout.add_widget(substation_spinner)
        
        # Maintenance Type
        content_layout.add_widget(Label(text='Τύπος Συντήρησης:', size_hint_y=None, height=35))
        maint_type_default = maintenance_record[4] if maintenance_record and maintenance_record[4] else 'Επαναληπτική συντήρηση'
        if not maintenance_id and prefill_data.get('maintenance_type'):
            maint_type_default = prefill_data.get('maintenance_type')
        maintenance_type_spinner = Spinner(
            text=maint_type_default,
            values=['Επαναληπτική συντήρηση', 'Βλάβη', 'Οπτικός έλεγχος'],
            size_hint_y=None,
            height=35
        )
        content_layout.add_widget(maintenance_type_spinner)
        
        # Date/Time (auto-filled with current)
        from datetime import datetime
        content_layout.add_widget(Label(text='Ημερομηνία & Ώρα:', size_hint_y=None, height=35))
        datetime_default = maintenance_record[2] if maintenance_record and maintenance_record[2] else datetime.now().strftime('%Y-%m-%d %H:%M')
        if not maintenance_id and prefill_data.get('date_time'):
            datetime_default = prefill_data.get('date_time')
        datetime_input = TextInput(
            text=datetime_default,
            hint_text='YYYY-MM-DD HH:MM',
            size_hint_y=None,
            height=35,
            multiline=False
        )
        content_layout.add_widget(datetime_input)
        
        # Responsible person (mandatory)
        c.execute("SELECT id, name, role FROM people WHERE active=1 ORDER BY name")
        people = c.fetchall()
        if not people:
            show_message_popup('Σφάλμα', 'Δεν υπάρχουν καταχωρημένα άτομα. Παρακαλώ προσθέστε προσωπικό.', callback=lambda: self.show_people_management(None))
            return

        content_layout.add_widget(Label(text='Υπεύθυνος Συντήρησης (υποχρεωτικό):', size_hint_y=None, height=35))
        people_map = {f"{p[1]} ({p[2]})": p[0] for p in people}
        responsible_default_text = list(people_map.keys())[0] if people_map else ''
        if not maintenance_id and prefill_data.get('responsible_id'):
            responsible_person_id = prefill_data.get('responsible_id')
        if responsible_person_id:
            for label, pid in people_map.items():
                if pid == responsible_person_id:
                    responsible_default_text = label
                    break

        responsible_spinner = Spinner(
            text=responsible_default_text,
            values=list(people_map.keys()),
            size_hint_y=None,
            height=35
        )
        content_layout.add_widget(responsible_spinner)

        # Crew selection (optional)
        content_layout.add_widget(Label(text='Ομάδα Συντήρησης (προαιρετικό):', size_hint_y=None, height=35))

        crew_actions = BoxLayout(size_hint_y=None, height=30, spacing=5)
        select_all_btn = Button(text='Επιλογή Όλων', size_hint_x=0.5)
        clear_all_btn = Button(text='Καμία', size_hint_x=0.5)
        crew_actions.add_widget(select_all_btn)
        crew_actions.add_widget(clear_all_btn)
        content_layout.add_widget(crew_actions)

        crew_container = GridLayout(cols=1, spacing=3, size_hint_y=None, padding=5)
        crew_container.bind(minimum_height=crew_container.setter('height'))
        crew_checks = {}
        crew_ids = {pid for pid, role in maintenance_people if role == 'crew'}
        if not maintenance_id and prefill_data.get('crew_ids'):
            crew_ids = set(prefill_data.get('crew_ids'))
        for pid, name, role in people:
            row = BoxLayout(size_hint_y=None, height=28, spacing=5)
            cb = CheckBox(size_hint_x=0.1, color=self.theme.get('primary', (0.05, 0.18, 0.36, 1)))
            if pid in crew_ids:
                cb.active = True
            row.add_widget(cb)
            row.add_widget(Label(text=f'{name} ({role})', size_hint_x=0.9))
            crew_container.add_widget(row)
            crew_checks[pid] = cb

        crew_scroll = ScrollView(bar_width=8, scroll_type=['bars', 'content'], size_hint_y=None)
        crew_scroll.height = min(220, max(60, len(people) * 30 + 10))
        crew_scroll.add_widget(crew_container)
        content_layout.add_widget(crew_scroll)

        def set_all_crew(value):
            for cb in crew_checks.values():
                cb.active = value

        select_all_btn.bind(on_press=lambda x: set_all_crew(True))
        clear_all_btn.bind(on_press=lambda x: set_all_crew(False))
        
        # Overall comments
        content_layout.add_widget(Label(text='Γενικά Σχόλια Συντήρησης:', size_hint_y=None, height=35))
        comments_default = maintenance_record[3] if maintenance_record and maintenance_record[3] else ''
        if not maintenance_id and prefill_data.get('overall_comments'):
            comments_default = prefill_data.get('overall_comments')
        overall_comments = TextInput(
            hint_text='Γενικά σχόλια για την συντήρηση...',
            text=comments_default,
            size_hint_y=None,
            height=60,
            multiline=True
        )
        def _resize_comments(_instance=None, _value=None):
            lines = overall_comments.text.count('\n') + 1
            overall_comments.height = max(60, min(320, 24 * lines + 20))

        overall_comments.bind(text=_resize_comments)
        _resize_comments()
        content_layout.add_widget(overall_comments)
        
        # Elements selection area
        content_layout.add_widget(Label(text='Στοιχεία που συντηρήθηκαν (τουλάχιστον 1):', size_hint_y=None, height=40))
        
        # Container for element checkboxes (no longer in a separate ScrollView)
        elements_container = GridLayout(cols=1, spacing=5, size_hint_y=None, padding=5)
        elements_container.bind(minimum_height=elements_container.setter('height'))
        content_layout.add_widget(elements_container)
        
        # Dictionary to store element widgets
        element_widgets = {}
        
        def load_elements(substation_name):
            """Load elements for selected substation"""
            elements_container.clear_widgets()
            element_widgets.clear()
            
            substation_id = substation_map[substation_name]
            c = self.conn.cursor()
            c.execute("""
                  SELECT e.id, e.element_type, e.name, e.serial_number, e.gate, e.is_main_switch,
                       e.breaker_category, e.manufacturer, e.model, e.operations_count,
                       em.manufacturer as model_manufacturer, em.model_name
                FROM elements e
                LEFT JOIN element_models em ON e.element_model_id = em.id
                WHERE e.substation_id=?
                  ORDER BY e.gate
            """, (substation_id,))
            elements = c.fetchall()
            
            if not elements:
                elements_container.add_widget(Label(
                    text='Δεν υπάρχουν στοιχεία σε αυτόν τον υποσταθμό',
                    size_hint_y=None,
                    height=40
                ))
                return

            element_ids = [elem[0] for elem in elements]
            last_ops_map = {}
            if element_ids:
                placeholders = ','.join(['?'] * len(element_ids))
                c.execute(f"""
                    SELECT me.element_id, me.operations_count, m.date_time
                    FROM maintenance_elements me
                    JOIN maintenance m ON me.maintenance_id = m.id
                    WHERE me.element_id IN ({placeholders})
                    ORDER BY m.date_time DESC
                """, element_ids)
                for elem_id, ops_count_val, _date_time in c.fetchall():
                    if elem_id not in last_ops_map:
                        last_ops_map[elem_id] = ops_count_val
            
            # Define sort priority for element types
            def get_element_priority(elem):
                elem_id, elem_type, elem_name, serial_number, gate, is_main_switch, breaker_category, manufacturer, model, operations_count, model_manufacturer, model_name = elem
                
                # Priority order: HV breaker, Transformer, Motor Drive, MV main breaker, MV interconnection breaker, MV line breaker, MV capacitor breaker, rest
                if elem_type == 'Διακόπτης ΥΤ':
                    return (1, elem_name)
                elif elem_type == 'Μετασχηματιστής 150/20KV':
                    return (2, elem_name)
                elif elem_type == 'Motor Drive':
                    return (3, elem_name)
                elif elem_type == 'Διακόπτης ΜΤ' and is_main_switch == 1:  # Main breaker
                    return (4, elem_name)
                elif elem_type == 'Διακόπτης ΜΤ' and is_main_switch == 2:  # Interconnection breaker
                    return (5, elem_name)
                elif elem_type == 'Διακόπτης ΜΤ' and is_main_switch == 0:  # Line breaker
                    return (6, elem_name)
                elif elem_type == 'Διακόπτης ΜΤ' and is_main_switch == 3:  # Capacitor breaker
                    return (7, elem_name)
                else:
                    return (8, elem_name)
            
            # Group elements by gate
            gates_dict = {}
            for elem in elements:
                elem_id, elem_type, elem_name, serial_number, gate, is_main_switch, breaker_category, manufacturer, model, operations_count, model_manufacturer, model_name = elem
                
                gate_key = gate if gate else '(Μη καταχωρημένο)'
                if gate_key not in gates_dict:
                    gates_dict[gate_key] = []
                gates_dict[gate_key].append(elem)
            
            # Sort elements within each gate according to priority
            for gate_key in gates_dict:
                gates_dict[gate_key].sort(key=get_element_priority)
            
            # Display elements grouped by gate
            # Show gates in order: ΠΥΛΗ 1, ΠΥΛΗ 2, etc., then unassigned
            sorted_gates = sorted([g for g in gates_dict.keys() if g.startswith('ΠΥΛΗ')])
            if '(Μη καταχωρημένο)' in gates_dict:
                sorted_gates.append('(Μη καταχωρημένο)')
            
            # Display elements grouped by gate
            for gate_name in sorted_gates:
                gate_elements = gates_dict[gate_name]
                
                # Gate header with count
                element_count = len(gate_elements)
                gate_label = Label(
                    text=f'{gate_name} ({element_count} στοιχεία)',
                    size_hint_y=None,
                    height=35,
                    bold=True,
                    color=(0.2, 0.6, 1, 1)  # Blue color for gate headers
                )
                elements_container.add_widget(gate_label)
                
                # Display elements in this gate
                for elem_id, elem_type, elem_name, serial_number, gate, is_main_switch, breaker_category, manufacturer, model, operations_count, model_manufacturer, model_name in gate_elements:
                    # Determine if this is a circuit breaker for showing measurement fields
                    is_breaker = (elem_type in ['Διακόπτης ΜΤ', 'Διακόπτης ΥΤ'])
                    
                    # Build element display text with breaker type, manufacturer, and model
                    elem_display = f'[b]{elem_name}[/b] - {elem_type}'
                    if breaker_category:
                        elem_display += f' ({breaker_category})'
                    elem_display += f'\nS/N: {serial_number or "-"}'
                    
                    # Add manufacturer and model info
                    mfr = model_manufacturer or manufacturer or '-'
                    mdl = model_name or model or '-'
                    elem_display += f' | Κατ.: {mfr} | Μοντ.: {mdl}'
                    
                    # Element container - initially just checkbox and label
                    elem_box = BoxLayout(size_hint_y=None, spacing=5, orientation='vertical')
                    elem_box.bind(minimum_height=elem_box.setter('height'))
                    
                    # Checkbox and name (always visible)
                    checkbox_layout = BoxLayout(size_hint_y=None, height=50, spacing=5)
                    checkbox = CheckBox(size_hint_x=0.08, color=self.theme.get('primary', (0.05, 0.18, 0.36, 1)))
                    checkbox_layout.add_widget(checkbox)
                    
                    elem_label = Label(
                        text=elem_display,
                        size_hint_x=0.92,
                        markup=True
                    )
                    checkbox_layout.add_widget(elem_label)
                    elem_box.add_widget(checkbox_layout)

                    details_container = None
                    elem_comments = None
                    measurements = None

                    def build_details():
                        nonlocal details_container, elem_comments, measurements
                        if details_container is not None:
                            return

                        details_container = BoxLayout(size_hint_y=None, spacing=5, orientation='vertical')
                        details_container.bind(minimum_height=details_container.setter('height'))

                        elem_comments = TextInput(
                            hint_text='Σχόλια για αυτό το στοιχείο...',
                            size_hint_y=None,
                            height=30,
                            multiline=False
                        )
                        details_container.add_widget(elem_comments)

                        measurements = {}

                        if is_breaker:
                            details_container.add_widget(Label(
                                text='ΜΕΤΡΗΣΗ ΑΝΤΙΣΤΑΣΗΣ ΜΟΝΩΣΗΣ - ΔΙΑΚΟΠΤΗΣ ΚΛΕΙΣΤΟΣ (Φ-ΓΗ):',
                                size_hint_y=None,
                                height=25,
                                bold=True
                            ))

                            closed_fa_layout = BoxLayout(size_hint_y=None, height=30, spacing=3)
                            closed_fa_layout.add_widget(Label(text='ΦΑ-ΓΗ:', size_hint_x=0.15))
                            ins_closed_fa = TextInput(hint_text='0.0', size_hint_x=0.35, multiline=False)
                            closed_fa_layout.add_widget(ins_closed_fa)
                            ins_closed_fa_unit = Spinner(text='GΩ', values=['MΩ', 'GΩ', 'TΩ'], size_hint_x=0.15)
                            closed_fa_layout.add_widget(ins_closed_fa_unit)
                            closed_fa_layout.add_widget(Label(text='', size_hint_x=0.35))
                            details_container.add_widget(closed_fa_layout)

                            closed_fb_layout = BoxLayout(size_hint_y=None, height=30, spacing=3)
                            closed_fb_layout.add_widget(Label(text='ΦΒ-ΓΗ:', size_hint_x=0.15))
                            ins_closed_fb = TextInput(hint_text='0.0', size_hint_x=0.35, multiline=False)
                            closed_fb_layout.add_widget(ins_closed_fb)
                            ins_closed_fb_unit = Spinner(text='GΩ', values=['MΩ', 'GΩ', 'TΩ'], size_hint_x=0.15)
                            closed_fb_layout.add_widget(ins_closed_fb_unit)
                            closed_fb_layout.add_widget(Label(text='', size_hint_x=0.35))
                            details_container.add_widget(closed_fb_layout)

                            closed_fc_layout = BoxLayout(size_hint_y=None, height=30, spacing=3)
                            closed_fc_layout.add_widget(Label(text='ΦΓ-ΓΗ:', size_hint_x=0.15))
                            ins_closed_fc = TextInput(hint_text='0.0', size_hint_x=0.35, multiline=False)
                            closed_fc_layout.add_widget(ins_closed_fc)
                            ins_closed_fc_unit = Spinner(text='GΩ', values=['MΩ', 'GΩ', 'TΩ'], size_hint_x=0.15)
                            closed_fc_layout.add_widget(ins_closed_fc_unit)
                            closed_fc_layout.add_widget(Label(text='', size_hint_x=0.35))
                            details_container.add_widget(closed_fc_layout)

                            details_container.add_widget(Label(
                                text='ΜΕΤΡΗΣΗ ΑΝΤΙΣΤΑΣΗΣ ΜΟΝΩΣΗΣ - ΔΙΑΚΟΠΤΗΣ ΑΝΟΙΧΤΟΣ (Φ-Φ):',
                                size_hint_y=None,
                                height=25,
                                bold=True
                            ))

                            open_fa_layout = BoxLayout(size_hint_y=None, height=30, spacing=3)
                            open_fa_layout.add_widget(Label(text='ΦΑ-ΦΑ:', size_hint_x=0.15))
                            ins_open_fa = TextInput(hint_text='0.0', size_hint_x=0.35, multiline=False)
                            open_fa_layout.add_widget(ins_open_fa)
                            ins_open_fa_unit = Spinner(text='GΩ', values=['MΩ', 'GΩ', 'TΩ'], size_hint_x=0.15)
                            open_fa_layout.add_widget(ins_open_fa_unit)
                            open_fa_layout.add_widget(Label(text='', size_hint_x=0.35))
                            details_container.add_widget(open_fa_layout)

                            open_fb_layout = BoxLayout(size_hint_y=None, height=30, spacing=3)
                            open_fb_layout.add_widget(Label(text='ΦΒ-ΦΒ:', size_hint_x=0.15))
                            ins_open_fb = TextInput(hint_text='0.0', size_hint_x=0.35, multiline=False)
                            open_fb_layout.add_widget(ins_open_fb)
                            ins_open_fb_unit = Spinner(text='GΩ', values=['MΩ', 'GΩ', 'TΩ'], size_hint_x=0.15)
                            open_fb_layout.add_widget(ins_open_fb_unit)
                            open_fb_layout.add_widget(Label(text='', size_hint_x=0.35))
                            details_container.add_widget(open_fb_layout)

                            open_fc_layout = BoxLayout(size_hint_y=None, height=30, spacing=3)
                            open_fc_layout.add_widget(Label(text='ΦΓ-ΦΓ:', size_hint_x=0.15))
                            ins_open_fc = TextInput(hint_text='0.0', size_hint_x=0.35, multiline=False)
                            open_fc_layout.add_widget(ins_open_fc)
                            ins_open_fc_unit = Spinner(text='GΩ', values=['MΩ', 'GΩ', 'TΩ'], size_hint_x=0.15)
                            open_fc_layout.add_widget(ins_open_fc_unit)
                            open_fc_layout.add_widget(Label(text='', size_hint_x=0.35))
                            details_container.add_widget(open_fc_layout)

                            details_container.add_widget(Label(
                                text='ΑΝΤΙΣΤΑΣΗ ΔΙΕΛΕΥΣΗΣ (μΩ) - ΔΙΑΚΟΠΤΗΣ ΚΛΕΙΣΤΟΣ:',
                                size_hint_y=None,
                                height=25,
                                bold=True
                            ))

                            contact_layout = BoxLayout(size_hint_y=None, height=30, spacing=3)
                            contact_layout.add_widget(Label(text='ΦΑ-ΦΑ:', size_hint_x=0.15))
                            cont_fa = TextInput(hint_text='0.0', size_hint_x=0.25, multiline=False)
                            contact_layout.add_widget(cont_fa)
                            contact_layout.add_widget(Label(text='ΦΒ-ΦΒ:', size_hint_x=0.15))
                            cont_fb = TextInput(hint_text='0.0', size_hint_x=0.25, multiline=False)
                            contact_layout.add_widget(cont_fb)
                            contact_layout.add_widget(Label(text='ΦΓ-ΦΓ:', size_hint_x=0.15))
                            cont_fc = TextInput(hint_text='0.0', size_hint_x=0.25, multiline=False)
                            contact_layout.add_widget(cont_fc)
                            details_container.add_widget(contact_layout)

                            details_container.add_widget(Label(
                                text='ΜΕΤΡΗΤΗΣ ΧΕΙΡΙΣΜΩΝ:',
                                size_hint_y=None,
                                height=25,
                                bold=True
                            ))

                            ops_layout = BoxLayout(size_hint_y=None, height=30, spacing=3)
                            ops_layout.add_widget(Label(text='Αριθμός Χειρισμών:', size_hint_x=0.3))
                            ops_count_input = TextInput(
                                text='',
                                hint_text=f'Τελευταία τιμή: {operations_count}' if operations_count else '0',
                                size_hint_x=0.2,
                                multiline=False
                            )
                            ops_layout.add_widget(ops_count_input)

                            last_ops_val = last_ops_map.get(elem_id)
                            ops_diff = ''
                            if last_ops_val is not None:
                                try:
                                    current = int(ops_count_input.text) if ops_count_input.text else 0
                                    diff = current - last_ops_val
                                    ops_diff = f'(Διαφορά από τελευταία: +{diff})' if diff >= 0 else f'(Διαφορά από τελευταία: {diff})'
                                except:
                                    pass

                            ops_diff_label = Label(text=ops_diff, size_hint_x=0.5, font_size='10sp')
                            ops_layout.add_widget(ops_diff_label)
                            details_container.add_widget(ops_layout)

                            sf6_widgets = {}
                            vidar_widgets = {}

                            if breaker_category == 'SF6':
                                sf6_leakage_input = TextInput(hint_text='kg', size_hint_x=0.25, multiline=False)
                                sf6_methodology_input = TextInput(hint_text='Μεθοδολογία', size_hint_x=0.55, multiline=False)
                                leak_layout = BoxLayout(size_hint_y=None, height=30, spacing=3)
                                leak_layout.add_widget(Label(text='Διαρροή SF6 (kg):', size_hint_x=0.45))
                                leak_layout.add_widget(sf6_leakage_input)
                                leak_layout.add_widget(Label(text='', size_hint_x=0.3))
                                details_container.add_widget(leak_layout)

                                method_layout = BoxLayout(size_hint_y=None, height=30, spacing=3)
                                method_layout.add_widget(Label(text='Πλήρωση/Αντικατάσταση (Μεθοδολογία):', size_hint_x=0.45))
                                method_layout.add_widget(sf6_methodology_input)
                                details_container.add_widget(method_layout)

                                details_container.add_widget(Label(
                                    text='ΠΟΙΟΤΗΤΑ ΑΕΡΙΟΥ SF6:',
                                    size_hint_y=None,
                                    height=25,
                                    bold=True
                                ))

                                sf6_header = BoxLayout(size_hint_y=None, height=25, spacing=3)
                                sf6_header.add_widget(Label(text='', size_hint_x=0.15))
                                sf6_header.add_widget(Label(text='SF6/N2 (%)', size_hint_x=0.28, bold=True))
                                sf6_header.add_widget(Label(text='H2O (°C atm)', size_hint_x=0.28, bold=True))
                                sf6_header.add_widget(Label(text='SO2 (ppm)', size_hint_x=0.28, bold=True))
                                details_container.add_widget(sf6_header)

                                sf6_fa_layout = BoxLayout(size_hint_y=None, height=30, spacing=3)
                                sf6_fa_layout.add_widget(Label(text='ΦΑ:', size_hint_x=0.15))
                                sf6_n2_fa = TextInput(hint_text='0.0', size_hint_x=0.28, multiline=False)
                                sf6_fa_layout.add_widget(sf6_n2_fa)
                                h2o_fa = TextInput(hint_text='0.0', size_hint_x=0.28, multiline=False)
                                sf6_fa_layout.add_widget(h2o_fa)
                                so2_fa = TextInput(hint_text='0.0', size_hint_x=0.28, multiline=False)
                                sf6_fa_layout.add_widget(so2_fa)
                                details_container.add_widget(sf6_fa_layout)

                                sf6_fb_layout = BoxLayout(size_hint_y=None, height=30, spacing=3)
                                sf6_fb_layout.add_widget(Label(text='ΦΒ:', size_hint_x=0.15))
                                sf6_n2_fb = TextInput(hint_text='0.0', size_hint_x=0.28, multiline=False)
                                sf6_fb_layout.add_widget(sf6_n2_fb)
                                h2o_fb = TextInput(hint_text='0.0', size_hint_x=0.28, multiline=False)
                                sf6_fb_layout.add_widget(h2o_fb)
                                so2_fb = TextInput(hint_text='0.0', size_hint_x=0.28, multiline=False)
                                sf6_fb_layout.add_widget(so2_fb)
                                details_container.add_widget(sf6_fb_layout)

                                sf6_fc_layout = BoxLayout(size_hint_y=None, height=30, spacing=3)
                                sf6_fc_layout.add_widget(Label(text='ΦΓ:', size_hint_x=0.15))
                                sf6_n2_fc = TextInput(hint_text='0.0', size_hint_x=0.28, multiline=False)
                                sf6_fc_layout.add_widget(sf6_n2_fc)
                                h2o_fc = TextInput(hint_text='0.0', size_hint_x=0.28, multiline=False)
                                sf6_fc_layout.add_widget(h2o_fc)
                                so2_fc = TextInput(hint_text='0.0', size_hint_x=0.28, multiline=False)
                                sf6_fc_layout.add_widget(so2_fc)
                                details_container.add_widget(sf6_fc_layout)

                                sf6_widgets = {
                                    'sf6_n2_fa': sf6_n2_fa, 'h2o_fa': h2o_fa, 'so2_fa': so2_fa,
                                    'sf6_n2_fb': sf6_n2_fb, 'h2o_fb': h2o_fb, 'so2_fb': so2_fb,
                                    'sf6_n2_fc': sf6_n2_fc, 'h2o_fc': h2o_fc, 'so2_fc': so2_fc
                                }
                            else:
                                sf6_leakage_input = None
                                sf6_methodology_input = None

                            if breaker_category in ['Vacuum', 'Κενού']:
                                details_container.add_widget(Label(
                                    text='ΕΛΕΓΧΟΣ ΚΕΝΟΥ (VIDAR):',
                                    size_hint_y=None,
                                    height=25,
                                    bold=True
                                ))

                                vidar_layout = BoxLayout(size_hint_y=None, height=30, spacing=3)
                                vidar_layout.add_widget(Label(text='ΦΑ-ΦΑ:', size_hint_x=0.15))
                                vidar_fa = TextInput(hint_text='0.0', size_hint_x=0.25, multiline=False)
                                vidar_layout.add_widget(vidar_fa)
                                vidar_layout.add_widget(Label(text='ΦΒ-ΦΒ:', size_hint_x=0.15))
                                vidar_fb = TextInput(hint_text='0.0', size_hint_x=0.25, multiline=False)
                                vidar_layout.add_widget(vidar_fb)
                                vidar_layout.add_widget(Label(text='ΦΓ-ΦΓ:', size_hint_x=0.15))
                                vidar_fc = TextInput(hint_text='0.0', size_hint_x=0.25, multiline=False)
                                vidar_layout.add_widget(vidar_fc)
                                details_container.add_widget(vidar_layout)

                                vidar_widgets = {
                                    'vidar_fa': vidar_fa,
                                    'vidar_fb': vidar_fb,
                                    'vidar_fc': vidar_fc
                                }

                            measurements = {
                                'ins_closed_fa': ins_closed_fa,
                                'ins_closed_fa_unit': ins_closed_fa_unit,
                                'ins_closed_fb': ins_closed_fb,
                                'ins_closed_fb_unit': ins_closed_fb_unit,
                                'ins_closed_fc': ins_closed_fc,
                                'ins_closed_fc_unit': ins_closed_fc_unit,
                                'ins_open_fa': ins_open_fa,
                                'ins_open_fa_unit': ins_open_fa_unit,
                                'ins_open_fb': ins_open_fb,
                                'ins_open_fb_unit': ins_open_fb_unit,
                                'ins_open_fc': ins_open_fc,
                                'ins_open_fc_unit': ins_open_fc_unit,
                                'cont_fa': cont_fa,
                                'cont_fb': cont_fb,
                                'cont_fc': cont_fc,
                                'ops_count': ops_count_input,
                                'sf6': sf6_widgets,
                                'sf6_leakage': sf6_leakage_input,
                                'sf6_leak_methodology': sf6_methodology_input,
                                'vidar': vidar_widgets
                            }

                        element_widgets[elem_id]['details_container'] = details_container
                        element_widgets[elem_id]['comments'] = elem_comments
                        element_widgets[elem_id]['measurements'] = measurements

                    def ensure_details(elem_box=elem_box):
                        build_details()
                        if details_container not in elem_box.children:
                            elem_box.add_widget(details_container)

                    def toggle_details(_checkbox_instance, value, elem_box=elem_box):
                        if value:
                            ensure_details(elem_box)
                        else:
                            if details_container in elem_box.children:
                                elem_box.remove_widget(details_container)

                    checkbox.bind(active=toggle_details)

                    elements_container.add_widget(elem_box)

                    spacing = Label(text='', size_hint_y=None, height=5)
                    elements_container.add_widget(spacing)

                    element_widgets[elem_id] = {
                        'checkbox': checkbox,
                        'label': elem_label,
                        'display': elem_display,
                        'comments': None,
                        'measurements': None,
                        'elem_type': elem_type,
                        'details_container': None,
                        'ensure_details': ensure_details
                    }

            if not maintenance_id and prefill_data.get('element_ids'):
                prefill_elements = set(prefill_data.get('element_ids'))
                incomplete_elements = set(prefill_data.get('incomplete_elements') or [])
                for elem_id in prefill_elements:
                    if elem_id not in element_widgets:
                        continue
                    widgets = element_widgets[elem_id]
                    widgets['checkbox'].active = True
                    widgets['ensure_details']()
                    if elem_id in incomplete_elements:
                        widgets['label'].text = f"[color=ff3333]{widgets['display']}[/color]"

            if maintenance_id and existing_elements_data:
                for elem_id, data in existing_elements_data.items():
                    if elem_id not in element_widgets:
                        continue
                    widgets = element_widgets[elem_id]
                    widgets['checkbox'].active = True
                    widgets['ensure_details']()
                    widgets['comments'].text = data.get('element_comments', '')

                    measurements = widgets['measurements'] or {}
                    if measurements:
                        if data.get('ins_closed_fa') is not None:
                            measurements['ins_closed_fa'].text = str(data.get('ins_closed_fa'))
                        measurements['ins_closed_fa_unit'].text = data.get('ins_closed_fa_unit', 'GΩ')
                        if data.get('ins_closed_fb') is not None:
                            measurements['ins_closed_fb'].text = str(data.get('ins_closed_fb'))
                        measurements['ins_closed_fb_unit'].text = data.get('ins_closed_fb_unit', 'GΩ')
                        if data.get('ins_closed_fc') is not None:
                            measurements['ins_closed_fc'].text = str(data.get('ins_closed_fc'))
                        measurements['ins_closed_fc_unit'].text = data.get('ins_closed_fc_unit', 'GΩ')

                        if data.get('ins_open_fa') is not None:
                            measurements['ins_open_fa'].text = str(data.get('ins_open_fa'))
                        measurements['ins_open_fa_unit'].text = data.get('ins_open_fa_unit', 'GΩ')
                        if data.get('ins_open_fb') is not None:
                            measurements['ins_open_fb'].text = str(data.get('ins_open_fb'))
                        measurements['ins_open_fb_unit'].text = data.get('ins_open_fb_unit', 'GΩ')
                        if data.get('ins_open_fc') is not None:
                            measurements['ins_open_fc'].text = str(data.get('ins_open_fc'))
                        measurements['ins_open_fc_unit'].text = data.get('ins_open_fc_unit', 'GΩ')

                        if data.get('cont_fa') is not None:
                            measurements['cont_fa'].text = str(data.get('cont_fa'))
                        if data.get('cont_fb') is not None:
                            measurements['cont_fb'].text = str(data.get('cont_fb'))
                        if data.get('cont_fc') is not None:
                            measurements['cont_fc'].text = str(data.get('cont_fc'))

                        if data.get('ops_count') is not None:
                            measurements['ops_count'].text = str(data.get('ops_count'))

                        if measurements.get('sf6'):
                            for key, widget in measurements['sf6'].items():
                                if data['sf6'].get(key) is not None:
                                    widget.text = str(data['sf6'].get(key))

                        if measurements.get('sf6_leakage') and data.get('sf6_leakage_kg') is not None:
                            measurements['sf6_leakage'].text = str(data.get('sf6_leakage_kg'))

                        if measurements.get('sf6_leak_methodology') and data.get('sf6_leak_methodology'):
                            measurements['sf6_leak_methodology'].text = str(data.get('sf6_leak_methodology'))

                        if measurements.get('vidar'):
                            for key, widget in measurements['vidar'].items():
                                if data['vidar'].get(key) is not None:
                                    widget.text = str(data['vidar'].get(key))
        
        # Load initial elements
        load_elements(substation_spinner.text)
        
        # Update elements when substation changes
        substation_spinner.bind(text=lambda spinner, text: load_elements(text))
        
        # Add scrollable content to scroll view
        scroll_view.add_widget(content_layout)
        
        # Main layout with scroll view and buttons
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        main_layout.add_widget(scroll_view)
        
        # Add-element button row (inside scrollable area, below elements)
        add_element_row = BoxLayout(size_hint_y=None, height=45, spacing=10)

        def add_element_from_maintenance():
            substation_name = substation_spinner.text
            substation_id = substation_map.get(substation_name)
            if not substation_id:
                show_message_popup('Σφάλμα', 'Δεν βρέθηκε υποσταθμός.')
                return
            self.show_add_element_popup_for_substation(substation_id, substation_name, popup)

        add_element_btn = Button(text='Προσθήκη Στοιχείου', size_hint_x=None)
        add_element_btn.bind(on_press=lambda x: add_element_from_maintenance())

        def _resize_add_button(*_args):
            min_width = Window.width * 0.3
            text_width = add_element_btn.texture_size[0] + 40
            add_element_btn.width = max(min_width, text_width)

        add_element_btn.bind(texture_size=lambda *_: _resize_add_button())
        Window.bind(size=lambda *_: _resize_add_button())
        _resize_add_button()

        add_element_row.add_widget(Widget())
        add_element_row.add_widget(add_element_btn)
        add_element_row.add_widget(Widget())

        content_layout.add_widget(add_element_row)

        # Buttons at the bottom (not scrollable)
        buttons_layout = BoxLayout(size_hint_y=None, height=50, spacing=10)
        
        def save_maintenance():
            nonlocal maintenance_id
            # Validate at least one element selected
            selected_elements = [(eid, widgets) for eid, widgets in element_widgets.items() 
                                if widgets['checkbox'].active]
            
            if not selected_elements:
                show_message_popup('Σφάλμα', 'Πρέπει να επιλέξετε τουλάχιστον ένα στοιχείο!')
                return
            
            if not datetime_input.text.strip():
                show_message_popup('Σφάλμα', 'Η ημερομηνία είναι υποχρεωτική!')
                return

            if not responsible_spinner.text:
                show_message_popup('Σφάλμα', 'Ο υπεύθυνος συντήρησης είναι υποχρεωτικός!')
                return
            
            # Insert/update maintenance record with type and user
            substation_id = substation_map[substation_spinner.text]
            maintenance_date = datetime_input.text.strip()
            maintenance_type = maintenance_type_spinner.text
            user_name = ''
            maintenance_name = self._build_maintenance_name(substation_spinner.text, maintenance_date)
            responsible_id = people_map.get(responsible_spinner.text)

            if maintenance_id:
                c.execute(
                    """UPDATE maintenance
                       SET substation_id=?, name=?, date_time=?, overall_comments=?, maintenance_type=?, user_name=?, responsible_id=?
                       WHERE id=?""",
                    (substation_id, maintenance_name, maintenance_date, overall_comments.text.strip(), maintenance_type, user_name, responsible_id, maintenance_id)
                )
                c.execute("DELETE FROM maintenance_people WHERE maintenance_id=?", (maintenance_id,))
                c.execute("DELETE FROM maintenance_elements WHERE maintenance_id=?", (maintenance_id,))
            else:
                c.execute(
                    "INSERT INTO maintenance (substation_id, name, date_time, overall_comments, maintenance_type, user_name, responsible_id) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (substation_id, maintenance_name, maintenance_date, overall_comments.text.strip(), maintenance_type, user_name, responsible_id)
                )
                maintenance_id = c.lastrowid

            # Store responsible and crew in maintenance_people
            if responsible_id:
                c.execute(
                    "INSERT INTO maintenance_people (maintenance_id, person_id, role) VALUES (?, ?, ?)",
                    (maintenance_id, responsible_id, 'responsible')
                )

            for pid, cb in crew_checks.items():
                if cb.active and pid != responsible_id:
                    c.execute(
                        "INSERT INTO maintenance_people (maintenance_id, person_id, role) VALUES (?, ?, ?)",
                        (maintenance_id, pid, 'crew')
                    )
            
            # Insert maintenance elements and update their maintenance_date
            for elem_id, widgets in selected_elements:
                # Prepare measurement values
                measurements = widgets['measurements']
                
                if measurements:  # Circuit breaker with measurements
                    # Helper to parse float or None
                    def parse_float(val):
                        try:
                            return float(val.strip()) if val.strip() else None
                        except:
                            return None
                    
                    # Parse operations count
                    ops_count = None
                    try:
                        ops_count = int(measurements['ops_count'].text) if measurements['ops_count'].text.strip() else None
                    except:
                        pass
                    
                    # Parse SF6 measurements if present
                    sf6_vals = {}
                    if measurements['sf6']:
                        for key, widget in measurements['sf6'].items():
                            sf6_vals[key] = parse_float(widget.text)

                    sf6_leakage_val = None
                    if measurements.get('sf6_leakage'):
                        sf6_leakage_val = parse_float(measurements['sf6_leakage'].text)

                    sf6_leak_methodology_val = None
                    if measurements.get('sf6_leak_methodology'):
                        sf6_leak_methodology_val = measurements['sf6_leak_methodology'].text.strip() or None

                    if sf6_leakage_val is not None and not sf6_leak_methodology_val:
                        show_message_popup('Σφάλμα', 'Για διαρροή SF6 απαιτείται συμπλήρωση μεθοδολογίας (Πλήρωση/Αντικατάσταση).')
                        return
                    
                    # Parse VIDAR measurements if present
                    vidar_vals = {}
                    if measurements['vidar']:
                        for key, widget in measurements['vidar'].items():
                            vidar_vals[key] = parse_float(widget.text)
                    
                    c.execute(
                        """INSERT INTO maintenance_elements 
                        (maintenance_id, element_id, element_comments,
                         insulation_closed_fa_ground, insulation_closed_fa_unit,
                         insulation_closed_fb_ground, insulation_closed_fb_unit,
                         insulation_closed_fc_ground, insulation_closed_fc_unit,
                         insulation_open_fa_fa, insulation_open_fa_unit,
                         insulation_open_fb_fb, insulation_open_fb_unit,
                         insulation_open_fc_fc, insulation_open_fc_unit,
                         contact_resistance_fa_fa, contact_resistance_fb_fb, contact_resistance_fc_fc,
                         operations_count,
                        sf6_leakage_kg, sf6_leak_methodology,
                         sf6_n2_fa, h2o_fa, so2_fa, sf6_n2_fb, h2o_fb, so2_fb, sf6_n2_fc, h2o_fc, so2_fc,
                         vidar_fa, vidar_fb, vidar_fc)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (maintenance_id, elem_id, widgets['comments'].text.strip(),
                         parse_float(measurements['ins_closed_fa'].text), measurements['ins_closed_fa_unit'].text,
                         parse_float(measurements['ins_closed_fb'].text), measurements['ins_closed_fb_unit'].text,
                         parse_float(measurements['ins_closed_fc'].text), measurements['ins_closed_fc_unit'].text,
                         parse_float(measurements['ins_open_fa'].text), measurements['ins_open_fa_unit'].text,
                         parse_float(measurements['ins_open_fb'].text), measurements['ins_open_fb_unit'].text,
                         parse_float(measurements['ins_open_fc'].text), measurements['ins_open_fc_unit'].text,
                         parse_float(measurements['cont_fa'].text),
                         parse_float(measurements['cont_fb'].text),
                         parse_float(measurements['cont_fc'].text),
                         ops_count,
                        sf6_leakage_val, sf6_leak_methodology_val,
                         sf6_vals.get('sf6_n2_fa'), sf6_vals.get('h2o_fa'), sf6_vals.get('so2_fa'),
                         sf6_vals.get('sf6_n2_fb'), sf6_vals.get('h2o_fb'), sf6_vals.get('so2_fb'),
                         sf6_vals.get('sf6_n2_fc'), sf6_vals.get('h2o_fc'), sf6_vals.get('so2_fc'),
                         vidar_vals.get('vidar_fa'), vidar_vals.get('vidar_fb'), vidar_vals.get('vidar_fc'))
                    )
                    
                    # Update element's operations_count
                    if ops_count is not None:
                        c.execute("UPDATE elements SET operations_count=? WHERE id=?", (ops_count, elem_id))
                else:  # Other element types without measurements
                    c.execute(
                        "INSERT INTO maintenance_elements (maintenance_id, element_id, element_comments) VALUES (?, ?, ?)",
                        (maintenance_id, elem_id, widgets['comments'].text.strip())
                    )
                
                # Update element's maintenance_date
                c.execute(
                    "UPDATE elements SET maintenance_date=? WHERE id=?",
                    (maintenance_date, elem_id)
                )
            
            # Update substation's last maintenance date
            c.execute(
                "UPDATE substations SET last_maintenance=? WHERE id=?",
                (maintenance_date, substation_id)
            )
            
            self.conn.commit()
            popup.dismiss()
            success_msg = 'Η συντήρηση ενημερώθηκε!' if maintenance_record else 'Η συντήρηση καταχωρήθηκε!'
            if after_save_callback:
                show_message_popup('Επιτυχία', success_msg, callback=lambda: after_save_callback())
            else:
                show_message_popup('Επιτυχία', success_msg)
        
        save_btn = Button(text='Αποθήκευση')
        save_btn.bind(on_press=lambda x: save_maintenance())
        buttons_layout.add_widget(save_btn)
        
        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        buttons_layout.add_widget(cancel_btn)
        
        main_layout.add_widget(buttons_layout)
        popup.content = main_layout
        popup.open()
    
    def show_maintenance_menu_for_substation(self, substation_id, substation_name, parent_popup=None):
        """Wrapper to show maintenance menu with preselected substation
        
        Args:
            substation_id: ID of the substation (for compatibility, not used)
            substation_name: Name of the substation to preselect
            parent_popup: Parent popup to dismiss when opening this one
        """
        # Simply call the main function with the preselected substation
        self.show_maintenance_menu(
            instance=None,
            preselected_substation_name=substation_name,
            parent_popup=parent_popup
        )
    
    def show_maintenance_history(self, instance):
        """Show maintenance history"""
        font_kwargs = self._get_ui_font_kwargs()
        c = self.conn.cursor()
        c.execute("SELECT id, name FROM substations")
        substation_map = {row[1]: row[0] for row in c.fetchall()}

        c.execute('''
            SELECT m.id, s.name, m.name, m.date_time, m.overall_comments
            FROM maintenance m
            JOIN substations s ON m.substation_id = s.id
            ORDER BY m.date_time DESC
        ''')
        all_records = c.fetchall()
        
        if not all_records:
            show_message_popup('Πληροφορία', 'Δεν υπάρχουν καταχωρημένες συντηρήσεις')
            return
        
        popup = Popup(title='Ιστορικό Συντήρησης', size_hint=(0.95, 0.9))
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        filter_bar = BoxLayout(size_hint_y=None, height=40, spacing=10)
        filter_bar.add_widget(Label(text='Φίλτρο Υποσταθμού:', size_hint_x=0.3))
        substation_values = ['(Όλα)'] + sorted(substation_map.keys())
        substation_spinner = Spinner(text='(Όλα)', values=substation_values, size_hint_x=0.7)
        filter_bar.add_widget(substation_spinner)
        main_layout.add_widget(filter_bar)

        scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        grid = GridLayout(cols=1, spacing=10, size_hint_y=None, padding=10)
        grid.bind(minimum_height=grid.setter('height'))

        def fetch_records(selected_substation):
            if selected_substation and selected_substation != '(Όλα)':
                c.execute('''
                    SELECT m.id, s.name, m.name, m.date_time, m.overall_comments
                    FROM maintenance m
                    JOIN substations s ON m.substation_id = s.id
                    WHERE s.name = ?
                    ORDER BY m.date_time DESC
                ''', (selected_substation,))
            else:
                c.execute('''
                    SELECT m.id, s.name, m.name, m.date_time, m.overall_comments
                    FROM maintenance m
                    JOIN substations s ON m.substation_id = s.id
                    ORDER BY m.date_time DESC
                ''')
            return c.fetchall()

        def render_records(selected_substation):
            grid.clear_widgets()
            maintenance_records = fetch_records(selected_substation)
            if not maintenance_records:
                grid.add_widget(Label(text='Δεν υπάρχουν καταχωρημένες συντηρήσεις', size_hint_y=None, height=40))
                return

            for maint_id, sub_name, maint_name, date_time, overall_comments in maintenance_records:
                # Maintenance card
                card = BoxLayout(orientation='vertical', size_hint_y=None, padding=5, spacing=5)
                card.bind(minimum_height=card.setter('height'))

                # Header
                header = BoxLayout(size_hint_y=None, height=40, spacing=5)
                display_name = maint_name or self._build_maintenance_name(sub_name, date_time)
                header.add_widget(Label(
                    text=f'Συντήρηση: {display_name}',
                    bold=True,
                    size_hint_x=0.45
                ))
                header.add_widget(Label(
                    text=f'Ημ/νία: {date_time}',
                    size_hint_x=0.2
                ))
                edit_btn = Button(
                    text='Επεξ.',
                    size_hint_x=0.11
                )
                email_btn = Button(
                    text='Email',
                    size_hint_x=0.12
                )
                delete_btn = Button(
                    text='Διαγραφή',
                    size_hint_x=0.12
                )
                def make_delete_handler(m_id, p):
                    return lambda x: self.confirm_delete_maintenance(m_id, p)
                def make_email_handler(m_id):
                    return lambda x: self.send_maintenance_email_report(m_id)
                def make_edit_handler(m_id, p):
                    return lambda x: self.show_maintenance_menu(None, None, p, m_id, lambda: self.show_maintenance_history(None))
                delete_btn.bind(on_press=make_delete_handler(maint_id, popup))
                email_btn.bind(on_press=make_email_handler(maint_id))
                edit_btn.bind(on_press=make_edit_handler(maint_id, popup))
                header.add_widget(edit_btn)
                header.add_widget(delete_btn)
                header.add_widget(email_btn)
                card.add_widget(header)

                # Responsible and crew
                responsible, crew = self._get_maintenance_people(maint_id)
                if responsible or crew:
                    crew_text = ', '.join(crew) if crew else '-'
                    resp_text = responsible if responsible else '-'
                    people_label = Label(
                        text=f'Υπεύθυνος: {resp_text} | Ομάδα: {crew_text}',
                        size_hint_y=None,
                        height=25
                    )
                    people_label.bind(
                        width=lambda instance, value: setattr(instance, 'text_size', (value, None)),
                        texture_size=lambda instance, value: setattr(instance, 'height', value[1] + 6)
                    )
                    card.add_widget(people_label)

                # Overall comments
                if overall_comments:
                    comment_label = Label(
                        text=f'Σχόλια: {overall_comments}',
                        size_hint_y=None,
                        height=30
                    )
                    comment_label.bind(
                        width=lambda instance, value: setattr(instance, 'text_size', (value, None)),
                        texture_size=lambda instance, value: setattr(instance, 'height', value[1] + 6)
                    )
                    card.add_widget(comment_label)

                # Get elements for this maintenance
                c.execute('''
                    SELECT e.id, e.element_type, e.name, e.serial_number, me.element_comments, e.breaker_category
                    FROM maintenance_elements me
                    JOIN elements e ON me.element_id = e.id
                    WHERE me.maintenance_id = ?
                ''', (maint_id,))
                elements = c.fetchall()

                # Elements list
                elements_label = Label(
                    text='Στοιχεία που συντηρήθηκαν:',
                    size_hint_y=None,
                    height=25,
                    bold=True
                )
                card.add_widget(elements_label)

                for elem_id, elem_type, elem_name, serial_num, elem_comments, breaker_category in elements:
                    # Element info with optional PDF button
                    elem_row = BoxLayout(size_hint_y=None, height=40, spacing=5)
                    elem_row.bind(minimum_height=elem_row.setter('height'))

                    elem_text = f'  • {elem_type}: {elem_name} (S/N: {serial_num or "-"})'
                    if elem_comments:
                        elem_text += f'\n    Σχόλια: {elem_comments}'

                    elem_label = Label(
                        text=elem_text,
                        size_hint_x=0.6,
                        size_hint_y=None
                    )
                    elem_label.bind(
                        width=lambda instance, value: setattr(instance, 'text_size', (value, None)),
                        texture_size=lambda instance, value: (
                            setattr(instance, 'height', value[1] + 6),
                            setattr(elem_row, 'height', max(40, value[1] + 10))
                        )
                    )
                    elem_row.add_widget(elem_label)

                    # Add PDF button for circuit breakers (check Greek names from BREAKER_CATEGORIES_ALL)
                    buttons_container = BoxLayout(size_hint_x=0.4, spacing=5)

                    view_btn = Button(
                        text='Εμφ.',
                        size_hint_x=0.34,
                        size_hint_y=None,
                        height=35,
                        **font_kwargs
                    )

                    def make_view_handler(m_id, e_id, e_name):
                        return lambda x: self.show_maintenance_element_details(m_id, e_id, e_name)

                    view_btn.bind(on_press=make_view_handler(maint_id, elem_id, elem_name))
                    buttons_container.add_widget(view_btn)

                    if 'Διακόπτης' in elem_type and breaker_category in self.BREAKER_CATEGORIES_ALL:
                        pdf_btn = Button(
                            text='PDF',
                            size_hint_x=0.5,
                            size_hint_y=None,
                            height=35,
                            **font_kwargs
                        )

                        def make_pdf_handler(m_id, e_id, e_name):
                            return lambda x: self.generate_pdf_report(m_id, e_id, e_name)

                        pdf_btn.bind(on_press=make_pdf_handler(maint_id, elem_id, elem_name))
                        buttons_container.add_widget(pdf_btn)
                    else:
                        buttons_container.add_widget(Label(text='', size_hint_x=0.5))

                    elem_row.add_widget(buttons_container)

                    card.add_widget(elem_row)

                grid.add_widget(card)

        render_records(substation_spinner.text)
        substation_spinner.bind(text=lambda _spinner, text: render_records(text))
        
        scroll.add_widget(grid)
        main_layout.add_widget(scroll)
        
        # Close button
        close_btn = Button(text='Κλείσιμο', size_hint_y=0.1)
        close_btn.bind(on_press=popup.dismiss)
        main_layout.add_widget(close_btn)
        
        popup.content = main_layout
        popup.open()
    
    def show_substation_maintenance_history(self, substation_id, substation_name, parent_display_popup=None):
        """Show maintenance history for a specific substation"""
        font_kwargs = self._get_ui_font_kwargs()
        c = self.conn.cursor()
        c.execute('''
            SELECT m.id, m.name, m.date_time, m.overall_comments
            FROM maintenance m
            WHERE m.substation_id = ?
            ORDER BY m.date_time DESC
        ''', (substation_id,))
        maintenance_records = c.fetchall()
        
        popup = Popup(title=f'Ιστορικό Συντήρησης: {substation_name}', size_hint=(0.95, 0.9))
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # Add Maintenance button at the top
        add_maint_btn = Button(text='+ Προσθήκη Νέας Συντήρησης', size_hint_y=0.1)
        add_maint_btn.bind(on_press=lambda x: self.show_maintenance_menu_for_substation(substation_id, substation_name, popup))
        main_layout.add_widget(add_maint_btn)
        
        if not maintenance_records:
            # Show message but still allow adding maintenance
            no_records_label = Label(
                text=f'Δεν υπάρχουν καταχωρημένες συντηρήσεις για τον υποσταθμό "{substation_name}".\nΧρησιμοποιήστε το κουμπί παραπάνω για να προσθέσετε.',
                size_hint_y=0.7
            )
            main_layout.add_widget(no_records_label)
        else:
            scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
            grid = GridLayout(cols=1, spacing=10, size_hint_y=None, padding=10)
            grid.bind(minimum_height=grid.setter('height'))
        
        for maint_id, maint_name, date_time, overall_comments in maintenance_records:
            # Maintenance card
            card = BoxLayout(orientation='vertical', size_hint_y=None, padding=5, spacing=5)
            card.bind(minimum_height=card.setter('height'))
            
            # Header
            header = BoxLayout(size_hint_y=None, height=40, spacing=5)
            display_name = maint_name or self._build_maintenance_name(substation_name, date_time)
            header.add_widget(Label(
                text=f'Συντήρηση: {display_name}',
                bold=True,
                size_hint_x=0.6
            ))
            edit_btn = Button(
                text='Επεξ.',
                size_hint_x=0.12
            )
            email_btn = Button(
                text='Email',
                size_hint_x=0.13
            )
            delete_btn = Button(
                text='Διαγραφή',
                size_hint_x=0.15
            )
            def make_delete_handler(m_id, p):
                return lambda x: self.confirm_delete_maintenance_for_substation(m_id, p, substation_id, substation_name, parent_display_popup)
            delete_btn.bind(on_press=make_delete_handler(maint_id, popup))
            def make_email_handler(m_id):
                return lambda x: self.send_maintenance_email_report(m_id)
            email_btn.bind(on_press=make_email_handler(maint_id))
            def make_edit_handler(m_id, p):
                return lambda x: self.show_maintenance_menu(None, substation_name, p, m_id, lambda: self.show_substation_maintenance_history(substation_id, substation_name, parent_display_popup))
            edit_btn.bind(on_press=make_edit_handler(maint_id, popup))
            header.add_widget(edit_btn)
            header.add_widget(email_btn)
            header.add_widget(delete_btn)
            card.add_widget(header)

            # Responsible and crew
            responsible, crew = self._get_maintenance_people(maint_id)
            if responsible or crew:
                crew_text = ', '.join(crew) if crew else '-'
                resp_text = responsible if responsible else '-'
                people_label = Label(
                    text=f'Υπεύθυνος: {resp_text} | Ομάδα: {crew_text}',
                    size_hint_y=None,
                    height=25
                )
                people_label.bind(
                    width=lambda instance, value: setattr(instance, 'text_size', (value, None)),
                    texture_size=lambda instance, value: setattr(instance, 'height', value[1] + 6)
                )
                card.add_widget(people_label)
            
            # Overall comments
            if overall_comments:
                comment_label = Label(
                    text=f'Σχόλια: {overall_comments}',
                    size_hint_y=None,
                    height=30
                )
                comment_label.bind(
                    width=lambda instance, value: setattr(instance, 'text_size', (value, None)),
                    texture_size=lambda instance, value: setattr(instance, 'height', value[1] + 6)
                )
                card.add_widget(comment_label)
            
            # Get elements for this maintenance
            c.execute('''
                SELECT e.id, e.element_type, e.name, e.serial_number, me.element_comments, e.breaker_category
                FROM maintenance_elements me
                JOIN elements e ON me.element_id = e.id
                WHERE me.maintenance_id = ?
            ''', (maint_id,))
            elements = c.fetchall()
            
            # Elements list
            elements_label = Label(
                text='Στοιχεία που συντηρήθηκαν:',
                size_hint_y=None,
                height=25,
                bold=True
            )
            card.add_widget(elements_label)
            
            for elem_id, elem_type, elem_name, serial_num, elem_comments, breaker_category in elements:
                # Element info with optional PDF button
                elem_row = BoxLayout(size_hint_y=None, height=40, spacing=5)
                elem_row.bind(minimum_height=elem_row.setter('height'))
                
                elem_text = f'  • {elem_type}: {elem_name} (S/N: {serial_num or "-"})'
                if elem_comments:
                    elem_text += f'\n    Σχόλια: {elem_comments}'
                
                elem_label = Label(
                    text=elem_text,
                    size_hint_x=0.6,
                    size_hint_y=None
                )
                elem_label.bind(
                    width=lambda instance, value: setattr(instance, 'text_size', (value, None)),
                    texture_size=lambda instance, value: (
                        setattr(instance, 'height', value[1] + 6),
                        setattr(elem_row, 'height', max(40, value[1] + 10))
                    )
                )
                elem_row.add_widget(elem_label)
                
                # Add PDF button for circuit breakers (check Greek names from BREAKER_CATEGORIES_ALL)
                buttons_container = BoxLayout(size_hint_x=0.4, spacing=5)

                view_btn = Button(
                    text='Εμφ.',
                    size_hint_x=0.34,
                    size_hint_y=None,
                    height=35,
                    **font_kwargs
                )

                def make_view_handler(m_id, e_id, e_name):
                    return lambda x: self.show_maintenance_element_details(m_id, e_id, e_name)

                view_btn.bind(on_press=make_view_handler(maint_id, elem_id, elem_name))
                buttons_container.add_widget(view_btn)

                if 'Διακόπτης' in elem_type and breaker_category in self.BREAKER_CATEGORIES_ALL:
                    pdf_btn = Button(
                        text='PDF',
                        size_hint_x=0.5,
                        size_hint_y=None,
                        height=35,
                        **font_kwargs
                    )

                    def make_pdf_handler(m_id, e_id, e_name):
                        return lambda x: self.generate_pdf_report(m_id, e_id, e_name)

                    pdf_btn.bind(on_press=make_pdf_handler(maint_id, elem_id, elem_name))
                    buttons_container.add_widget(pdf_btn)
                else:
                    buttons_container.add_widget(Label(text='', size_hint_x=0.5))

                elem_row.add_widget(buttons_container)
                
                card.add_widget(elem_row)
            
            grid.add_widget(card)
        
            scroll.add_widget(grid)
            main_layout.add_widget(scroll)
        
        # Close button
        close_btn = Button(text='Κλείσιμο', size_hint_y=0.1)
        close_btn.bind(on_press=popup.dismiss)
        main_layout.add_widget(close_btn)
        
        popup.content = main_layout
        popup.open()

    def show_substation_inspection_history(self, substation_id, substation_name, parent_display_popup=None):
        """Show inspection history for a specific substation."""
        font_kwargs = self._get_ui_font_kwargs()
        c = self.conn.cursor()
        c.execute('''
            SELECT id, inspection_date, data_json
            FROM inspections
            WHERE substation_id = ?
            ORDER BY inspection_date DESC
        ''', (substation_id,))
        inspection_records = c.fetchall()

        popup = Popup(title=f'Ιστορικό Επιθεωρήσεων: {substation_name}', size_hint=(0.95, 0.9))
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        add_insp_btn = Button(text='+ Προσθήκη Νέας Επιθεώρησης', size_hint_y=0.1)
        add_insp_btn.bind(on_press=lambda x: self.show_inspection_entry_popup(None, substation_name, popup))
        main_layout.add_widget(add_insp_btn)

        if not inspection_records:
            no_records_label = Label(
                text=f'Δεν υπάρχουν καταχωρημένες επιθεωρήσεις για τον υποσταθμό "{substation_name}".\nΧρησιμοποιήστε το κουμπί παραπάνω για να προσθέσετε.',
                size_hint_y=0.7
            )
            main_layout.add_widget(no_records_label)
        else:
            scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
            grid = GridLayout(cols=1, spacing=10, size_hint_y=None, padding=10)
            grid.bind(minimum_height=grid.setter('height'))

            for insp_id, inspection_date, data_json in inspection_records:
                card = BoxLayout(orientation='horizontal', size_hint_y=None, height=40, spacing=5)
                card.add_widget(Label(
                    text=f'Ημερομηνία: {inspection_date}',
                    size_hint_x=0.6
                ))

                buttons_box = BoxLayout(size_hint_x=0.4, spacing=5)
                view_btn = Button(text='Εμφ.', size_hint_x=0.34, **font_kwargs)
                pdf_btn = Button(text='PDF', size_hint_x=0.33, **font_kwargs)
                email_btn = Button(text='Email', size_hint_x=0.33)
                view_btn.bind(on_press=lambda x, iid=insp_id: self.show_inspection_details(iid))
                pdf_btn.bind(on_press=lambda x, iid=insp_id: self.generate_inspection_pdf(iid, substation_name))
                email_btn.bind(on_press=lambda x, iid=insp_id: self.send_inspection_email_report(iid))
                buttons_box.add_widget(view_btn)
                buttons_box.add_widget(pdf_btn)
                buttons_box.add_widget(email_btn)
                card.add_widget(buttons_box)

                grid.add_widget(card)

            scroll.add_widget(grid)
            main_layout.add_widget(scroll)

        close_btn = Button(text='Κλείσιμο', size_hint_y=0.1)
        close_btn.bind(on_press=popup.dismiss)
        main_layout.add_widget(close_btn)

        popup.content = main_layout
        popup.open()

    def show_inspection_details(self, inspection_id):
        """Display inspection details inside the app."""
        c = self.conn.cursor()
        c.execute("""
            SELECT substation_name, inspection_date, data_json
            FROM inspections
            WHERE id = ?
        """, (inspection_id,))
        row = c.fetchone()

        if not row:
            show_message_popup('Σφάλμα', 'Δεν βρέθηκε η επιθεώρηση.')
            return

        substation_name, inspection_date, data_json = row

        try:
            data = json.loads(data_json or '{}')
        except Exception:
            data = {}
        fields = data.get('fields', [])

        popup = Popup(title='Προβολή Επιθεώρησης', size_hint=(0.95, 0.9))
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        header_text = f'Υποσταθμός: {substation_name or "-"} | Ημερομηνία: {inspection_date}'
        main_layout.add_widget(Label(text=header_text, size_hint_y=None, height=30))

        scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        content = GridLayout(cols=1, spacing=8, size_hint_y=None, padding=10)
        content.bind(minimum_height=content.setter('height'))

        def make_wrapped_label(text, size_hint_x):
            lbl = Label(text=text, size_hint_x=size_hint_x, size_hint_y=None)
            lbl.bind(width=lambda inst, val: setattr(inst, 'text_size', (val, None)))
            lbl.bind(texture_size=lambda inst, val: setattr(inst, 'height', max(30, val[1])))
            return lbl

        def build_inspection_sections():
            fallback = self._get_inspection_fallback_fields()
            sections = []
            current_title = 'Στοιχεία Επιθεώρησης'
            current_labels = []
            for item in fallback:
                if isinstance(item, dict) and item.get('type') == 'section':
                    if current_labels:
                        sections.append((current_title, list(current_labels)))
                    current_title = item.get('title', '')
                    current_labels = []
                elif isinstance(item, str):
                    current_labels.append(item)
            if current_labels:
                sections.append((current_title, list(current_labels)))
            label_to_section = {}
            for title, labels in sections:
                for label in labels:
                    label_to_section[label] = title
            return sections, label_to_section

        if not fields:
            content.add_widget(Label(text='Δεν υπάρχουν διαθέσιμα δεδομένα επιθεώρησης.', size_hint_y=None, height=30))
        else:
            sections, label_to_section = build_inspection_sections()
            section_items = {title: [] for title, _labels in sections}
            other_items = []

            for field in fields:
                if not isinstance(field, dict):
                    continue
                label = field.get('label', '')
                if label in label_to_section:
                    section_items[label_to_section[label]].append(field)
                else:
                    other_items.append(field)

            for title, _labels in sections:
                items = section_items.get(title) or []
                if not items:
                    continue
                section_title = Label(text=f'[b]{title}[/b]', markup=True, size_hint_y=None, height=26)
                content.add_widget(section_title)
                grid = GridLayout(cols=2, spacing=6, size_hint_y=None)
                grid.bind(minimum_height=grid.setter('height'))
                for field in items:
                    label = field.get('label', '')
                    value = field.get('value', '')
                    grid.add_widget(make_wrapped_label(str(label), 0.35))
                    grid.add_widget(make_wrapped_label(str(value), 0.65))
                content.add_widget(grid)

            if other_items:
                section_title = Label(text='[b]Λοιπά[/b]', markup=True, size_hint_y=None, height=26)
                content.add_widget(section_title)
                grid = GridLayout(cols=2, spacing=6, size_hint_y=None)
                grid.bind(minimum_height=grid.setter('height'))
                for field in other_items:
                    label = field.get('label', '')
                    value = field.get('value', '')
                    grid.add_widget(make_wrapped_label(str(label), 0.35))
                    grid.add_widget(make_wrapped_label(str(value), 0.65))
                content.add_widget(grid)

        scroll.add_widget(content)
        main_layout.add_widget(scroll)

        close_btn = Button(text='Κλείσιμο', size_hint_y=None, height=40)
        close_btn.bind(on_press=popup.dismiss)
        main_layout.add_widget(close_btn)

        popup.content = main_layout
        popup.open()

    def show_maintenance_element_details(self, maintenance_id, element_id, element_name):
        """Show stored comments/measurements for a maintenance element entry."""
        c = self.conn.cursor()
        c.execute("""
            SELECT m.name, m.date_time, m.overall_comments, m.maintenance_type, m.user_name,
                   s.name as substation_name, s.location, s.division,
                   e.element_type, e.name, e.serial_number, e.manufacturer, e.model,
                   e.breaker_category, e.voltage_level, e.gate, e.manufacture_year,
                   em.model_name, em.manufacturer as model_manufacturer
            FROM maintenance m
            JOIN substations s ON m.substation_id = s.id
            JOIN elements e ON e.id = ?
            LEFT JOIN element_models em ON e.element_model_id = em.id
            WHERE m.id = ?
            LIMIT 1
        """, (element_id, maintenance_id))
        header_row = c.fetchone()

        c.execute("""
            SELECT me.element_comments,
                   me.insulation_closed_fa_ground, me.insulation_closed_fa_unit,
                   me.insulation_closed_fb_ground, me.insulation_closed_fb_unit,
                   me.insulation_closed_fc_ground, me.insulation_closed_fc_unit,
                   me.insulation_open_fa_fa, me.insulation_open_fa_unit,
                   me.insulation_open_fb_fb, me.insulation_open_fb_unit,
                   me.insulation_open_fc_fc, me.insulation_open_fc_unit,
                   me.contact_resistance_fa_fa, me.contact_resistance_fb_fb, me.contact_resistance_fc_fc,
                   me.operations_count,
                                         me.sf6_leakage_kg, me.sf6_leak_methodology,
                   me.sf6_n2_fa, me.h2o_fa, me.so2_fa,
                   me.sf6_n2_fb, me.h2o_fb, me.so2_fb,
                   me.sf6_n2_fc, me.h2o_fc, me.so2_fc,
                   me.vidar_fa, me.vidar_fb, me.vidar_fc,
                   e.breaker_category
            FROM maintenance_elements me
            JOIN elements e ON me.element_id = e.id
            WHERE me.maintenance_id = ? AND me.element_id = ?
            LIMIT 1
        """, (maintenance_id, element_id))
        row = c.fetchone()

        if not row:
            show_message_popup('Πληροφορία', 'Δεν βρέθηκαν στοιχεία για το στοιχείο.')
            return

        (
            element_comments,
            ins_closed_fa, ins_closed_fa_unit,
            ins_closed_fb, ins_closed_fb_unit,
            ins_closed_fc, ins_closed_fc_unit,
            ins_open_fa, ins_open_fa_unit,
            ins_open_fb, ins_open_fb_unit,
            ins_open_fc, ins_open_fc_unit,
            cont_fa, cont_fb, cont_fc,
            ops_count,
            sf6_leakage_kg, sf6_leak_methodology,
            sf6_n2_fa, h2o_fa, so2_fa,
            sf6_n2_fb, h2o_fb, so2_fb,
            sf6_n2_fc, h2o_fc, so2_fc,
            vidar_fa, vidar_fb, vidar_fc,
            breaker_category
        ) = row

        def fmt(val, unit=None):
            if val is None or val == '':
                return '-'
            return f"{val} {unit}" if unit else f"{val}"

        def make_wrapped_label(text, bold=False, size_hint_x=1, font_size='14sp'):
            lbl = Label(
                text=f"[b]{text}[/b]" if bold else str(text),
                markup=bold,
                size_hint_x=size_hint_x,
                size_hint_y=None,
                font_size=font_size,
                halign='left',
                valign='top'
            )
            lbl.bind(
                width=lambda inst, val: setattr(inst, 'text_size', (val, None)),
                texture_size=lambda inst, val: setattr(inst, 'height', val[1] + 6)
            )
            return lbl

        def add_kv_row(grid, label_text, value_text):
            grid.add_widget(make_wrapped_label(label_text, bold=True, size_hint_x=0.38))
            grid.add_widget(make_wrapped_label(value_text, bold=False, size_hint_x=0.62))

        def add_section(title):
            content.add_widget(make_wrapped_label(title, bold=True, font_size='15sp'))

        has_measurements = any([
            ins_closed_fa, ins_closed_fb, ins_closed_fc,
            ins_open_fa, ins_open_fb, ins_open_fc,
            cont_fa, cont_fb, cont_fc,
            ops_count,
            sf6_n2_fa, h2o_fa, so2_fa,
            sf6_leakage_kg,
            sf6_n2_fb, h2o_fb, so2_fb,
            sf6_n2_fc, h2o_fc, so2_fc,
            vidar_fa, vidar_fb, vidar_fc
        ])

        if not has_measurements and not element_comments:
            show_message_popup('Πληροφορία', 'Δεν υπάρχουν καταχωρημένα στοιχεία για αυτό το στοιχείο.')
            return

        popup = Popup(title=f'Μετρήσεις: {element_name}', size_hint=(0.9, 0.9))
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        scroll = ScrollView(bar_width=10, scroll_type=['bars', 'content'])
        content = GridLayout(cols=1, spacing=6, size_hint_y=None, padding=10)
        content.bind(minimum_height=content.setter('height'))

        if header_row:
            (
                maint_name, maint_date, maint_comments, maint_type, maint_user,
                sub_name, sub_location, division,
                elem_type, elem_name, serial_number, manufacturer, model,
                breaker_cat, voltage_level, gate, manufacture_year,
                model_name, model_manufacturer
            ) = header_row

            add_section('Στοιχεία Συντήρησης')
            grid = GridLayout(cols=2, spacing=6, size_hint_y=None)
            grid.bind(minimum_height=grid.setter('height'))
            add_kv_row(grid, 'Υποσταθμός', sub_name or '-')
            add_kv_row(grid, 'Ημερομηνία', maint_date or '-')
            add_kv_row(grid, 'Τύπος Συντήρησης', maint_type or '-')
            add_kv_row(grid, 'Χειριστής', maint_user or '-')
            add_kv_row(grid, 'Τομέας', division or '-')
            add_kv_row(grid, 'Τοποθεσία', sub_location or '-')
            content.add_widget(grid)

            add_section('Στοιχεία Διακόπτη')
            grid = GridLayout(cols=2, spacing=6, size_hint_y=None)
            grid.bind(minimum_height=grid.setter('height'))
            add_kv_row(grid, 'Τύπος', elem_type or '-')
            add_kv_row(grid, 'Όνομα', elem_name or '-')
            add_kv_row(grid, 'S/N', serial_number or '-')
            add_kv_row(grid, 'Κατασκευαστής', manufacturer or '-')
            if model_name or model_manufacturer:
                add_kv_row(grid, 'Μοντέλο (Βάση)', f"{model_name or '-'} / {model_manufacturer or '-'}")
            add_kv_row(grid, 'Μοντέλο (Στοιχείο)', model or '-')
            add_kv_row(grid, 'Κατηγορία Διακόπτη', breaker_cat or '-')
            add_kv_row(grid, 'Τάση', voltage_level or '-')
            add_kv_row(grid, 'Πύλη', gate or '-')
            add_kv_row(grid, 'Έτος Κατασκευής', manufacture_year or '-')
            content.add_widget(grid)

            add_section('Σχόλια Συντήρησης')
            content.add_widget(make_wrapped_label(maint_comments or '-', bold=False))

        add_section('Σχόλια Στοιχείου')
        content.add_widget(make_wrapped_label(element_comments or '-', bold=False))

        if has_measurements:
            add_section('Αντίσταση Μόνωσης - Διακόπτης Κλειστός (Γη)')
            grid = GridLayout(cols=2, spacing=6, size_hint_y=None)
            grid.bind(minimum_height=grid.setter('height'))
            add_kv_row(grid, 'ΦΑ-Γη', fmt(ins_closed_fa, ins_closed_fa_unit))
            add_kv_row(grid, 'ΦΒ-Γη', fmt(ins_closed_fb, ins_closed_fb_unit))
            add_kv_row(grid, 'ΦΓ-Γη', fmt(ins_closed_fc, ins_closed_fc_unit))
            content.add_widget(grid)

            add_section('Αντίσταση Μόνωσης - Διακόπτης Ανοικτός (Φάση-Φάση)')
            grid = GridLayout(cols=2, spacing=6, size_hint_y=None)
            grid.bind(minimum_height=grid.setter('height'))
            add_kv_row(grid, 'ΦΑ-ΦΑ', fmt(ins_open_fa, ins_open_fa_unit))
            add_kv_row(grid, 'ΦΒ-ΦΒ', fmt(ins_open_fb, ins_open_fb_unit))
            add_kv_row(grid, 'ΦΓ-ΦΓ', fmt(ins_open_fc, ins_open_fc_unit))
            content.add_widget(grid)

            add_section('Αντίσταση Διέλευσης (μΩ)')
            grid = GridLayout(cols=2, spacing=6, size_hint_y=None)
            grid.bind(minimum_height=grid.setter('height'))
            add_kv_row(grid, 'ΦΑ-ΦΑ', fmt(cont_fa))
            add_kv_row(grid, 'ΦΒ-ΦΒ', fmt(cont_fb))
            add_kv_row(grid, 'ΦΓ-ΦΓ', fmt(cont_fc))
            content.add_widget(grid)

            add_section('Μετρητής Χειρισμών')
            grid = GridLayout(cols=2, spacing=6, size_hint_y=None)
            grid.bind(minimum_height=grid.setter('height'))
            add_kv_row(grid, 'Αριθμός Χειρισμών', fmt(ops_count))
            content.add_widget(grid)

        if has_measurements and breaker_category == 'SF6' and (sf6_leakage_kg is not None or any([sf6_n2_fa, h2o_fa, so2_fa, sf6_n2_fb, h2o_fb, so2_fb, sf6_n2_fc, h2o_fc, so2_fc])):
            add_section('Ποιότητα Αερίου SF6')
            grid = GridLayout(cols=2, spacing=6, size_hint_y=None)
            grid.bind(minimum_height=grid.setter('height'))
            add_kv_row(grid, 'Διαρροή SF6 (kg)', fmt(sf6_leakage_kg))
            add_kv_row(grid, 'Πλήρωση/Αντικατάσταση', sf6_leak_methodology or '-')
            add_kv_row(grid, 'ΦΑ', f"SF6/N2 {fmt(sf6_n2_fa)} | H2O {fmt(h2o_fa)} | SO2 {fmt(so2_fa)}")
            add_kv_row(grid, 'ΦΒ', f"SF6/N2 {fmt(sf6_n2_fb)} | H2O {fmt(h2o_fb)} | SO2 {fmt(so2_fb)}")
            add_kv_row(grid, 'ΦΓ', f"SF6/N2 {fmt(sf6_n2_fc)} | H2O {fmt(h2o_fc)} | SO2 {fmt(so2_fc)}")
            content.add_widget(grid)

        if has_measurements and breaker_category in ['Vacuum', 'Κενού'] and any([vidar_fa, vidar_fb, vidar_fc]):
            add_section('Έλεγχος Κενού (VIDAR)')
            grid = GridLayout(cols=2, spacing=6, size_hint_y=None)
            grid.bind(minimum_height=grid.setter('height'))
            add_kv_row(grid, 'ΦΑ-ΦΑ', fmt(vidar_fa))
            add_kv_row(grid, 'ΦΒ-ΦΒ', fmt(vidar_fb))
            add_kv_row(grid, 'ΦΓ-ΦΓ', fmt(vidar_fc))
            content.add_widget(grid)

        scroll.add_widget(content)
        main_layout.add_widget(scroll)

        close_btn = Button(text='Κλείσιμο', size_hint_y=None, height=40)
        close_btn.bind(on_press=popup.dismiss)
        main_layout.add_widget(close_btn)

        popup.content = main_layout
        popup.open()
    
    def generate_pdf_report(self, maintenance_id, element_id, element_name):
        """Generate PDF maintenance report for a circuit breaker"""
        try:
            # Generate the PDF
            pdf_path = generate_maintenance_report(self.conn, maintenance_id, element_id)
            
            # Show success message and offer to open
            from kivy.uix.popup import Popup
            from kivy.uix.boxlayout import BoxLayout
            from kivy.uix.button import Button
            from kivy.uix.label import Label
            
            confirm_popup = Popup(title='PDF Δημιουργήθηκε', size_hint=(0.6, 0.4))
            layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
            
            msg_label = Label(
                text=f'Το αρχείο PDF για το στοιχείο "{element_name}"\nδημιουργήθηκε επιτυχώς!',
                size_hint_y=0.5
            )
            layout.add_widget(msg_label)
            
            path_label = Label(
                text=f'Αποθηκεύτηκε στο:\n{pdf_path}',
                size_hint_y=0.3,
                font_size='10sp'
            )
            layout.add_widget(path_label)
            
            buttons_layout = BoxLayout(size_hint_y=0.2, spacing=10)
            
            def open_pdf():
                import subprocess
                import sys
                if sys.platform == 'win32':
                    os.startfile(pdf_path)
                elif sys.platform == 'darwin':
                    subprocess.call(['open', pdf_path])
                else:
                    subprocess.call(['xdg-open', pdf_path])
                confirm_popup.dismiss()
            
            open_btn = Button(text='Άνοιγμα PDF')
            open_btn.bind(on_press=lambda x: open_pdf())
            buttons_layout.add_widget(open_btn)
            
            close_btn = Button(text='Κλείσιμο')
            close_btn.bind(on_press=confirm_popup.dismiss)
            buttons_layout.add_widget(close_btn)
            
            layout.add_widget(buttons_layout)
            confirm_popup.content = layout
            confirm_popup.open()
            
        except Exception as e:
            show_message_popup('Σφάλμα', f'Αποτυχία δημιουργίας PDF:\n{str(e)}')
    
    def confirm_delete_maintenance(self, maintenance_id, parent_popup):
        """Confirm before deleting a maintenance record."""
        confirm_popup = Popup(title='Επιβεβαίωση Διαγραφής', size_hint=(0.6, 0.3))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        warning_label = Label(
            text='Είστε σίγουροι ότι θέλετε να διαγράψετε\nαυτή τη συντήρηση;',
            size_hint_y=0.6
        )
        layout.add_widget(warning_label)

        buttons_layout = BoxLayout(size_hint_y=0.3, spacing=10)

        def confirm():
            confirm_popup.dismiss()
            self.delete_maintenance(maintenance_id, parent_popup)

        yes_btn = Button(text='ΝΑΙ', color=(1, 0, 0, 1))
        yes_btn.bind(on_press=lambda x: confirm())
        buttons_layout.add_widget(yes_btn)

        no_btn = Button(text='ΟΧΙ')
        no_btn.bind(on_press=confirm_popup.dismiss)
        buttons_layout.add_widget(no_btn)

        layout.add_widget(buttons_layout)
        confirm_popup.content = layout
        confirm_popup.open()

    def delete_maintenance(self, maintenance_id, parent_popup):
        """Delete a maintenance record and update related last maintenance dates"""
        c = self.conn.cursor()
        
        # Get substation_id and affected elements before deletion
        c.execute("SELECT substation_id FROM maintenance WHERE id=?", (maintenance_id,))
        result = c.fetchone()
        if not result:
            show_message_popup('Σφάλμα', 'Η συντήρηση δεν βρέθηκε!')
            return
        substation_id = result[0]
        
        c.execute("SELECT element_id FROM maintenance_elements WHERE maintenance_id=?", (maintenance_id,))
        affected_elements = [row[0] for row in c.fetchall()]
        
        # Explicitly delete maintenance_elements records first
        c.execute("DELETE FROM maintenance_elements WHERE maintenance_id=?", (maintenance_id,))
        
        # Delete the maintenance record
        c.execute("DELETE FROM maintenance WHERE id=?", (maintenance_id,))
        
        # Update last maintenance date for each affected element
        for element_id in affected_elements:
            c.execute("""
                SELECT m.date_time 
                FROM maintenance m
                JOIN maintenance_elements me ON m.id = me.maintenance_id
                WHERE me.element_id = ?
                ORDER BY m.date_time DESC
                LIMIT 1
            """, (element_id,))
            result = c.fetchone()
            new_date = result[0] if result else ''
            c.execute("UPDATE elements SET maintenance_date=? WHERE id=?", (new_date, element_id))
        
        # Update last maintenance date for the substation
        c.execute("""
            SELECT MAX(date_time) 
            FROM maintenance 
            WHERE substation_id=?
        """, (substation_id,))
        result = c.fetchone()
        new_sub_date = result[0] if result and result[0] else ''
        c.execute("UPDATE substations SET last_maintenance=? WHERE id=?", (new_sub_date, substation_id))
        
        self.conn.commit()
        parent_popup.dismiss()
        
        # Refresh both maintenance history and main records view
        def on_close():
            self.show_maintenance_history(None)
            self.show_records(None)
        
        show_message_popup('Ολοκληρώθηκε', 'Η συντήρηση διαγράφηκε!', callback=on_close)
    
    def confirm_delete_maintenance_for_substation(self, maintenance_id, parent_popup, substation_id, substation_name, parent_display_popup=None):
        """Confirm before deleting a maintenance record for a substation."""
        confirm_popup = Popup(title='Επιβεβαίωση Διαγραφής', size_hint=(0.6, 0.3))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        warning_label = Label(
            text=f'Είστε σίγουροι ότι θέλετε να διαγράψετε\nτη συντήρηση του υποσταθμού "{substation_name}";',
            size_hint_y=0.6
        )
        layout.add_widget(warning_label)

        buttons_layout = BoxLayout(size_hint_y=0.3, spacing=10)

        def confirm():
            confirm_popup.dismiss()
            self.delete_maintenance_for_substation(maintenance_id, parent_popup, substation_id, substation_name, parent_display_popup)

        yes_btn = Button(text='ΝΑΙ', color=(1, 0, 0, 1))
        yes_btn.bind(on_press=lambda x: confirm())
        buttons_layout.add_widget(yes_btn)

        no_btn = Button(text='ΟΧΙ')
        no_btn.bind(on_press=confirm_popup.dismiss)
        buttons_layout.add_widget(no_btn)

        layout.add_widget(buttons_layout)
        confirm_popup.content = layout
        confirm_popup.open()

    def delete_maintenance_for_substation(self, maintenance_id, parent_popup, substation_id, substation_name, parent_display_popup=None):
        """Delete a maintenance record, update last maintenance dates, and refresh substation-specific view"""
        c = self.conn.cursor()
        
        # Get affected elements before deletion
        c.execute("SELECT element_id FROM maintenance_elements WHERE maintenance_id=?", (maintenance_id,))
        affected_elements = [row[0] for row in c.fetchall()]
        
        # Explicitly delete maintenance_elements records first
        c.execute("DELETE FROM maintenance_elements WHERE maintenance_id=?", (maintenance_id,))
        
        # Delete the maintenance record
        c.execute("DELETE FROM maintenance WHERE id=?", (maintenance_id,))
        
        # Update last maintenance date for each affected element
        for element_id in affected_elements:
            c.execute("""
                SELECT m.date_time 
                FROM maintenance m
                JOIN maintenance_elements me ON m.id = me.maintenance_id
                WHERE me.element_id = ?
                ORDER BY m.date_time DESC
                LIMIT 1
            """, (element_id,))
            result = c.fetchone()
            new_date = result[0] if result else None
            c.execute("UPDATE elements SET maintenance_date=? WHERE id=?", (new_date, element_id))
        
        # Update last maintenance date for the substation
        c.execute("""
            SELECT MAX(date_time) 
            FROM maintenance 
            WHERE substation_id=?
        """, (substation_id,))
        result = c.fetchone()
        new_sub_date = result[0] if (result and result[0] is not None) else None
        c.execute("UPDATE substations SET last_maintenance=? WHERE id=?", (new_sub_date, substation_id))
        
        self.conn.commit()
        
        # Close both popups - maintenance history and parent display
        parent_popup.dismiss()
        if parent_display_popup:
            parent_display_popup.dismiss()
    
    def show_isolation_requests(self, instance=None):
        """Show isolation requests in calendar view"""
        from datetime import datetime, timedelta
        from calendar import monthrange

        font_kwargs = self._get_ui_font_kwargs()
        
        popup = Popup(title='Αιτήσεις Απομόνωσης', size_hint=(0.95, 0.95))
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # Current month/year state
        current_date = datetime.now()
        current_month = [current_date.month]
        current_year = [current_date.year]
        
        # Top controls
        controls_layout = BoxLayout(size_hint_y=0.1, spacing=10)
        
        prev_btn = Button(text='◀ Προηγούμενος', **font_kwargs)
        next_btn = Button(text='Επόμενος ▶', **font_kwargs)
        today_btn = Button(text='Σήμερα', **font_kwargs)
        add_btn = Button(text='+ Νέα Αίτηση', **font_kwargs)
        
        controls_layout.add_widget(prev_btn)
        controls_layout.add_widget(today_btn)
        controls_layout.add_widget(next_btn)
        controls_layout.add_widget(add_btn)
        
        main_layout.add_widget(controls_layout)
        
        # Month/Year header
        header_label = Label(
            text='',
            size_hint_y=0.08,
            font_size='20sp',
            bold=True
        )
        main_layout.add_widget(header_label)
        
        # Calendar container (will be recreated on month change)
        calendar_container = BoxLayout(orientation='vertical')
        main_layout.add_widget(calendar_container)
        
        # Legend
        legend_layout = BoxLayout(size_hint_y=0.08, spacing=10, padding=[10, 5])
        legend_layout.add_widget(Label(text='', size_hint_x=0.3, **font_kwargs))
        legend_layout.add_widget(Label(text='● Αιτήθηκε', size_hint_x=0.2, color=(1, 0.85, 0, 1), **font_kwargs))
        legend_layout.add_widget(Label(text='● Εγκρίθηκε', size_hint_x=0.2, color=(0.2, 0.8, 0.2, 1), **font_kwargs))
        legend_layout.add_widget(Label(text='● Ακυρώθηκε', size_hint_x=0.2, color=(0.9, 0.2, 0.2, 1), **font_kwargs))
        legend_layout.add_widget(Label(text='', size_hint_x=0.1, **font_kwargs))
        main_layout.add_widget(legend_layout)
        
        def load_calendar():
            """Load calendar for current month/year"""
            calendar_container.clear_widgets()
            
            month = current_month[0]
            year = current_year[0]
            
            # Update header
            month_names = ['', 'Ιανουάριος', 'Φεβρουάριος', 'Μάρτιος', 'Απρίλιος', 'Μάιος', 'Ιούνιος',
                          'Ιούλιος', 'Αύγουστος', 'Σεπτέμβριος', 'Οκτώβριος', 'Νοέμβριος', 'Δεκέμβριος']
            header_label.text = f'{month_names[month]} {year}'
            
            # Get all isolation requests for this month
            c = self.conn.cursor()
            first_day = f'{year}-{month:02d}-01 00:00'
            last_day_num = monthrange(year, month)[1]
            last_day = f'{year}-{month:02d}-{last_day_num} 23:59'
            
            c.execute("""
                SELECT ir.id, ir.substation_id, s.name, ir.start_datetime, ir.end_datetime, 
                       ir.status, ir.notes
                FROM isolation_requests ir
                JOIN substations s ON ir.substation_id = s.id
                WHERE (ir.start_datetime <= ? AND ir.end_datetime >= ?)
                   OR (ir.start_datetime >= ? AND ir.start_datetime <= ?)
                ORDER BY ir.start_datetime
            """, (last_day, first_day, first_day, last_day))
            requests = c.fetchall()
            
            # Group requests by day
            requests_by_day = {}
            for req_id, sub_id, sub_name, start_dt, end_dt, status, notes in requests:
                try:
                    start = datetime.strptime(start_dt, '%Y-%m-%d %H:%M')
                    end = datetime.strptime(end_dt, '%Y-%m-%d %H:%M')
                    
                    # Add to all days in range within this month
                    current = start
                    while current <= end:
                        if current.year == year and current.month == month:
                            day = current.day
                            if day not in requests_by_day:
                                requests_by_day[day] = []
                            # Avoid duplicates
                            if not any(r[0] == req_id for r in requests_by_day[day]):
                                requests_by_day[day].append((req_id, sub_id, sub_name, start_dt, end_dt, status, notes))
                        current += timedelta(days=1)
                except Exception:
                    pass
            
            # Create calendar grid
            calendar_grid = GridLayout(cols=7, spacing=2)
            
            # Day headers
            day_names = ['Δευ', 'Τρί', 'Τετ', 'Πέμ', 'Παρ', 'Σάβ', 'Κυρ']
            for day_name in day_names:
                calendar_grid.add_widget(Label(
                    text=day_name,
                    size_hint_y=None,
                    height=30,
                    bold=True
                ))
            
            # Get first day of month (0=Monday, 6=Sunday)
            first_weekday = datetime(year, month, 1).weekday()
            days_in_month = monthrange(year, month)[1]
            
            # Add empty cells for days before month starts
            for _ in range(first_weekday):
                calendar_grid.add_widget(Label(text=''))
            
            # Add days of month
            for day in range(1, days_in_month + 1):
                day_box = BoxLayout(orientation='vertical', size_hint_y=None, height=100)
                
                # Day number
                day_label = Label(
                    text=str(day),
                    size_hint_y=0.3,
                    bold=True
                )
                day_box.add_widget(day_label)
                
                # Requests for this day
                if day in requests_by_day:
                    scroll = ScrollView(size_hint_y=0.7)
                    requests_layout = GridLayout(cols=1, size_hint_y=None, spacing=2, padding=2)
                    requests_layout.bind(minimum_height=requests_layout.setter('height'))
                    
                    for req_id, sub_id, sub_name, start_dt, end_dt, status, notes in requests_by_day[day]:
                        # Color based on status
                        if status == 'Accepted':
                            color = (0.2, 0.8, 0.2, 1)  # Green
                            symbol = '●'
                        elif status == 'Cancelled':
                            color = (0.8, 0.2, 0.2, 1)  # Red
                            symbol = '●'
                        else:  # Requested
                            color = (0.8, 0.8, 0.2, 1)  # Yellow
                            symbol = '●'
                        
                        req_btn = Button(
                            text=f'{symbol} {sub_name[:15]}',
                            size_hint_y=None,
                            height=30,
                            background_color=color,
                            **font_kwargs
                        )
                        
                        def make_request_handler(r_id):
                            return lambda x: self.show_isolation_request_details(r_id, popup)
                        
                        req_btn.bind(on_press=make_request_handler(req_id))
                        requests_layout.add_widget(req_btn)
                    
                    scroll.add_widget(requests_layout)
                    day_box.add_widget(scroll)
                else:
                    day_box.add_widget(Label(text='', size_hint_y=0.7))
                
                calendar_grid.add_widget(day_box)
            
            calendar_container.add_widget(calendar_grid)
        
        def go_prev_month(instance):
            if current_month[0] == 1:
                current_month[0] = 12
                current_year[0] -= 1
            else:
                current_month[0] -= 1
            load_calendar()
        
        def go_next_month(instance):
            if current_month[0] == 12:
                current_month[0] = 1
                current_year[0] += 1
            else:
                current_month[0] += 1
            load_calendar()
        
        def go_today(instance):
            today = datetime.now()
            current_month[0] = today.month
            current_year[0] = today.year
            load_calendar()
        
        def add_request(instance):
            self.show_add_isolation_request(popup)
        
        prev_btn.bind(on_press=go_prev_month)
        next_btn.bind(on_press=go_next_month)
        today_btn.bind(on_press=go_today)
        add_btn.bind(on_press=add_request)
        
        # Load initial calendar
        load_calendar()
        
        # Close button
        close_btn = Button(text='Κλείσιμο', size_hint_y=0.08)
        close_btn.bind(on_press=popup.dismiss)
        main_layout.add_widget(close_btn)
        
        popup.content = main_layout
        popup.open()
    
    def show_add_isolation_request(self, parent_popup):
        """Show dialog to add new isolation request"""
        from datetime import datetime, timedelta
        
        # Get list of substations
        c = self.conn.cursor()
        c.execute("SELECT id, name FROM substations ORDER BY name")
        substations = c.fetchall()
        
        if not substations:
            show_message_popup('Σφάλμα', 'Δεν υπάρχουν υποσταθμοί!')
            return
        
        popup = Popup(title='Νέα Αίτηση Απομόνωσης', size_hint=(0.7, 0.75))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # Substation selection
        layout.add_widget(Label(text='Υποσταθμός:', size_hint_y=None, height=30))
        substation_map = {s[1]: s[0] for s in substations}
        substation_spinner = Spinner(
            text=substations[0][1],
            values=[s[1] for s in substations],
            size_hint_y=None,
            height=40
        )
        layout.add_widget(substation_spinner)
        
        # Start datetime
        layout.add_widget(Label(text='Ημ/νία & Ώρα Έναρξης:', size_hint_y=None, height=30))
        start_input = TextInput(
            text=datetime.now().strftime('%Y-%m-%d %H:%M'),
            hint_text='YYYY-MM-DD HH:MM',
            size_hint_y=None,
            height=35,
            multiline=False
        )
        layout.add_widget(start_input)
        
        # Quick presets for start
        start_presets = BoxLayout(size_hint_y=None, height=35, spacing=5)
        
        def set_start_today_morning():
            start_input.text = datetime.now().strftime('%Y-%m-%d 08:00')
        
        def set_start_today_evening():
            start_input.text = datetime.now().strftime('%Y-%m-%d 18:00')
        
        today_morning_btn = Button(text='Σήμερα 08:00')
        today_morning_btn.bind(on_press=lambda x: set_start_today_morning())
        start_presets.add_widget(today_morning_btn)
        
        today_evening_btn = Button(text='Σήμερα 18:00')
        today_evening_btn.bind(on_press=lambda x: set_start_today_evening())
        start_presets.add_widget(today_evening_btn)
        
        layout.add_widget(start_presets)
        
        # End datetime
        layout.add_widget(Label(text='Ημ/νία & Ώρα Λήξης:', size_hint_y=None, height=30))
        end_input = TextInput(
            text=(datetime.now() + timedelta(hours=4)).strftime('%Y-%m-%d %H:%M'),
            hint_text='YYYY-MM-DD HH:MM',
            size_hint_y=None,
            height=35,
            multiline=False
        )
        layout.add_widget(end_input)
        
        # Quick duration presets
        duration_presets = BoxLayout(size_hint_y=None, height=35, spacing=5)
        
        def set_duration_hours(hours):
            try:
                start = datetime.strptime(start_input.text, '%Y-%m-%d %H:%M')
                end = start + timedelta(hours=hours)
                end_input.text = end.strftime('%Y-%m-%d %H:%M')
            except:
                pass
        
        dur_2h_btn = Button(text='2 ώρες')
        dur_2h_btn.bind(on_press=lambda x: set_duration_hours(2))
        duration_presets.add_widget(dur_2h_btn)
        
        dur_4h_btn = Button(text='4 ώρες')
        dur_4h_btn.bind(on_press=lambda x: set_duration_hours(4))
        duration_presets.add_widget(dur_4h_btn)
        
        dur_1day_btn = Button(text='1 ημέρα')
        dur_1day_btn.bind(on_press=lambda x: set_duration_hours(24))
        duration_presets.add_widget(dur_1day_btn)
        
        layout.add_widget(duration_presets)
        
        # Status
        layout.add_widget(Label(text='Κατάσταση:', size_hint_y=None, height=30))
        status_spinner = Spinner(
            text='Requested',
            values=['Requested', 'Accepted', 'Cancelled'],
            size_hint_y=None,
            height=40
        )
        layout.add_widget(status_spinner)
        
        # Notes
        layout.add_widget(Label(text='Σημειώσεις:', size_hint_y=None, height=30))
        notes_input = TextInput(
            hint_text='Πρόσθετες πληροφορίες...',
            size_hint_y=None,
            height=80,
            multiline=True
        )
        layout.add_widget(notes_input)
        
        # Buttons
        buttons_layout = BoxLayout(size_hint_y=None, height=50, spacing=10)
        
        def save_request():
            substation_id = substation_map[substation_spinner.text]
            start_dt = start_input.text.strip()
            end_dt = end_input.text.strip()
            status = status_spinner.text
            notes = notes_input.text.strip()
            
            # Validate dates
            try:
                start = datetime.strptime(start_dt, '%Y-%m-%d %H:%M')
                end = datetime.strptime(end_dt, '%Y-%m-%d %H:%M')
                
                if end <= start:
                    show_message_popup('Σφάλμα', 'Η ημερομηνία λήξης πρέπει να είναι μετά την έναρξη!')
                    return
            except ValueError:
                show_message_popup('Σφάλμα', 'Μη έγκυρη μορφή ημερομηνίας! Χρησιμοποιήστε: YYYY-MM-DD HH:MM')
                return
            
            # Insert request
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            c = self.conn.cursor()
            c.execute("""
                INSERT INTO isolation_requests 
                (substation_id, start_datetime, end_datetime, status, notes, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (substation_id, start_dt, end_dt, status, notes, now, now))
            self.conn.commit()
            
            popup.dismiss()
            parent_popup.dismiss()
            show_message_popup('Επιτυχία', 'Η αίτηση απομόνωσης καταχωρήθηκε!', 
                             callback=lambda: self.show_isolation_requests(None))
        
        save_btn = Button(text='Αποθήκευση')
        save_btn.bind(on_press=lambda x: save_request())
        buttons_layout.add_widget(save_btn)
        
        cancel_btn = Button(text='Ακύρωση')
        cancel_btn.bind(on_press=popup.dismiss)
        buttons_layout.add_widget(cancel_btn)
        
        layout.add_widget(buttons_layout)
        popup.content = layout
        popup.open()
    
    def show_isolation_request_details(self, request_id, parent_popup):
        """Show details of an isolation request with edit/delete options"""
        from datetime import datetime
        
        c = self.conn.cursor()
        c.execute("""
            SELECT ir.id, ir.substation_id, s.name, ir.start_datetime, ir.end_datetime,
                   ir.status, ir.notes, ir.created_at, ir.updated_at
            FROM isolation_requests ir
            JOIN substations s ON ir.substation_id = s.id
            WHERE ir.id = ?
        """, (request_id,))
        request = c.fetchone()
        
        if not request:
            show_message_popup('Σφάλμα', 'Η αίτηση δεν βρέθηκε!')
            return
        
        req_id, sub_id, sub_name, start_dt, end_dt, status, notes, created_at, updated_at = request
        
        popup = Popup(title=f'Αίτηση Απομόνωσης - {sub_name}', size_hint=(0.7, 0.8))
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        scroll = ScrollView()
        details_layout = BoxLayout(orientation='vertical', size_hint_y=None, spacing=10, padding=5)
        details_layout.bind(minimum_height=details_layout.setter('height'))
        
        # Substation
        details_layout.add_widget(Label(text='Υποσταθμός:', size_hint_y=None, height=30, bold=True))
        substation_label = Label(text=sub_name, size_hint_y=None, height=30)
        details_layout.add_widget(substation_label)
        
        # Start datetime
        details_layout.add_widget(Label(text='Έναρξη:', size_hint_y=None, height=30, bold=True))
        start_input = TextInput(
            text=start_dt,
            size_hint_y=None,
            height=35,
            multiline=False
        )
        details_layout.add_widget(start_input)
        
        # End datetime
        details_layout.add_widget(Label(text='Λήξη:', size_hint_y=None, height=30, bold=True))
        end_input = TextInput(
            text=end_dt,
            size_hint_y=None,
            height=35,
            multiline=False
        )
        details_layout.add_widget(end_input)
        
        # Status
        details_layout.add_widget(Label(text='Κατάσταση:', size_hint_y=None, height=30, bold=True))
        status_spinner = Spinner(
            text=status,
            values=['Requested', 'Accepted', 'Cancelled'],
            size_hint_y=None,
            height=40
        )
        details_layout.add_widget(status_spinner)
        
        # Notes
        details_layout.add_widget(Label(text='Σημειώσεις:', size_hint_y=None, height=30, bold=True))
        notes_input = TextInput(
            text=notes or '',
            size_hint_y=None,
            height=80,
            multiline=True
        )
        details_layout.add_widget(notes_input)
        
        # Metadata
        details_layout.add_widget(Label(
            text=f'Δημιουργήθηκε: {created_at}',
            size_hint_y=None,
            height=25,
            font_size='11sp'
        ))
        details_layout.add_widget(Label(
            text=f'Τελευταία ενημέρωση: {updated_at}',
            size_hint_y=None,
            height=25,
            font_size='11sp'
        ))
        
        scroll.add_widget(details_layout)
        layout.add_widget(scroll)
        
        # Buttons
        buttons_layout = BoxLayout(size_hint_y=None, height=50, spacing=10)
        
        def update_request():
            start_new = start_input.text.strip()
            end_new = end_input.text.strip()
            status_new = status_spinner.text
            notes_new = notes_input.text.strip()
            
            # Validate dates
            try:
                start = datetime.strptime(start_new, '%Y-%m-%d %H:%M')
                end = datetime.strptime(end_new, '%Y-%m-%d %H:%M')
                
                if end <= start:
                    show_message_popup('Σφάλμα', 'Η ημερομηνία λήξης πρέπει να είναι μετά την έναρξη!')
                    return
            except ValueError:
                show_message_popup('Σφάλμα', 'Μη έγκυρη μορφή ημερομηνίας!')
                return
            
            # Update request
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            c.execute("""
                UPDATE isolation_requests
                SET start_datetime=?, end_datetime=?, status=?, notes=?, updated_at=?
                WHERE id=?
            """, (start_new, end_new, status_new, notes_new, now, req_id))
            self.conn.commit()
            
            popup.dismiss()
            parent_popup.dismiss()
            show_message_popup('Επιτυχία', 'Η αίτηση ενημερώθηκε!',
                             callback=lambda: self.show_isolation_requests(None))
        
        def delete_request():
            # Confirmation
            confirm_popup = Popup(title='Επιβεβαίωση', size_hint=(0.5, 0.3))
            confirm_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
            
            confirm_layout.add_widget(Label(text='Είστε σίγουροι ότι θέλετε να διαγράψετε\nαυτήν την αίτηση απομόνωσης;'))
            
            confirm_buttons = BoxLayout(size_hint_y=0.3, spacing=10)
            
            def do_delete():
                c.execute("DELETE FROM isolation_requests WHERE id=?", (req_id,))
                self.conn.commit()
                confirm_popup.dismiss()
                popup.dismiss()
                parent_popup.dismiss()
                show_message_popup('Επιτυχία', 'Η αίτηση διαγράφηκε!',
                                 callback=lambda: self.show_isolation_requests(None))
            
            yes_btn = Button(text='Ναι')
            yes_btn.bind(on_press=lambda x: do_delete())
            confirm_buttons.add_widget(yes_btn)
            
            no_btn = Button(text='Όχι')
            no_btn.bind(on_press=confirm_popup.dismiss)
            confirm_buttons.add_widget(no_btn)
            
            confirm_layout.add_widget(confirm_buttons)
            confirm_popup.content = confirm_layout
            confirm_popup.open()
        
        update_btn = Button(text='Ενημέρωση')
        update_btn.bind(on_press=lambda x: update_request())
        buttons_layout.add_widget(update_btn)
        
        delete_btn = Button(text='Διαγραφή', background_color=(0.8, 0.2, 0.2, 1))
        delete_btn.bind(on_press=lambda x: delete_request())
        buttons_layout.add_widget(delete_btn)
        
        close_btn = Button(text='Κλείσιμο')
        close_btn.bind(on_press=popup.dismiss)
        buttons_layout.add_widget(close_btn)
        
        layout.add_widget(buttons_layout)
        popup.content = layout
        popup.open()
    
    def show_models_management(self, instance):
        """Show model management interface"""
        show_models_management(self)

SubstationApp().run()